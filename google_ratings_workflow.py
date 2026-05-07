"""
google_ratings_workflow.py
──────────────────────────
CI-compatible Google ratings scraper designed for GitHub Actions.

1. Scans artifacts/ for every batch_*.xlsx file downloaded from the
   dental_scraper workflow (supports all 60 batches / 6000 rows).
2. Reads every unique practice website across all batch files.
3. Searches Google Maps (headless Playwright or Places API) for
   star rating + review count.
4. Writes merged output: google_ratings_output.xlsx

Run locally:
    python google_ratings_workflow.py

Run via GitHub Actions (CI=true set automatically):
    Headless Playwright, no CAPTCHA input() pause, shorter delays.

Environment variables:
    CI=true                  headless mode (auto-set by GitHub Actions)
    GOOGLE_PLACES_API_KEY    optional; uses Places API first if set
"""

import os
import re
import glob
import time
import random
import logging
import warnings
from urllib.parse import quote_plus

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", message="Unverified HTTPS request")

# Import all extraction / search helpers from the main ratings script
import google_ratings_v12 as _g

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ── Config ────────────────────────────────────────────────────────────────────
IS_CI        = os.environ.get("CI", "").lower() in ("true", "1")
ARTIFACT_DIR = "artifacts"
OUTPUT_FILE  = "google_ratings_output.xlsx"

GOOGLE_PLACES_API_KEY = os.environ.get("GOOGLE_PLACES_API_KEY", "") or _g.GOOGLE_PLACES_API_KEY

# Override module-level constants so imported functions pick up CI values
_g.GOOGLE_PLACES_API_KEY = GOOGLE_PLACES_API_KEY
_g.DELAY_MIN  = 3.0  if IS_CI else 6.0
_g.DELAY_MAX  = 7.0  if IS_CI else 13.0
_g.BATCH_PAUSE = 20  if IS_CI else 45

# In CI: replace _wait_captcha so it never calls input()
if IS_CI:
    def _ci_captcha_noop(pw_page):
        log.warning("  ⚠  CAPTCHA detected in CI — skipping (mark for manual retry)")
    _g._wait_captcha = _ci_captcha_noop

# ── Column indices (same as 100data.xlsx / dental_scraper output) ─────────────
C_INDEX    = 1
C_PRACTICE = 2
C_DOCTOR   = 3
C_ADDRESS  = 4
C_CITY     = 5
C_STATE    = 6
C_ZIP      = 7
C_WEBSITE  = 8
C_GOOGLE_R = 42
C_GOOGLE_N = 43
HDR_ROW    = 2
DATA_START = 3

PLACEHOLDER = {"Not Found", "See Website", "See Profile", "Blocked", "ERROR", "", "N/A"}


# ── Discover batch files ───────────────────────────────────────────────────────

def find_batch_files() -> list:
    """
    Find all batch xlsx files.
    Searches artifacts/**/*.xlsx first, then falls back to batch_*.xlsx
    in the current directory (for local testing).
    """
    files = sorted(glob.glob(os.path.join(ARTIFACT_DIR, "**", "*.xlsx"), recursive=True))
    if not files:
        files = sorted(glob.glob("batch_*.xlsx"))
    log.info(f"Found {len(files)} batch file(s):")
    for f in files:
        log.info(f"  {f}")
    return files


# ── Read practices from all files ─────────────────────────────────────────────

def read_all_practices(files: list) -> dict:
    """
    Read every row from every batch file.
    Deduplicates by website domain — each unique domain is scraped once
    and the rating is written back to every row that shares that domain.
    Returns: {domain_key: practice_dict}
    """
    practices = {}
    total_rows = 0

    for fpath in files:
        try:
            wb   = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            ws   = wb.active
            file_rows = 0
            for r in range(DATA_START, (ws.max_row or 0) + 1):
                if ws.cell(r, C_INDEX).value is None:
                    continue
                website = str(ws.cell(r, C_WEBSITE).value or "").strip()
                if not website or website in ("None", "nan"):
                    continue

                domain = _g.get_domain(website)
                key    = domain or website

                row_snapshot = {
                    "file":     fpath,
                    "row":      r,
                    "index":    ws.cell(r, C_INDEX).value,
                    "practice": str(ws.cell(r, C_PRACTICE).value or "").strip(),
                    "doctor":   str(ws.cell(r, C_DOCTOR).value   or "").strip(),
                    "address":  str(ws.cell(r, C_ADDRESS).value  or "").strip(),
                    "city":     str(ws.cell(r, C_CITY).value     or "").strip(),
                    "state":    str(ws.cell(r, C_STATE).value    or "").strip(),
                    "zip":      str(ws.cell(r, C_ZIP).value      or "").strip(),
                    "website":  website,
                    "google_r": str(ws.cell(r, C_GOOGLE_R).value or "").strip(),
                    "google_n": str(ws.cell(r, C_GOOGLE_N).value or "").strip(),
                }

                if key not in practices:
                    practices[key] = {
                        **row_snapshot,
                        "domain": domain,
                        "rows":   [row_snapshot],
                    }
                else:
                    practices[key]["rows"].append(row_snapshot)

                file_rows += 1

            wb.close()
            total_rows += file_rows
            log.info(f"  {os.path.basename(fpath)}: {file_rows} rows")

        except Exception as e:
            log.warning(f"  Could not read {fpath}: {e}")

    log.info(f"Total: {total_rows} rows | {len(practices)} unique websites")
    return practices


# ── Playwright launch ─────────────────────────────────────────────────────────

def launch_playwright():
    """
    CI mode  → headless Chromium with no-sandbox flags (required on Linux/Actions).
    Local    → visible persistent browser (same as google_ratings_v12.py).
    Returns (_pw, context_or_browser, page).
    """
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        log.error("Playwright not installed.  Run: pip install playwright && playwright install chromium")
        return None, None, None

    _pw = sync_playwright().__enter__()

    if IS_CI:
        log.info("Launching headless Playwright (CI / GitHub Actions mode)…")
        browser = _pw.chromium.launch(
            headless=True,
            args=[
                "--no-sandbox",
                "--disable-setuid-sandbox",
                "--disable-dev-shm-usage",
                "--disable-gpu",
                "--disable-blink-features=AutomationControlled",
                "--lang=en-US",
            ],
        )
        ctx = browser.new_context(
            locale="en-US",
            timezone_id="America/New_York",
            geolocation={"latitude": 40.7128, "longitude": -74.0060},
            permissions=["geolocation"],
            user_agent=random.choice(_g._USER_AGENTS),
        )
        _g.apply_stealth(ctx)
        page = ctx.new_page()
        page.set_extra_http_headers({"Accept-Language": "en-US,en;q=0.9"})
        return _pw, ctx, page

    else:
        log.info("Launching visible Playwright (local mode)…")
        pu = "/tmp/pw_google_profile"
        os.makedirs(pu, exist_ok=True)
        ctx = _pw.chromium.launch_persistent_context(
            user_data_dir=pu,
            headless=False,
            slow_mo=150,
            ignore_https_errors=True,
            locale="en-US",
            timezone_id="America/New_York",
            geolocation={"latitude": 40.7128, "longitude": -74.0060},
            permissions=["geolocation"],
            args=[
                "--disable-blink-features=AutomationControlled",
                "--lang=en-US",
            ],
            user_agent=_g._USER_AGENTS[0],
        )
        _g.apply_stealth(ctx)
        page = ctx.new_page()
        page.set_extra_http_headers({"Accept-Language": "en-US,en;q=0.9"})
        _g._prewarm_browser(page)
        return _pw, ctx, page


# ── Output writer ─────────────────────────────────────────────────────────────

def write_output(practices: dict, results: dict):
    wb = Workbook()
    ws = wb.active
    ws.title = "Google Ratings"

    hdr_font  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    ctr       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft       = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    data_font = Font(name="Arial", size=10)
    alt_fill  = PatternFill("solid", fgColor="DCE6F1")
    found_fill = PatternFill("solid", fgColor="E2EFDA")   # light green for found ratings

    headers = [
        "#", "Practice Name", "Doctor Name", "Address",
        "City", "State", "Zip", "Website",
        "Google Reviews Ranking", "Total # of Google Reviews",
        "Source Batch File",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = ctr

    # Sort by original index so output order matches input
    def _sort_key(item):
        try:
            return int(float(str(item[1].get("index") or 0)))
        except Exception:
            return 0

    r_idx = 2
    for key, p in sorted(practices.items(), key=_sort_key):
        rating, count = results.get(key, ("Not Found", ""))
        row_fill_base = found_fill if rating not in ("Not Found", "") else None

        for row_data in p["rows"]:
            rf = row_fill_base or (alt_fill if r_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF"))
            vals = [
                row_data["index"],
                row_data["practice"],
                row_data["doctor"],
                row_data["address"],
                row_data["city"],
                row_data["state"],
                row_data["zip"],
                row_data["website"],
                rating,
                count,
                os.path.basename(row_data["file"]),
            ]
            for c_idx, val in enumerate(vals, 1):
                cell           = ws.cell(r_idx, c_idx, val)
                cell.font      = data_font
                cell.fill      = rf
                cell.alignment = lft if c_idx in (2, 3, 4, 8, 11) else ctr
            r_idx += 1

    col_widths = {1:6, 2:28, 3:26, 4:28, 5:14, 6:7, 7:8, 8:34, 9:16, 10:18, 11:30}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    wb.save(OUTPUT_FILE)
    found = sum(1 for r, c in results.values() if r not in ("Not Found", ""))
    log.info(f"  Saved → {OUTPUT_FILE}  ({r_idx - 2} rows | {found}/{len(results)} rated)")


# ── Main ──────────────────────────────────────────────────────────────────────

def run():
    log.info("=" * 60)
    log.info(f"  Google Ratings Workflow  ({'CI headless' if IS_CI else 'local visible'})")
    log.info(f"  Places API: {'enabled' if GOOGLE_PLACES_API_KEY else 'disabled (no key)'}")
    log.info("=" * 60)

    # ── 1. Discover and read all batch files ──────────────────────────────────
    files = find_batch_files()
    if not files:
        log.error(
            f"No batch xlsx files found in '{ARTIFACT_DIR}/' or current directory.\n"
            "Make sure artifacts are downloaded before running this script."
        )
        return

    practices = read_all_practices(files)
    if not practices:
        log.error("No practices with websites found in batch files.")
        return

    # ── 2. Launch Playwright ──────────────────────────────────────────────────
    _pw, ctx, pw_page = launch_playwright()
    if pw_page is None:
        return

    results = {}
    total   = len(practices)

    try:
        for i, (key, p) in enumerate(practices.items(), 1):
            log.info(f"\n[{i}/{total}]  {p['practice']}  |  {p['domain']}")

            # Skip if ALL rows for this practice already have a valid rating
            existing_r = p.get("google_r", "")
            if existing_r not in PLACEHOLDER:
                log.info(f"  Already rated: {existing_r} ★  ({p.get('google_n', '')})")
                results[key] = (existing_r, p.get("google_n", ""))
                continue

            rating, count = "", ""

            # ── Method 1: Google Places API by domain ─────────────────────────
            if GOOGLE_PLACES_API_KEY and p["domain"]:
                try:
                    rating, count = _g.places_by_website(p["domain"], p["city"], p["state"])
                    if rating:
                        log.info(f"  ✓ Places/domain: {rating} ★  ({count})")
                except Exception as e:
                    log.debug(f"  Places/domain error: {e}")

            # ── Method 2: Google Places API by name ───────────────────────────
            if not rating and GOOGLE_PLACES_API_KEY:
                try:
                    rating, count = _g.places_by_name(p["practice"], p["city"], p["state"], p["zip"])
                    if rating:
                        log.info(f"  ✓ Places/name: {rating} ★  ({count})")
                except Exception as e:
                    log.debug(f"  Places/name error: {e}")

            # ── Method 3: Google Maps via Playwright by domain ────────────────
            if not rating and p["domain"]:
                try:
                    rating, count = _g.search_maps_by_website(
                        p["domain"], p["practice"], p["city"], p["state"], pw_page
                    )
                    if rating:
                        log.info(f"  ✓ Maps/domain: {rating} ★  ({count})")
                except Exception as e:
                    log.debug(f"  Maps/domain error: {e}")

            # ── Method 4: Google Maps via Playwright by name ──────────────────
            if not rating:
                try:
                    rating, count = _g.search_maps_by_name(
                        p["practice"], p["doctor"],
                        p["address"], p["city"], p["state"], p["zip"],
                        pw_page,
                    )
                    if rating:
                        log.info(f"  ✓ Maps/name: {rating} ★  ({count})")
                except Exception as e:
                    log.debug(f"  Maps/name error: {e}")

            if not rating:
                log.info("  ✗ Not found after all methods")

            results[key] = (rating or "Not Found", count or "")

            # Save progress every 10 practices
            if i % 10 == 0:
                write_output(practices, results)
                log.info(f"  💾 Progress saved ({i}/{total})")
                # Batch pause (longer locally, shorter in CI)
                if i < total:
                    pause = _g.BATCH_PAUSE
                    log.info(f"  ⏸  Batch pause {pause}s…")
                    time.sleep(pause)

    finally:
        try:
            ctx.close()
            _pw.__exit__(None, None, None)
        except Exception:
            pass

    write_output(practices, results)
    found = sum(1 for r, c in results.values() if r not in ("Not Found", ""))
    log.info(f"\n🎉 Done — {found}/{total} ratings found  →  {OUTPUT_FILE}")


if __name__ == "__main__":
    run()
