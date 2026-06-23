"""
refresh_doctors.py
==================
Re-extracts Doctor Name, Doctor Specialty, and Associations/Memberships
for each practice using the page_cache/ from a previous scraper run.
Uses the latest dental_scraper code — picks up all recent fixes including:
  • Plain-name secondary pass (catches "Colton Crane" style cards)
  • Expanded association map (AARD, ACD, ICD, HDA, WCLI, etc.)
  • University affiliations (OSU, CWRU, NYU, etc.) with edu-context guard
  • Bio page name upgrade (Dr. + middle initial + credentials)

No full website re-crawl needed.  Only individual bio URLs are fetched
live (one small GET per doctor), everything else uses cached HTML.

Outputs:
    <input>_doctors_refreshed.xlsx   — updated xlsx
    <input>_doctors_comparison.xlsx  — before/after report (changed practices)

Usage:
    python3 refresh_doctors.py batch_7_deduped.xlsx
    python3 refresh_doctors.py batch_7_deduped.xlsx --cache-dir page_cache
    python3 refresh_doctors.py input.xlsx --cache-dir page_cache --output out.xlsx
"""

import argparse
import glob
import json
import logging
import os
import re
import shutil
import sys
import time

import openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import dental_scraper as ds

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

HERE        = os.path.dirname(os.path.abspath(__file__))
CACHE_DIR   = os.path.join(HERE, "page_cache")
DATA_START  = 3   # row 1 = group label, row 2 = headers, row 3+ = data

# Column positions (1-based) — match dental_scraper.write_output
C_IDX       = 1
C_NAME      = 2
C_DOCNAME   = 3
C_WEBSITE   = 8
C_ASSOC     = 40
C_SPECIALTY = 41


# ── Cache helpers (mirrors refresh_tech_services.py) ─────────────────────────

def _find_cache_folder(idx: int, cache_dir: str) -> str:
    """Return the cache folder path for a practice index, or empty string."""
    prefix = f"{idx:03d}_"
    try:
        for name in sorted(os.listdir(cache_dir)):
            if name.startswith(prefix) and os.path.isdir(os.path.join(cache_dir, name)):
                return os.path.join(cache_dir, name)
    except FileNotFoundError:
        pass
    return ""


def _load_pages(folder: str) -> list:
    """
    Load all cached HTML pages for a practice.
    Returns list of (page_type, url, html_text).
    """
    url_map = {}
    manifest_path = os.path.join(folder, "manifest.json")
    if os.path.exists(manifest_path):
        try:
            with open(manifest_path, encoding="utf-8") as f:
                m = json.load(f)
            for ptype, info in m.get("pages", {}).items():
                url_map[info.get("file", "")] = (ptype, info.get("url", ""))
        except Exception:
            pass

    pages = []
    seen  = set()
    for fpath in sorted(glob.glob(os.path.join(folder, "*.html"))):
        fname = os.path.basename(fpath)
        if fname in seen:
            continue
        seen.add(fname)
        try:
            with open(fpath, encoding="utf-8", errors="replace") as f:
                html = f.read()
        except Exception:
            continue
        if len(html) < 200:
            continue
        ptype, url = url_map.get(fname, (fname.replace(".html", ""), ""))
        pages.append((ptype, url, html))
    return pages


# ── Doctor extraction from cached pages ──────────────────────────────────────

def _extract_doctors(pages: list, website_url: str):
    """
    Re-run scrape_doctors_full() on cached pages + live bio GETs.
    Returns (doctors_list, hygienist_count).
    doctors_list is a list of {"name": str, "specialty": str, "associations": str}.
    """
    if not pages:
        return [], None

    homepage_soup = None
    all_text      = ""
    all_soups     = []

    for ptype, url, html in pages:
        soup = BeautifulSoup(html, "lxml")
        if ptype in ("homepage", "pw_homepage") and homepage_soup is None:
            homepage_soup = soup
            all_text = soup.get_text(separator=" ", strip=True)[:50_000]
        all_soups.append((ptype, soup))

    if not homepage_soup and all_soups:
        homepage_soup = all_soups[0][1]
        all_text = all_soups[0][1].get_text(separator=" ", strip=True)[:50_000]

    if not homepage_soup:
        return [], None

    try:
        doctors, hyg = ds.scrape_doctors_full(
            homepage_soup    = homepage_soup,
            base_url         = website_url or "",
            all_text         = all_text,
            pw_page          = None,        # no browser in refresh mode
            all_soups_for_team = all_soups,
        )
        return doctors or [], hyg
    except Exception as e:
        log.warning(f"  Doctor extraction error: {e}")
        return [], None


# ── Main ──────────────────────────────────────────────────────────────────────

def refresh(input_xlsx: str, cache_dir: str = CACHE_DIR, output: str = ""):
    if not output:
        base, ext = os.path.splitext(input_xlsx)
        output = base + "_doctors_refreshed" + ext
    comp_path = os.path.splitext(output)[0] + "_comparison.xlsx"

    shutil.copy2(input_xlsx, output)
    log.info("Input  : %s", input_xlsx)
    log.info("Output : %s", output)
    log.info("Cache  : %s", cache_dir)

    if not os.path.isdir(cache_dir):
        log.warning("Cache dir not found: %s  (bio pages will still be fetched live)", cache_dir)

    # ── Read existing xlsx ────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(output, data_only=True)
    ws = wb.active

    # Detect header row (might be row 1 or row 2 depending on sheet)
    data_start = DATA_START
    for r in (1, 2):
        for c in range(1, min(ws.max_column + 1, 50)):
            v = ws.cell(r, c).value
            if v and "doctor name" in str(v).lower():
                data_start = r + 1
                break

    # Group rows by practice index
    idx_rows:    dict = {}  # idx(int) → [row_num, ...]
    idx_name:    dict = {}  # idx → practice name string
    idx_website: dict = {}  # idx → website url

    for row in range(data_start, ws.max_row + 1):
        raw = ws.cell(row, C_IDX).value
        if raw is None:
            continue
        try:
            idx = int(str(raw).strip())
        except (TypeError, ValueError):
            continue
        idx_rows.setdefault(idx, []).append(row)
        if idx not in idx_name:
            idx_name[idx]    = str(ws.cell(row, C_NAME).value or "")
            w_raw            = str(ws.cell(row, C_WEBSITE).value or "").strip()
            idx_website[idx] = "" if w_raw in ("None", "nan", "not found", "n/a", "") else w_raw

    log.info("Practices in xlsx: %d", len(idx_rows))

    # Snapshot before-values for comparison report
    idx_before: dict = {}
    for idx, rows in idx_rows.items():
        names  = [ws.cell(r, C_DOCNAME).value   for r in rows]
        assocs = [ws.cell(r, C_ASSOC).value      for r in rows]
        specs  = [ws.cell(r, C_SPECIALTY).value  for r in rows]
        idx_before[idx] = {"names": names, "assocs": assocs, "specs": specs}

    wb.close()

    # ── Re-extract doctors from cache ─────────────────────────────────────────
    new_doctor_data: dict = {}  # idx → list of {"name", "specialty", "associations"}
    n_cache = n_skip = n_no_change = 0

    for i, idx in enumerate(sorted(idx_rows.keys()), 1):
        name    = idx_name.get(idx, "")
        website = idx_website.get(idx, "")

        log.info("[%d/%d]  idx=%03d  %s", i, len(idx_rows), idx, name[:50])

        folder = _find_cache_folder(idx, cache_dir)
        if not folder:
            log.info("  No cache folder — skipping")
            n_skip += 1
            continue

        pages = _load_pages(folder)
        if not pages:
            log.info("  Cache empty — skipping")
            n_skip += 1
            continue

        log.info("  Cache: %d pages  website: %s", len(pages), website[:55])
        doctors, _hyg = _extract_doctors(pages, website)

        if not doctors:
            log.info("  No doctors extracted — keeping existing data")
            n_no_change += 1
            continue

        n_cache += 1
        new_doctor_data[idx] = doctors
        log.info("  → %d doctor(s): %s",
                 len(doctors), ", ".join(d["name"] for d in doctors))

    log.info("─" * 60)
    log.info("Cache hits: %d  Skipped: %d  Unchanged: %d  Total: %d",
             n_cache, n_skip, n_no_change, len(idx_rows))

    if not new_doctor_data:
        log.warning("No updated doctor data extracted — output unchanged.")
        return output, None

    # ── Rewrite xlsx ──────────────────────────────────────────────────────────
    # Strategy: read ALL data rows into memory, rebuild with updated doctor rows,
    # rewrite the data section.  Header rows (1..data_start-1) are untouched.
    wb2 = openpyxl.load_workbook(output, data_only=True)
    ws2 = wb2.active

    # Read all data rows into memory
    all_data: list = []  # list of (idx, row_values_list)
    for row in range(data_start, ws2.max_row + 1):
        raw = ws2.cell(row, C_IDX).value
        if raw is None:
            continue
        try:
            idx = int(str(raw).strip())
        except (TypeError, ValueError):
            continue
        row_vals = [ws2.cell(row, c).value for c in range(1, ws2.max_column + 1)]
        all_data.append((idx, row_vals))

    # Group consecutive rows by practice
    from itertools import groupby
    groups = [(k, [v for _, v in grp])
              for k, grp in groupby(all_data, key=lambda x: x[0])]

    # Clear data section
    for row in range(data_start, ws2.max_row + 2):
        for col in range(1, ws2.max_column + 1):
            ws2.cell(row, col).value = None

    # Rewrite with updated doctor data
    cur_row        = data_start
    rows_added     = 0
    practices_updated = 0

    for idx, existing_rows in groups:
        template = existing_rows[0]   # use first row for practice-level columns

        if idx in new_doctor_data:
            new_docs = new_doctor_data[idx]
            practices_updated += 1
            if len(new_docs) > len(existing_rows):
                rows_added += len(new_docs) - len(existing_rows)

            for di, doc in enumerate(new_docs):
                row_vals = list(template)
                row_vals[C_DOCNAME   - 1] = doc.get("name")         or "Not Found"
                row_vals[C_ASSOC     - 1] = doc.get("associations") or "Not Found"
                row_vals[C_SPECIALTY - 1] = doc.get("specialty")    or "Not Found"
                for col, val in enumerate(row_vals, 1):
                    ws2.cell(cur_row, col).value = val
                cur_row += 1
        else:
            # No new data — keep all existing rows unchanged
            for row_vals in existing_rows:
                for col, val in enumerate(row_vals, 1):
                    ws2.cell(cur_row, col).value = val
                cur_row += 1

    wb2.save(output)
    log.info("Practices updated  : %d", practices_updated)
    log.info("Extra rows added   : %d", rows_added)
    log.info("Saved → %s", output)

    # ── Comparison report ─────────────────────────────────────────────────────
    _write_comparison(comp_path, idx_name, idx_before, new_doctor_data)
    return output, comp_path


def _write_comparison(path: str, idx_name: dict, before: dict, after: dict):
    """Write a before/after comparison xlsx showing changed practices."""
    wb   = openpyxl.Workbook()
    ws   = wb.active
    ws.title = "Doctor Data Comparison"

    YELLOW = PatternFill("solid", fgColor="FFFACD")
    GREEN  = PatternFill("solid", fgColor="C6EFCE")
    GREY   = PatternFill("solid", fgColor="E8E8E8")
    HDR    = Font(bold=True, size=9)
    DATA   = Font(size=9)
    ctr    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft    = Alignment(horizontal="left",   vertical="center")

    headers = ["Index", "Practice Name",
               "Old Doctor Names", "New Doctor Names",
               "Old Specialty",    "New Specialty",
               "Old Associations", "New Associations",
               "Doctors: Before → After"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c)
        cell.value = h
        cell.font  = HDR
        cell.fill  = GREY
        cell.alignment = ctr

    row = 2
    changed_count = 0
    for idx in sorted(after.keys()):
        b   = before.get(idx, {})
        new = after[idx]

        old_names  = " | ".join(str(n or "") for n in b.get("names", []))
        new_names  = " | ".join(d["name"]                       for d in new)
        old_spec   = " | ".join(str(s or "") for s in b.get("specs", []))
        new_spec   = " | ".join(d.get("specialty", "")          for d in new)
        old_assoc  = " | ".join(str(a or "") for a in b.get("assocs", []))
        new_assoc  = " | ".join(d.get("associations", "")       for d in new)
        summary    = f"{len(b.get('names', []))} → {len(new)}"

        changed = (old_names.strip() != new_names.strip()
                   or old_assoc.strip() != new_assoc.strip()
                   or old_spec.strip() != new_spec.strip())
        fill = GREEN if changed else None
        changed_count += 1 if changed else 0

        vals = [idx, idx_name.get(idx, ""),
                old_names, new_names,
                old_spec,  new_spec,
                old_assoc, new_assoc,
                summary]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row, c)
            cell.value     = v
            cell.font      = DATA
            cell.alignment = lft if c in (2, 3, 4, 5, 6, 7, 8) else ctr
            if changed and fill:
                cell.fill = fill
        row += 1

    col_widths = [8, 30, 35, 35, 30, 30, 35, 35, 16]
    for c, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"

    wb.save(path)
    log.info("Comparison report  : %d changed practices → %s", changed_count, path)


# ── CLI entry point ───────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(
        description="Re-extract doctor data from page_cache into an existing batch xlsx."
    )
    p.add_argument("xlsx",        help="Input xlsx file (batch deduped/rated)")
    p.add_argument("--cache-dir", default=CACHE_DIR, help="Path to page_cache directory")
    p.add_argument("--output",    default="",         help="Output xlsx path")
    args = p.parse_args()

    if not os.path.exists(args.xlsx):
        log.error("File not found: %s", args.xlsx)
        sys.exit(1)

    refresh(args.xlsx, cache_dir=args.cache_dir, output=args.output)


if __name__ == "__main__":
    main()
