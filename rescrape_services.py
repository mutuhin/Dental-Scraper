#!/usr/bin/env python3
"""
rescrape_services.py
====================
Live re-scrapes Technology in Practice, Services (# of Mentions), and
Testimonials for every practice in a batch xlsx file, then writes the
results back.

Uses the same keyword lists and extraction logic as dental_scraper.py.

Usage:
    python3 rescrape_services.py <input.xlsx> [output.xlsx]

    If output.xlsx is omitted, writes to <input>_rescraped.xlsx.
"""

import os, sys, re, time, logging
from urllib.parse import urljoin, urlparse
from collections import deque

import requests
from bs4 import BeautifulSoup
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import dental_scraper as ds

logging.basicConfig(level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger(__name__)

MAX_PAGES   = 50
REQ_TIMEOUT = 12
DELAY       = 0.5

SKIP_EXTS = frozenset({
    ".pdf", ".jpg", ".jpeg", ".png", ".gif", ".svg", ".webp",
    ".mp4", ".mp3", ".zip", ".doc", ".docx", ".xls", ".xlsx", ".ico",
})

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

# ── Column positions (1-based) ────────────────────────────────────────────────
C_IDX   = 1;  C_WEBSITE = 8
C_CEREC = 23; C_CBCT    = 24; C_LASER = 25; C_AI    = 26; C_INTRA = 27
C_INV   = 28  # col 29 = InvTier (skip)
C_CLEAR = 30; C_VEN     = 31; C_IMPL  = 32; C_SMILE = 33; C_WHITE = 34
C_SED   = 35; C_HOL     = 36; C_PLAN  = 37; C_CANC  = 38
C_TESTI = 46
DATA_START = 3

_TEST_RE = re.compile(
    r"(testimonial|review|quote|patient.story|patient.review|"
    r"feedback|client.say|what.people|slider|swiper|slick|"
    r"rating.block|star.review)", re.I
)


# ── Link collector ────────────────────────────────────────────────────────────

def _collect_links(html: str, base_url: str, domain: str) -> list:
    soup = BeautifulSoup(html, "lxml")
    seen, links = set(), []

    def _add(href):
        if not href or href.startswith(("#", "mailto:", "tel:", "javascript:")):
            return
        full = urljoin(base_url, href).split("#")[0]
        if not full.startswith("http"):
            return
        if urlparse(full).netloc.lstrip("www.") != domain.lstrip("www."):
            return
        if any(full.lower().endswith(e) for e in SKIP_EXTS):
            return
        if full in seen:
            return
        seen.add(full)
        links.append(full)

    for container in (soup.find_all(["nav", "header"]) +
                      soup.find_all(class_=re.compile(
                          r'\b(menu|navigation|navbar|nav|megamenu|primary.?nav)\b', re.I
                      ))):
        for a in container.find_all("a", href=True):
            _add(a["href"])
    for a in soup.find_all("a", href=True):
        _add(a["href"])
    return links


# ── Per-site scraper ──────────────────────────────────────────────────────────

def scrape_site(start_url: str) -> dict:
    """
    BFS-crawl the site. Returns dict with tech, service counts, and testimonials.
    Uses ds.SERVICE_KEYWORDS, ds.TECH_KEYWORDS, ds.count_keyword_capped.
    """
    parsed = urlparse(start_url)
    domain = parsed.netloc
    sess   = requests.Session()
    sess.headers.update(HEADERS)

    queue   = deque([start_url.split("#")[0]])
    visited = set()
    pages   = []   # list of (body_text, full_text, soup)

    while queue and len(visited) < MAX_PAGES:
        url = queue.popleft()
        if url in visited:
            continue
        visited.add(url)
        try:
            r = sess.get(url, timeout=REQ_TIMEOUT, allow_redirects=True)
            if r.status_code != 200:
                continue
            html = r.text
        except Exception:
            continue

        bt   = ds.extract_body_text(html)
        ft   = ds.extract_text(html)
        soup = BeautifulSoup(html, "lxml")
        pages.append((bt, ft, soup))

        for link in _collect_links(html, url, domain):
            if link not in visited:
                queue.append(link)
        time.sleep(DELAY)

    log.info("    Crawled %d pages", len(visited))

    # ── Tech detection ────────────────────────────────────────────────────
    all_text = " ".join(ft for _, ft, _ in pages)
    tf = set()
    for kw, tn in ds.TECH_KEYWORDS.items():
        if kw in all_text:
            tf.add(tn)
    if "AI" not in tf and re.search(r"\bai\b", all_text):
        tf.add("AI")

    # ── Service counts (body-text primary, full-text fallback) ────────────
    svc_b = {k: 0 for k in ds.SERVICE_KEYWORDS.values()}
    svc_f = {k: 0 for k in ds.SERVICE_KEYWORDS.values()}
    svc_b = dict.fromkeys(set(ds.SERVICE_KEYWORDS.values()), 0)
    svc_f = dict.fromkeys(set(ds.SERVICE_KEYWORDS.values()), 0)
    for bt, ft, _ in pages:
        for kw, cat in ds.SERVICE_KEYWORDS.items():
            svc_b[cat] += ds.count_keyword_capped(bt, kw, cap=5)
            svc_f[cat] += ds.count_keyword_capped(ft, kw, cap=3)
    svc = {cat: (svc_b[cat] if svc_b[cat] > 0 else svc_f[cat])
           for cat in svc_b}

    # ── Testimonials ──────────────────────────────────────────────────────
    seen_t, tt = set(), 0
    for _, _, soup in pages:
        for blk in soup.find_all(["div", "section", "article", "blockquote", "li"],
                                  class_=_TEST_RE):
            k = blk.get_text(separator=" ", strip=True)[:80]
            if k and k not in seen_t:
                seen_t.add(k); tt += 1
    if tt == 0:
        for _, _, soup in pages:
            for bq in soup.find_all("blockquote"):
                k = bq.get_text(separator=" ", strip=True)[:80]
                if k and k not in seen_t:
                    seen_t.add(k); tt += 1

    return {
        "cerec":           "X" if "CEREC"              in tf else "",
        "cbct":            "X" if "CBCT"               in tf else "",
        "lasers":          "X" if "Lasers"             in tf else "",
        "ai":              "X" if "AI"                 in tf else "",
        "intraoral":       "X" if "Intraoral Scanners" in tf else "",
        "invisalign":      svc.get("Invisalign",        0),
        "clear_aligners":  svc.get("Clear Aligners",    0),
        "veneers":         svc.get("Veneers",           0),
        "implants":        svc.get("Implants",          0),
        "smile_makeovers": svc.get("Smile Makeovers",   0),
        "whitening":       svc.get("Teeth Whitening",   0),
        "sedation":        svc.get("Sedation Dentistry",0),
        "holistic":        svc.get("Holistic Dentistry",0),
        "dental_plan":     "Mentioned" if svc.get("Dental Plan", 0) > 0 else "",
        "cancer_screening":svc.get("Cancer Screening",  0),
        "testimonials":    str(tt),
    }


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    args = sys.argv[1:]
    if not args:
        print("Usage: python3 rescrape_services.py <input.xlsx> [output.xlsx]")
        sys.exit(1)

    input_file = args[0]
    if len(args) >= 2:
        output_file = args[1]
    else:
        base, ext = os.path.splitext(input_file)
        output_file = base + "_rescraped" + ext

    log.info("Input : %s", input_file)
    log.info("Output: %s", output_file)

    import shutil
    shutil.copy2(input_file, output_file)

    wb = openpyxl.load_workbook(output_file, data_only=True)
    ws = wb.active

    # Collect unique (idx → url)
    practices = {}
    for row in ws.iter_rows(min_row=DATA_START, values_only=True):
        idx = row[C_IDX - 1]
        url = row[C_WEBSITE - 1]
        if idx is None or not url:
            continue
        try:
            idx = int(idx)
        except (ValueError, TypeError):
            continue
        if idx not in practices:
            u = str(url).strip()
            practices[idx] = u if u.startswith("http") else "https://" + u

    log.info("Unique practices: %d", len(practices))

    results = {}
    for idx in sorted(practices):
        url = practices[idx]
        log.info("[%03d] %s", idx, url)
        try:
            r = scrape_site(url)
            results[idx] = r
            log.info("      CEREC=%-2s  Inv=%-3s  Impl=%-3s  White=%-3s  Testi=%s",
                     r["cerec"] or "-", r["invisalign"], r["implants"],
                     r["whitening"], r["testimonials"])
        except Exception as e:
            log.warning("      FAILED: %s — keeping original values", e)
            results[idx] = None

    # Write back
    updates = 0
    for row_num in range(DATA_START, ws.max_row + 1):
        try:
            idx = int(ws.cell(row_num, C_IDX).value)
        except (TypeError, ValueError):
            continue
        r = results.get(idx)
        if r is None:
            continue
        ws.cell(row_num, C_CEREC).value = r["cerec"]
        ws.cell(row_num, C_CBCT ).value = r["cbct"]
        ws.cell(row_num, C_LASER).value = r["lasers"]
        ws.cell(row_num, C_AI   ).value = r["ai"]
        ws.cell(row_num, C_INTRA).value = r["intraoral"]
        ws.cell(row_num, C_INV  ).value = r["invisalign"]
        ws.cell(row_num, C_CLEAR).value = r["clear_aligners"]
        ws.cell(row_num, C_VEN  ).value = r["veneers"]
        ws.cell(row_num, C_IMPL ).value = r["implants"]
        ws.cell(row_num, C_SMILE).value = r["smile_makeovers"]
        ws.cell(row_num, C_WHITE).value = r["whitening"]
        ws.cell(row_num, C_SED  ).value = r["sedation"]
        ws.cell(row_num, C_HOL  ).value = r["holistic"]
        ws.cell(row_num, C_PLAN ).value = r["dental_plan"]
        ws.cell(row_num, C_CANC ).value = r["cancer_screening"]
        ws.cell(row_num, C_TESTI).value = r["testimonials"]
        updates += 1

    wb.save(output_file)
    log.info("Wrote %d rows → %s", updates, output_file)


if __name__ == "__main__":
    main()
