"""
social_scraper.py
-----------------
Standalone social-media stat scraper for dental practice xlsx files.
Reads FB / Instagram / TikTok URLs from the batch xlsx and fills in
Posts + Followers counts for any row that still shows 'Not Found' or blank.

Usage (local):
    python social_scraper.py input_batch.xlsx [output.xlsx]

GitHub Actions: called by social_scraper.yml — pass input file as first arg.

Column layout (1-indexed, matches batch xlsx):
    11 = Facebook URL       12 = FB # Posts     13 = FB Followers
    14 = Instagram URL      15 = IG # Posts     16 = IG Followers
    17 = TikTok URL         18 = TT # Posts     19 = TT Followers
Header: row 1 = group, row 2 = sub-header. Data starts row 3.
"""

import json
import logging
import os
import random
import re
import sys
import time
from urllib.parse import urlparse

import openpyxl
from bs4 import BeautifulSoup

# ─────────────────────────────────────────────────────────────────────────────
# Logging
# ─────────────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("social")

# ─────────────────────────────────────────────────────────────────────────────
# curl_cffi
# ─────────────────────────────────────────────────────────────────────────────
try:
    import curl_cffi.requests as cffi_requests
    _CFFI_AVAILABLE = True
    log.info("curl_cffi loaded OK")
except ImportError:
    cffi_requests = None
    _CFFI_AVAILABLE = False
    log.warning("curl_cffi not available — using requests fallback (less effective)")

try:
    import requests as _req
    _REQUESTS_AVAILABLE = True
except ImportError:
    _req = None
    _REQUESTS_AVAILABLE = False

# ─────────────────────────────────────────────────────────────────────────────
# Proxy (Oxylabs residential)
# ─────────────────────────────────────────────────────────────────────────────
def _load_proxy() -> "str | None":
    user = os.environ.get("OXYLABS_USER", "").strip()
    pwd  = os.environ.get("OXYLABS_PASS", "").strip()
    if user and pwd:
        proxy = f"http://{user}:country-US,{pwd}@pr.oxylabs.io:7777"
        log.info("Oxylabs proxy loaded from env vars")
        return proxy
    # local fallback: proxies.txt
    try:
        with open("proxies.txt") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#"):
                    log.info("Proxy loaded from proxies.txt")
                    return line
    except FileNotFoundError:
        pass
    log.warning("No proxy configured — scraping without proxy (expect high block rate)")
    return None

_PROXY = _load_proxy()

def _fresh_proxy() -> "dict | None":
    """Return a rotating residential proxy dict with a fresh IP per call."""
    if not _PROXY:
        return None
    user = os.environ.get("OXYLABS_USER", "").strip()
    pwd  = os.environ.get("OXYLABS_PASS", "").strip()
    if user and pwd:
        # session ID makes each call rotate to a different IP
        sid = random.randint(10000000, 99999999)
        p = f"http://{user}-cc-US-sid-{sid}:{pwd}@pr.oxylabs.io:7777"
        return {"http": p, "https": p}
    return {"http": _PROXY, "https": _PROXY}

# ─────────────────────────────────────────────────────────────────────────────
# Common helpers
# ─────────────────────────────────────────────────────────────────────────────
_CFFI_PROFILES = ["chrome136", "chrome124", "chrome133a", "chrome110", "safari260", "safari17_2"]
_CFFI_IG_PROF  = ["chrome124", "chrome136", "safari260"]

_SOCIAL_UA_LIST = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
]

def _sleep(lo=1.0, hi=2.5):
    time.sleep(random.uniform(lo, hi))

def _compact(n: "str | int | None") -> str:
    """Format a raw number as K/M compact string if large."""
    if n is None or n == "":
        return ""
    try:
        v = int(str(n).replace(",", "").strip())
        if v >= 1_000_000:
            return f"{v/1_000_000:.1f}M"
        if v >= 1_000:
            return f"{v/1_000:.1f}K"
        return str(v)
    except ValueError:
        return str(n)

MISSING = {"", "not found", "n/a", "none", "false", "no", "blocked", "0"}

def _is_missing(v) -> bool:
    return v is None or str(v).strip().lower() in MISSING

# ─────────────────────────────────────────────────────────────────────────────
# Facebook
# ─────────────────────────────────────────────────────────────────────────────
def _parse_fb_followers(html: str) -> str:
    for pat in (
        r'"followers_count"\s*:\s*(\d+)',
        r'"fan_count"\s*:\s*(\d+)',
        r'"subscriber_count"\s*:\s*(\d+)',
        r'"like_count"\s*:\s*(\d+)',
        r'"page_likers"\s*:\s*\{"count"\s*:\s*(\d+)',
        r'"likers"\s*:\s*\{"count"\s*:\s*(\d+)',
        r'"profile_plus_follower_count"\s*:\s*(\d+)',
        r'"social_context"[^}]*?"(\d+)\s*(?:people\s+)?follow',
    ):
        m = re.search(pat, html, re.I)
        if m:
            val = m.group(1)
            if val.isdigit() and int(val) > 0:
                return _compact(val)
    m = re.search(r"([\d,]+(?:\.\d+)?[KMB]?)\s*(?:people\s+)?follow(?:ers?)?", html, re.I)
    if m:
        return m.group(1).replace(",", "")
    m = re.search(r"([\d,]+(?:\.\d+)?[KMB]?)\s*(?:people\s+)?like(?:\s+this)?", html, re.I)
    if m:
        return m.group(1).replace(",", "")
    return ""

_FB_HARD_LOGIN = (
    "log in to facebook",
    "log into facebook",
    "you must log in",
    "you must be logged in",
    '"loginrequired":true',
    "checkpoint/block",
)

def _fb_is_hard_login_wall(html: str) -> bool:
    lc = html.lower()
    return any(s in lc for s in _FB_HARD_LOGIN)

def _fb_mbasic_url(url: str) -> str:
    return re.sub(r"https?://(?:www\.|m\.)?facebook\.com", "https://mbasic.facebook.com", url, flags=re.I)

def _fb_mobile_url(url: str) -> str:
    return re.sub(r"https?://(?:www\.)?facebook\.com", "https://m.facebook.com", url, flags=re.I)

def _make_fb_cookies():
    import base64, os as _os
    datr = base64.b64encode(_os.urandom(12)).decode("ascii").rstrip("=")
    fbp  = f"fb.1.{int(time.time()*1000)}.{random.randint(100000000,999999999)}"
    return {"datr": datr, "_fbp": fbp, "locale": "en_US"}

def scrape_fb(url: str) -> tuple:
    """Returns (posts, followers). posts is always 'See Page' — FB doesn't expose post count."""
    if not url:
        return "", ""

    # Filter out non-profile FB URLs
    parsed = urlparse(url)
    if parsed.netloc and "facebook.com" not in parsed.netloc.lower():
        return "", ""
    path = parsed.path.strip("/")
    if not path or path in ("pages", "groups") or path.startswith("2008/"):
        return "", ""  # generic/template URL

    _desktop_ua = _SOCIAL_UA_LIST[0]
    _mobile_ua  = _SOCIAL_UA_LIST[1]

    def _hdrs(ua):
        return {
            "User-Agent":      ua,
            "Accept-Language": "en-US,en;q=0.9",
            "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Connection":      "close",
        }

    urls_to_try = [
        (_fb_mbasic_url(url), _hdrs(_mobile_ua),  ["chrome136", "chrome124"]),
        (url,                  _hdrs(_desktop_ua), ["chrome136", "chrome124"]),
        (_fb_mobile_url(url),  _hdrs(_mobile_ua),  ["chrome136", "safari260"]),
    ]

    if _CFFI_AVAILABLE:
        for _url, _hdrs_v, _profiles in urls_to_try:
            for _profile in _profiles:
                try:
                    _px  = _fresh_proxy()
                    sess = cffi_requests.Session(impersonate=_profile)
                    r = sess.get(_url, headers=_hdrs_v, cookies=_make_fb_cookies(),
                                 proxies=_px, timeout=25, allow_redirects=True)
                    if r.status_code != 200 or len(r.text) < 500:
                        continue
                    followers = _parse_fb_followers(r.text)
                    if followers:
                        log.info(f"  FB hit → {followers} followers ({_profile})")
                        return "See Page", followers
                    if _fb_is_hard_login_wall(r.text):
                        log.debug(f"  FB hard login wall at {_url}")
                        break
                except Exception as e:
                    log.debug(f"  FB cffi error ({_profile}): {e}")
            _sleep(0.5, 1.2)

    # Fallback: plain requests via proxy
    if _REQUESTS_AVAILABLE and _PROXY:
        try:
            r = _req.get(_fb_mbasic_url(url),
                         headers={"User-Agent": _mobile_ua, "Accept-Language": "en-US,en;q=0.9"},
                         proxies={"http": _PROXY, "https": _PROXY},
                         timeout=20, allow_redirects=True)
            if r.status_code == 200:
                followers = _parse_fb_followers(r.text)
                if followers:
                    log.info(f"  FB requests fallback → {followers} followers")
                    return "See Page", followers
        except Exception as e:
            log.debug(f"  FB requests error: {e}")

    return "", ""

# ─────────────────────────────────────────────────────────────────────────────
# Instagram
# ─────────────────────────────────────────────────────────────────────────────
_IG_LOGIN_SIGNALS = (
    "log in to instagram",
    "log into instagram",
    '"requiresLogin":true',
    "loginandsignuppage",
    "this page isn't available",
    "sorry, this page",
)

def _ig_is_login_wall(html: str) -> bool:
    lc = html.lower()
    return any(s.lower() in lc for s in _IG_LOGIN_SIGNALS)

def _parse_ig_html(html: str) -> tuple:
    # 1. window._sharedData
    m = re.search(r'window\._sharedData\s*=\s*(\{.+?\});\s*</script>', html, re.S)
    if m:
        try:
            sd   = json.loads(m.group(1))
            user = (sd.get("entry_data", {}).get("ProfilePage", [{}])[0]
                      .get("graphql", {}).get("user", {}))
            pm   = user.get("edge_owner_to_timeline_media", {}).get("count", "")
            fm   = user.get("edge_followed_by", {}).get("count", "")
            if pm != "" or fm != "":
                return str(int(pm)) if pm != "" else "", str(int(fm)) if fm != "" else ""
        except Exception:
            pass

    # 2. GraphQL count patterns
    for p_pat, f_pat in (
        (r'"edge_owner_to_timeline_media".*?"count":(\d+)', r'"edge_followed_by".*?"count":(\d+)'),
        (r'"media_count"\s*:\s*(\d+)',    r'"follower_count"\s*:\s*(\d+)'),
        (r'"post_count"\s*:\s*(\d+)',     r'"followers"\s*:\s*(\d+)'),
        (r'"mediaCount"\s*:\s*(\d+)',     r'"followerCount"\s*:\s*(\d+)'),
        (r'"postsCount"\s*:\s*(\d+)',     r'"followersCount"\s*:\s*(\d+)'),
    ):
        pm = re.search(p_pat, html, re.I)
        fm = re.search(f_pat, html, re.I)
        if pm or fm:
            return (pm.group(1) if pm else ""), (fm.group(1) if fm else "")

    soup = BeautifulSoup(html, "lxml")

    # 3. Meta description
    for attr in ({"name": "description"}, {"property": "og:description"}):
        meta = soup.find("meta", attrs=attr)
        if meta:
            desc = meta.get("content", "")
            pmd  = re.search(r"([\d,.]+[KMBkmb]?)\s*Posts?", desc, re.I)
            fmd  = re.search(r"([\d,.]+[KMBkmb]?)\s*Followers?", desc, re.I)
            if pmd or fmd:
                return (
                    pmd.group(1).replace(",", "") if pmd else "",
                    fmd.group(1).replace(",", "") if fmd else "",
                )

    # 4. Schema.org JSON-LD
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            ld = json.loads(script.string or "")
            ic = ld.get("interactionStatistic", [])
            if isinstance(ic, dict):
                ic = [ic]
            posts_ld = follow_ld = ""
            for stat in ic:
                itype = stat.get("interactionType", "")
                val   = str(stat.get("userInteractionCount", ""))
                if "WriteAction" in itype or "Post" in itype:
                    posts_ld = val
                elif "Follow" in itype:
                    follow_ld = val
            if posts_ld or follow_ld:
                return posts_ld, follow_ld
        except Exception:
            pass

    # 5. Visible text fallback
    text = soup.get_text(" ", strip=True)
    pm3  = re.search(r"([\d,.]+[KMBkmb]?)\s+posts?", text, re.I)
    fm3  = re.search(r"([\d,.]+[KMBkmb]?)\s+followers?", text, re.I)
    if pm3 or fm3:
        return (
            pm3.group(1).replace(",", "") if pm3 else "",
            fm3.group(1).replace(",", "") if fm3 else "",
        )

    return "", ""

def _make_ig_cookies():
    import base64 as _b64, uuid as _uuid, os as _os
    return {
        "csrftoken": "".join(random.choices(
            "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789", k=32)),
        "ig_did": str(_uuid.uuid4()).upper(),
        "mid":    _b64.b64encode(_os.urandom(18)).decode("ascii").rstrip("="),
        "datr":   _b64.b64encode(_os.urandom(12)).decode("ascii").rstrip("="),
    }

def scrape_ig(url: str) -> tuple:
    """Returns (posts, followers)."""
    if not url or "instagram.com" not in url.lower():
        return "", ""
    if "squarespace" in url.lower():
        return "", ""

    try:
        path     = urlparse(url).path.strip("/").rstrip("#")
        username = path.split("/")[0].lstrip("@")
        if not username:
            return "", ""
    except Exception:
        return "", ""

    _ua = random.choice(_SOCIAL_UA_LIST)
    _html_headers = {
        "User-Agent":      _ua,
        "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Referer":         "https://www.instagram.com/",
        "sec-fetch-site":  "none",
        "sec-fetch-mode":  "navigate",
        "sec-fetch-dest":  "document",
    }

    if _CFFI_AVAILABLE:
        # Strategy 1: profile page HTML
        for _profile in random.sample(_CFFI_IG_PROF, len(_CFFI_IG_PROF)):
            try:
                _px  = _fresh_proxy()
                sess = cffi_requests.Session(impersonate=_profile)
                r = sess.get(f"https://www.instagram.com/{username}/",
                             headers=_html_headers, cookies=_make_ig_cookies(),
                             proxies=_px, timeout=25, allow_redirects=True)
                if r.status_code == 200 and len(r.text) > 1000:
                    posts, followers = _parse_ig_html(r.text)
                    if posts or followers:
                        log.info(f"  IG HTML hit → {posts} posts, {followers} followers @{username}")
                        return posts, followers
                    if _ig_is_login_wall(r.text):
                        log.debug(f"  IG HTML login wall @{username} ({_profile})")
                        continue
                log.debug(f"  IG HTML no data @{username} ({_profile})")
            except Exception as e:
                log.debug(f"  IG HTML error @{username} ({_profile}): {e}")
        _sleep(0.5, 1.5)

        # Strategy 2: internal API
        api_url = f"https://www.instagram.com/api/v1/users/web_profile_info/?username={username}"
        api_hdrs = {
            "x-ig-app-id":      "936619743392459",
            "User-Agent":       _ua,
            "Accept":           "*/*",
            "Accept-Language":  "en-US,en;q=0.9",
            "Referer":          f"https://www.instagram.com/{username}/",
            "X-Requested-With": "XMLHttpRequest",
            "sec-fetch-site":   "same-origin",
            "sec-fetch-mode":   "cors",
            "sec-fetch-dest":   "empty",
        }
        for _profile in random.sample(_CFFI_IG_PROF, len(_CFFI_IG_PROF)):
            try:
                _px  = _fresh_proxy()
                sess = cffi_requests.Session(impersonate=_profile)
                r = sess.get(api_url, headers=api_hdrs, cookies=_make_ig_cookies(),
                             proxies=_px, timeout=20)
                if r.status_code == 200:
                    data = r.json()
                    user = data.get("data", {}).get("user", {})
                    if user:
                        pm   = user.get("edge_owner_to_timeline_media", {}).get("count", "")
                        fm   = user.get("edge_followed_by", {}).get("count", "")
                        ps   = str(int(pm)) if pm != "" else ""
                        fs   = str(int(fm)) if fm != "" else ""
                        if ps or fs:
                            log.info(f"  IG API hit → {ps} posts, {fs} followers @{username}")
                            return ps, fs
                log.debug(f"  IG API status {r.status_code} @{username} ({_profile})")
            except Exception as e:
                log.debug(f"  IG API error @{username} ({_profile}): {e}")
        _sleep(0.5, 1.5)

        # Strategy 3: oEmbed (public, no auth — returns some metadata)
        try:
            _px  = _fresh_proxy()
            sess = cffi_requests.Session(impersonate="chrome136")
            r = sess.get(f"https://graph.facebook.com/instagram_oembed?url=https://www.instagram.com/{username}/&access_token=",
                         timeout=10, proxies=_px)
            # oEmbed won't return follower counts but worth trying for any data
        except Exception:
            pass

    return "", ""

# ─────────────────────────────────────────────────────────────────────────────
# TikTok
# ─────────────────────────────────────────────────────────────────────────────
def scrape_tiktok(url: str) -> tuple:
    """Returns (videos, followers)."""
    if not url or "tiktok.com" not in url.lower():
        return "", ""

    try:
        path     = urlparse(url).path.strip("/")
        username = path.split("/")[0].lstrip("@")
        if not username:
            return "", ""
    except Exception:
        return "", ""

    _sleep(0.5, 1.5)
    tiktok_url = f"https://www.tiktok.com/@{username}"
    headers = {
        "User-Agent":      random.choice(_SOCIAL_UA_LIST),
        "Accept-Language": "en-US,en;q=0.9",
        "Accept":          "text/html,application/xhtml+xml,*/*;q=0.8",
        "Referer":         "https://www.tiktok.com/",
        "sec-fetch-site":  "none",
        "sec-fetch-mode":  "navigate",
        "sec-fetch-dest":  "document",
    }

    def _parse(html: str) -> tuple:
        # Method 1: __UNIVERSAL_DATA_FOR_REHYDRATION__
        m = re.search(
            r'id=["\']__UNIVERSAL_DATA_FOR_REHYDRATION__["\'][^>]*>(.*?)</script>',
            html, re.S)
        if m:
            try:
                data  = json.loads(m.group(1))
                stats = (data["__DEFAULT_SCOPE__"]["webapp.user-detail"]
                              ["userInfo"].get("stats", {}))
                followers = str(int(stats["followerCount"])) if "followerCount" in stats else ""
                videos    = str(int(stats["videoCount"]))    if "videoCount"    in stats else ""
                if followers or videos:
                    return videos, followers
            except Exception:
                pass

        # Method 2: SIGI_STATE
        m2 = re.search(r'id=["\']SIGI_STATE["\'][^>]*>(.*?)</script>', html, re.S)
        if m2:
            try:
                data2 = json.loads(m2.group(1))
                users = data2.get("UserPage", {}).get("uniqueId", {})
                if not users:
                    users = data2.get("UserModule", {}).get("users", {})
                if users:
                    udata = next(iter(users.values()), {})
                    stats = udata.get("stats", {})
                    followers = str(int(stats["followerCount"])) if "followerCount" in stats else ""
                    videos    = str(int(stats["videoCount"]))    if "videoCount"    in stats else ""
                    if followers or videos:
                        return videos, followers
            except Exception:
                pass

        # Method 3: raw regex
        mf = re.search(r'"followerCount"\s*:\s*(\d+)', html)
        mv = re.search(r'"videoCount"\s*:\s*(\d+)', html)
        return (mv.group(1) if mv else ""), (mf.group(1) if mf else "")

    if _CFFI_AVAILABLE:
        for _px in [_fresh_proxy(), None]:
            for profile in random.sample(_CFFI_PROFILES, min(3, len(_CFFI_PROFILES))):
                try:
                    sess = cffi_requests.Session(impersonate=profile)
                    r = sess.get(tiktok_url, headers=headers, proxies=_px,
                                 timeout=25, allow_redirects=True)
                    if r.status_code != 200:
                        continue
                    videos, followers = _parse(r.text)
                    if videos or followers:
                        log.info(f"  TikTok hit → {videos} videos, {followers} followers @{username}")
                        return videos, followers
                except Exception as e:
                    log.debug(f"  TikTok error ({profile}): {e}")
                _sleep(0.4, 1.0)

    return "", ""

# ─────────────────────────────────────────────────────────────────────────────
# XLSX column indexes (0-based)
# ─────────────────────────────────────────────────────────────────────────────
COL_FB_URL  = 10   # col 11: Facebook URL
COL_FB_POST = 11   # col 12: FB # Posts
COL_FB_FOL  = 12   # col 13: FB Followers
COL_IG_URL  = 13   # col 14: Instagram URL
COL_IG_POST = 14   # col 15: IG # Posts
COL_IG_FOL  = 15   # col 16: IG Followers
COL_TT_URL  = 16   # col 17: TikTok URL
COL_TT_POST = 17   # col 18: TT # Posts
COL_TT_FOL  = 18   # col 19: TT Followers

DATA_START_ROW = 3  # rows 1+2 are headers

# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────
def run(input_path: str, output_path: str):
    log.info(f"Loading: {input_path}")
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    total_rows = 0
    fb_done = ig_done = tt_done = 0
    fb_skip = ig_skip = tt_skip = 0

    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value is None:
            continue
        total_rows += 1

        practice = ws.cell(row=row_idx, column=2).value or f"row {row_idx}"

        # ── Facebook ──────────────────────────────────────────────────────────
        fb_url  = ws.cell(row=row_idx, column=COL_FB_URL  + 1).value
        fb_fol  = ws.cell(row=row_idx, column=COL_FB_FOL  + 1).value

        if fb_url and _is_missing(fb_fol):
            log.info(f"[{row_idx}] {practice} — FB: {fb_url}")
            posts, followers = scrape_fb(str(fb_url).strip())
            if followers:
                ws.cell(row=row_idx, column=COL_FB_POST + 1).value = posts or "See Page"
                ws.cell(row=row_idx, column=COL_FB_FOL  + 1).value = followers
                fb_done += 1
            else:
                log.info(f"  → FB not found")
        else:
            if fb_fol and not _is_missing(fb_fol):
                fb_skip += 1  # already has data

        # ── Instagram ─────────────────────────────────────────────────────────
        ig_url  = ws.cell(row=row_idx, column=COL_IG_URL  + 1).value
        ig_fol  = ws.cell(row=row_idx, column=COL_IG_FOL  + 1).value

        if ig_url and _is_missing(ig_fol):
            log.info(f"[{row_idx}] {practice} — IG: {ig_url}")
            posts, followers = scrape_ig(str(ig_url).strip())
            if posts or followers:
                ws.cell(row=row_idx, column=COL_IG_POST + 1).value = posts or "See Page"
                ws.cell(row=row_idx, column=COL_IG_FOL  + 1).value = followers or "See Page"
                ig_done += 1
            else:
                log.info(f"  → IG not found")
        else:
            if ig_fol and not _is_missing(ig_fol):
                ig_skip += 1

        # ── TikTok ────────────────────────────────────────────────────────────
        tt_url  = ws.cell(row=row_idx, column=COL_TT_URL  + 1).value
        tt_fol  = ws.cell(row=row_idx, column=COL_TT_FOL  + 1).value

        if tt_url and _is_missing(tt_fol):
            log.info(f"[{row_idx}] {practice} — TikTok: {tt_url}")
            videos, followers = scrape_tiktok(str(tt_url).strip())
            if videos or followers:
                ws.cell(row=row_idx, column=COL_TT_POST + 1).value = videos or "See Page"
                ws.cell(row=row_idx, column=COL_TT_FOL  + 1).value = followers or "See Page"
                tt_done += 1
            else:
                log.info(f"  → TikTok not found")
        else:
            if tt_fol and not _is_missing(tt_fol):
                tt_skip += 1

    wb.save(output_path)
    log.info("=" * 60)
    log.info(f"Rows processed : {total_rows}")
    log.info(f"FB  filled     : {fb_done}  (already had: {fb_skip})")
    log.info(f"IG  filled     : {ig_done}  (already had: {ig_skip})")
    log.info(f"TikTok filled  : {tt_done}  (already had: {tt_skip})")
    log.info(f"Saved → {output_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python social_scraper.py <input.xlsx> [output.xlsx]")
        sys.exit(1)
    _input  = sys.argv[1]
    _output = sys.argv[2] if len(sys.argv) > 2 else _input.replace(".xlsx", "_social.xlsx")
    run(_input, _output)
