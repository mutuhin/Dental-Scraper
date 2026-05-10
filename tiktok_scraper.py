"""
tiktok_scraper.py
──────────────────
Reads TikTok URLs from a batch xlsx file, scrapes follower count
(and video count if available), and writes results back.

HOW TO RUN:
    python3 tiktok_scraper.py
    python3 tiktok_scraper.py batch_01_rows1_100_deduped.xlsx
"""

import re
import sys
import json
import time
import random
import logging
from urllib.parse import urlparse

import openpyxl
from curl_cffi import requests as cf_requests

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

INPUT_FILE  = sys.argv[1] if len(sys.argv) > 1 else "batch_01_rows1_100_deduped.xlsx"
OUTPUT_FILE = INPUT_FILE

C_TT_URL  = 17   # TikTok URL column (1-based)
C_TT_POST = 18   # TT # Posts
C_TT_FOLL = 19   # TT Followers
HDR_ROW   = 2
DATA_START = 3

DELAY_MIN = 3.0
DELAY_MAX = 7.0

_USER_AGENTS = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_4_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
]

_TT_PROFILES = ["chrome124", "chrome136", "chrome133a", "safari260"]


# ── Helpers ────────────────────────────────────────────────────────────────────

def _username(tt_url: str) -> str:
    """Extract TikTok username from URL. '@artofdental' → 'artofdental'"""
    if not tt_url:
        return ""
    try:
        path = urlparse(tt_url).path.strip("/")
        name = path.split("/")[0].lstrip("@")
        return name if name else ""
    except Exception:
        return ""


def _fmt(n) -> str:
    if n is None:
        return ""
    try:
        return str(int(n))
    except Exception:
        return str(n)


# ── Scrape one TikTok profile ──────────────────────────────────────────────────

def scrape_tiktok(username: str) -> tuple[str, str]:
    """
    Fetch TikTok profile page and parse follower + video count from
    the embedded __UNIVERSAL_DATA_FOR_REHYDRATION__ JSON.
    Returns (videos, followers) strings — either may be "" if not found.
    """
    url = f"https://www.tiktok.com/@{username}"
    headers = {
        "User-Agent":      random.choice(_USER_AGENTS),
        "Accept-Language": "en-US,en;q=0.9",
        "Accept":          "text/html,application/xhtml+xml,*/*;q=0.8",
        "Referer":         "https://www.tiktok.com/",
    }

    for attempt, profile in enumerate(random.sample(_TT_PROFILES, len(_TT_PROFILES)), 1):
        try:
            sess = cf_requests.Session(impersonate=profile)
            r = sess.get(url, headers=headers, timeout=20, allow_redirects=True)

            if r.status_code != 200:
                log.debug(f"  [{attempt}] HTTP {r.status_code} with profile={profile}")
                time.sleep(1)
                continue

            html = r.text

            # ── Method 1: __UNIVERSAL_DATA_FOR_REHYDRATION__ JSON ─────────────
            m = re.search(
                r'id=["\']__UNIVERSAL_DATA_FOR_REHYDRATION__["\'][^>]*>(.*?)</script>',
                html, re.S
            )
            if m:
                try:
                    data  = json.loads(m.group(1))
                    info  = data["__DEFAULT_SCOPE__"]["webapp.user-detail"]["userInfo"]
                    stats = info.get("stats", {})
                    followers = _fmt(stats.get("followerCount"))
                    videos    = _fmt(stats.get("videoCount"))
                    if followers or videos:
                        return videos, followers
                except Exception as e:
                    log.debug(f"  JSON parse error: {e}")

            # ── Method 2: Raw regex fallback ──────────────────────────────────
            followers = videos = ""
            mf = re.search(r'"followerCount"\s*:\s*(\d+)', html)
            mv = re.search(r'"videoCount"\s*:\s*(\d+)', html)
            if mf:
                followers = mf.group(1)
            if mv:
                videos = mv.group(1)
            if followers or videos:
                return videos, followers

            log.debug(f"  [{attempt}] No stats found in page ({len(html):,} bytes)")

        except Exception as e:
            log.debug(f"  [{attempt}] Error with profile={profile}: {e}")
        time.sleep(random.uniform(0.5, 1.5))

    return "", ""


# ── Read xlsx ─────────────────────────────────────────────────────────────────

def read_tt_urls(filepath: str) -> dict:
    """Returns {username: [row_numbers]} for all unique TikTok URLs."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    mapping: dict = {}
    for r in range(DATA_START, (ws.max_row or 0) + 1):
        tt_url = str(ws.cell(r, C_TT_URL).value or "").strip()
        if not tt_url or "tiktok.com" not in tt_url:
            continue
        uname = _username(tt_url)
        if not uname:
            continue
        mapping.setdefault(uname, []).append(r)
    wb.close()
    return mapping


# ── Write results back ────────────────────────────────────────────────────────

def write_results(filepath: str, results: dict):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    for r in range(DATA_START, (ws.max_row or 0) + 1):
        tt_url = str(ws.cell(r, C_TT_URL).value or "").strip()
        if not tt_url:
            continue
        uname = _username(tt_url)
        if uname not in results:
            continue
        videos, followers = results[uname]
        if videos:
            ws.cell(r, C_TT_POST).value = videos
        if followers:
            ws.cell(r, C_TT_FOLL).value = followers
    wb.save(filepath)
    log.info(f"Saved → {filepath}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    log.info("=" * 55)
    log.info(f"  TikTok Scraper  →  {INPUT_FILE}")
    log.info("=" * 55)

    mapping = read_tt_urls(INPUT_FILE)
    unique_users = list(mapping.keys())
    log.info(f"Found {len(unique_users)} unique TikTok profiles to scrape")

    if not unique_users:
        log.warning("No TikTok URLs found in file.")
        return

    results: dict = {}
    total = len(unique_users)

    for i, username in enumerate(unique_users, 1):
        log.info(f"[{i}/{total}]  @{username}")
        videos, followers = scrape_tiktok(username)

        if followers or videos:
            log.info(f"  ✓  Videos={videos or '?'}  Followers={followers or '?'}")
        else:
            log.warning(f"  ✗  Could not retrieve data")

        results[username] = (videos, followers)

        if i % 10 == 0:
            write_results(INPUT_FILE, results)
            log.info(f"  Progress saved ({i}/{total})")

        time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))

    write_results(INPUT_FILE, results)
    found = sum(1 for v, f in results.values() if v or f)
    log.info(f"\nDone — {found}/{total} profiles scraped  →  {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
