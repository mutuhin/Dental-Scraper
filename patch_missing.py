#!/usr/bin/env python3
"""
patch_missing.py
================
Automatically scrape and fill missing Specialty & Memberships columns
in a batch Excel output file.

USAGE
-----
  python patch_missing.py --show INPUT.xlsx
        List every doctor missing Specialty or Memberships (with website URL).

  python patch_missing.py --fill INPUT.xlsx [--output OUTPUT.xlsx]
        Scrape each practice website, find doctor bio pages, extract data,
        and write results back to the Excel file.
        Saves a checkpoint after every practice — safe to interrupt.

DEPENDENCIES
------------
  pip install requests beautifulsoup4 openpyxl lxml
  pip install playwright && playwright install chromium   # optional JS fallback
"""

import argparse
import os
import re
import sys
import time
import warnings
from urllib.parse import urljoin, urlparse, urlencode

warnings.filterwarnings("ignore")

try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl")

try:
    import requests
    from bs4 import BeautifulSoup
    _HTTP_OK = True
except ImportError:
    _HTTP_OK = False
    print("WARNING: pip install requests beautifulsoup4 lxml  — HTTP scraping disabled")

try:
    from playwright.sync_api import sync_playwright
    _PW_OK = True
except ImportError:
    _PW_OK = False

# ── Excel column map (1-based) ────────────────────────────────────────────────
COL_INDEX    = 1
COL_PRACTICE = 2
COL_DOCTOR   = 3
COL_WEBSITE  = 8
COL_ASSOC    = 40   # Associations / Memberships
COL_SPEC     = 41   # Doctor Specialty
HEADERS_ROW  = 2
DATA_START   = 3

EMPTY_VALS = {"", "not found", "n/a", "none", "error", "-"}

_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,*/*;q=0.8",
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def _is_empty(val):
    return str(val or "").strip().lower() in EMPTY_VALS


def _norm(s):
    return re.sub(r'\s+', ' ', str(s or "").strip().lower())


def _slug(name):
    """'Dr. Sarah Choi DDS' → 'sarah choi' (drop titles)"""
    name = re.sub(r'\b(Dr\.?|DDS|DMD|DPM|MD|PhD|FACS|MSD)\b', '', name, flags=re.I)
    return re.sub(r'\s+', ' ', name.strip().lower())


def _last_name(name):
    parts = _slug(name).split()
    return parts[-1] if parts else ""


def _same_origin(a, b):
    return urlparse(a).netloc == urlparse(b).netloc


# ── HTTP helpers ──────────────────────────────────────────────────────────────

_SESSION = None

def _get_session():
    global _SESSION
    if _SESSION is None:
        _SESSION = requests.Session()
        _SESSION.headers.update(_HEADERS)
    return _SESSION


def safe_get(url, timeout=12):
    if not _HTTP_OK:
        return None
    try:
        r = _get_session().get(url, timeout=timeout, verify=False,
                               allow_redirects=True)
        return r if r.ok else None
    except Exception:
        return None


def get_soup(url, timeout=12):
    r = safe_get(url, timeout)
    if not r:
        return None, url
    return BeautifulSoup(r.text, "lxml"), r.url


# ── Playwright fallback ───────────────────────────────────────────────────────

_pw_browser = None
_pw_context = None
_pw         = None

def _pw_start():
    global _pw_browser, _pw_context, _pw
    if _pw_browser:
        return True
    if not _PW_OK:
        return False
    try:
        _pw = sync_playwright().__enter__()
        _pw_browser = _pw.chromium.launch(headless=True)
        _pw_context = _pw_browser.new_context(
            user_agent=_HEADERS["User-Agent"],
            locale="en-US",
        )
        return True
    except Exception as e:
        print(f"  [PW] Could not start Playwright: {e}")
        return False


def _pw_stop():
    global _pw_browser, _pw_context, _pw
    try:
        if _pw_context: _pw_context.close()
        if _pw_browser: _pw_browser.close()
        if _pw:         _pw.__exit__(None, None, None)
    except Exception:
        pass
    _pw_browser = _pw_context = _pw = None


def pw_get_text(url, timeout=15000):
    """Fetch page with Playwright, return (html_text, final_url)."""
    if not _pw_start():
        return None, url
    try:
        page = _pw_context.new_page()
        page.goto(url, timeout=timeout, wait_until="domcontentloaded")
        try:
            page.wait_for_load_state("networkidle", timeout=5000)
        except Exception:
            pass
        html = page.content()
        final = page.url
        page.close()
        return html, final
    except Exception as e:
        print(f"  [PW] {url}: {e}")
        return None, url


def get_soup_pw(url):
    """Try static request first; fall back to Playwright."""
    soup, final = get_soup(url)
    if soup and len(soup.get_text()) > 300:
        return soup, final
    if _PW_OK:
        html, final = pw_get_text(url)
        if html:
            return BeautifulSoup(html, "lxml"), final
    return None, url


# ── Team-page discovery ───────────────────────────────────────────────────────

_TEAM_PATHS = [
    "/team", "/our-team", "/meet-the-team", "/meet-our-team",
    "/doctors", "/our-doctors", "/meet-our-doctors",
    "/providers", "/our-providers",
    "/staff", "/our-staff",
    "/about", "/about-us",
    "/dentists", "/meet-our-dentists",
]

_TEAM_KW = (
    "team", "doctor", "provider", "staff", "dentist",
    "meet", "our-doctor", "about", "physician", "specialist",
)

_SKIP_EXT = ('.pdf', '.jpg', '.jpeg', '.png', '.gif', '.svg', '.zip', '.mp4')


def find_team_url(website, soup):
    """
    Given the homepage soup, return the best candidate team/doctor page URL.
    Falls back to trying common path patterns.
    """
    base = website.rstrip('/')

    # 1. Nav links containing team keywords
    for a in soup.find_all("a", href=True):
        href = a["href"]
        href_l = href.lower()
        if any(ext in href_l for ext in _SKIP_EXT):
            continue
        full = urljoin(website, href)
        if not _same_origin(website, full):
            continue
        if any(kw in href_l for kw in _TEAM_KW):
            return full

    # 2. Try known path patterns
    for path in _TEAM_PATHS:
        url = base + path
        r = safe_get(url, timeout=8)
        if r and r.ok:
            return r.url

    return website   # fallback: use homepage itself


# ── Doctor bio-link discovery ─────────────────────────────────────────────────

_BIO_KW = (
    "bio", "profile", "doctor", "team", "provider", "dentist",
    "meet", "about", "dr-", "/dr/", "staff",
)

_DR_RE = re.compile(r'\b(Dr\.?|DDS|DMD)\b', re.I)


def _link_score(href_l, link_text, doctor_slug, last):
    """Score a link for how likely it is to be the doctor's bio page."""
    score = 0
    if last and last in href_l:
        score += 10
    if last and last in link_text.lower():
        score += 8
    # Full slug match
    slug_parts = doctor_slug.split()
    if len(slug_parts) >= 2 and all(p in href_l for p in slug_parts):
        score += 15
    if any(k in href_l for k in _BIO_KW):
        score += 3
    if _DR_RE.search(link_text):
        score += 2
    return score


def find_doctor_bio_links(team_url, team_soup, doctor_names):
    """
    Return {doctor_name: bio_url} for each doctor we can match.
    """
    result = {}
    candidates = []

    for a in team_soup.find_all("a", href=True):
        href = a["href"]
        href_l = href.lower()
        if any(ext in href_l for ext in _SKIP_EXT):
            continue
        full = urljoin(team_url, href)
        if not _same_origin(team_url, full):
            continue
        link_text = a.get_text(" ", strip=True)
        candidates.append((full, href_l, link_text))

    for name in doctor_names:
        slug = _slug(name)
        last = _last_name(name)
        best_url, best_score = None, 0
        for full, href_l, link_text in candidates:
            s = _link_score(href_l, link_text, slug, last)
            if s > best_score:
                best_score = s
                best_url = full
        if best_score >= 8:   # require a decent match
            result[name] = best_url

    return result


# ── Bio text extraction from a page ──────────────────────────────────────────

def _extract_bio_text(soup, doctor_name):
    """
    Try to extract the portion of the page that is this doctor's bio.
    Falls back to full page text.
    """
    last = _last_name(doctor_name)
    full_text = soup.get_text(" ", strip=True)

    if not last:
        return full_text

    # Look for a section/div/article that contains the doctor's name
    for tag in soup.find_all(["section", "article", "div"], limit=200):
        tag_text = tag.get_text(" ", strip=True)
        if last.lower() in tag_text.lower() and len(tag_text) > 100:
            # Prefer a reasonably sized block that's not the entire body
            if len(tag_text) < len(full_text) * 0.85:
                return tag_text

    return full_text


def _scrape_bio_page(url, doctor_name):
    """Fetch a doctor bio page and return its text. Tries Playwright if needed."""
    soup, _ = get_soup_pw(url)
    if soup:
        return _extract_bio_text(soup, doctor_name)
    return ""


# ── Specialty extraction (mirrors dental_scraper.py) ─────────────────────────

def _extract_specialty_phrase(text):
    _FILLER = {
        "dentistry", "dental care", "dental health", "patients", "all patients",
        "our patients", "patient care", "the best care", "quality care",
        "comprehensive care", "the community",
    }
    _PATTERNS = [
        r'(?:earned?|completed?|received?|holds?|has)\s+certificate(?:s|ation)?\s+in\s+([^.;\n]{5,120})',
        r'certificate(?:s|ation)?\s+in\s+([^.;\n]{5,120})',
        r'certif(?:ied|ication)\s+in\s+([^.;\n]{5,100})',
        r'additional\s+training\s+in\s+([^.;\n]{5,100})',
        r'(?:advanced\s+)?training\s+in\s+([^.;\n]{5,100})',
        r'continuing\s+education\s+in\s+([^.;\n]{5,100})',
        r'areas?\s+of\s+(?:specialty|specialization|focus|interest|expertise)[:\s]+(?:include[s]?\s+)?([^.;\n]{5,90})',
        r'specializ(?:es?|ing|ation)\s+in\s+([^.;\n]{5,80})',
        r'specialty\s+(?:is\b|includes?\b|:)\s*([^.;\n]{5,70})',
        r'specialist\s+in\s+([^.;\n]{5,70})',
        r'board[- ]certified\s+(?:in\s+)?([^.;\n]{5,70})',
        r'expertise\s+in\s+([^.;\n]{5,70})',
    ]
    _VALID_SPEC = re.compile(
        r'\b(implant|cosmetic|esthetic|orthodont|invisalign|aligner|'
        r'periodon|endodon|prosthodon|oral\s+surg|pediat|'
        r'sleep\s+apnea|tmj|root\s+canal|extract|veneer|whitening|'
        r'crown|bridge|laser|sedation|restorat|preventi|preventa|'
        r'biomimetic|holistic|biolog|clear\s+aligner|dental\s+implant)\b',
        re.I,
    )
    _REJECT = re.compile(
        r'\b(?:Dr\.|DDS|DMD|MD\b|C-FNP|NP\b|PA\b|LISW|LCSW|APRN|'
        r'View\s+Profile|Healthsource|Schedule|Appointment|'
        r'Aspects\b|Every\s+Age|Every\s+Patient|'
        r'Highest\s+Standard|Highest\s+Level|Constant\s+Pursuit|'
        r'Committed\s+To|Commitment\s+To|Continuing\s+Education\b|'
        r'Patient\s+Care|Our\s+Team|Our\s+Practice|'
        r'And\s+Is\s+Committ|And\s+Has\s+Spoken)\b',
        re.I,
    )
    text_l = text.lower()
    for pat in _PATTERNS:
        m = re.search(pat, text_l, re.IGNORECASE)
        if m:
            phrase = m.group(1).strip().rstrip(' ,;.')
            if len(phrase) < 5 or phrase.lower() in _FILLER:
                continue
            phrase = re.split(
                r'\s+(?:while|as well as|in addition to|for every|'
                r'when\s+(?:he|she|they)\b|and\s+is\b|and\s+has\b|and\s+was\b)',
                phrase, flags=re.I
            )[0]
            phrase = phrase[:150].strip().rstrip(' ,;.')
            if _REJECT.search(phrase):
                continue
            if not _VALID_SPEC.search(phrase):
                continue
            if phrase:
                return phrase.title() if phrase == phrase.lower() else phrase
    return ""


def find_specialty(text):
    specialty_map = [
        ("Cosmetic",         ["cosmetic dent", "esthetic dent", "smile makeover", "cosmetic smile"]),
        ("Restorative",      ["restorative dent", "dental restoration", "full mouth restoration",
                              "full mouth reconstruction", "dental rebuild"]),
        ("Implants",         ["dental implant", "implant specialist", "implant dentist",
                              "tooth implant", "all-on-4", "all on 4", "all-on-x"]),
        ("Orthodontics",     ["orthodontist", "orthodontic"]),
        ("Clear Aligners",   ["invisalign", "clear aligner", "clear correct"]),
        ("Pediatric",        ["pediatric dent", "children's dent", "kids dent", "child dent"]),
        ("Periodontics",     ["periodontist", "periodontal", "gum disease specialist"]),
        ("Endodontics",      ["endodontist", "root canal specialist", "root canal therapy"]),
        ("Oral Surgery",     ["oral surgeon", "oral surgery", "wisdom teeth", "wisdom tooth",
                              "third molar", "tooth extraction", "jaw surgery", "maxillofacial"]),
        ("Prosthodontics",   ["prosthodontist", "prosthodontic"]),
        ("Sleep Apnea / TMJ",["tmj", "sleep apnea", "sleep dentistry", "snoring treatment",
                              "temporomandibular"]),
        ("Laser",            ["laser dent", "laser treatment", "laser therapy", "soft tissue laser",
                              "biolase", "waterlase"]),
        ("Sedation",         ["sedation dent", "sedation specialist", "iv sedation", "nitrous oxide"]),
        ("Holistic",         ["holistic dent", "biological dent", "mercury-free", "mercury free",
                              "biocompatible"]),
        ("Family",           ["family dent", "family practice", "comprehensive dental",
                              "general and family"]),
        ("General",          ["general dent", "general dentist"]),
    ]
    text_lower = text.lower()
    found = []
    seen = set()
    for label, kws in specialty_map:
        if label in seen:
            continue
        if any(kw in text_lower for kw in kws):
            found.append(label)
            seen.add(label)

    phrase = _extract_specialty_phrase(text)
    if found and phrase:
        return phrase
    if found:
        return " / ".join(found)
    if phrase:
        return phrase
    return ""


# ── Associations extraction ───────────────────────────────────────────────────

def find_associations(text):
    assoc_map = {
        "ADA":    "American Dental Association",
        "AGD":    "Academy of General Dentistry",
        "FAGD":   "Fellow of the Academy of General Dentistry",
        "MAGD":   "Master of the Academy of General Dentistry",
        "AACD":   "American Academy of Cosmetic Dentistry",
        "AAED":   "American Academy of Esthetic Dentistry",
        "AAO":    "American Association of Orthodontists",
        "AAE":    "American Association of Endodontists",
        "AAP":    "American Academy of Periodontology",
        "AAPD":   "American Academy of Pediatric Dentistry",
        "AAOMS":  "American Association of Oral and Maxillofacial Surgeons",
        "ACP":    "American College of Prosthodontists",
        "AAID":   "American Academy of Implant Dentistry",
        "ICOI":   "International Congress of Oral Implantologists",
        "FICOI":  "Fellow of the International Congress of Oral Implantologists",
        "MICOI":  "Master of the International Congress of Oral Implantologists",
        "ABOI":   "American Board of Oral Implantology",
        "AO":     "Academy of Osseointegration",
        "ITI":    "International Team for Implantology",
        "AADSM":  "American Academy of Dental Sleep Medicine",
        "FACD":   "Fellow of the American College of Dentists",
        "FICD":   "Fellow of the International College of Dentists",
        "IABDM":  "International Academy of Biological Dentistry and Medicine",
        "IAOMT":  "International Academy of Oral Medicine and Toxicology",
        "OKU":    "Omicron Kappa Upsilon",
        "AEGD":   "Academy of Education in General Dentistry",
        "ISDS":   "Illinois State Dental Society",
        "IDIA":   "International Dental Implant Association",
        "CDA":    "California Dental Association",
        "TDA":    "Texas Dental Association",
        "FDC":    "Florida Dental Association",
        "NDA":    "National Dental Association",
        "ASDA":   "American Student Dental Association",
        "SPEAR":  "Spear Education",
        "PANKEY": "L.D. Pankey Institute",
    }

    text_upper = text.upper()
    found = []
    for abbr, full in assoc_map.items():
        if re.search(rf'\b{re.escape(abbr)}\b', text_upper) or full.upper() in text_upper:
            found.append(abbr)

    # Free-text fallback — only fires when no abbreviations matched.
    # Requires the phrase to contain a known org-type word AND not contain
    # address/contact markers (digits, "reach out", "schedule", etc.).
    if not found:
        _ORG_MARKER = re.compile(
            r'\b(academy|association|society|college|institute|board|congress|'
            r'federation|council|foundation|alliance|organization|university)\b',
            re.I,
        )
        _MEMBER_PAT = re.compile(
            r'(?:member|fellow|diplomate|affiliate)\s+of\s+(?:the\s+)?([^.;\n]{5,70})',
            re.I,
        )
        _STOP = re.compile(
            r'\s+\d'                                    # address / zip code
            r'|\band\s+(?:has|is|was|he|she|they)\b'   # "and has spoken at…"
            r'|\bwhere\b|\bwho\b|\bwhich\b'            # relative clause
            r'|schedule|appointment|reach\s+out'       # contact-form noise
            r'|our\s+team|your\s+care',
            re.I,
        )
        for m in _MEMBER_PAT.finditer(text):
            org = m.group(1)
            # Truncate at stop markers
            stop = _STOP.search(org)
            if stop:
                org = org[: stop.start()]
            org = org.strip().rstrip('.,; ')
            # Only keep if it looks like a real org name
            if len(org) > 8 and _ORG_MARKER.search(org):
                found.append(org)

    return ", ".join(found) if found else ""


# ── Per-practice scraping ─────────────────────────────────────────────────────

def scrape_practice(website, doctors_needed):
    """
    Visit website → find team page → find bio pages → extract data.
    Returns {doctor_name: {"specialty": str, "associations": str}}
    """
    results = {d: {"specialty": "", "associations": ""} for d in doctors_needed}

    print(f"  → {website}")

    # 1. Fetch homepage
    home_soup, home_url = get_soup_pw(website)
    if not home_soup:
        print(f"    ✗ Could not fetch homepage")
        return results

    # 2. Find team page
    team_url = find_team_url(home_url, home_soup)
    team_soup, team_url = get_soup_pw(team_url)
    if not team_soup:
        print(f"    ✗ Could not fetch team page")
        return results

    # If homepage == team page, try sub-nav for individual doctor links
    print(f"    Team page: {team_url}")

    # 3. Find bio links for all doctors in one pass
    bio_links = find_doctor_bio_links(team_url, team_soup, list(doctors_needed))
    print(f"    Bio links found: {len(bio_links)}/{len(doctors_needed)}")

    # 4. Scrape each bio page
    for name, needed in doctors_needed.items():
        bio_url = bio_links.get(name)

        # Try to extract from team page text first (quick win)
        team_text = _extract_bio_text(team_soup, name)
        spec  = find_specialty(team_text)    if needed.get("spec")  else ""
        assoc = find_associations(team_text) if needed.get("assoc") else ""

        # Fetch individual bio page if still missing
        if (not spec and needed.get("spec")) or (not assoc and needed.get("assoc")):
            if bio_url and bio_url != team_url:
                print(f"    Fetching bio: {bio_url}")
                bio_text = _scrape_bio_page(bio_url, name)
                if bio_text:
                    if not spec  and needed.get("spec"):
                        spec  = find_specialty(bio_text)
                    if not assoc and needed.get("assoc"):
                        assoc = find_associations(bio_text)
                time.sleep(0.4)
            else:
                # Try a direct URL guess: /team/dr-firstname-lastname
                slug_url = _try_slug_urls(website, name)
                if slug_url:
                    print(f"    Guessed bio: {slug_url}")
                    bio_text = _scrape_bio_page(slug_url, name)
                    if bio_text:
                        if not spec  and needed.get("spec"):
                            spec  = find_specialty(bio_text)
                        if not assoc and needed.get("assoc"):
                            assoc = find_associations(bio_text)
                    time.sleep(0.4)

        results[name]["specialty"]     = spec
        results[name]["associations"]  = assoc
        status = []
        if spec:  status.append(f"Spec='{spec[:40]}'")
        if assoc: status.append(f"Assoc='{assoc}'")
        print(f"    {name}: {', '.join(status) if status else 'no data found'}")

    time.sleep(0.5)
    return results


def _try_slug_urls(website, doctor_name):
    """
    Try common URL patterns for a doctor bio page and return the first that responds.
    """
    slug = _slug(doctor_name)
    parts = slug.split()
    if len(parts) < 1:
        return None

    first = parts[0]  if len(parts) >= 1 else ""
    last  = parts[-1] if len(parts) >= 1 else ""
    dash_slug = "-".join(parts)

    candidates = [
        f"/team/dr-{last}",
        f"/team/dr-{dash_slug}",
        f"/team/{dash_slug}",
        f"/doctors/dr-{last}",
        f"/doctors/dr-{dash_slug}",
        f"/doctors/{dash_slug}",
        f"/providers/dr-{last}",
        f"/about/team/dr-{last}",
        f"/dr-{last}",
        f"/dr-{first}-{last}",
    ]

    base = website.rstrip('/')
    for path in candidates:
        url = base + path
        r = safe_get(url, timeout=6)
        if r and r.ok and len(r.text) > 500:
            return r.url

    return None


# ── Excel I/O ─────────────────────────────────────────────────────────────────

def iter_missing_rows(ws):
    """
    Yield (row, index, doctor, website, needs_spec, needs_assoc).
    Only rows where at least one of spec/assoc is missing.
    """
    for r in range(DATA_START, ws.max_row + 1):
        idx   = ws.cell(r, COL_INDEX).value
        doc   = ws.cell(r, COL_DOCTOR).value
        web   = ws.cell(r, COL_WEBSITE).value
        assoc = ws.cell(r, COL_ASSOC).value
        spec  = ws.cell(r, COL_SPEC).value
        if idx is None and doc is None:
            continue
        needs_spec  = _is_empty(spec)
        needs_assoc = _is_empty(assoc)
        if needs_spec or needs_assoc:
            yield r, idx, str(doc or ""), str(web or ""), needs_spec, needs_assoc


# ── Mode: --show ──────────────────────────────────────────────────────────────

def cmd_show(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = list(iter_missing_rows(ws))
    if not rows:
        print("No missing Specialty/Memberships rows found.")
        return
    print(f"{'Row':<5} {'Index':<6} {'Doctor':<35} {'Missing':<20} {'Website'}")
    print("-" * 110)
    for r, idx, doc, web, ns, na in rows:
        miss = "+".join(["Specialty" if ns else "", "Memberships" if na else ""]
                        ).strip("+").replace("+", " & ")
        print(f"{r:<5} {str(idx):<6} {doc:<35} {miss:<20} {web}")
    print(f"\nTotal: {len(rows)} rows missing data")


# ── Mode: --fill ──────────────────────────────────────────────────────────────

def cmd_fill(input_path, output_path):
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    # Group rows by website
    groups = {}   # website → {doctor_name: {"row": r, "spec": bool, "assoc": bool}}
    for r, idx, doc, web, ns, na in iter_missing_rows(ws):
        if not web or web.lower() in EMPTY_VALS:
            continue
        web = web.strip()
        if not web.startswith("http"):
            web = "https://" + web
        groups.setdefault(web, {})[doc] = {"row": r, "spec": ns, "assoc": na}

    if not groups:
        print("No rows with missing data and a website URL found.")
        return

    total     = len(groups)
    filled    = 0
    not_found = 0

    print(f"Processing {total} practices...\n")

    for i, (website, doctors) in enumerate(groups.items(), 1):
        print(f"[{i}/{total}] {website}  ({len(doctors)} doctor(s) missing)")

        # Scrape
        try:
            data = scrape_practice(website, doctors)
        except Exception as e:
            print(f"  ERROR: {e}")
            data = {}

        # Write results back to worksheet
        for name, info in data.items():
            row_info = doctors.get(name, {})
            r        = row_info.get("row")
            if not r:
                continue
            changed = False
            if info["specialty"] and row_info.get("spec"):
                ws.cell(r, COL_SPEC).value = info["specialty"]
                changed = True
            if info["associations"] and row_info.get("assoc"):
                ws.cell(r, COL_ASSOC).value = info["associations"]
                changed = True
            if changed:
                filled += 1
            else:
                not_found += 1

        # Save checkpoint after every practice
        wb.save(output_path)

    _pw_stop()
    print(f"\n{'='*60}")
    print(f"Done. Filled: {filled}  |  Not found: {not_found}")
    print(f"Saved → {output_path}")


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(
        description="Auto-scrape missing Specialty & Memberships from practice websites.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument("--show", metavar="INPUT.xlsx",
                   help="List all rows missing Specialty or Memberships")
    p.add_argument("--fill", metavar="INPUT.xlsx",
                   help="Scrape and fill missing data, save to --output")
    p.add_argument("--output", "-o", metavar="OUTPUT.xlsx",
                   help="Output file (default: INPUT_filled.xlsx)")
    args = p.parse_args()

    if args.show:
        cmd_show(args.show)
    elif args.fill:
        out = args.output or args.fill.replace(".xlsx", "_filled.xlsx")
        cmd_fill(args.fill, out)
    else:
        p.print_help()


if __name__ == "__main__":
    main()
