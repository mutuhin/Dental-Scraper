"""
linkedin_scraper.py
────────────────────
Reads LinkedIn URLs from a batch xlsx file, scrapes follower count
(and post count if available), and writes results back.

URL types handled:
  - /company/{slug}  → company page (followers + posts attempted)
  - /in/{slug}       → personal profile (followers attempted)
  - /shareArticle    → skipped (share links, not profiles)

HOW TO RUN:
    python3 linkedin_scraper.py
    python3 linkedin_scraper.py batch_01_rows1_100_deduped.xlsx

Strategy (tried in order):
  1. curl_cffi TLS impersonation → parse embedded JSON from public page
  2. Playwright visible browser → login once, reuse persistent session
"""

import os
import re
import sys
import time
import random
import logging
from urllib.parse import urlparse

import openpyxl
from curl_cffi import requests as cf_requests

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

INPUT_FILE  = sys.argv[1] if len(sys.argv) > 1 else "batch_01_rows1_100_deduped.xlsx"
OUTPUT_FILE = INPUT_FILE

C_LI_URL  = 20   # LinkedIn URL column (1-based)
C_LI_POST = 21   # LI # Posts
C_LI_FOLL = 22   # LI Followers
HDR_ROW   = 2
DATA_START = 3

DELAY_MIN = 4.0
DELAY_MAX = 9.0

_PW_USER_DATA = "/tmp/pw_li_profile"

_USER_AGENTS = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_4_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
]

_STEALTH_JS = """
(() => {
  Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
  Object.defineProperty(navigator, 'plugins',   {get: () => [1,2,3,4,5]});
  Object.defineProperty(navigator, 'languages', {get: () => ['en-US','en']});
  window.chrome = { runtime: {}, loadTimes: function(){}, csi: function(){} };
  delete window.__playwright;
})();
"""

_CF_PROFILES = ["chrome124", "chrome136", "chrome133a", "safari260"]


# ── Helpers ────────────────────────────────────────────────────────────────────

def _parse_li_url(li_url: str) -> tuple[str, str]:
    """
    Returns (url_type, slug).
    url_type: 'company' | 'personal' | 'skip'
    """
    if not li_url:
        return "skip", ""
    if "shareArticle" in li_url or "share?" in li_url or "/shareArticle" in li_url:
        return "skip", ""
    try:
        path = urlparse(li_url).path.strip("/")
        parts = [p for p in path.split("/") if p]
        if not parts:
            return "skip", ""
        if parts[0] == "company" and len(parts) >= 2:
            return "company", parts[1]
        elif parts[0] == "in" and len(parts) >= 2:
            return "personal", parts[1]
        else:
            return "skip", ""
    except Exception:
        return "skip", ""


def _fmt(n) -> str:
    if n is None:
        return ""
    try:
        return str(int(str(n).replace(",", "").replace(".", "")))
    except Exception:
        return str(n)


def _extract_from_html(html: str) -> tuple[str, str]:
    """
    Try multiple patterns to extract (followers, posts) from raw HTML/JSON.
    Returns ("", "") if nothing found.
    """
    followers = posts = ""

    # Pattern 1: followerCount in embedded JSON (company pages)
    for pat in [
        r'"followerCount"\s*:\s*(\d+)',
        r'"followersCount"\s*:\s*(\d+)',
        r'"num_followers"\s*:\s*(\d+)',
    ]:
        m = re.search(pat, html)
        if m:
            followers = m.group(1)
            break

    # Pattern 2: "X followers" in text (rendered page)
    if not followers:
        m = re.search(r'([\d,]+)\s+followers', html, re.I)
        if m:
            followers = m.group(1).replace(",", "")

    # Pattern 3: posts count (company pages sometimes have this)
    for pat in [
        r'"postsCount"\s*:\s*(\d+)',
        r'"totalCount"\s*:\s*(\d+)',
    ]:
        m = re.search(pat, html)
        if m:
            # Avoid picking up large unrelated counts
            val = int(m.group(1))
            if val < 100_000:
                posts = str(val)
                break

    return followers, posts


# ── Method 1: curl_cffi fetch ──────────────────────────────────────────────────

def _curl_fetch(url: str) -> tuple[str, str]:
    """Try curl_cffi with TLS impersonation. Returns (followers, posts)."""
    headers = {
        "User-Agent":      random.choice(_USER_AGENTS),
        "Accept-Language": "en-US,en;q=0.9",
        "Accept":          "text/html,application/xhtml+xml,*/*;q=0.8",
        "Referer":         "https://www.linkedin.com/",
    }
    for profile in random.sample(_CF_PROFILES, len(_CF_PROFILES)):
        try:
            sess = cf_requests.Session(impersonate=profile)
            r = sess.get(url, headers=headers, timeout=20, allow_redirects=True)
            if r.status_code == 200:
                followers, posts = _extract_from_html(r.text)
                if followers:
                    return followers, posts
                # LinkedIn redirected to login page
                if "authwall" in r.url or "login" in r.url.lower():
                    log.debug(f"  curl: redirected to auth wall ({profile})")
                    return "", ""
            log.debug(f"  curl: HTTP {r.status_code} ({profile})")
        except Exception as e:
            log.debug(f"  curl error ({profile}): {e}")
        time.sleep(0.5)
    return "", ""


# ── Method 2: Playwright browser ──────────────────────────────────────────────

def _launch_browser():
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        log.warning("Playwright not installed — only curl_cffi method available")
        return None, None, None
    os.makedirs(_PW_USER_DATA, exist_ok=True)
    _pw = sync_playwright().__enter__()
    ctx = _pw.chromium.launch_persistent_context(
        user_data_dir=_PW_USER_DATA,
        headless=False,
        slow_mo=80,
        locale="en-US",
        args=["--disable-blink-features=AutomationControlled", "--lang=en-US"],
        user_agent=random.choice(_USER_AGENTS),
    )
    try:
        ctx.add_init_script(_STEALTH_JS)
    except Exception:
        pass
    page = ctx.new_page()
    page.set_extra_http_headers({"Accept-Language": "en-US,en;q=0.9"})
    return _pw, ctx, page


def _is_logged_in(page) -> bool:
    """Check if the current Playwright page/session is logged into LinkedIn."""
    try:
        url = page.url
        if "linkedin.com/feed" in url or "linkedin.com/in/" in url or "linkedin.com/company/" in url:
            content = page.content()
            # Logged-in pages have the global nav with profile photo
            if 'global-nav' in content or 'feed-identity-module' in content:
                return True
        # Check for presence of authenticated nav element
        el = page.query_selector(".global-nav__me-photo, .feed-identity-module, [data-test-id='nav-settings__account-type-label']")
        return el is not None
    except Exception:
        return False


def _ensure_logged_in(page):
    """Navigate to LinkedIn and wait for the user to log in if needed."""
    try:
        page.goto("https://www.linkedin.com/feed/", timeout=30_000, wait_until="domcontentloaded")
        time.sleep(2)

        if _is_logged_in(page):
            log.info("  Already logged in to LinkedIn ✓")
            return True

        log.warning("=" * 55)
        log.warning("  ACTION REQUIRED: Log in to LinkedIn in the browser")
        log.warning("  window that just opened, then the scraper will")
        log.warning("  continue automatically.")
        log.warning("=" * 55)

        # Wait up to 3 minutes for the user to log in
        page.wait_for_url("**/feed/**", timeout=180_000)
        time.sleep(3)
        return True
    except Exception as e:
        log.warning(f"  Login wait error: {e}")
        return False


def _pw_fetch(url: str, page, url_type: str) -> tuple[str, str]:
    """Scrape LinkedIn profile via Playwright. Returns (followers, posts)."""
    try:
        page.goto(url, timeout=35_000, wait_until="domcontentloaded")
        time.sleep(random.uniform(2.5, 4.0))

        # If redirected to auth wall, we need login
        if "authwall" in page.url or "login" in page.url:
            log.debug("  Redirected to auth wall — session may have expired")
            return "", ""

        content = page.content()

        # Try JSON in page source first
        followers, posts = _extract_from_html(content)
        if followers:
            return followers, posts

        # Try visible text for followers
        if url_type == "company":
            selectors = [
                ".org-top-card-summary-info-list__info-item",
                "[data-test-id='followers-count']",
                ".org-top-card__followers-count",
                "span.t-bold",
                ".org-top-card__primary-content li",
            ]
        else:
            selectors = [
                ".pv-top-card--list li",
                ".pvs-header__subtitle span",
                "span.t-bold",
                ".profile-top-card__connections",
            ]

        for sel in selectors:
            try:
                for el in page.query_selector_all(sel):
                    txt = (el.inner_text() or "").strip()
                    m = re.search(r'([\d,]+)\s*followers', txt, re.I)
                    if m:
                        followers = m.group(1).replace(",", "")
                        return followers, posts
                    m2 = re.search(r'([\d.]+[KkMm]?)\s*followers', txt, re.I)
                    if m2:
                        followers = m2.group(1)
                        return followers, posts
            except Exception:
                pass

        # Last resort: regex on full page text
        try:
            body_text = page.inner_text("body")
            m = re.search(r'([\d,]+)\s+followers', body_text, re.I)
            if m:
                followers = m.group(1).replace(",", "")
                return followers, posts
            m2 = re.search(r'([\d.]+[KkMm]?)\s+followers', body_text, re.I)
            if m2:
                followers = m2.group(1)
                return followers, posts
        except Exception:
            pass

        return followers, posts

    except Exception as e:
        log.warning(f"  Playwright error: {e}")
        return "", ""


# ── Read xlsx ─────────────────────────────────────────────────────────────────

def read_li_urls(filepath: str) -> dict:
    """Returns {(url_type, slug): [row_numbers]} for all valid LinkedIn URLs."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    mapping: dict = {}
    for r in range(DATA_START, (ws.max_row or 0) + 1):
        li_url = str(ws.cell(r, C_LI_URL).value or "").strip()
        if not li_url or "linkedin.com" not in li_url:
            continue
        url_type, slug = _parse_li_url(li_url)
        if url_type == "skip" or not slug:
            continue
        key = (url_type, slug)
        mapping.setdefault(key, []).append(r)
    wb.close()
    return mapping


# ── Write results back ────────────────────────────────────────────────────────

def write_results(filepath: str, results: dict):
    """results = {(url_type, slug): (followers, posts)}"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    for r in range(DATA_START, (ws.max_row or 0) + 1):
        li_url = str(ws.cell(r, C_LI_URL).value or "").strip()
        if not li_url:
            continue
        url_type, slug = _parse_li_url(li_url)
        key = (url_type, slug)
        if key not in results:
            continue
        followers, posts = results[key]
        if posts:
            ws.cell(r, C_LI_POST).value = posts
        if followers:
            ws.cell(r, C_LI_FOLL).value = followers
    wb.save(filepath)
    log.info(f"Saved → {filepath}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    log.info("=" * 55)
    log.info(f"  LinkedIn Scraper  →  {INPUT_FILE}")
    log.info("=" * 55)

    mapping = read_li_urls(INPUT_FILE)
    unique_profiles = list(mapping.keys())
    log.info(f"Found {len(unique_profiles)} unique LinkedIn profiles to scrape")
    company_count = sum(1 for t, _ in unique_profiles if t == "company")
    personal_count = sum(1 for t, _ in unique_profiles if t == "personal")
    log.info(f"  Company pages: {company_count}  |  Personal profiles: {personal_count}")

    if not unique_profiles:
        log.warning("No LinkedIn URLs found in file.")
        return

    _pw, ctx, pw_page = _launch_browser()

    # Ensure we're logged in before starting
    if pw_page:
        _ensure_logged_in(pw_page)

    results: dict = {}
    total = len(unique_profiles)

    try:
        for i, (url_type, slug) in enumerate(unique_profiles, 1):
            if url_type == "company":
                url = f"https://www.linkedin.com/company/{slug}/"
            else:
                url = f"https://www.linkedin.com/in/{slug}/"

            log.info(f"[{i}/{total}]  [{url_type}]  {slug}")

            # Method 1: curl_cffi
            followers, posts = _curl_fetch(url)

            # Method 2: Playwright fallback
            if not followers and pw_page:
                log.info("  curl failed — trying Playwright…")
                followers, posts = _pw_fetch(url, pw_page, url_type)

            if followers or posts:
                log.info(f"  ✓  Followers={followers or '?'}  Posts={posts or 'N/A'}")
            else:
                log.warning("  ✗  Could not retrieve data")

            results[(url_type, slug)] = (followers, posts)

            if i % 5 == 0:
                write_results(INPUT_FILE, results)
                log.info(f"  Progress saved ({i}/{total})")

            time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))

    finally:
        if ctx:
            try:
                ctx.close()
                _pw.__exit__(None, None, None)
            except Exception:
                pass

    write_results(INPUT_FILE, results)
    found = sum(1 for f, p in results.values() if f or p)
    log.info(f"\nDone — {found}/{total} profiles scraped  →  {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
