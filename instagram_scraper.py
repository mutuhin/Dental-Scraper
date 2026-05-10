"""
instagram_scraper.py
─────────────────────
Reads Instagram URLs from a batch xlsx file, scrapes follower count
and post count for each unique profile, and writes the results back.

HOW TO RUN:
    python3 instagram_scraper.py
    python3 instagram_scraper.py batch_01_rows1_100.xlsx

Strategy (tried in order):
  1. Instagram internal API endpoint (fast, no login needed for public profiles)
  2. Playwright visible browser (fallback when API is rate-limited)

OUTPUT:
    Overwrites the input file with updated IG # Posts and IG Followers columns.
"""

import re
import os
import sys
import time
import random
import logging
from urllib.parse import urlparse

import openpyxl
from curl_cffi import requests as cf_requests

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

INPUT_FILE   = sys.argv[1] if len(sys.argv) > 1 else "batch_01_rows1_100.xlsx"
OUTPUT_FILE  = INPUT_FILE   # overwrite in place

# Column indices (1-based)
C_INDEX   = 1
C_IG_URL  = 14
C_IG_POST = 15
C_IG_FOLL = 16
HDR_ROW   = 2
DATA_START = 3

DELAY_MIN = 3.0
DELAY_MAX = 7.0

_USER_AGENTS = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_4_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
]

_PW_USER_DATA = "/tmp/pw_ig_profile"
_STEALTH_JS = """
(() => {
  Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
  Object.defineProperty(navigator, 'plugins',   {get: () => [1,2,3,4,5]});
  Object.defineProperty(navigator, 'languages', {get: () => ['en-US','en']});
  window.chrome = { runtime: {}, loadTimes: function(){}, csi: function(){} };
  delete window.__playwright;
})();
"""


# ── Helpers ────────────────────────────────────────────────────────────────────

def _username(ig_url: str) -> str:
    """Extract Instagram username from URL."""
    if not ig_url:
        return ""
    try:
        path = urlparse(ig_url).path.strip("/")
        return path.split("/")[0] if path else ""
    except Exception:
        return ""


def _fmt(n) -> str:
    if n is None:
        return ""
    try:
        return str(int(n))
    except Exception:
        return str(n)


# ── Method 1: Instagram internal API ──────────────────────────────────────────

def _api_fetch(username: str) -> tuple[str, str]:
    """
    Fetch follower + post count via Instagram's internal web API.
    Returns (posts, followers) strings or ("", "") on failure.
    """
    url = f"https://www.instagram.com/api/v1/users/web_profile_info/?username={username}"
    headers = {
        "x-ig-app-id": "936619743392459",
        "User-Agent":   random.choice(_USER_AGENTS),
        "Accept":       "*/*",
        "Accept-Language": "en-US,en;q=0.9",
        "Referer":      f"https://www.instagram.com/{username}/",
        "X-Requested-With": "XMLHttpRequest",
    }
    try:
        sess = cf_requests.Session(impersonate=random.choice(["chrome124", "chrome136", "safari260"]))
        r = sess.get(url, headers=headers, timeout=15)
        if r.status_code == 200:
            data = r.json()
            user = data.get("data", {}).get("user", {})
            if user:
                posts     = user.get("edge_owner_to_timeline_media", {}).get("count", "")
                followers = user.get("edge_followed_by", {}).get("count", "")
                return _fmt(posts), _fmt(followers)
        if r.status_code == 401:
            log.debug(f"  API 401 (login required) for @{username}")
        elif r.status_code == 404:
            log.debug(f"  API 404 (profile not found) for @{username}")
        else:
            log.debug(f"  API status {r.status_code} for @{username}")
    except Exception as e:
        log.debug(f"  API error for @{username}: {e}")
    return "", ""


# ── Method 2: Playwright browser ──────────────────────────────────────────────

def _launch_browser():
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        log.warning("Playwright not installed — only API method available")
        return None, None, None
    os.makedirs(_PW_USER_DATA, exist_ok=True)
    _pw = sync_playwright().__enter__()
    ctx = _pw.chromium.launch_persistent_context(
        user_data_dir=_PW_USER_DATA,
        headless=False,
        slow_mo=100,
        locale="en-US",
        args=["--disable-blink-features=AutomationControlled", "--lang=en-US"],
        user_agent=random.choice(_USER_AGENTS),
    )
    try:
        ctx.add_init_script(_STEALTH_JS)
    except Exception:
        pass
    page = ctx.new_page()
    page.set_extra_http_headers({"Accept-Language": "en-US,en;q=0.9"})
    return _pw, ctx, page


def _pw_fetch(username: str, page) -> tuple[str, str]:
    """Scrape Instagram profile via Playwright. Returns (posts, followers)."""
    try:
        page.goto(f"https://www.instagram.com/{username}/",
                  timeout=30_000, wait_until="domcontentloaded")
        time.sleep(random.uniform(2, 3))

        # Check for login wall
        content = page.content()
        if "Log in" in content and "edge_followed_by" not in content:
            log.warning(f"  @{username}: login wall — need to log in to Instagram in the browser")
            input("  Log in to Instagram in the browser window, then press Enter to continue… ")
            time.sleep(2)
            content = page.content()

        # Parse counts from page JSON
        posts_str = followers_str = ""

        # Try meta tags first
        for meta in page.query_selector_all("meta[name='description']"):
            desc = meta.get_attribute("content") or ""
            m_foll = re.search(r'([\d,.KMk]+)\s*Followers', desc, re.I)
            m_post = re.search(r'([\d,.KMk]+)\s*Posts', desc, re.I)
            if m_foll:
                followers_str = m_foll.group(1).replace(",", "")
            if m_post:
                posts_str = m_post.group(1).replace(",", "")

        # Try JSON in page source
        if not followers_str:
            m = re.search(r'"edge_followed_by"\s*:\s*\{"count"\s*:\s*(\d+)', content)
            if m:
                followers_str = m.group(1)
        if not posts_str:
            m = re.search(r'"edge_owner_to_timeline_media"\s*:\s*\{"count"\s*:\s*(\d+)', content)
            if m:
                posts_str = m.group(1)

        # Try aria-labels on stat elements
        if not followers_str:
            for el in page.query_selector_all("a[href*='followers'] span, span[title]"):
                try:
                    txt = el.get_attribute("title") or el.inner_text()
                    if re.search(r'\d', txt):
                        followers_str = txt.replace(",", "").strip()
                        break
                except Exception:
                    pass

        return posts_str, followers_str

    except Exception as e:
        log.warning(f"  Playwright error for @{username}: {e}")
        return "", ""


# ── Read xlsx and collect unique IG URLs ──────────────────────────────────────

def read_ig_urls(filepath: str) -> dict:
    """
    Returns {username: [row_numbers]} mapping.
    Row numbers are 1-based xlsx row indices.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    mapping: dict = {}
    for r in range(DATA_START, (ws.max_row or 0) + 1):
        ig_url = str(ws.cell(r, C_IG_URL).value or "").strip()
        if not ig_url or "instagram.com" not in ig_url:
            continue
        uname = _username(ig_url)
        if not uname:
            continue
        mapping.setdefault(uname, []).append(r)
    wb.close()
    return mapping


# ── Write results back to xlsx ────────────────────────────────────────────────

def write_results(filepath: str, results: dict):
    """
    results = {username: (posts_str, followers_str)}
    Writes IG # Posts (col 15) and IG Followers (col 16) for every
    matching row.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    for r in range(DATA_START, (ws.max_row or 0) + 1):
        ig_url = str(ws.cell(r, C_IG_URL).value or "").strip()
        if not ig_url:
            continue
        uname = _username(ig_url)
        if uname not in results:
            continue
        posts, followers = results[uname]
        if posts:
            ws.cell(r, C_IG_POST).value = posts
        if followers:
            ws.cell(r, C_IG_FOLL).value = followers

    wb.save(filepath)
    log.info(f"Saved → {filepath}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    log.info("=" * 55)
    log.info(f"  Instagram Scraper  →  {INPUT_FILE}")
    log.info("=" * 55)

    mapping = read_ig_urls(INPUT_FILE)
    unique_users = list(mapping.keys())
    log.info(f"Found {len(unique_users)} unique Instagram profiles to scrape")

    if not unique_users:
        log.warning("No Instagram URLs found in file.")
        return

    # Launch Playwright as fallback
    _pw, ctx, pw_page = _launch_browser()

    results: dict = {}
    total = len(unique_users)

    try:
        for i, username in enumerate(unique_users, 1):
            rows = mapping[username]
            log.info(f"[{i}/{total}]  @{username}  (rows: {rows[0]}…)")

            # Method 1: API
            posts, followers = _api_fetch(username)

            # Method 2: Playwright fallback
            if (not posts and not followers) and pw_page:
                log.info(f"  API failed — trying Playwright…")
                posts, followers = _pw_fetch(username, pw_page)

            if posts or followers:
                log.info(f"  ✓  Posts={posts}  Followers={followers}")
                results[username] = (posts, followers)
            else:
                log.warning(f"  ✗  Could not retrieve data")
                results[username] = ("", "")

            # Save progress every 10
            if i % 10 == 0:
                write_results(INPUT_FILE, results)
                log.info(f"  Progress saved ({i}/{total})")

            time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))

    finally:
        if ctx:
            try:
                ctx.close()
                _pw.__exit__(None, None, None)
            except Exception:
                pass

    write_results(INPUT_FILE, results)
    found = sum(1 for p, f in results.values() if p or f)
    log.info(f"\nDone — {found}/{total} profiles scraped  →  {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
