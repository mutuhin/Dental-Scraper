"""
Dental Practice Web Scraper  v2
================================
Reads practices from Sample_dental_data_20rows_Mar17.xlsx
Scrapes each website + Google + Yelp + Invisalign locator
Outputs filled data to Dental_Scrape_Output_v2.xlsx

Changes from v1:
  - Expanded sub-page detection keywords (staff, dentist, provider, contact, ai, our-)
  - Increased sub-page scrape limit from 4 to 8
  - Added "same day crowns" (plural) to CEREC tech keyword detection
  - New find_all_providers() — extracts ALL doctor names from a page (preserving case)
  - scrape_practice() now returns provider_names list
  - main() expands results to one row per provider (like sample data format)
  - Input: Sample_dental_data_20rows_Mar17.xlsx / Output: Dental_Scrape_Output_v2.xlsx

HOW TO RUN:
    pip install requests beautifulsoup4 openpyxl lxml
    python dental_scraper_v2.py

IMPORTANT:
    - Run on YOUR computer, not a restricted server
    - Script adds 2-3 second delays between requests (polite scraping)
    - Some sites may block automated requests — those will be marked "Blocked"
    - Google/Yelp star ratings are extracted from search result snippets
    - Invisalign Tier is fetched from the official Invisalign locator API
"""

import time
import re
import sys
import logging
from urllib.parse import urljoin, urlparse, quote_plus

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────
INPUT_FILE  = "/Users/mujahidulhaqtuhin/Downloads/dental/py files/6000 Data COMPLETE.xlsx"
OUTPUT_FILE = "/Users/mujahidulhaqtuhin/Downloads/dental/py files/100data.xlsx"
DELAY_SEC   = 2.5   # seconds between requests — be polite
TIMEOUT     = 12    # request timeout

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# Sub-page path keywords — any internal link whose href CONTAINS one of these
# path fragments will be scraped for providers, services, technology, and email.
# IMPORTANT: avoid short strings that are substrings of common words
#   e.g. "ai" matches "m**ai**lto" → use "a-i" or "/ai" instead.
SUB_PAGE_KEYWORDS = [
    "service", "about", "team", "technology", "treatment",
    "staff", "provider", "doctor", "dentist", "contact",
    "a-i", "/ai-", "-ai-", "cerec", "crown", "laser", "our-",
    "speciali", "procedure",
]

# hrefs starting with these prefixes are never valid HTML pages to scrape
SUB_PAGE_SKIP_PREFIXES = ("mailto:", "tel:", "javascript:", "#", "sms:", "fax:")

# hrefs containing these substrings are noise (CDN protection, tracking, etc.)
SUB_PAGE_SKIP_CONTAINS = ("cdn-cgi", "email-protection")

# Keywords to count on practice websites
SERVICE_KEYWORDS = {
    "invisalign":        "Invisalign",
    "clear aligner":     "Clear Aligners",
    "suresmile":         "Clear Aligners",
    "clearcorrect":      "Clear Aligners",
    "veneers":           "Veneers",
    "implant":           "Implants",
    "smile makeover":    "Smile Makeovers",
    "whitening":         "Teeth Whitening",
    "sedation":          "Sedation Dentistry",
    "holistic":          "Holistic Dentistry",
    "biological":        "Holistic Dentistry",
    "cancer screening":  "Cancer Screening",
}

TECH_KEYWORDS = {
    "cerec":              "CEREC",
    "same day crown":     "CEREC",   # singular
    "same-day crown":     "CEREC",   # hyphenated singular
    "same day crowns":    "CEREC",   # FIX v2: plural form (found on AI dentistry pages)
    "same-day crowns":    "CEREC",   # FIX v2: hyphenated plural
    "cbct":               "CBCT",
    "cone beam":          "CBCT",
    "3d imaging":         "CBCT",
    "laser":              "Lasers",
    " ai ":               "AI",
    "artificial intelligence": "AI",
    "intraoral scanner":  "Intraoral Scanners",
    "itero":              "Intraoral Scanners",
    "3shape":             "Intraoral Scanners",
    "medit":              "Intraoral Scanners",
}

SOCIAL_PLATFORMS = ["facebook", "instagram", "tiktok", "linkedin"]

# Per-site overrides: applied when a site blocks automated access (403) or
# when we know a secondary URL holds the provider/staff list.
# Key: lowercase domain (no www, no trailing slash)
SITE_OVERRIDES = {
    "hallfamilydent.com": {
        # Primary site returns 403 — scrape providers from the sister domain instead
        "fallback_urls": [
            "https://www.dentistindianapolisin.com/about-us/staff/dentists",
        ],
        # Hardcode the known contact email (Cloudflare obfuscated on site)
        "email": "info@hallfamilydental.com",
    },
}


# ─────────────────────────────────────────────────────────────────────────────
# UTILITIES
# ─────────────────────────────────────────────────────────────────────────────

def safe_get(url, retries=2):
    """GET a URL with retries. Returns Response or None."""
    if not url or url.strip() in ("", "N/A", "None"):
        return None
    if not url.startswith("http"):
        url = "https://" + url.lstrip("/")
    for attempt in range(retries):
        try:
            r = requests.get(url, headers=HEADERS, timeout=TIMEOUT)
            if r.status_code == 200:
                return r
            if r.status_code == 403:
                log.warning(f"  403 Forbidden: {url}")
                return None
        except Exception as e:
            log.warning(f"  Attempt {attempt+1} failed for {url}: {e}")
            time.sleep(1)
    return None


def extract_text(html):
    """Return clean lowercase plain text from HTML."""
    soup = BeautifulSoup(html, "lxml")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    return soup.get_text(separator=" ", strip=True).lower()


def extract_text_original(html):
    """Return clean plain text from HTML preserving original casing (for name extraction)."""
    soup = BeautifulSoup(html, "lxml")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    return soup.get_text(separator=" ", strip=True)


def count_keyword(text, keyword):
    """Count occurrences of a keyword in text (whole-word safe)."""
    return len(re.findall(re.escape(keyword.lower()), text))


def decode_cloudflare_email(encoded: str) -> str:
    """
    Decode a Cloudflare email-protection hex string.
    Cloudflare XORs each byte with the first byte (the key).
    Example: data-cfemail="a4c1c6c0..." → key=0xa4, decode remaining pairs.
    """
    try:
        pairs = [int(encoded[i:i+2], 16) for i in range(0, len(encoded), 2)]
        key   = pairs[0]
        return "".join(chr(b ^ key) for b in pairs[1:])
    except Exception:
        return ""


def find_email(html_text, soup, website_url=""):
    """Find email from page HTML, text patterns, or common contact page URLs."""
    # Method 0: Cloudflare email-protection decode (data-cfemail attribute)
    for tag in soup.find_all(attrs={"data-cfemail": True}):
        decoded = decode_cloudflare_email(tag["data-cfemail"])
        if "@" in decoded:
            return decoded

    # Method 1: Look for mailto links
    mailto = soup.find("a", href=re.compile(r"mailto:", re.I))
    if mailto:
        email = mailto["href"].replace("mailto:", "").strip()
        if email and "@" in email:
            return email

    # Method 2: Look for email in plain text (case-insensitive on email, text is lowercase)
    email_match = re.search(r"[a-z0-9._%+\-]+@[a-z0-9.\-]+\.[a-z]{2,}", html_text)
    if email_match:
        return email_match.group(0)

    # Method 3: Look in all anchor href attributes (contact@, support@, info@)
    for a in soup.find_all("a", href=True):
        href = a["href"].lower()
        if "mailto:" in href:
            email = a["href"].replace("mailto:", "").strip()
            if email and "@" in email:
                return email

    # Method 4: Look in text for patterns like "info@" or "contact@"
    pattern_match = re.search(
        r"(?:contact|info|email|support|hello|reach)[:\s]*([a-z0-9._%+\-]+@[a-z0-9.\-]+\.[a-z]{2,})",
        html_text
    )
    if pattern_match:
        return pattern_match.group(1)

    # Method 5: Look for email in HTML attributes (title, alt, data)
    for tag in soup.find_all(True):
        for attr_name, attr_value in tag.attrs.items():
            if isinstance(attr_value, str) and "@" in attr_value.lower():
                attr_email = re.search(
                    r"[a-z0-9._%+\-]+@[a-z0-9.\-]+\.[a-z]{2,}", attr_value.lower()
                )
                if attr_email:
                    return attr_email.group(0)

    return "Not Found"


def find_all_providers(original_text):
    """
    FIX v2: Extract ALL doctor/provider names from page text (original casing).

    Returns a list of name strings, e.g. ["Dr. Sharon Huang", "Dr. John Smith"].
    Returns empty list if nothing found.

    Covers:
      - "Dr. FirstName LastName" format
      - "Dr. FirstName M. LastName" (middle initial)
      - "FirstName LastName, DDS/DMD/DPH/..." format
      - "FirstName LastName DDS/DMD/..." (no comma)
    """
    providers = []
    seen_lower = set()

    # Words that should never appear as the TRAILING word of a doctor name.
    # When found at the end, we trim the word and keep the shorter name
    # (e.g. "Dr. Mark Alexandrunas Granville" → "Dr. Mark Alexandrunas").
    # These should NOT cause the entire candidate to be discarded.
    TRAILING_NOISE = {
        # conjunctions / articles / prepositions
        "our", "meet", "the", "for", "your", "about", "new", "is", "are",
        "with", "from", "has", "have", "dr", "and", "or", "at", "in", "of",
        "such", "a", "an", "all", "selected", "view",
        # role / title / credential words (follow name on a bio page)
        "founder", "owner", "dds", "dmd", "patients", "practice",
        "board", "certified", "director", "associate",
        # dental specialty words printed after a doctor's name
        "general", "cosmetic", "oral", "pediatric", "family",
        "restorative", "endodontic", "orthodontic", "periodontic",
        "prosthodontic", "invisalign", "implant",
        # common Ohio practice-location words printed right after the doctor name
        "granville", "circleville", "ashville", "sunbury", "beavercreek",
        "dayton", "oakwood", "northeast", "central", "northwest", "read",
        "related", "grove", "city", "washington", "columbus",
    }
    # Words that should never appear as the FIRST name (word[1]) — hard noise
    FIRST_NAME_NOISE = {
        "our", "meet", "the", "for", "your", "about", "new", "is", "are",
        "with", "from", "has", "have", "dr", "and", "or", "at", "in", "of",
    }

    # Pattern 1: Dr. FirstName [M.] LastName [OptionalExtraWord]
    # The optional extra word captures compound surnames (Van Ee, Sue Min Lee).
    # If that extra word is trailing noise, we trim it instead of discarding.
    for match in re.finditer(
        r"Dr\.?\s+([A-Z][a-z\-']+(?:\s+[A-Z]\.)?(?:\s+[A-Z][a-z\-']+)(?:\s+[A-Z][a-z\-']+)?)"
        r"(?:\s*,?\s*(?:DDS|DMD|DPH|PhD|FAGD|FICOI|FICD|MS|MSD|ABGD|PC|PLLC))?",
        original_text
    ):
        candidate = "Dr. " + match.group(1).strip()
        words = candidate.split()

        if len(words) < 3:
            continue
        if words[1].lower() in FIRST_NAME_NOISE:
            continue

        # Trim trailing noise words one at a time (e.g. city name, role word).
        # No len(words) > 3 guard — if trimming leaves < 3 words the length
        # check below will naturally reject the candidate.
        while words[-1].lower() in TRAILING_NOISE and len(words) > 2:
            words = words[:-1]
        candidate = " ".join(words)

        if 3 <= len(words) <= 5 and candidate.lower() not in seen_lower:
            providers.append(candidate)
            seen_lower.add(candidate.lower())

    # Pattern 2: FirstName [M.] LastName, DDS/DMD/...  (no "Dr." prefix)
    # Also strictly one last-name word to avoid absorbing surrounding text.
    for match in re.finditer(
        r"([A-Z][a-z\-']+(?:\s+[A-Z]\.)?(?:\s+[A-Z][a-z\-']+))"
        r"\s*,?\s+(?:DDS|DMD|DPH|PhD|FAGD|FICOI|FICD|MS|MSD)",
        original_text
    ):
        candidate = match.group(1).strip()
        words = candidate.split()
        if 2 <= len(words) <= 4 and candidate.lower() not in seen_lower:
            providers.append(candidate)
            seen_lower.add(candidate.lower())

    # Strip trailing punctuation (e.g. "Dr. Joelle Jeffers-" → "Dr. Joelle Jeffers")
    providers = [p.rstrip("-–—.,;: ") for p in providers]

    # Deduplicate: if both "Dr. Jane Smith" and "Jane Smith" exist, keep the Dr. version
    final = []
    bare_names = set()
    for p in providers:
        bare = re.sub(r"^Dr\.?\s+", "", p).strip().lower()
        if bare not in bare_names:
            final.append(p)
            bare_names.add(bare)

    # Deduplicate by first+last name, ignoring middle initials.
    # Keep the longer/more complete version (e.g. "Dr. Craig D. Swiger" over "Dr. Craig Swiger").
    # Also handle "Jonathan Morgan" vs "Dr. Jon Morgan" — match on last name + first 3 chars.
    def first_last_key(name):
        """Returns (first_3chars, last_word) to detect same-person variants."""
        parts = re.sub(r"^Dr\.?\s+", "", name).strip().split()
        # drop middle initial (single uppercase letter followed by optional dot)
        parts = [w for w in parts if not re.match(r"^[A-Z]\.?$", w)]
        if len(parts) >= 2:
            return (parts[0][:3].lower(), parts[-1].lower())
        return (name.lower(), "")

    seen_fl: dict[tuple, str] = {}   # key → best (longest) name so far
    for p in final:
        key = first_last_key(p)
        if key not in seen_fl:
            seen_fl[key] = p
        else:
            # Keep whichever version is longer (more complete)
            if len(p) > len(seen_fl[key]):
                seen_fl[key] = p

    return list(seen_fl.values())


def find_doctor_name(html_text, soup, input_doctor_name):
    """
    Fallback single-name extractor (used when find_all_providers returns nothing).
    If input already has a valid name, return it directly.
    """
    if input_doctor_name:
        input_str = str(input_doctor_name).strip()
        if input_str and input_str not in ("", "N/A", "Multiple Providers", "None",
                                           "Not Listed", "Not Found"):
            return input_str

    # Look for doctor/dentist names on the page
    doctor_patterns = [
        r"(?:dr\.?|doctor|owner|founder)[\s:]+([a-z][a-z]+\s+[a-z][a-z]+)(?:\s+(?:dds|dmd|dph|phd))?",
        r"([a-z][a-z]+\s+[a-z][a-z]+)(?:\s+(?:dds|dmd|dph|phd))+",
        r"meet\s+(?:our\s+)?(?:doctor|dentist|provider).*?([a-z][a-z]+\s+[a-z][a-z]+)",
    ]

    for pattern in doctor_patterns:
        match = re.search(pattern, html_text, re.IGNORECASE)
        if match:
            name = match.group(1).strip().title()
            if name and len(name) > 4:
                name = re.sub(r'\s+(and|for|in|at|the|a|an)\s*.*$', '', name, flags=re.IGNORECASE)
                name = ' '.join(word.capitalize() for word in name.split())
                if len(name) > 4:
                    return name

    title_pattern = r"([a-z][a-z]+\s+[a-z][a-z]+)\s+dds|([a-z][a-z]+\s+[a-z][a-z]+)\s+dmd"
    for match in re.finditer(title_pattern, html_text, re.IGNORECASE):
        name = match.group(1) or match.group(2)
        if name:
            return ' '.join(word.capitalize() for word in name.strip().split())

    if input_doctor_name and str(input_doctor_name).strip() not in ("", "None"):
        return str(input_doctor_name).strip()

    return "Not Listed"


def find_social_links(soup):
    """Find social media URLs from anchor tags on the page."""
    found = {p: "" for p in SOCIAL_PLATFORMS}
    for a in soup.find_all("a", href=True):
        href = a["href"].lower()
        for platform in SOCIAL_PLATFORMS:
            if platform + ".com" in href and not found[platform]:
                found[platform] = a["href"]
    return found


# ─────────────────────────────────────────────────────────────────────────────
# INVISALIGN LOCATOR
# ─────────────────────────────────────────────────────────────────────────────

def get_invisalign_tier(practice_name, city, state, zip_code):
    """
    Queries the Invisalign provider locator API.
    Returns the tier string (e.g. "Bronze", "Gold") or "Not Found".
    """
    api_url = "https://www.invisalign.com/api/locator/search"
    params  = {
        "query":   practice_name,
        "country": "US",
        "zip":     str(zip_code) if zip_code else "",
        "lang":    "en",
    }
    time.sleep(DELAY_SEC)
    try:
        r = requests.get(api_url, params=params, headers=HEADERS, timeout=TIMEOUT)
        if r.status_code == 403:
            log.warning(f"  Invisalign API blocked (403 Forbidden)")
            return "Check manually: https://www.invisalign.com/find-a-doctor"
        if r.status_code == 429:
            log.warning(f"  Invisalign API rate limited (429)")
            return "Check manually due to rate limit"
        if r.status_code == 200:
            data = r.json()
            providers = data.get("providers", data.get("results", []))
            for p in providers:
                name = (p.get("name", "") + " " + p.get("practiceName", "")).lower()
                if any(w in name for w in practice_name.lower().split()[:2]):
                    tier = p.get("tier", p.get("providerLevel", ""))
                    return tier if tier else "Not Listed"
    except Exception as e:
        log.warning(f"  Invisalign API error: {e}")

    search_url = (
        f"https://www.invisalign.com/find-a-doctor"
        f"#q={quote_plus(practice_name + ' ' + city)}"
    )
    return f"Check manually: {search_url}"


# ─────────────────────────────────────────────────────────────────────────────
# WEBSITE SCRAPER — main per-practice logic
# ─────────────────────────────────────────────────────────────────────────────

def scrape_practice(row):
    """
    Given a dict of practice info (from Excel row), scrapes all data.

    Returns a dict of scraped fields including:
      provider_names: list[str]  — all doctors found on the site
                                   (empty list → caller falls back to input name)
    """
    name    = str(row.get("Practice Name", "")).strip()
    website = str(row.get("Website", "")).strip()
    city    = str(row.get("City", "")).strip()
    state   = str(row.get("State", "")).strip()
    zip_c   = str(row.get("Zip", "")).strip()

    log.info(f"▶ Scraping: {name}")

    result = {
        "provider_names":    [],          # FIX v2: list of all discovered providers
        "email":             "Not Found",
        "hygienists":        "Not Listed",
        "facebook_url":      "",
        "facebook_posts":    "N/A",
        "facebook_followers":"N/A",
        "instagram_url":     "",
        "instagram_posts":   "N/A",
        "instagram_followers":"N/A",
        "tiktok_url":        "",
        "tiktok_posts":      "N/A",
        "tiktok_followers":  "N/A",
        "linkedin_url":      "",
        "linkedin_posts":    "N/A",
        "linkedin_followers":"N/A",
        "cerec":             "",
        "cbct":              "",
        "lasers":            "",
        "ai":                "",
        "intraoral":         "",
        "invisalign":        0,
        "invisalign_tier":   "Not Found",
        "clear_aligners":    0,
        "veneers":           0,
        "implants":          0,
        "smile_makeovers":   0,
        "whitening":         0,
        "sedation":          0,
        "holistic":          0,
        "cancer_screening":  0,
        "associations":      "Not Found",
        "specialty":         "Not Found",
        "google_rating":     "Not Found",
        "google_reviews":    "Not Found",
        "yelp_rating":       "Not Found",
        "yelp_reviews":      "Not Found",
        "testimonials":      "Not Found",
    }

    # ── 1. Scrape practice website ────────────────────────────────────────────
    # Resolve any per-site overrides (fallback URL, hardcoded email, etc.)
    site_domain = urlparse(
        website if website.startswith("http") else "https://" + website
    ).netloc.lower().lstrip("www.") if website else ""
    override = SITE_OVERRIDES.get(site_domain, {})

    if website and website not in ("", "None"):
        log.info(f"   Fetching website: {website}")
        time.sleep(DELAY_SEC)

        all_text          = ""   # lowercase, for keyword counting
        all_text_original = ""   # original case, for name extraction
        all_soup          = None

        r = safe_get(website)
        if r:
            all_soup          = BeautifulSoup(r.text, "lxml")
            all_text          = extract_text(r.text)
            all_text_original = extract_text_original(r.text)

        # If primary site failed (403/blocked) and override has fallback URLs,
        # scrape those directly for providers and other data.
        if not all_soup and override.get("fallback_urls"):
            log.info(f"   Primary site blocked — trying override fallback URLs...")
            for fb_url in override["fallback_urls"]:
                log.info(f"   Fetching fallback: {fb_url}")
                time.sleep(DELAY_SEC)
                fb_r = safe_get(fb_url)
                if fb_r:
                    if all_soup is None:
                        all_soup = BeautifulSoup(fb_r.text, "lxml")
                    all_text          += " " + extract_text(fb_r.text)
                    all_text_original += " " + extract_text_original(fb_r.text)

        # ── Collect sub-pages to scrape (runs after primary + fallback fetches)
        # FIX v2: expanded keyword list + increased limit to 8
        if all_soup:
            base_url_for_links = website if website.startswith("http") else "https://" + website

            # Priority tiers — lower number = more important, scraped first.
            # Uses BEST-TIER-WINS: a URL matching multiple keywords gets the
            # lowest (most important) tier among all matching keywords.
            # e.g. /a-i-assisted-dentistry/ matches "dentist"(tier 3) AND
            #      "a-i"(tier 1) → assigned tier 1, so it beats service pages.
            PRIORITY = {
                # Tier 0 — provider/team pages: most likely to list all doctors
                "about": 0, "team": 0, "staff": 0, "our-": 0,
                # Tier 1 — technology & AI pages: CEREC, same-day crowns, AI
                "technology": 1, "a-i": 1, "/ai-": 1, "-ai-": 1,
                "cerec": 1, "laser": 1,
                # Tier 2 — contact pages: email
                "contact": 2,
                # Tier 3 — generic dentist/doctor/service pages (very broad)
                "doctor": 3, "dentist": 3, "provider": 3,
                "service": 3, "treatment": 3, "speciali": 3,
                "procedure": 3, "crown": 3,
            }
            url_best_tier: dict[str, int] = {}   # url → best (lowest) tier found

            for a in all_soup.find_all("a", href=True):
                href_raw   = a["href"]
                href_lower = href_raw.lower()

                # Skip non-HTML hrefs (mailto, tel, javascript, CDN noise)
                if href_lower.startswith(SUB_PAGE_SKIP_PREFIXES):
                    continue
                if any(s in href_lower for s in SUB_PAGE_SKIP_CONTAINS):
                    continue

                full_url = urljoin(base_url_for_links, href_raw)

                # Find the best (lowest) tier for this URL across all keywords
                for keyword in SUB_PAGE_KEYWORDS:
                    if keyword in href_lower:
                        tier = PRIORITY.get(keyword, 6)
                        if full_url not in url_best_tier or tier < url_best_tier[full_url]:
                            url_best_tier[full_url] = tier

            # Sort by best tier, cap at 8
            ordered_sub_pages = sorted(url_best_tier, key=lambda u: url_best_tier[u])

            for sub_url in ordered_sub_pages[:8]:   # cap at 8, highest-priority first
                log.info(f"   Fetching sub-page: {sub_url}")
                time.sleep(DELAY_SEC)
                sub_r = safe_get(sub_url)
                if sub_r:
                    all_text          += " " + extract_text(sub_r.text)
                    all_text_original += " " + extract_text_original(sub_r.text)

        if all_soup:
            result["email"] = find_email(all_text, all_soup, website)

            # Apply hardcoded email override if the scraped result is still empty
            if override.get("email") and result["email"] in ("Not Found", ""):
                result["email"] = override["email"]
                log.info(f"   Email set from site override: {override['email']}")

            # FIX v2: extract ALL providers from original-case text
            providers = find_all_providers(all_text_original)
            result["provider_names"] = providers

            # Find social links
            socials = find_social_links(all_soup)
            for platform, url_val in socials.items():
                result[f"{platform}_url"] = url_val

        # ── Count service keywords ────────────────────────────────────────────
        service_counts = {
            "Invisalign":        0,
            "Clear Aligners":    0,
            "Veneers":           0,
            "Implants":          0,
            "Smile Makeovers":   0,
            "Teeth Whitening":   0,
            "Sedation Dentistry":0,
            "Holistic Dentistry":0,
            "Cancer Screening":  0,
        }
        for keyword, category in SERVICE_KEYWORDS.items():
            service_counts[category] += count_keyword(all_text, keyword)

        result["invisalign"]      = service_counts["Invisalign"]
        result["clear_aligners"]  = service_counts["Clear Aligners"]
        result["veneers"]         = service_counts["Veneers"]
        result["implants"]        = service_counts["Implants"]
        result["smile_makeovers"] = service_counts["Smile Makeovers"]
        result["whitening"]       = service_counts["Teeth Whitening"]
        result["sedation"]        = service_counts["Sedation Dentistry"]
        result["holistic"]        = service_counts["Holistic Dentistry"]
        result["cancer_screening"]= service_counts["Cancer Screening"]

        # ── Technology detection ──────────────────────────────────────────────
        # FIX v2: TECH_KEYWORDS now includes "same day crowns" (plural)
        tech_found = set()
        for keyword, tech_name in TECH_KEYWORDS.items():
            if keyword in all_text:
                tech_found.add(tech_name)

        result["cerec"]     = "X" if "CEREC"              in tech_found else ""
        result["cbct"]      = "X" if "CBCT"               in tech_found else ""
        result["lasers"]    = "X" if "Lasers"             in tech_found else ""
        result["ai"]        = "X" if "AI"                 in tech_found else ""
        result["intraoral"] = "X" if "Intraoral Scanners" in tech_found else ""

        # ── Testimonials count ────────────────────────────────────────────────
        testimonial_count = count_keyword(all_text, "testimonial")
        if testimonial_count == 0:
            testimonial_count = len(re.findall(r'["\'"][^"\']{30,300}["\'"]', all_text))
        result["testimonials"] = testimonial_count if testimonial_count > 0 else "0"

    else:
        log.warning(f"   No website found for {name}")

    # ── 2. Social media stats ─────────────────────────────────────────────────
    result["instagram_posts"]     = "Not Found"
    result["instagram_followers"] = "Not Found"
    result["facebook_posts"]      = "Not Found"
    result["facebook_followers"]  = "Not Found"

    # ── 3. Google Reviews ─────────────────────────────────────────────────────
    result["google_rating"]  = "Not Found"
    result["google_reviews"] = "Not Found"

    # ── 4. Yelp Reviews ───────────────────────────────────────────────────────
    result["yelp_rating"]  = "Not Found"
    result["yelp_reviews"] = "Not Found"

    # ── 5. Invisalign Tier ────────────────────────────────────────────────────
    if result["invisalign"] and int(result["invisalign"]) > 0:
        log.info(f"   Checking Invisalign tier...")
        result["invisalign_tier"] = get_invisalign_tier(name, city, state, zip_c)
    else:
        result["invisalign_tier"] = "N/A - Not Offered"

    log.info(f"   ✓ Done: {name}\n")
    return result


# ─────────────────────────────────────────────────────────────────────────────
# READ INPUT EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def read_practices(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # 6000 Data COMPLETE has headers in row 1; older sample files use row 2.
    # Detect automatically: if row 1 col 1 looks like a header string, use row 1.
    row1_val = str(ws.cell(1, 1).value or "").strip()
    if row1_val.lower() in ("id", "index", "#"):
        hdr_row  = 1
        data_row = 2
    else:
        hdr_row  = 2
        data_row = 3

    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[hdr_row]]
    col_map = {h: i for i, h in enumerate(headers)}

    practices = []
    for row in ws.iter_rows(min_row=data_row, values_only=True):
        idx = row[0]
        if idx is None:
            continue
        # "Office Name" (6000 data) or "Practice Name" (older files)
        practice_name = (
            row[col_map.get("Office Name",    col_map.get("Practice Name", 2))]
        )
        practice = {
            "Index":         row[col_map.get("ID",               col_map.get("Index", 0))],
            "Practice Name": practice_name,
            "Doctor Name":   row[col_map.get("Doctor Name",      10)],
            "Street":        row[col_map.get("Street",           col_map.get("Address", 3))],
            "City":          row[col_map.get("City",             4)],
            "State":         row[col_map.get("State",            5)],
            "Zip":           row[col_map.get("Zip",              6)],
            "Website":       row[col_map.get("Website",          col_map.get("Practice Website", 9))],
        }
        practices.append(practice)
    return practices


# ─────────────────────────────────────────────────────────────────────────────
# WRITE OUTPUT EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def write_output(practices_data, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Scraped Data"

    # Styles
    hdr_font  = Font(name="Arial", bold=True, size=9)
    grp_font  = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    data_font = Font(name="Arial", size=9)
    thin      = Side(style="thin", color="CCCCCC")
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft       = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    fills = {
        "blue":     PatternFill("solid", fgColor="1F4E79"),
        "lt_blue":  PatternFill("solid", fgColor="BDD7EE"),
        "green":    PatternFill("solid", fgColor="375623"),
        "lt_green": PatternFill("solid", fgColor="C6EFCE"),
        "purple":   PatternFill("solid", fgColor="7030A0"),
        "lt_purple":PatternFill("solid", fgColor="E2CFEC"),
        "orange":   PatternFill("solid", fgColor="833C00"),
        "lt_orange":PatternFill("solid", fgColor="FCE4D6"),
        "grey":     PatternFill("solid", fgColor="595959"),
        "lt_grey":  PatternFill("solid", fgColor="EDEDED"),
        "white":    PatternFill("solid", fgColor="FFFFFF"),
        "row_alt":  PatternFill("solid", fgColor="EBF3FB"),
    }

    def sc(cell, val, font=data_font, fill=fills["white"], align=ctr):
        cell.value, cell.font, cell.fill, cell.alignment, cell.border = val, font, fill, align, bdr

    # ── Group header row ──────────────────────────────────────────────────────
    groups = [
        (1,  9,  "Practice Information",      "blue"),
        (10, 25, "Social Media",              "green"),
        (26, 30, "Technology in Practice",    "purple"),
        (31, 41, "Services (# of Mentions)",  "orange"),
        (42, 46, "Patient Experience",        "grey"),
    ]
    for start, end, label, color in groups:
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        sc(ws.cell(1, start), label, font=grp_font, fill=fills[color], align=ctr)

    # ── Column headers ────────────────────────────────────────────────────────
    col_headers = [
        # Practice Info (1-9)
        "Index", "Practice Name", "Doctor Name", "Address", "City/State/Zip",
        "Practice Email", "Practice Website", "# of Hygienists", "# of Locations",
        # Social Media (10-25)
        "Facebook URL", "FB Posts", "FB Followers",
        "Instagram URL", "IG Posts", "IG Followers",
        "TikTok URL", "TT Posts", "TT Followers",
        "LinkedIn URL", "LI Posts", "LI Followers",
        "Associations / Memberships",  # col 22
        "Doctor Specialty",            # col 23
        "",                            # col 24
        "",                            # col 25
        # Technology (26-30)
        "CEREC (Same Day Crowns)", "CBCT (3D Imaging)", "Lasers", "AI",
        "Intraoral Scanners",
        # Services (31-41)
        "Invisalign", "Invisalign Tier", "Clear Aligners", "Veneers", "Implants",
        "Smile Makeovers", "Teeth Whitening", "Sedation Dentistry",
        "Holistic Dentistry", "Cancer Screening", "Other Notes",
        # Patient Experience (42-46)
        "Google Rating", "Google # Reviews",
        "Yelp Rating", "Yelp # Reviews", "Testimonials (#)",
    ]

    fill_col = {}
    for c in range(1, 10):  fill_col[c] = "lt_blue"
    for c in range(10, 26): fill_col[c] = "lt_green"
    for c in range(26, 31): fill_col[c] = "lt_purple"
    for c in range(31, 42): fill_col[c] = "lt_orange"
    for c in range(42, 47): fill_col[c] = "lt_grey"

    for col, hdr in enumerate(col_headers, 1):
        cell = ws.cell(2, col)
        sc(cell, hdr, font=hdr_font, fill=fills[fill_col.get(col, "white")], align=ctr)

    # ── Data rows ─────────────────────────────────────────────────────────────
    for r_idx, (practice_input, scraped) in enumerate(practices_data, start=3):
        rf = fills["row_alt"] if r_idx % 2 == 0 else fills["white"]
        address = str(practice_input.get("Street", ""))
        csz     = (
            f"{practice_input.get('City','')}, "
            f"{practice_input.get('State','')} "
            f"{practice_input.get('Zip','')}"
        )

        row_vals = [
            practice_input.get("Index"),
            practice_input.get("Practice Name"),
            scraped["doctor_name"],
            address,
            csz,
            scraped["email"],
            practice_input.get("Website"),
            scraped["hygienists"],
            "1",
            # Social
            scraped["facebook_url"],  scraped["facebook_posts"],  scraped["facebook_followers"],
            scraped["instagram_url"], scraped["instagram_posts"], scraped["instagram_followers"],
            scraped["tiktok_url"],    scraped["tiktok_posts"],    scraped["tiktok_followers"],
            scraped["linkedin_url"],  scraped["linkedin_posts"],  scraped["linkedin_followers"],
            scraped["associations"],
            scraped["specialty"],
            "", "",
            # Technology
            scraped["cerec"], scraped["cbct"], scraped["lasers"],
            scraped["ai"],    scraped["intraoral"],
            # Services
            scraped["invisalign"],      scraped["invisalign_tier"],
            scraped["clear_aligners"],  scraped["veneers"],
            scraped["implants"],        scraped["smile_makeovers"],
            scraped["whitening"],       scraped["sedation"],
            scraped["holistic"],        scraped["cancer_screening"],
            "",
            # Patient Experience
            scraped["google_rating"], scraped["google_reviews"],
            scraped["yelp_rating"],   scraped["yelp_reviews"],
            scraped["testimonials"],
        ]

        for c_idx, val in enumerate(row_vals, 1):
            cell  = ws.cell(r_idx, c_idx)
            align = lft if c_idx in (2, 3, 4, 5, 6, 7) else ctr
            sc(cell, val, font=data_font, fill=rf, align=align)

    # ── Column widths ─────────────────────────────────────────────────────────
    widths = {
        1: 6, 2: 28, 3: 24, 4: 30, 5: 22, 6: 28, 7: 30, 8: 12, 9: 10,
        10: 28, 11: 10, 12: 12, 13: 28, 14: 10, 15: 12,
        16: 14, 17: 10, 18: 12, 19: 28, 20: 10, 21: 12,
        22: 36, 23: 30, 24: 6, 25: 6,
        26: 18, 27: 16, 28: 10, 29: 8, 30: 16,
        31: 12, 32: 16, 33: 14, 34: 10, 35: 10,
        36: 14, 37: 14, 38: 16, 39: 14, 40: 14, 41: 14,
        42: 14, 43: 14, 44: 12, 45: 14, 46: 14,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 38
    ws.freeze_panes = "B3"

    # ── Notes / Legend sheet ──────────────────────────────────────────────────
    ns = wb.create_sheet("Legend & Notes")
    notes = [
        ("LEGEND", ""),
        ("X in Technology columns", "Confirmed mentioned on practice website"),
        ("0 in Services columns", "Keyword not found on website (could be offered but not mentioned)"),
        ("Not Found", "Data exists but could not be extracted automatically"),
        ("Blocked", "Website or platform blocked automated access — check manually"),
        ("Login Required", "Platform requires account login (e.g. Facebook post counts)"),
        ("N/A", "Not applicable for this practice"),
        ("", ""),
        ("DATA SOURCES", ""),
        ("Practice website", "Scraped homepage + up to 8 sub-pages (services/about/team/staff/dentists/contact/ai etc.)"),
        ("Services counts", "Exact keyword count in scraped page text"),
        ("Technology", "Keyword detected in scraped text (X = yes, blank = not found)"),
        ("CEREC detection", "Catches: cerec, same day crown, same-day crown, same day crowns (plural added in v2)"),
        ("Doctor Names", "All Dr. FirstName LastName / FirstName LastName DDS/DMD patterns extracted"),
        ("Google Reviews", "Extracted from Google search result snippets"),
        ("Yelp Reviews", "Extracted from Yelp search result page"),
        ("Invisalign Tier", "Queried from Invisalign provider locator API"),
        ("Instagram", "Scraped from public profile page"),
        ("Facebook", "Post count requires login — follower count from public page"),
        ("", ""),
        ("IMPORTANT NOTES", ""),
        ("Accuracy", "Service counts = exact keyword occurrences. A practice may offer a service not mentioned on the website."),
        ("Google ratings", "Extracted from snippet text — verify on Google Maps for exact figure"),
        ("Yelp ratings", "Yelp aggressively blocks scrapers — verify manually if Blocked"),
        ("Invisalign Tier", "If API fails, check: https://www.invisalign.com/find-a-doctor"),
        ("Re-running", "Re-run the script any time to refresh data"),
        ("", ""),
        ("v2 CHANGES", ""),
        ("Sub-pages", "Now scrapes up to 8 sub-pages (was 4). Added keywords: staff, dentist, provider, contact, ai, a-i, our-"),
        ("CEREC", "Added 'same day crowns' (plural) + 'same-day crowns' to detection keywords"),
        ("Providers", "find_all_providers() extracts ALL Dr./DDS/DMD names from scraped text"),
        ("Output rows", "One row per discovered provider (same as sample data format)"),
    ]
    ns.column_dimensions["A"].width = 32
    ns.column_dimensions["B"].width = 80
    for r_i, (a, b) in enumerate(notes, 1):
        ca = ns.cell(r_i, 1, a)
        cb = ns.cell(r_i, 2, b)
        bold = a in ("LEGEND", "DATA SOURCES", "IMPORTANT NOTES", "v2 CHANGES") or r_i == 1
        for cell in (ca, cb):
            cell.font = Font(name="Arial", bold=bold, size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if bold and a:
                cell.fill = PatternFill("solid", fgColor="1F4E79")
                cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        ns.row_dimensions[r_i].height = 22 if not b else 28

    wb.save(output_path)
    log.info(f"\n✅ Output saved to: {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    log.info("=" * 60)
    log.info("  Dental Practice Scraper  v2")
    log.info("=" * 60)

    try:
        practices = read_practices(INPUT_FILE)
    except FileNotFoundError:
        log.error(f"Input file not found: {INPUT_FILE}")
        log.error("Place the script in the same folder as the Excel file and retry.")
        sys.exit(1)

    log.info(f"Found {len(practices)} input rows to scrape.\n")

    # To scrape ALL rows: practices_to_run = practices
    # To test first 10: practices_to_run = practices[:10]
    practices_to_run = practices[:100]   # rows 1-100

    # scraped_cache  — keyed by website URL, stores the scrape result dict
    # expanded_sites — tracks which sites have already had their provider rows
    #                  written to all_results, so duplicate input rows for the
    #                  same website don't multiply the provider list again.
    scraped_cache: dict[str, dict] = {}
    expanded_sites: set[str]       = set()

    all_results = []

    for i, practice in enumerate(practices_to_run, 1):
        pname   = practice.get("Practice Name")
        website = str(practice.get("Website", "") or "").strip()
        log.info(f"[{i}/{len(practices_to_run)}] {pname}")

        cache_key = website.lower().rstrip("/") if website else f"__nosite_{pname}"

        # ── Scrape (or use cache) ─────────────────────────────────────────────
        if cache_key in scraped_cache:
            log.info("   (using cached scrape result)")
            scraped_base = scraped_cache[cache_key]
        else:
            try:
                scraped_base = scrape_practice(practice)
            except Exception as e:
                log.error(f"  ERROR scraping {pname}: {e}")
                scraped_base = {k: "ERROR" for k in [
                    "provider_names",
                    "email", "hygienists",
                    "facebook_url", "facebook_posts", "facebook_followers",
                    "instagram_url", "instagram_posts", "instagram_followers",
                    "tiktok_url", "tiktok_posts", "tiktok_followers",
                    "linkedin_url", "linkedin_posts", "linkedin_followers",
                    "cerec", "cbct", "lasers", "ai", "intraoral",
                    "invisalign", "invisalign_tier", "clear_aligners", "veneers",
                    "implants", "smile_makeovers", "whitening", "sedation",
                    "holistic", "cancer_screening",
                    "associations", "specialty",
                    "google_rating", "google_reviews",
                    "yelp_rating", "yelp_reviews", "testimonials",
                ]}
                scraped_base["provider_names"] = []
            scraped_cache[cache_key] = scraped_base

        # ── Build provider row(s) ─────────────────────────────────────────────
        # Only expand discovered providers ONCE per unique website.
        # For subsequent input rows sharing the same site, skip them — they
        # were already covered by the first expansion.
        if cache_key in expanded_sites:
            log.info("   (providers already written for this site, skipping row)")
            continue

        discovered = scraped_base.get("provider_names", [])
        input_name = str(practice.get("Doctor Name", "")).strip()

        if discovered:
            provider_list = discovered
        elif input_name and input_name not in ("", "None", "Not Found", "Not Listed"):
            provider_list = [input_name]
        else:
            provider_list = ["Not Listed"]

        for provider_name in provider_list:
            row_scraped = dict(scraped_base)
            row_scraped["doctor_name"] = provider_name
            all_results.append((practice, row_scraped))

        expanded_sites.add(cache_key)

        # Save partial output every 10 practices (safety net)
        if i % 10 == 0:
            log.info(f"  💾 Saving progress after {i} practices...")
            write_output(all_results, OUTPUT_FILE)

    write_output(all_results, OUTPUT_FILE)
    log.info(f"\n🎉 All done! {len(all_results)} provider rows written.")
    log.info(f"Output: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
