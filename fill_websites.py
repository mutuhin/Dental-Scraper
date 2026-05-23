#!/usr/bin/env python3
"""
fill_websites.py
================
Fills blank Website cells in "6000 Data COMPLETE.xlsx" using website data
from "6000 Office List to Send (1).xlsx", matched by ID (column A).

Usage:
    python fill_websites.py
    python fill_websites.py --source "6000 Office List to Send (1).xlsx" \
                            --target "6000 Data COMPLETE.xlsx" \
                            --output "6000 Data COMPLETE_filled.xlsx"
"""

import argparse
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl")

COL_ID      = 1   # Column A — ID (1-based)
COL_WEBSITE = 10  # Column J — Website (1-based)


def main():
    p = argparse.ArgumentParser(description="Fill blank websites from Office List into Data COMPLETE.")
    p.add_argument("--source",  default="6000 Office List to Send (1).xlsx",
                   help="Office List file (has the websites)")
    p.add_argument("--target",  default="6000 Data COMPLETE.xlsx",
                   help="Data COMPLETE file (has the blank website cells)")
    p.add_argument("--output",  default="6000 Data COMPLETE_filled.xlsx",
                   help="Output file path")
    args = p.parse_args()

    # ── Step 1: Build ID → website map from the source file ──────────────────
    print(f"Reading websites from: {args.source}")
    wb_src = openpyxl.load_workbook(args.source, read_only=True, data_only=True)
    ws_src = wb_src.active
    source_map = {}   # {id: website_url}
    for row in ws_src.iter_rows(min_row=2, values_only=True):
        rid = row[COL_ID - 1]
        web = str(row[COL_WEBSITE - 1] or "").strip()
        if rid is not None and web:
            source_map[rid] = web
    wb_src.close()
    print(f"  → {len(source_map)} IDs with websites loaded")

    # ── Step 2: Open target file and fill blanks ──────────────────────────────
    print(f"\nProcessing: {args.target}")
    wb_tgt = openpyxl.load_workbook(args.target)
    ws_tgt = wb_tgt.active

    filled   = 0
    skipped  = 0   # already has a website
    no_match = 0   # blank but no website in source

    for row_idx in range(2, ws_tgt.max_row + 1):
        rid = ws_tgt.cell(row_idx, COL_ID).value
        existing_web = str(ws_tgt.cell(row_idx, COL_WEBSITE).value or "").strip()

        if existing_web:
            skipped += 1
            continue

        if rid in source_map:
            ws_tgt.cell(row_idx, COL_WEBSITE).value = source_map[rid]
            office_name = ws_tgt.cell(row_idx, 3).value  # Column C = Office Name
            print(f"  ✓ Row {row_idx:>5} | ID {rid:<6} | {office_name or ''}  →  {source_map[rid]}")
            filled += 1
        else:
            no_match += 1

    # ── Step 3: Save ──────────────────────────────────────────────────────────
    wb_tgt.save(args.output)

    print(f"\n{'='*60}")
    print(f"Filled   : {filled} rows")
    print(f"No match : {no_match} rows (blank in both files)")
    print(f"Had data : {skipped} rows (already had a website)")
    print(f"\nSaved → {args.output}")


if __name__ == "__main__":
    main()
