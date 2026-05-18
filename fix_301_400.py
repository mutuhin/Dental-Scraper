#!/usr/bin/env python3
"""
fix_301_400.py
──────────────
Scrapes the 9 flagged indexes from the 301-400 range that have missing data.
Reads from 6000 Data COMPLETE.xlsx, runs the full dental scraper on each,
and writes a new output Excel: fix_301_400_output.xlsx

Target indexes: 304, 316, 330, 342, 345, 354, 378, 394, 397

Usage:
    python3 fix_301_400.py
"""

import sys, logging
sys.path.insert(0, "/Users/mujahidulhaqtuhin/Downloads/dental/py files")

import dental_scraper as ds
from playwright.sync_api import sync_playwright

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

INPUT_FILE  = "/Users/mujahidulhaqtuhin/Downloads/dental/py files/6000 Data COMPLETE.xlsx"
OUTPUT_FILE = "/Users/mujahidulhaqtuhin/Downloads/dental/py files/fix_301_400_output.xlsx"

# Indexes flagged "Need to capture" / "Need to fix"
TARGETS = {304, 316, 330, 342, 345, 354, 378, 394, 397}

# Index 342 stored URL is a doctor profile sub-page; glenoaksdental.com
# redirects to centennialfamilydentalmn.com — use the real homepage.
HOMEPAGE_OVERRIDE = {
    342: "https://www.centennialfamilydentalmn.com",
}

# Known emails confirmed from previous scrape runs (Cloudflare-blocked sites
# where website is inaccessible but email was retrieved via Facebook About page).
# Use None to force "Not Found" (e.g., scraper found a developer/template email).
EMAIL_OVERRIDE = {
    316: "office@mountainempiredental.com",
    397: None,   # micah@micahrich.com is the Wix site developer's email, not the practice
}


def load_target_practices():
    import openpyxl
    wb = openpyxl.load_workbook(INPUT_FILE, read_only=True)
    ws = wb.active

    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers)}

    practices = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            idx = int(row[col.get("ID", 0)])
        except Exception:
            continue
        if idx not in TARGETS:
            continue

        website = str(row[col.get("Website", 9)] or "").strip()
        website = HOMEPAGE_OVERRIDE.get(idx, website)

        practices.append({
            "Index":         idx,
            "Practice Name": row[col.get("Office Name", 2)],
            "Doctor Name":   row[col.get("Doctor Name", 10)],
            "Street":        row[col.get("Street", 3)],
            "City":          row[col.get("City", 4)],
            "State":         row[col.get("State", 5)],
            "Zip":           row[col.get("Zip", 6)],
            "Website":       website,
        })

    wb.close()
    practices.sort(key=lambda p: p["Index"])
    log.info("Loaded %d target practices: %s", len(practices), [p["Index"] for p in practices])
    return practices


def main():
    practices = load_target_practices()
    if not practices:
        log.error("No target practices found — check INPUT_FILE path")
        sys.exit(1)

    all_results = []

    with sync_playwright() as pw:
        browser = pw.chromium.launch_persistent_context(
            user_data_dir="",
            headless=True,
            ignore_https_errors=True,
            args=["--disable-blink-features=AutomationControlled"],
            user_agent=ds.HEADERS["User-Agent"],
        )
        pw_page = browser.new_page()
        pw_page.set_extra_http_headers({"Accept-Language": "en-US,en;q=0.9"})

        for i, practice in enumerate(practices, 1):
            idx  = practice["Index"]
            name = practice.get("Practice Name", "")
            url  = practice.get("Website", "")
            log.info("")
            log.info("[%d/%d] Index %d — %s  →  %s", i, len(practices), idx, name, url)

            try:
                scraped = ds.scrape_practice(practice, pw_page=pw_page)
            except Exception as exc:
                log.error("  ERROR on index %d: %s", idx, exc, exc_info=True)
                scraped = dict(ds.EMPTY_SCRAPED)
                # Recover Playwright if page crashed
                try:
                    if pw_page.is_closed():
                        pw_page = browser.new_page()
                        pw_page.set_extra_http_headers({"Accept-Language": "en-US,en;q=0.9"})
                except Exception:
                    pass

            # Apply email overrides: restore known-good emails or clear bad ones
            if idx in EMAIL_OVERRIDE:
                override = EMAIL_OVERRIDE[idx]
                if override is None:
                    scraped["email"] = "Not Found"
                    log.info("  Cleared developer/template email for index %d", idx)
                elif scraped.get("email") in ("Not Found", "", None):
                    scraped["email"] = override
                    log.info("  Applied email override for index %d: %s", idx, override)

            all_results.append((practice, scraped))
            log.info(
                "  doctors=%s  email=%s  hygienists=%s  google=%s/%s",
                len(scraped.get("doctors") or []),
                scraped.get("email", "-"),
                scraped.get("hygienists", "-"),
                scraped.get("google_rating", "-"),
                scraped.get("google_reviews", "-"),
            )

        try:
            browser.close()
        except Exception:
            pass

    log.info("")
    log.info("Writing output → %s", OUTPUT_FILE)
    ds.write_output(all_results, OUTPUT_FILE)
    log.info("Done — %d practices written", len(all_results))


if __name__ == "__main__":
    main()
