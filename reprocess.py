"""
Reprocess Cached Pages → Excel
===============================
Reads page_cache/ (written by dental_scraper.py) and re-runs all
extraction logic on the saved HTML — WITHOUT hitting any websites.

Use this when you fix a parsing bug (doctors, services, tech, etc.) and
want to regenerate the Excel without re-scraping.

HOW TO RUN:
    python3 reprocess.py

    # Re-run specific practice indices only:
    python3 reprocess.py 5 6 13

    # Deduplicate doctor names in an existing xlsx file:
    python3 reprocess.py --dedup batch_01_rows1_100.xlsx

OUTPUT:
    Dental_Scrape_Output_reprocessed.xlsx   (cache mode)
    <input>_deduped.xlsx                    (dedup mode)
"""

import sys
import os
import json
import re
import time
import logging
import openpyxl
from openpyxl import load_workbook

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ── Import all extraction helpers from dental_scraper ────────────────────────
# dental_scraper uses  if __name__ == "__main__": main()  so importing is safe
import dental_scraper as ds

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
CACHE_DIR   = "page_cache"
OUTPUT_FILE = "Dental_Scrape_Output_reprocessed.xlsx"

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def load_cache(folder: str):
    """
    Load all HTML files for one practice from its cache folder.
    Returns (manifest_dict, combined_html_list, result_dict).
    combined_html_list = [(page_type, url, html_text), ...]
    """
    manifest_path = os.path.join(folder, "manifest.json")
    result_path   = os.path.join(folder, "result.json")

    manifest = {}
    if os.path.exists(manifest_path):
        with open(manifest_path, encoding="utf-8") as f:
            manifest = json.load(f)

    cached_result = {}
    if os.path.exists(result_path):
        with open(result_path, encoding="utf-8") as f:
            data = json.load(f)
            cached_result = data.get("result", {})
            cached_result["doctors"] = data.get("doctors", [])
            cached_result["_practice"] = data.get("practice", {})

    pages = []
    for page_type, info in manifest.get("pages", {}).items():
        fpath = os.path.join(folder, info.get("file", f"{page_type}.html"))
        if os.path.exists(fpath):
            with open(fpath, encoding="utf-8", errors="replace") as f:
                html = f.read()
            pages.append((page_type, info.get("url", ""), html))

    return manifest, pages, cached_result


def reextract(pages: list, cached_result: dict) -> dict:
    """
    Re-run all extraction functions on the cached HTML pages.
    Social stats (FB/IG followers, Google rating) are kept from cached_result
    since they require live network calls.
    """
    from bs4 import BeautifulSoup

    # ── Build combined all_text and merged soup ───────────────────────────────
    all_text  = ""
    all_soup  = None
    all_soups = []
    raw_htmls = []   # keep raw HTML for body-text service counting

    for page_type, url, html in pages:
        text = ds.extract_text(html)
        all_text += " " + text
        soup = BeautifulSoup(html, "lxml")
        all_soups.append((page_type, soup))
        raw_htmls.append(html)
        if page_type == "homepage" and all_soup is None:
            all_soup = soup

    if all_soup is None and all_soups:
        all_soup = all_soups[0][1]

    # ── Rebuild result dict starting from cached network data ─────────────────
    result = {
        # Keep network-fetched data (social stats, ratings) from cached result
        "email":               cached_result.get("email", "Not Found"),
        "hygienists":          cached_result.get("hygienists", ""),
        "locations_count":     cached_result.get("locations_count", "1"),
        "facebook_url":        cached_result.get("facebook_url", ""),
        "facebook_posts":      cached_result.get("facebook_posts", ""),
        "facebook_followers":  cached_result.get("facebook_followers", "Not Found"),
        "instagram_url":       cached_result.get("instagram_url", ""),
        "instagram_posts":     cached_result.get("instagram_posts", ""),
        "instagram_followers": cached_result.get("instagram_followers", "Not Found"),
        "tiktok_url":          cached_result.get("tiktok_url", ""),
        "tiktok_posts":        cached_result.get("tiktok_posts", ""),
        "tiktok_followers":    cached_result.get("tiktok_followers", "Not Found"),
        "linkedin_url":        cached_result.get("linkedin_url", ""),
        "linkedin_posts":      cached_result.get("linkedin_posts", ""),
        "linkedin_followers":  cached_result.get("linkedin_followers", "Not Found"),
        "google_rating":       cached_result.get("google_rating", "Not Found"),
        "google_reviews":      cached_result.get("google_reviews", "Not Found"),
        "testimonials":        cached_result.get("testimonials", "0"),
        "invisalign_tier":     cached_result.get("invisalign_tier", "N/A"),
        # Re-extracted below
        "cerec": "", "cbct": "", "lasers": "", "ai": "", "intraoral": "",
        "invisalign": 0, "clear_aligners": 0, "veneers": 0, "implants": 0,
        "smile_makeovers": 0, "whitening": 0, "sedation": 0, "holistic": 0,
        "dental_plan": "", "cancer_screening": 0,
        "associations": "Not Found", "specialty": "Not Found",
        "scraped_doctor_names": "Not Found",
        "doctors": [],
    }

    if not all_text.strip():
        return result

    # ── Re-extract email if still not found ──────────────────────────────────
    if result["email"] in ("Not Found", "ERROR", "", None) and all_soup:
        result["email"] = ds.find_email(all_text, all_soup)

    # ── Locations, testimonials ───────────────────────────────────────────────
    if all_soup:
        result["locations_count"] = ds.find_locations_count(all_text, all_soup)
        # Count testimonials across ALL pages (not just homepage), dedup by text
        import re as _re
        _test_seen: set = set()
        _test_total = 0
        for _, _sp in all_soups:
            for _blk in _sp.find_all(
                ["div", "section", "article", "blockquote"],
                class_=_re.compile(r"(testimonial|review|quote|patient.story|patient.review)", _re.I),
            ):
                _key = _blk.get_text(separator=" ", strip=True)[:80]
                if _key and _key not in _test_seen:
                    _test_seen.add(_key)
                    _test_total += 1
        if _test_total == 0:
            for _, _sp in all_soups:
                for _bq in _sp.find_all("blockquote"):
                    _key = _bq.get_text(separator=" ", strip=True)[:80]
                    if _key and _key not in _test_seen:
                        _test_seen.add(_key)
                        _test_total += 1
        result["testimonials"] = str(_test_total) if _test_total > 0 else "0"

    # ── Associations / specialty ──────────────────────────────────────────────
    result["associations"] = ds.find_associations(all_text)
    result["specialty"]    = ds.find_specialty(all_text)

    # ── Hygienists ────────────────────────────────────────────────────────────
    hyg = ds.find_hygienists(all_text)
    if not result["hygienists"] or result["hygienists"] in ("N/A", "ERROR", "See Website"):
        result["hygienists"] = hyg
    # Also count from team page
    for page_type, soup in all_soups:
        hyg_n = ds._count_hygienists_from_team(soup)
        if hyg_n > 0:
            result["hygienists"] = str(hyg_n)
            break

    # ── Services (per-page: body-text primary, full-text fallback) ───────────
    svc_counts = {k: 0 for k in [
        "Invisalign", "Clear Aligners", "Veneers", "Implants",
        "Smile Makeovers", "Teeth Whitening", "Sedation Dentistry",
        "Holistic Dentistry", "Dental Plan", "Cancer Screening",
    ]}
    _svc_body = {k: 0 for k in svc_counts}
    _svc_full = {k: 0 for k in svc_counts}
    _seen_svc_urls: set = set()
    for page_type, url, html in pages:
        url_base = url.split("#")[0] if url else page_type
        if url_base in _seen_svc_urls:
            continue
        _seen_svc_urls.add(url_base)
        bt = ds.extract_body_text(html)
        ft = ds.extract_text(html)
        for keyword, category in ds.SERVICE_KEYWORDS.items():
            _svc_body[category] += ds.count_keyword_capped(bt, keyword, cap=3)
            _svc_full[category] += ds.count_keyword_capped(ft, keyword, cap=2)
    # If body-text gives 0, fall back to full-text (handles nav-only service listings)
    for cat in svc_counts:
        svc_counts[cat] = _svc_body[cat] if _svc_body[cat] > 0 else _svc_full[cat]

    result["invisalign"]      = svc_counts["Invisalign"]
    result["clear_aligners"]  = svc_counts["Clear Aligners"]
    result["veneers"]         = svc_counts["Veneers"]
    result["implants"]        = svc_counts["Implants"]
    result["smile_makeovers"] = svc_counts["Smile Makeovers"]
    result["whitening"]       = svc_counts["Teeth Whitening"]
    result["sedation"]        = svc_counts["Sedation Dentistry"]
    result["holistic"]        = svc_counts["Holistic Dentistry"]
    result["dental_plan"]     = "Mentioned" if svc_counts["Dental Plan"] > 0 else ""
    result["cancer_screening"]= svc_counts["Cancer Screening"]

    # ── Technology ────────────────────────────────────────────────────────────
    tech_found = set()
    for keyword, tech_name in ds.TECH_KEYWORDS.items():
        if keyword in all_text:
            tech_found.add(tech_name)

    result["cerec"]     = "X" if "CEREC"              in tech_found else ""
    result["cbct"]      = "X" if "CBCT"               in tech_found else ""
    result["lasers"]    = "X" if "Lasers"             in tech_found else ""
    result["ai"]        = "X" if "AI"                 in tech_found else ""
    result["intraoral"] = "X" if "Intraoral Scanners" in tech_found else ""

    # ── Doctors — re-parse from team/about pages ──────────────────────────────
    # Prefer the page that yields the most doctor *sections*; fall back to
    # most raw names if no page produces structured sections.
    # Use scrape_doctors_full with all soups for comprehensive dedup + merging.
    homepage_soup = all_soup

    doctors, hyg_count = ds.scrape_doctors_full(
        homepage_soup=homepage_soup,
        base_url="",       # no live fetching in reprocess mode
        all_text=all_text,
        pw_page=None,      # no browser in reprocess mode
        all_soups_for_team=all_soups,
    )

    result["doctors"] = doctors
    if doctors:
        result["scraped_doctor_names"] = ", ".join(d["name"] for d in doctors)

    # Hygienist count from team pages
    if hyg_count and hyg_count > 0:
        result["hygienists"] = str(hyg_count)

    return result


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────────────────────
# DOCTOR DEDUPLICATION  (works directly on an xlsx file)
# ─────────────────────────────────────────────────────────────────────────────

_CRED_RE = re.compile(
    r'\b(DDS|DMD|DPH|PhD|MS|FAGD|FICOI|PC|PLLC|PA|LLC|Inc\.?|MD|DO|RDH|AAACD)\b',
    re.I
)

def _norm(name: str) -> str:
    """Lowercase, strip Dr. prefix and credentials for comparison."""
    n = str(name or '').strip()
    n = re.sub(r"^Dr\.?\s+", "", n, flags=re.I)
    n = _CRED_RE.sub("", n)
    n = re.sub(r"[,.\s]+", " ", n).strip().lower()
    return n

def _score(name: str) -> int:
    """More words after normalisation = fuller name = higher score."""
    return len(_norm(name).split())

def _same_person(a: str, b: str) -> bool:
    na, nb = _norm(a), _norm(b)
    if not na or not nb:
        return False
    if na == nb or na in nb or nb in na:
        return True
    pa, pb = na.split(), nb.split()
    if not pa or not pb:
        return False
    last_a, last_b = pa[-1], pb[-1]
    # Last-name match (allow 1-char typo suffix: sottosanti / sottosantis)
    last_match = (
        last_a == last_b
        or last_a in last_b or last_b in last_a
        or (len(last_a) > 4 and last_a[:-1] == last_b)
        or (len(last_b) > 4 and last_b[:-1] == last_a)
    )
    if not last_match:
        return False
    # First-name prefix match (Kris / Kristan)
    if len(pa) > 1 and len(pb) > 1:
        fa, fb = pa[0], pb[0]
        return fa.startswith(fb) or fb.startswith(fa)
    return True


def dedup_doctors_in_xlsx(input_path: str) -> str:
    """
    Read an xlsx batch file, remove duplicate doctor rows within each practice
    (keeping the fullest name), and write <name>_deduped.xlsx.
    Returns the output path.
    """
    wb = load_workbook(input_path, data_only=True)
    ws = wb.active

    # Collect all rows grouped by practice index
    groups: dict = {}   # index -> [full_row_tuple, ...]
    header_rows = []
    for ri, row in enumerate(ws.iter_rows(values_only=True), 1):
        if ri <= 2:
            header_rows.append(list(row))
            continue
        idx = row[0]
        if idx is None:
            continue
        groups.setdefault(str(idx), []).append(list(row))

    kept_rows = []
    removed_total = 0

    for idx_key in sorted(groups, key=lambda x: float(x)):
        rows = groups[idx_key]
        practice_name = str(rows[0][1] or "")

        # Collect all doctor names (col 3, 0-based index 2)
        raw_names = [r[2] for r in rows]

        # Remove practice-name-as-doctor if real doctor names exist
        pnorm = _norm(practice_name)
        real = [n for n in raw_names if _norm(str(n or "")) != pnorm]
        if real:
            raw_names = real

        # Cluster duplicates and keep the fullest name per cluster
        clusters: list = []
        for name in raw_names:
            placed = False
            for cluster in clusters:
                if any(_same_person(str(name or ""), str(ex or "")) for ex in cluster):
                    cluster.append(name)
                    placed = True
                    break
            if not placed:
                clusters.append([name])

        kept_names = []
        for cluster in clusters:
            best = max(cluster, key=lambda n: (_score(str(n or "")), len(str(n or ""))))
            kept_names.append(best)

        removed = len(raw_names) - len(kept_names)
        removed_total += removed
        if removed:
            log.info(f"  [{idx_key}] {practice_name}: {len(raw_names)} → {len(kept_names)} doctors  (-{removed} dupes)")

        # Build a merged template: for each column take the first non-empty value
        # across ALL rows so that Google ratings (or any other data) stored on a
        # later row is not lost when we drop earlier rows.
        _EMPTY = {None, "", "Not Found", "ERROR"}
        template = list(rows[0])
        for row in rows[1:]:
            for ci, val in enumerate(row):
                if str(template[ci]).strip() in _EMPTY and str(val or "").strip() not in _EMPTY:
                    template[ci] = val

        for name in kept_names:
            new_row = list(template)
            new_row[2] = name
            kept_rows.append(new_row)

    # Write output workbook
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb_out = load_workbook(input_path, data_only=True)
    ws_out = wb_out.active

    # Delete all data rows then re-append
    for ri in range(ws_out.max_row, 2, -1):
        ws_out.delete_rows(ri)
    for row in kept_rows:
        ws_out.append(row)

    base, ext = os.path.splitext(input_path)
    out_path = base + "_deduped" + ext
    wb_out.save(out_path)

    log.info(f"\nRemoved {removed_total} duplicate doctor rows")
    log.info(f"Result : {len(kept_rows)} rows  →  {out_path}")
    return out_path


def main():
    # ── Dedup mode: python3 reprocess.py --dedup <file.xlsx> ─────────────────
    if len(sys.argv) > 1 and sys.argv[1] == "--dedup":
        if len(sys.argv) < 3:
            log.error("Usage: python3 reprocess.py --dedup <input.xlsx>")
            sys.exit(1)
        dedup_doctors_in_xlsx(sys.argv[2])
        return

    # Optional: filter to specific practice indices passed on the command line
    filter_indices = set()
    if len(sys.argv) > 1:
        for a in sys.argv[1:]:
            try:
                filter_indices.add(int(a))
            except ValueError:
                pass

    if not os.path.isdir(CACHE_DIR):
        log.error(f"Cache directory not found: {CACHE_DIR}")
        log.error("Run dental_scraper.py first to build the cache.")
        sys.exit(1)

    # Discover practice folders (sorted by idx prefix)
    folders = sorted(
        [os.path.join(CACHE_DIR, d) for d in os.listdir(CACHE_DIR)
         if os.path.isdir(os.path.join(CACHE_DIR, d))],
        key=lambda p: os.path.basename(p).split("_")[0]
    )

    if not folders:
        log.error("No cached practices found.")
        sys.exit(1)

    all_results = []

    for folder in folders:
        basename = os.path.basename(folder)
        # Extract idx from folder name like "005_Wince__Stacy"
        try:
            idx = int(basename.split("_")[0])
        except ValueError:
            idx = 0

        if filter_indices and idx not in filter_indices:
            continue

        log.info(f"Reprocessing: {basename}")
        manifest, pages, cached_result = load_cache(folder)

        if not pages:
            log.warning(f"  No cached pages found — skipping {basename}")
            continue

        practice_info = cached_result.get("_practice", {})
        if not practice_info:
            # Try to reconstruct minimal practice info from manifest
            practice_info = manifest.get("practice", {
                "Index": idx, "Practice Name": basename,
            })

        result = reextract(pages, cached_result)

        # Log doctor count
        n_doc = len(result["doctors"])
        log.info(f"  Doctors found: {n_doc}  |  "
                 f"Invisalign: {result['invisalign']}  |  "
                 f"CBCT: {result['cbct'] or '-'}")

        all_results.append((practice_info, result))

    if not all_results:
        log.warning("Nothing to write.")
        sys.exit(0)

    ds.write_output(all_results, OUTPUT_FILE)
    log.info(f"\n✅ Reprocessed {len(all_results)} practices → {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
