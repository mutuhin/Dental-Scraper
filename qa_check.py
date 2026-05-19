#!/usr/bin/env python3
"""
qa_check.py — Dental Scraper QA System v2
──────────────────────────────────────────
Reads any scraped, deduped, or refresh-tech-services Excel and produces a
5-sheet colour-coded QA report.

Sheets:
  1. QA Results       — every practice × every field, colour-coded; flags LOW DATA /
                        WEBSITE NOT SCRAPED in the status column
  2. Issues Report    — only practices that have any flag; lists all problems per row
  3. Missing Data     — field fill rates split by "has website" vs "no website";
                        shows which fields are blank DESPITE a website existing
  4. Bad Doctor Names — every doctor name cell that looks like a non-name
  5. Field Stats      — per-field fill rate across all practices

Status logic:
  FAIL          — any critical field missing (Doctor Name, Email, Google Rating/Reviews)
  NOT SCRAPED   — has a website URL but all key scraped fields are empty
  LOW DATA      — fill score is ≥1 std-dev below the batch average (relative flag)
  REVIEW        — ≥2 important fields missing OR no tech found OR no services found
  PASS          — everything looks good

Usage:
    python3 qa_check.py batch_1_deduped.xlsx
    python3 qa_check.py refresh_tech_batch1.xlsx --out my_qa.xlsx
"""

import sys, os, re, argparse, collections, statistics
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Field definitions ──────────────────────────────────────────────────────────
# tier: critical | important | optional | tech | service | info
# info = tracked/displayed but not included in any score
QA_FIELDS = [
    ("Doctor Name",                      "critical"),
    ("Practice Email",                   "critical"),
    ("Google Reviews Ranking",           "critical"),
    ("Total # of Google Reviews",        "critical"),
    ("# of Hygienists",                  "important"),
    ("Facebook URL",                     "important"),
    ("Associations / Memberships",       "important"),
    ("Doctor Specialty",                 "important"),
    ("Instagram URL",                    "optional"),
    ("TikTok URL",                       "optional"),
    ("LinkedIn URL",                     "optional"),
    ("Yelp Rating",                      "optional"),
    ("Total # of Yelp Reviews",          "optional"),
    ("Testimonials (Number of)",         "optional"),
    ("# of Locations",                   "optional"),
    # Social engagement stats (info — not scored)
    ("FB # Posts",                       "info"),
    ("FB Followers",                     "info"),
    ("IG # Posts",                       "info"),
    ("IG Followers",                     "info"),
    ("TT # Posts",                       "info"),
    ("TT Followers",                     "info"),
    ("LI # Posts",                       "info"),
    ("LI Followers",                     "info"),
    # Technology (scored as a group — expect ≥1)
    ("CEREC (Same Day Crowns)",          "tech"),
    ("CBCT (3D Imaging)",                "tech"),
    ("Lasers",                           "tech"),
    ("AI",                               "tech"),
    ("Intraoral Scanners",               "tech"),
    # Services (scored as a group — expect ≥1)
    ("Invisalign (Mentions)",            "service"),
    ("Invisalign Tier (check manually)", "info"),
    ("Clear Aligners",                   "service"),
    ("Veneers",                          "service"),
    ("Implants",                         "service"),
    ("Smile Makeovers",                  "service"),
    ("Teeth Whitening",                  "service"),
    ("Sedation Dentistry",               "service"),
    ("Holistic Dentistry",               "service"),
    ("Dental Plan (Membership Plan)",    "service"),
    ("Cancer Screening",                 "service"),
]

EMPTY_VALUES = {
    None, "", "Not Found", "N/A", "nan", "None", "none",
    "ERROR", "N/A – Not Offered", "0", "not found", "n/a",
}

# Business/generic terms that disqualify a string from being a person's name
_BIZ_TERMS = frozenset({
    "dental", "dentistry", "dentist", "care", "group", "center", "centre",
    "clinic", "practice", "office", "associates", "health", "wellness",
    "smile", "smiles", "studio", "family", "general", "orthodontic",
    "implant", "cosmetic", "pediatric", "oral", "surgery", "specialist",
    "specialists", "advanced", "premier", "excellence", "professional",
    "and", "the", "of", "for", "at", "your", "our", "team", "services",
})

# ── Colours / styles ───────────────────────────────────────────────────────────
F_GREEN   = PatternFill("solid", fgColor="C6EFCE")
F_YELLOW  = PatternFill("solid", fgColor="FFEB9C")
F_RED     = PatternFill("solid", fgColor="FFC7CE")
F_ORANGE  = PatternFill("solid", fgColor="FCE4D6")
F_GREY    = PatternFill("solid", fgColor="EDEDED")
F_BLUE    = PatternFill("solid", fgColor="BDD7EE")
F_LTBLUE  = PatternFill("solid", fgColor="DEEAF1")
F_DKBLUE  = PatternFill("solid", fgColor="1F4E79")
F_DKRED   = PatternFill("solid", fgColor="9C0006")
F_DKGRN   = PatternFill("solid", fgColor="375623")
F_DKORG   = PatternFill("solid", fgColor="833C00")
F_PURPLE  = PatternFill("solid", fgColor="7030A0")
F_DKGREY  = PatternFill("solid", fgColor="595959")
F_AMBER   = PatternFill("solid", fgColor="FF6600")

FONT_HDR  = Font(name="Arial", bold=True, size=9, color="FFFFFF")
FONT_DATA = Font(name="Arial", size=9)
FONT_BOLD = Font(name="Arial", bold=True, size=9)
FONT_RED  = Font(name="Arial", bold=True, size=9, color="9C0006")
FONT_GRN  = Font(name="Arial", bold=True, size=9, color="375623")
FONT_ORG  = Font(name="Arial", bold=True, size=9, color="833C00")
FONT_PRP  = Font(name="Arial", bold=True, size=9, color="7030A0")
ALIGN_CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LFT = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN  = Side(style="thin",   color="CCCCCC")
THICK = Side(style="medium", color="999999")
BDR   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _is_empty(val) -> bool:
    return str(val).strip() in EMPTY_VALUES or str(val).strip().lower() in EMPTY_VALUES


def _cell(ws, r, c, value="", font=FONT_DATA, fill=None, align=ALIGN_CTR, bold=False):
    cell = ws.cell(r, c, value)
    cell.font      = Font(name="Arial", bold=bold or font.bold,
                          size=font.size, color=font.color)
    cell.border    = BDR
    cell.alignment = align
    if fill:
        cell.fill = fill
    return cell


# ── Bad doctor name detection ─────────────────────────────────────────────────

def bad_name_reason(name) -> str:
    """Return a reason string if the doctor name looks wrong, else ''."""
    if name is None or str(name).strip() in EMPTY_VALUES:
        return "Empty / Not Found"
    n = str(name).strip()
    nl = n.lower()

    # URL or email accidentally in name field
    if "http" in nl or "@" in n or ".com" in nl or ".org" in nl:
        return "URL or email in name field"
    # Contains digits
    if re.search(r"\d", n):
        return "Contains digits"
    # All uppercase (not a name)
    alpha = re.sub(r"[^a-zA-Z]", "", n)
    if alpha and alpha == alpha.upper() and len(alpha) > 4:
        return "All uppercase"
    words = [w.strip(".,;") for w in n.split() if len(w.strip(".,;")) > 1]
    real_words = [w for w in words if not re.match(r"^[A-Z]\.$", w)]
    # Single meaningful word (no first + last)
    if len(real_words) == 1:
        return "Single word — no first + last name"
    # Too many words (≥6 meaningful words → likely practice name)
    if len(real_words) >= 6:
        return f"Too long ({len(real_words)} words) — possibly a practice/office name"
    # Business terms present
    biz_found = [w for w in real_words if w.lower() in _BIZ_TERMS]
    if biz_found:
        return f"Business term(s): {', '.join(biz_found)}"
    return ""


# ── Load input ─────────────────────────────────────────────────────────────────

def load_batch(path: str):
    """
    Returns:
        headers        list of all column names from the file
        practices      dict {index -> {field -> best_value}}
                         best_value = first non-empty value seen across doctor rows
        websites       dict {index -> website_url}
        practice_names dict {index -> Practice Name string}
        row_counts     dict {index -> number of doctor rows}
        doctor_rows    dict {index -> list of (doctor_name_str, row_dict)}
                         one entry per raw data row for bad-name checking
    """
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active

    # Find header row (row 1 or 2 — our files use row 2 with group labels in row 1)
    headers = []
    header_row_num = None
    for rn, row in enumerate(ws.iter_rows(min_row=1, max_row=4, values_only=True), 1):
        non_null = [v for v in row if v is not None and str(v).strip()]
        if "Index" in non_null or "Doctor Name" in non_null:
            headers = [str(v).strip() if v is not None else "" for v in row]
            header_row_num = rn
            break

    if not headers:
        raise ValueError("Could not find header row with 'Index' or 'Doctor Name' column")

    col_idx = {h: i for i, h in enumerate(headers) if h}

    qa_field_set = {f for f, _ in QA_FIELDS}
    practices      = collections.defaultdict(dict)
    websites       = {}
    practice_names = {}
    row_counts     = collections.Counter()
    doctor_rows    = collections.defaultdict(list)

    data_start = header_row_num + 1

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        raw_idx = row[col_idx.get("Index", 0)] if col_idx.get("Index") is not None else None
        try:
            idx = int(float(str(raw_idx)))
        except Exception:
            continue

        row_counts[idx] += 1

        if idx not in practice_names:
            pn_ci = col_idx.get("Practice Name")
            practice_names[idx] = str(row[pn_ci] or "") if pn_ci is not None else ""

        # Track website (first non-empty value wins)
        if idx not in websites:
            ws_ci = col_idx.get("Practice Website")
            if ws_ci is not None and not _is_empty(row[ws_ci]):
                websites[idx] = str(row[ws_ci]).strip()

        # Track doctor name rows for bad-name detection
        dn_ci = col_idx.get("Doctor Name")
        if dn_ci is not None:
            raw_dn = row[dn_ci]
            doctor_rows[idx].append(str(raw_dn).strip() if raw_dn is not None else "")

        # Collect best value per QA field
        for field in qa_field_set:
            ci = col_idx.get(field)
            if ci is None:
                continue
            val = row[ci]
            existing = practices[idx].get(field)
            if not _is_empty(val):
                # Prefer non-empty; for numeric service counts keep the largest
                if existing is None or _is_empty(existing):
                    practices[idx][field] = val
                else:
                    # Merge numeric counts (tech/service columns store counts)
                    try:
                        if int(float(str(val))) > int(float(str(existing))):
                            practices[idx][field] = val
                    except Exception:
                        pass  # keep existing non-empty string
            elif existing is None:
                practices[idx][field] = val

    wb.close()
    return headers, dict(practices), websites, practice_names, dict(row_counts), dict(doctor_rows)


# ── Scoring ────────────────────────────────────────────────────────────────────

def score_practice(data: dict, has_website: bool) -> dict:
    """
    Returns dict with:
        status         PASS | REVIEW | FAIL | NOT SCRAPED
        critical_miss  list of missing critical fields
        important_miss list of missing important fields
        optional_miss  list of missing optional fields
        tech_any       bool
        service_any    bool
        fill_pct       % of critical+important+optional fields that are filled
        not_scraped    bool — has website but key scraped data all empty
    """
    critical_miss  = []
    important_miss = []
    optional_miss  = []
    tech_vals      = []
    service_vals   = []

    for field, tier in QA_FIELDS:
        if tier == "info":
            continue
        val   = data.get(field)
        empty = _is_empty(val)
        if   tier == "critical":  (critical_miss  if empty else []).append(field) if empty else None
        elif tier == "important": (important_miss if empty else []).append(field) if empty else None
        elif tier == "optional":  (optional_miss  if empty else []).append(field) if empty else None
        elif tier == "tech":      tech_vals.append(not empty)
        elif tier == "service":   service_vals.append(not empty)

    # Re-do cleanly (the conditional append above is hard to read)
    critical_miss  = [f for f, t in QA_FIELDS if t == "critical"  and _is_empty(data.get(f))]
    important_miss = [f for f, t in QA_FIELDS if t == "important" and _is_empty(data.get(f))]
    optional_miss  = [f for f, t in QA_FIELDS if t == "optional"  and _is_empty(data.get(f))]
    tech_any    = any(not _is_empty(data.get(f)) for f, t in QA_FIELDS if t == "tech")
    service_any = any(not _is_empty(data.get(f)) for f, t in QA_FIELDS if t == "service")

    scoreable = [f for f, t in QA_FIELDS if t in ("critical", "important", "optional")]
    filled    = sum(1 for f in scoreable if not _is_empty(data.get(f)))
    fill_pct  = round(filled / len(scoreable) * 100) if scoreable else 0

    # NOT SCRAPED: has a website URL but key scraped data is all empty
    not_scraped = (
        has_website
        and all(_is_empty(data.get(f)) for f in (
            "Practice Email", "Google Reviews Ranking", "Total # of Google Reviews",
            "Doctor Specialty", "Associations / Memberships",
        ))
        and not tech_any
        and not service_any
    )

    if not_scraped:
        status = "NOT SCRAPED"
    elif critical_miss:
        status = "FAIL"
    elif len(important_miss) >= 2 or not tech_any or not service_any:
        status = "REVIEW"
    else:
        status = "PASS"

    return {
        "status":         status,
        "critical_miss":  critical_miss,
        "important_miss": important_miss,
        "optional_miss":  optional_miss,
        "tech_any":       tech_any,
        "service_any":    service_any,
        "fill_pct":       fill_pct,
        "not_scraped":    not_scraped,
    }


def compute_low_data_threshold(practices: dict, websites: dict) -> float:
    """
    Calculate the fill_pct threshold below which a practice is flagged LOW DATA.
    Uses mean - 1 std of fill_pcts across all practices with websites.
    Falls back to 40% if fewer than 5 practices.
    """
    scores = []
    for idx, data in practices.items():
        hw = idx in websites
        s  = score_practice(data, hw)
        scores.append(s["fill_pct"])
    if len(scores) < 5:
        return 40.0
    mean = statistics.mean(scores)
    std  = statistics.stdev(scores)
    return max(20.0, mean - std)   # never flag everything as low data


# ── Sheet 1: QA Results ────────────────────────────────────────────────────────

def write_qa_results(wb: Workbook, practices: dict, websites: dict,
                     practice_names: dict, row_counts: dict,
                     low_threshold: float):
    ws = wb.create_sheet("QA Results")
    tier_of = {f: t for f, t in QA_FIELDS}
    qa_cols = [f for f, _ in QA_FIELDS]   # in definition order

    # Group row 1 spans
    GROUP_DEFS = [
        ("Practice Info",   ["Doctor Name", "Practice Email", "# of Hygienists"],               F_DKBLUE),
        ("Social Media",    ["Facebook URL", "FB # Posts", "FB Followers",
                             "Instagram URL", "IG # Posts", "IG Followers",
                             "TikTok URL", "TT # Posts", "TT Followers",
                             "LinkedIn URL", "LI # Posts", "LI Followers"],                     F_DKGRN),
        ("Technology",      [f for f, t in QA_FIELDS if t == "tech"],                           F_PURPLE),
        ("Services",        [f for f, t in QA_FIELDS if t == "service"],                        F_DKORG),
        ("Inv. Tier",       ["Invisalign Tier (check manually)"],                               F_DKGREY),
        ("Doctor / Assoc.", ["Associations / Memberships", "Doctor Specialty"],                  F_DKGREY),
        ("Reviews",         ["Google Reviews Ranking", "Total # of Google Reviews",
                             "Yelp Rating", "Total # of Yelp Reviews",
                             "Testimonials (Number of)", "# of Locations"],                      F_DKBLUE),
    ]

    FIXED = ["Index", "Practice Name", "Website", "# Doctors", "Status", "Fill %", "Flags"]
    fixed_n = len(FIXED)

    # Row 1 — group label spans
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=fixed_n)
    _cell(ws, 1, 1, "Practice Identifiers", font=FONT_HDR, fill=F_DKBLUE)

    col_offset = fixed_n + 1
    for grp, fields, fill in GROUP_DEFS:
        count = sum(1 for f in fields if f in qa_cols)
        if count == 0:
            continue
        ws.merge_cells(start_row=1, start_column=col_offset,
                       end_row=1,   end_column=col_offset + count - 1)
        _cell(ws, 1, col_offset, grp, font=FONT_HDR, fill=fill)
        col_offset += count

    # Row 2 — column headers
    TIER_FILL = {
        "critical": F_DKBLUE, "important": F_DKBLUE,
        "optional": F_DKGREY, "tech": F_PURPLE, "service": F_DKORG, "info": F_DKGREY,
    }
    for c, h in enumerate(FIXED, 1):
        _cell(ws, 2, c, h, font=FONT_HDR, fill=F_DKBLUE)
    for c, h in enumerate(qa_cols, fixed_n + 1):
        _cell(ws, 2, c, h, font=FONT_HDR, fill=TIER_FILL.get(tier_of.get(h, ""), F_DKBLUE))

    ws.freeze_panes = f"{get_column_letter(fixed_n + 1)}3"

    counts = {"PASS": 0, "REVIEW": 0, "FAIL": 0, "NOT SCRAPED": 0}
    field_filled = collections.Counter()
    field_total  = collections.Counter()
    field_filled_w = collections.Counter()   # only practices WITH websites
    field_total_w  = collections.Counter()

    sorted_idxs = sorted(practices.keys())
    for row_num, idx in enumerate(sorted_idxs, 3):
        data       = practices[idx]
        name       = practice_names.get(idx, "")
        website    = websites.get(idx, "")
        has_website = bool(website)
        ndocs      = row_counts.get(idx, 1)
        score      = score_practice(data, has_website)
        status     = score["status"]
        fill_pct   = score["fill_pct"]
        low_data   = fill_pct < low_threshold and not score["not_scraped"]

        counts[status] = counts.get(status, 0) + 1

        # Build flags string
        flags = []
        if low_data:
            flags.append("LOW DATA")
        if score["not_scraped"]:
            flags.append("NOT SCRAPED")
        if not has_website:
            flags.append("NO WEBSITE")
        flag_str = " | ".join(flags) if flags else "—"

        even = row_num % 2 == 0
        row_bg = PatternFill("solid", fgColor="F7F7F7") if even else None

        STATUS_FILL = {
            "PASS":        F_GREEN,
            "REVIEW":      F_YELLOW,
            "FAIL":        F_RED,
            "NOT SCRAPED": F_ORANGE,
        }
        STATUS_FONT = {
            "PASS":        FONT_GRN,
            "REVIEW":      FONT_BOLD,
            "FAIL":        FONT_RED,
            "NOT SCRAPED": FONT_ORG,
        }
        sf = STATUS_FILL.get(status, F_GREY)
        fn = STATUS_FONT.get(status, FONT_DATA)

        _cell(ws, row_num, 1, idx,      font=FONT_DATA, fill=row_bg)
        _cell(ws, row_num, 2, name,     font=FONT_DATA, fill=row_bg, align=ALIGN_LFT)
        _cell(ws, row_num, 3, website,  font=FONT_DATA, fill=row_bg, align=ALIGN_LFT)
        _cell(ws, row_num, 4, ndocs,    font=FONT_DATA, fill=row_bg)
        _cell(ws, row_num, 5, status,   font=fn, fill=sf)
        pct_fill = F_GREEN if fill_pct >= 70 else (F_YELLOW if fill_pct >= 40 else F_RED)
        _cell(ws, row_num, 6, f"{fill_pct}%", font=FONT_BOLD, fill=pct_fill)
        flag_fill = F_RED if "NOT SCRAPED" in flag_str or "LOW DATA" in flag_str else (
                    F_YELLOW if "NO WEBSITE" in flag_str else row_bg)
        _cell(ws, row_num, 7, flag_str, font=FONT_DATA, fill=flag_fill, align=ALIGN_LFT)

        for c, field in enumerate(qa_cols, fixed_n + 1):
            tier  = tier_of.get(field, "")
            val   = data.get(field)
            empty = _is_empty(val)
            field_total[field] += 1
            if has_website:
                field_total_w[field] += 1

            if not empty:
                cell_fill = F_GREEN
                disp = str(val).strip()
                # For counted fields, show the number not just ✓
                try:
                    n = int(float(str(val)))
                    disp = str(n) if n > 0 else "✓"
                except Exception:
                    disp = "✓" if len(disp) > 20 else disp
                field_filled[field] += 1
                if has_website:
                    field_filled_w[field] += 1
            elif tier in ("tech", "service"):
                group_ok = score["tech_any"] if tier == "tech" else score["service_any"]
                cell_fill = F_YELLOW if group_ok else F_RED
                disp = "–"
            elif tier == "info":
                cell_fill = F_GREY
                disp = "–"
            else:
                cell_fill = F_RED
                disp = "✗"

            _cell(ws, row_num, c, disp, font=FONT_DATA, fill=cell_fill)

    # Summary footer row
    sum_r = ws.max_row + 2
    _cell(ws, sum_r, 1, "TOTALS", font=FONT_HDR, fill=F_DKBLUE, align=ALIGN_LFT)
    summary = (f"PASS {counts.get('PASS',0)}  |  REVIEW {counts.get('REVIEW',0)}  |"
               f"  FAIL {counts.get('FAIL',0)}  |  NOT SCRAPED {counts.get('NOT SCRAPED',0)}")
    _cell(ws, sum_r, 2, summary, font=FONT_HDR, fill=F_DKBLUE, align=ALIGN_LFT)
    ws.merge_cells(start_row=sum_r, start_column=2, end_row=sum_r, end_column=fixed_n)
    for c, field in enumerate(qa_cols, fixed_n + 1):
        total  = field_total.get(field, 1)
        filled = field_filled.get(field, 0)
        pct    = round(filled / total * 100) if total else 0
        fill   = F_GREEN if pct >= 80 else (F_YELLOW if pct >= 50 else F_RED)
        _cell(ws, sum_r, c, f"{pct}%", font=FONT_HDR, fill=fill)

    # Column widths
    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 9
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 22
    for c in range(fixed_n + 1, fixed_n + len(qa_cols) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 11
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 50

    return counts, field_filled, field_total, field_filled_w, field_total_w, low_threshold


# ── Sheet 2: Issues Report ─────────────────────────────────────────────────────

def write_issues_report(wb: Workbook, practices: dict, websites: dict,
                        practice_names: dict, doctor_rows: dict, low_threshold: float):
    ws = wb.create_sheet("Issues Report")

    headers = [
        "Index", "Practice Name", "Website", "Status", "Fill %",
        "Flags", "Missing Critical", "Missing Important", "Missing Optional",
        "Tech", "Services", "Bad Doctor Names",
    ]
    for c, h in enumerate(headers, 1):
        _cell(ws, 1, c, h, font=FONT_HDR, fill=F_DKBLUE)
    ws.freeze_panes = "A2"

    rows_written = 0
    for idx in sorted(practices.keys()):
        data       = practices[idx]
        name       = practice_names.get(idx, "")
        website    = websites.get(idx, "")
        has_website = bool(website)
        score      = score_practice(data, has_website)
        fill_pct   = score["fill_pct"]
        low_data   = fill_pct < low_threshold and not score["not_scraped"]
        status     = score["status"]

        # Detect bad doctor names for this index
        bad_names = []
        for dn in doctor_rows.get(idx, []):
            reason = bad_name_reason(dn)
            if reason and reason != "Empty / Not Found":
                bad_names.append(f"{dn!r} ({reason})")

        has_issue = (
            status in ("FAIL", "REVIEW", "NOT SCRAPED")
            or low_data
            or bool(bad_names)
        )
        if not has_issue:
            continue

        rows_written += 1
        r = rows_written + 1
        even = r % 2 == 0
        row_bg = PatternFill("solid", fgColor="F7F7F7") if even else None

        flags = []
        if low_data:           flags.append("LOW DATA")
        if score["not_scraped"]: flags.append("NOT SCRAPED")
        if not has_website:    flags.append("NO WEBSITE")
        if bad_names:          flags.append("BAD NAMES")

        STATUS_FILL = {"PASS": F_GREEN, "REVIEW": F_YELLOW,
                       "FAIL": F_RED,   "NOT SCRAPED": F_ORANGE}
        STATUS_FONT = {"PASS": FONT_GRN, "REVIEW": FONT_BOLD,
                       "FAIL": FONT_RED, "NOT SCRAPED": FONT_ORG}

        pct_fill = F_GREEN if fill_pct >= 70 else (F_YELLOW if fill_pct >= 40 else F_RED)

        _cell(ws, r, 1,  idx,   font=FONT_DATA, fill=row_bg)
        _cell(ws, r, 2,  name,  font=FONT_DATA, fill=row_bg, align=ALIGN_LFT)
        _cell(ws, r, 3,  website, font=FONT_DATA, fill=row_bg, align=ALIGN_LFT)
        _cell(ws, r, 4,  status, font=STATUS_FONT.get(status, FONT_DATA),
              fill=STATUS_FILL.get(status, F_GREY))
        _cell(ws, r, 5,  f"{fill_pct}%", font=FONT_BOLD, fill=pct_fill)
        _cell(ws, r, 6,  " | ".join(flags) if flags else "—",
              font=FONT_DATA, fill=F_RED if flags else row_bg, align=ALIGN_LFT)
        _cell(ws, r, 7,  ", ".join(score["critical_miss"])  or "—",
              font=FONT_RED if score["critical_miss"] else FONT_DATA,
              fill=row_bg, align=ALIGN_LFT)
        _cell(ws, r, 8,  ", ".join(score["important_miss"]) or "—",
              font=FONT_DATA, fill=row_bg, align=ALIGN_LFT)
        _cell(ws, r, 9,  ", ".join(score["optional_miss"][:4]) or "—",
              font=FONT_DATA, fill=row_bg, align=ALIGN_LFT)
        _cell(ws, r, 10, "✓" if score["tech_any"]    else "✗ None",
              font=FONT_GRN if score["tech_any"] else FONT_RED,
              fill=F_GREEN  if score["tech_any"] else F_RED)
        _cell(ws, r, 11, "✓" if score["service_any"] else "✗ None",
              font=FONT_GRN if score["service_any"] else FONT_RED,
              fill=F_GREEN  if score["service_any"] else F_RED)
        _cell(ws, r, 12, "; ".join(bad_names[:3]) if bad_names else "—",
              font=FONT_RED if bad_names else FONT_DATA, fill=row_bg, align=ALIGN_LFT)

    if rows_written == 0:
        _cell(ws, 2, 1, "✅ No issues found!", font=FONT_GRN, fill=F_GREEN, align=ALIGN_LFT)

    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 45
    ws.column_dimensions["H"].width = 40
    ws.column_dimensions["I"].width = 35
    ws.column_dimensions["J"].width = 10
    ws.column_dimensions["K"].width = 10
    ws.column_dimensions["L"].width = 50


# ── Sheet 3: Missing Data (Website vs No-Website) ─────────────────────────────

def write_missing_data(wb: Workbook, practices: dict, websites: dict,
                       field_filled: dict, field_total: dict,
                       field_filled_w: dict, field_total_w: dict):
    ws = wb.create_sheet("Missing Data")

    headers = [
        "Field", "Tier",
        "All — Filled", "All — Total", "All — Fill %",
        "With Website — Filled", "With Website — Total", "With Website — Fill %",
        "Gap (should be scraped but isn't)",
    ]
    for c, h in enumerate(headers, 1):
        _cell(ws, 1, c, h, font=FONT_HDR, fill=F_DKBLUE)
    ws.freeze_panes = "A2"

    tier_of = {f: t for f, t in QA_FIELDS}

    # Group separator rows
    prev_tier = None
    r = 2
    for field, tier in QA_FIELDS:
        if tier == "info":
            continue
        total    = field_total.get(field, 0)
        filled   = field_filled.get(field, 0)
        total_w  = field_total_w.get(field, 0)
        filled_w = field_filled_w.get(field, 0)

        pct_all = round(filled / total * 100)   if total   else 0
        pct_w   = round(filled_w / total_w * 100) if total_w else 0
        # Gap: practices WITH a website that are still missing the field
        gap     = total_w - filled_w

        even = r % 2 == 0
        row_bg = PatternFill("solid", fgColor="F7F7F7") if even else None

        pct_fill_all = F_GREEN if pct_all >= 80 else (F_YELLOW if pct_all >= 50 else F_RED)
        pct_fill_w   = F_GREEN if pct_w   >= 80 else (F_YELLOW if pct_w   >= 50 else F_RED)
        gap_fill     = F_RED if gap > 5 else (F_YELLOW if gap > 0 else F_GREEN)

        TIER_FILL2 = {
            "critical": F_DKBLUE, "important": F_BLUE, "optional": F_GREY,
            "tech": F_PURPLE, "service": F_DKORG,
        }
        tier_bg = PatternFill("solid", fgColor={
            "critical": "DEEAF1", "important": "EBF3FB",
            "optional": "F7F7F7", "tech": "F0E6F6", "service": "FDE9D9",
        }.get(tier, "FFFFFF"))

        _cell(ws, r, 1, field,   font=FONT_DATA, fill=tier_bg, align=ALIGN_LFT)
        _cell(ws, r, 2, tier,    font=FONT_BOLD, fill=TIER_FILL2.get(tier, F_GREY))
        _cell(ws, r, 3, filled,  font=FONT_DATA, fill=row_bg)
        _cell(ws, r, 4, total,   font=FONT_DATA, fill=row_bg)
        _cell(ws, r, 5, f"{pct_all}%", font=FONT_BOLD, fill=pct_fill_all)
        _cell(ws, r, 6, filled_w, font=FONT_DATA, fill=row_bg)
        _cell(ws, r, 7, total_w,  font=FONT_DATA, fill=row_bg)
        _cell(ws, r, 8, f"{pct_w}%",  font=FONT_BOLD, fill=pct_fill_w)
        gap_label = f"{gap} practices have website but field is empty" if gap > 0 else "—"
        _cell(ws, r, 9, gap_label, font=FONT_RED if gap > 5 else FONT_DATA,
              fill=gap_fill, align=ALIGN_LFT)
        r += 1

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 13
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["E"].width = 13
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 42
    ws.row_dimensions[1].height = 40


# ── Sheet 4: Bad Doctor Names ──────────────────────────────────────────────────

def write_bad_names(wb: Workbook, doctor_rows: dict, practice_names: dict,
                    websites: dict):
    ws = wb.create_sheet("Bad Doctor Names")

    headers = ["Index", "Practice Name", "Website", "Doctor Name (raw)", "Issue"]
    for c, h in enumerate(headers, 1):
        _cell(ws, 1, c, h, font=FONT_HDR, fill=F_DKRED)
    ws.freeze_panes = "A2"

    r = 2
    for idx in sorted(doctor_rows.keys()):
        names_seen = set()
        for dn in doctor_rows[idx]:
            if dn in names_seen:
                continue
            names_seen.add(dn)
            reason = bad_name_reason(dn)
            if not reason:
                continue
            even = r % 2 == 0
            row_bg = PatternFill("solid", fgColor="FFF2CC") if even else PatternFill("solid", fgColor="FFFFFF")
            _cell(ws, r, 1, idx,                        font=FONT_DATA, fill=row_bg)
            _cell(ws, r, 2, practice_names.get(idx,""), font=FONT_DATA, fill=row_bg, align=ALIGN_LFT)
            _cell(ws, r, 3, websites.get(idx,""),        font=FONT_DATA, fill=row_bg, align=ALIGN_LFT)
            _cell(ws, r, 4, dn,                          font=FONT_RED,  fill=F_RED,  align=ALIGN_LFT)
            _cell(ws, r, 5, reason,                      font=FONT_DATA, fill=F_YELLOW, align=ALIGN_LFT)
            r += 1

    if r == 2:
        _cell(ws, 2, 1, "✅ No bad doctor names detected", font=FONT_GRN, fill=F_GREEN, align=ALIGN_LFT)

    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 45


# ── Sheet 5: Field Stats ───────────────────────────────────────────────────────

def write_field_stats(wb: Workbook, field_filled: dict, field_total: dict):
    ws = wb.create_sheet("Field Stats")

    headers = ["Field", "Tier", "Filled", "Total", "Fill Rate", "Status"]
    for c, h in enumerate(headers, 1):
        _cell(ws, 1, c, h, font=FONT_HDR, fill=F_DKBLUE)
    ws.freeze_panes = "A2"

    TIER_FILL3 = {
        "critical": F_DKBLUE, "important": F_BLUE, "optional": F_GREY,
        "tech": F_PURPLE, "service": F_DKORG, "info": F_DKGREY,
    }

    for r, (field, tier) in enumerate(QA_FIELDS, 2):
        total  = field_total.get(field, 0)
        filled = field_filled.get(field, 0)
        pct    = round(filled / total * 100) if total else 0
        even   = r % 2 == 0
        row_bg = PatternFill("solid", fgColor="F7F7F7") if even else None

        if pct >= 80:    status, sf = "Good",   F_GREEN
        elif pct >= 50:  status, sf = "Review", F_YELLOW
        else:            status, sf = "Low",    F_RED

        _cell(ws, r, 1, field,  font=FONT_DATA, fill=row_bg, align=ALIGN_LFT)
        _cell(ws, r, 2, tier,   font=FONT_BOLD, fill=TIER_FILL3.get(tier, F_GREY))
        _cell(ws, r, 3, filled, font=FONT_DATA, fill=row_bg)
        _cell(ws, r, 4, total,  font=FONT_DATA, fill=row_bg)
        _cell(ws, r, 5, f"{pct}%", font=FONT_BOLD, fill=sf)
        _cell(ws, r, 6, status,    font=FONT_DATA, fill=sf)

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 9
    ws.column_dimensions["D"].width = 9
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 10


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Dental Scraper QA System")
    parser.add_argument("input",  help="Scraped batch or refresh-tech Excel file")
    parser.add_argument("--out",  help="Output QA report path (default: qa_<input>)")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"ERROR: File not found: {args.input}", file=sys.stderr)
        sys.exit(1)

    out_path = args.out or ("qa_" + os.path.basename(args.input))

    print(f"Loading:  {args.input}")
    headers, practices, websites, practice_names, row_counts, doctor_rows = load_batch(args.input)
    total = len(practices)
    print(f"Practices: {total}  |  With website: {sum(1 for i in practices if i in websites)}")

    low_threshold = compute_low_data_threshold(practices, websites)
    print(f"Low-data threshold: {low_threshold:.0f}% (mean − 1σ of batch fill scores)")

    wb = Workbook()
    wb.remove(wb.active)

    counts, ff, ft, ffw, ftw, _ = write_qa_results(
        wb, practices, websites, practice_names, row_counts, low_threshold)
    write_issues_report(wb, practices, websites, practice_names, doctor_rows, low_threshold)
    write_missing_data(wb, practices, websites, ff, ft, ffw, ftw)
    write_bad_names(wb, doctor_rows, practice_names, websites)
    write_field_stats(wb, ff, ft)

    wb.save(out_path)

    print(f"\n{'='*55}")
    print(f"QA REPORT — {total} practices")
    print(f"  PASS        : {counts.get('PASS',0)}")
    print(f"  REVIEW      : {counts.get('REVIEW',0)}")
    print(f"  FAIL        : {counts.get('FAIL',0)}")
    print(f"  NOT SCRAPED : {counts.get('NOT SCRAPED',0)}")
    low_count = sum(
        1 for idx, data in practices.items()
        if score_practice(data, idx in websites)["fill_pct"] < low_threshold
        and not score_practice(data, idx in websites)["not_scraped"]
    )
    print(f"  LOW DATA    : {low_count}  (below {low_threshold:.0f}% fill)")
    print(f"\nLowest fill-rate fields (all practices):")
    stats = [(f, ff.get(f, 0), ft.get(f, 1)) for f, t in QA_FIELDS if t != "info"]
    stats.sort(key=lambda x: x[1] / x[2])
    for field, filled, total_f in stats[:6]:
        print(f"  {field}: {round(filled/total_f*100)}%")
    print(f"\nSaved → {out_path}")


if __name__ == "__main__":
    main()
