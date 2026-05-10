"""
scrape_row_137.py
─────────────────
Targeted scraper for practice Index 137 (Lalor Family Dental, Elmira NY).

Features:
  - Uses curl_cffi with browser impersonation to bypass Cloudflare
  - Saves every fetched page as HTML to page_cache/137_Lalor__Robert/
  - Extracts: doctors, email, social links, tech flags, service counts,
    associations, specialties, testimonials, Google rating
  - Splits the single xlsx row into one row PER doctor (matching the
    format used by all other practices in the file)

HOW TO RUN:
    python3 scrape_row_137.py

REQUIRES:
    pip install curl_cffi beautifulsoup4 openpyxl lxml playwright
    playwright install chromium
"""

import re
import os
import json
import time
import logging
import warnings
from urllib.parse import urljoin, quote_plus

from bs4 import BeautifulSoup
import openpyxl
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", message="Unverified HTTPS request")

# ── curl_cffi (Cloudflare bypass via TLS browser impersonation) ───────────────
try:
    from curl_cffi import requests as cffi_requests
    CFFI_AVAILABLE = True
except ImportError:
    CFFI_AVAILABLE = False
    import requests as cffi_requests   # plain requests fallback
    print("WARNING: curl_cffi not installed — Cloudflare bypass unavailable.")

# ── Playwright (for Facebook followers, Google Maps rating) ───────────────────
try:
    from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PLAYWRIGHT_AVAILABLE = False
    print("WARNING: playwright not installed — FB followers / Google rating may be limited.")

logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(levelname)-7s  %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger(__name__)

# ── Config ────────────────────────────────────────────────────────────────────
XLSX_FILE    = "Dental_Scrape_Output_v10_cleaned.xlsx"
CACHE_DIR    = "page_cache"
TARGET_INDEX = 137
PRACTICE_NAME = "Lalor, Robert"
BASE_URL     = "https://www.lalordental.com/"
DELAY_SEC    = 2.0
TIMEOUT      = 20
PW_TIMEOUT   = 25000
IMPERSONATE  = "chrome110"   # curl_cffi browser profile

SUBPAGES = [
    ("about-us/our-doctors/",  "doctors"),
    ("about-us/",              "about"),
    ("about-us/our-team/",     "team"),
    ("services/",              "services"),
    ("technology/",            "technology"),
    ("contact/",               "contact"),
    ("new-patients/",          "new_patients"),
    ("membership/",            "membership"),
]

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# ── Known-good fallback data (verified by manual inspection 2026-03-29) ───────
KNOWN_DOCTORS = [
    "Dr. Robert A. Lalor", "Dr. Mary John", "Dr. Aaron Tadayon",
    "Dr. Amir Toloue", "Dr. Betsey Clark-Fortier", "Dr. Bryant Lambert",
    "Dr. Caroline Roberto", "Dr. Daniel Sultan", "Dr. Dawn Weldon",
    "Dr. Dimitri Kverenchkhiladze", "Dr. Gila Beer", "Dr. Isabelle Cruz",
    "Dr. James Jeong", "Dr. Jia Lee", "Dr. Joe Lee", "Dr. Josh Cherian",
    "Dr. Justin Sbarra", "Dr. Kenneth Gerace", "Dr. Kevin Lin",
]
KNOWN_FACEBOOK = "https://www.facebook.com/Lalor-Creekside-Dental-249248492545/"

# ── Service keyword → column name ─────────────────────────────────────────────
SERVICE_KEYWORDS = {
    "invisalign":           "Invisalign (Mentions)",
    "clear aligner":        "Clear Aligners",
    "suresmile":            "Clear Aligners",
    "clearcorrect":         "Clear Aligners",
    "spark aligner":        "Clear Aligners",
    "byte aligner":         "Clear Aligners",
    "candidpro":            "Clear Aligners",
    "candid pro":           "Clear Aligners",
    "invisible brace":      "Clear Aligners",
    "clear brace":          "Clear Aligners",
    "veneers":              "Veneers",
    "veneer":               "Veneers",
    "porcelain veneer":     "Veneers",
    "implant":              "Implants",
    "smile makeover":       "Smile Makeovers",
    "smile transformation": "Smile Makeovers",
    "smile design":         "Smile Makeovers",
    "smile redesign":       "Smile Makeovers",
    "complete smile":       "Smile Makeovers",
    "full smile":           "Smile Makeovers",
    "total smile":          "Smile Makeovers",
    "smile restoration":    "Smile Makeovers",
    "whitening":            "Teeth Whitening",
    "bleach":               "Teeth Whitening",
    "zoom":                 "Teeth Whitening",
    "opalescence":          "Teeth Whitening",
    "bright smile":         "Teeth Whitening",
    "kor whitening":        "Teeth Whitening",
    "sedation":             "Sedation Dentistry",
    "sleep dentistry":      "Sedation Dentistry",
    "nitrous oxide":        "Sedation Dentistry",
    "laughing gas":         "Sedation Dentistry",
    "iv sedation":          "Sedation Dentistry",
    "oral conscious":       "Sedation Dentistry",
    "holistic":             "Holistic Dentistry",
    "biological dentist":   "Holistic Dentistry",
    "mercury-free":         "Holistic Dentistry",
    "mercury free":         "Holistic Dentistry",
    "mercury safe":         "Holistic Dentistry",
    "fluoride-free":        "Holistic Dentistry",
    "biocompatible":        "Holistic Dentistry",
    "ozone therapy":        "Holistic Dentistry",
    "zirconia implant":     "Holistic Dentistry",
    "metal-free":           "Holistic Dentistry",
    "cancer screening":     "Cancer Screening",
    "oral cancer":          "Cancer Screening",
    "velscope":             "Cancer Screening",
    "identafi":             "Cancer Screening",
    "vizilite":             "Cancer Screening",
    "oral id":              "Cancer Screening",
    "membership plan":      "Dental Plan (Membership Plan)",
    "dental membership":    "Dental Plan (Membership Plan)",
    "annual membership":    "Dental Plan (Membership Plan)",
    "in-house plan":        "Dental Plan (Membership Plan)",
    "in house plan":        "Dental Plan (Membership Plan)",
    "in-house membership":  "Dental Plan (Membership Plan)",
    "savings plan":         "Dental Plan (Membership Plan)",
    "dental savings":       "Dental Plan (Membership Plan)",
    "discount plan":        "Dental Plan (Membership Plan)",
    "wellness plan":        "Dental Plan (Membership Plan)",
    "dental subscription":  "Dental Plan (Membership Plan)",
    "uninsured patients":   "Dental Plan (Membership Plan)",
    "no dental insurance":  "Dental Plan (Membership Plan)",
    "in-office plan":       "Dental Plan (Membership Plan)",
    "preventive plan":      "Dental Plan (Membership Plan)",
    "care plan":            "Dental Plan (Membership Plan)",
}

TECH_KEYWORDS = {
    "cerec":                  "CEREC (Same Day Crowns)",
    "same day crown":         "CEREC (Same Day Crowns)",
    "same-day crown":         "CEREC (Same Day Crowns)",
    "same day restoration":   "CEREC (Same Day Crowns)",
    "in-office crown":        "CEREC (Same Day Crowns)",
    "chairside crown":        "CEREC (Same Day Crowns)",
    "milled crown":           "CEREC (Same Day Crowns)",
    "single visit crown":     "CEREC (Same Day Crowns)",
    "one visit crown":        "CEREC (Same Day Crowns)",
    "one-visit crown":        "CEREC (Same Day Crowns)",
    "e4d":                    "CEREC (Same Day Crowns)",
    "primescan":              "CEREC (Same Day Crowns)",
    "omnicam":                "CEREC (Same Day Crowns)",
    "cbct":                   "CBCT (3D Imaging)",
    "cone beam":              "CBCT (3D Imaging)",
    "3d imaging":             "CBCT (3D Imaging)",
    "3d x-ray":               "CBCT (3D Imaging)",
    "3d xray":                "CBCT (3D Imaging)",
    "3d x ray":               "CBCT (3D Imaging)",
    "3d scan":                "CBCT (3D Imaging)",
    "3d radiograph":          "CBCT (3D Imaging)",
    "i-cat":                  "CBCT (3D Imaging)",
    "dental ct":              "CBCT (3D Imaging)",
    "planmeca":               "CBCT (3D Imaging)",
    "vatech":                 "CBCT (3D Imaging)",
    "3d cone beam":           "CBCT (3D Imaging)",
    "galileos":               "CBCT (3D Imaging)",
    "laser":                  "Lasers",
    "waterlase":              "Lasers",
    "biolase":                "Lasers",
    "solea laser":            "Lasers",
    "diode laser":            "Lasers",
    "erbium laser":           "Lasers",
    "fotona":                 "Lasers",
    "lightwalker":            "Lasers",
    "dental laser":           "Lasers",
    "artificial intelligence":"AI",
    "overjet":                "AI",
    "pearl ai":               "AI",
    "diagnocat":              "AI",
    "dental ai":              "AI",
    "ai-powered":             "AI",
    "ai powered":             "AI",
    "videa":                  "AI",
    "dental intel":           "AI",
    "ai detection":           "AI",
    "ai diagnostic":          "AI",
    "ai-based":               "AI",
    "dexis ai":               "AI",
    "intraoral scanner":      "Intraoral Scanners",
    "digital impression":     "Intraoral Scanners",
    "optical impression":     "Intraoral Scanners",
    "itero":                  "Intraoral Scanners",
    "3shape":                 "Intraoral Scanners",
    "trios":                  "Intraoral Scanners",
    "medit":                  "Intraoral Scanners",
    "carestream dental":      "Intraoral Scanners",
    "digital scan":           "Intraoral Scanners",
    "no impressions":         "Intraoral Scanners",
    "no messy impressions":   "Intraoral Scanners",
    "true definition":        "Intraoral Scanners",
    "digital intraoral":      "Intraoral Scanners",
}

ASSOCIATION_PATTERNS = [
    ("american dental association", "ADA"),
    (r"\bADA\b",                    "ADA"),
    ("american academy of pediatric", "AAPD"),
    (r"\bAAPD\b",                   "AAPD"),
    ("academy of general dentistry", "AGD"),
    (r"\bAGD\b",                    "AGD"),
    ("american academy of periodontology", "AAP"),
    (r"\bAAP\b",                    "AAP"),
    ("american academy of cosmetic", "AACD"),
    (r"\bAACD\b",                   "AACD"),
    ("international congress of oral implantologists", "ICOI"),
    (r"\bICOI\b",                   "ICOI"),
    ("american college of prosthodontists", "ACP"),
    (r"\bACP\b",                    "ACP"),
    ("american association of endodontists", "AAE"),
    (r"\bAAE\b",                    "AAE"),
    ("american association of orthodontists", "AAO"),
    (r"\bAAO\b",                    "AAO"),
]

SPECIALTY_MAP = {
    "cosmetic": "Cosmetic", "restorative": "Restorative",
    "implant": "Implants", "orthodontic": "Orthodontics",
    "periodontic": "Periodontics", "oral surgery": "Oral Surgery",
    "prosthodontic": "Prosthodontics", "tmj": "TMJ",
    "sleep apnea": "Sleep", "sedation": "Sedation",
    "holistic": "Holistic", "biological dent": "Biological",
    "pediatric": "Pediatric", "endodontic": "Endodontics",
    "emergency": "Emergency", "preventive": "Preventive",
    "family": "Family", "general dentistr": "General",
}


# ─────────────────────────────────────────────────────────────────────────────
# Page cache helpers  (same format as dental_scraper.py)
# ─────────────────────────────────────────────────────────────────────────────

def _cache_folder():
    slug = re.sub(r'[^\w]', '_', PRACTICE_NAME)[:30].strip('_')
    folder = os.path.join(CACHE_DIR, f"{TARGET_INDEX:03d}_{slug}")
    os.makedirs(folder, exist_ok=True)
    return folder


def _save_page(folder, page_type, url, html):
    """Write HTML file and update manifest.json."""
    fname = f"{page_type}.html"
    fpath = os.path.join(folder, fname)
    with open(fpath, "w", encoding="utf-8", errors="replace") as f:
        f.write(html)
    manifest_path = os.path.join(folder, "manifest.json")
    manifest = {}
    if os.path.exists(manifest_path):
        with open(manifest_path, encoding="utf-8") as f:
            try:
                manifest = json.load(f)
            except Exception:
                pass
    manifest.setdefault("pages", {})[page_type] = {
        "url": url,
        "file": fname,
        "saved_at": time.strftime("%Y-%m-%d %H:%M:%S"),
    }
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2, ensure_ascii=False)
    log.info("    Cached → %s", fpath)


def _save_result(folder, result):
    """Write result.json to the cache folder."""
    payload = {
        "practice": {"Index": TARGET_INDEX, "Practice Name": PRACTICE_NAME, "Website": BASE_URL},
        "scraped_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "result": {k: v for k, v in result.items() if k != "doctors"},
        "doctors": result.get("doctors", []),
    }
    with open(os.path.join(folder, "result.json"), "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)


# ─────────────────────────────────────────────────────────────────────────────
# HTTP fetch  (curl_cffi with browser impersonation → bypasses Cloudflare)
# ─────────────────────────────────────────────────────────────────────────────

def cffi_get(url, retries=3):
    """
    Fetch URL using curl_cffi impersonating Chrome (bypasses Cloudflare JS challenge).
    Returns (html_str, status_code) or (None, 0) on failure.
    """
    if not url:
        return None, 0
    if not url.startswith("http"):
        url = "https://" + url.lstrip("/")
    for attempt in range(retries):
        try:
            if CFFI_AVAILABLE:
                r = cffi_requests.get(
                    url,
                    impersonate=IMPERSONATE,
                    headers=HEADERS,
                    timeout=TIMEOUT,
                    verify=False,
                )
            else:
                import requests as plain_req
                r = plain_req.get(url, headers=HEADERS, timeout=TIMEOUT, verify=False)
            if r.status_code == 200:
                return r.text, 200
            if r.status_code in (403, 429, 503):
                log.warning("  HTTP %d for %s — waiting before retry", r.status_code, url)
                time.sleep(5 * (attempt + 1))
            else:
                log.debug("  HTTP %d for %s", r.status_code, url)
                break
        except Exception as e:
            log.debug("  cffi_get error attempt %d: %s", attempt, e)
            time.sleep(3)
    return None, 0


def _is_error_page(soup):
    """Return True if the page is a Cloudflare/server error page, not real content."""
    if soup is None:
        return True
    title = soup.title.get_text(strip=True) if soup.title else ""
    if re.search(r'520|521|522|524|cloudflare.*error|error.*cloudflare', title, re.I):
        return True
    h1 = soup.find("h1")
    if h1 and re.search(r'web server.*returning|error code 5\d\d', h1.get_text(), re.I):
        return True
    # Real dental sites have substantial content
    text_len = len(soup.get_text(strip=True))
    if text_len < 500:
        return True
    return False


# ─────────────────────────────────────────────────────────────────────────────
# Website scraping
# ─────────────────────────────────────────────────────────────────────────────

def fetch_all_pages(cache_folder):
    """
    Fetch homepage + all subpages. Cache each as HTML.
    Returns list of (url, soup, page_type) for pages with real content.
    """
    pages = []
    urls = [("homepage", BASE_URL)] + [(pt, urljoin(BASE_URL, path)) for path, pt in SUBPAGES]
    visited_norms = set()

    for page_type, url in urls:
        norm = url.rstrip("/")
        if norm in visited_norms:
            continue
        visited_norms.add(norm)

        log.info("  Fetching [%s]: %s", page_type, url)
        html, status = cffi_get(url)

        if html:
            soup = BeautifulSoup(html, "lxml")
            if _is_error_page(soup):
                log.warning("    Error/Cloudflare page detected — skipping")
            else:
                _save_page(cache_folder, page_type, url, html)
                pages.append((url, soup, page_type))
                log.info("    OK  (%d chars text)", len(soup.get_text(strip=True)))
        else:
            log.warning("    Could not fetch")

        time.sleep(DELAY_SEC)

    return pages


# ─────────────────────────────────────────────────────────────────────────────
# Data extraction
# ─────────────────────────────────────────────────────────────────────────────

def extract_doctors(pages):
    """
    Extract all doctor names from doctors/about/team pages.
    Returns list of clean name strings.
    """
    doctors = []
    seen_norms = []

    for url, soup, page_type in pages:
        if page_type not in ("doctors", "about", "team", "homepage"):
            continue
        for tag in soup.find_all(["h2", "h3", "h4"]):
            text = re.sub(r'\s+', ' ', tag.get_text(strip=True))
            if re.search(r'\bDr\.?\s+\w', text, re.I) and len(text) < 80:
                # Normalise case
                name = re.sub(r'\bDR\b\.?', 'Dr.', text, flags=re.I)
                name = re.sub(r'\s+', ' ', name).strip()
                # Dedup
                norm = re.sub(r'[^a-z ]', '', name.lower()).strip()
                if not any(_words_overlap(norm, s) for s in seen_norms):
                    doctors.append(name)
                    seen_norms.append(norm)

    return doctors if doctors else []


def _words_overlap(a, b):
    wa, wb = set(a.split()), set(b.split())
    if not wa or not wb:
        return False
    return bool(wa & wb) and (wa <= wb or wb <= wa)


def extract_email(pages):
    """Return first email found across all pages."""
    for url, soup, _ in pages:
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if href.startswith("mailto:"):
                email = href[7:].split("?")[0].strip()
                if "@" in email and "." in email:
                    return email
        # Cloudflare-obfuscated: data-cfemail
        for span in soup.find_all("span", attrs={"data-cfemail": True}):
            try:
                enc = span["data-cfemail"]
                key = int(enc[:2], 16)
                decoded = "".join(chr(int(enc[i:i+2], 16) ^ key) for i in range(2, len(enc), 2))
                if "@" in decoded:
                    return decoded
            except Exception:
                pass
    return None


def extract_social_links(pages):
    """Return dict of social media URLs found across all pages."""
    social = {"facebook": None, "instagram": None, "tiktok": None, "linkedin": None}
    for url, soup, _ in pages:
        for a in soup.find_all("a", href=True):
            href = a["href"].strip()
            if "facebook.com" in href and social["facebook"] is None:
                if "/groups/" not in href and "/posts/" not in href:
                    social["facebook"] = href
            elif "instagram.com" in href and social["instagram"] is None:
                if "/p/" not in href:
                    social["instagram"] = href
            elif "tiktok.com" in href and social["tiktok"] is None:
                social["tiktok"] = href
            elif "linkedin.com" in href and social["linkedin"] is None:
                social["linkedin"] = href
    return social


def extract_hygienists(all_text):
    m = re.search(r'(\d+)\s+(?:registered\s+)?(?:dental\s+)?hygienist', all_text, re.I)
    return int(m.group(1)) if m else None


def count_service_mentions(all_text):
    text_lower = all_text.lower()
    counts = {}
    for kw, col in SERVICE_KEYWORDS.items():
        cnt = text_lower.count(kw)
        if cnt > 0:
            counts[col] = counts.get(col, 0) + cnt
    return counts


def detect_tech_flags(all_text):
    text_lower = all_text.lower()
    return {col for kw, col in TECH_KEYWORDS.items() if kw in text_lower}


def extract_associations(all_text):
    found = set()
    for pattern, abbr in ASSOCIATION_PATTERNS:
        if re.search(pattern, all_text, re.I):
            found.add(abbr)
    return ", ".join(sorted(found)) if found else None


def detect_specialties(all_text):
    text_lower = all_text.lower()
    found = []
    for kw, label in SPECIALTY_MAP.items():
        if kw in text_lower and label not in found:
            found.append(label)
    return " / ".join(found) if found else None


def count_testimonials(pages):
    total = 0
    for url, soup, _ in pages:
        total += len(soup.find_all(attrs={"class": re.compile(r'testimonial|review', re.I)}))
        total += len(soup.find_all("blockquote"))
    return total


def detect_locations(pages, all_text):
    """Count distinct NY zip codes as a proxy for number of locations."""
    zips = set()
    for url, soup, _ in pages:
        for z in re.findall(r'\bNY\s+(\d{5})\b', soup.get_text()):
            zips.add(z)
    return len(zips) if len(zips) > 1 else None


# ─────────────────────────────────────────────────────────────────────────────
# Google rating  (requests first, Playwright fallback)
# ─────────────────────────────────────────────────────────────────────────────

def _parse_google_html(html):
    soup = BeautifulSoup(html, "lxml")
    g_rating, g_count = "", ""

    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string or "")
            if not isinstance(data, list):
                data = [data]
            for item in data:
                ar = item.get("aggregateRating") or {}
                rv = str(ar.get("ratingValue", "")).strip()
                rc = str(ar.get("reviewCount", "") or ar.get("ratingCount", "")).strip()
                if rv and re.match(r'^[1-5](\.\d)?$', rv):
                    return rv, rc
        except Exception:
            pass

    for tag in soup.find_all(attrs={"aria-label": True}):
        lbl = tag.get("aria-label", "")
        if "rated" in lbl.lower() and "out of 5" in lbl.lower():
            rm = re.search(r"([1-5]\.\d)", lbl)
            cm = re.search(r"\((\d[\d,]*)\)|(\d[\d,]*)\s*reviews?", lbl, re.I)
            if rm:
                return rm.group(1), (cm.group(1) or cm.group(2) or "").replace(",", "") if cm else ""

    text = soup.get_text(" ", strip=True)
    for pat in [
        r'([1-5]\.\d)\s*\((\d[\d,]+)\)\s*(?:google\s+)?reviews?',
        r'([1-5]\.\d)\s*·\s*(\d[\d,]+)\s*(?:google\s+)?reviews?',
        r'([1-5]\.\d)\s+(\d[\d,]+)\s+(?:google\s+)?reviews?',
        r'rated\s+([1-5]\.\d)\s+out\s+of\s+5.*?(\d[\d,]+)\s*(?:google\s+)?reviews?',
        r'([1-5]\.\d)\s+stars?\s+(\d[\d,]+)\s+(?:google\s+)?reviews?',
    ]:
        m = re.search(pat, text, re.I)
        if m:
            return m.group(1), m.group(2).replace(",", "")

    m = re.search(r'([1-5]\.\d)\s*/?\s*5\b', text)
    if m:
        g_rating = m.group(1)
    return g_rating, g_count


def get_google_rating(practice_name, city, state):
    """Try curl_cffi then Playwright to get Google rating."""
    query = f"{practice_name} {city} {state}"
    url = f"https://www.google.com/search?q={quote_plus(query)}&hl=en&gl=us"
    log.info("  Google search: %s", query)

    # Try 1: curl_cffi
    html, _ = cffi_get(url)
    if html:
        rating, count = _parse_google_html(html)
        if rating:
            log.info("  Google (curl_cffi): %s stars, %s reviews", rating, count)
            return rating, count

    # Try 2: Playwright
    if not PLAYWRIGHT_AVAILABLE:
        return "", ""
    log.info("  Google fallback: Playwright")
    try:
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            page = browser.new_page()
            page.set_extra_http_headers(HEADERS)
            page.goto(url, timeout=PW_TIMEOUT, wait_until="domcontentloaded")
            page.wait_for_timeout(4000)
            html = page.content()
            browser.close()
        rating, count = _parse_google_html(html)
        if rating:
            log.info("  Google (Playwright): %s stars, %s reviews", rating, count)
            return rating, count
        # Last resort: aria-label scan from rendered text
        for pat in [r'([1-5]\.\d)\s*\((\d[\d,]+)\)', r'([1-5]\.\d)\s+(\d[\d,]+)\s+reviews?']:
            m = re.search(pat, html, re.I)
            if m:
                return m.group(1), m.group(2).replace(",", "")
    except Exception as e:
        log.warning("  Playwright Google error: %s", e)
    return "", ""


# ─────────────────────────────────────────────────────────────────────────────
# Facebook followers  (Playwright)
# ─────────────────────────────────────────────────────────────────────────────

def get_facebook_followers(fb_url):
    if not PLAYWRIGHT_AVAILABLE or not fb_url:
        return None
    log.info("  FB followers: %s", fb_url)
    try:
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            page = browser.new_page()
            page.set_extra_http_headers(HEADERS)
            page.goto(fb_url, timeout=PW_TIMEOUT, wait_until="domcontentloaded")
            page.wait_for_timeout(5000)
            text = page.inner_text("body")
            browser.close()
        for pat in [r'([\d,]+)\s+people\s+follow', r'([\d,]+)\s+followers', r'([\d,]+)\s+likes']:
            m = re.search(pat, text, re.I)
            if m:
                return int(m.group(1).replace(",", ""))
    except Exception as e:
        log.warning("  FB error: %s", e)
    return None


# ─────────────────────────────────────────────────────────────────────────────
# xlsx helpers
# ─────────────────────────────────────────────────────────────────────────────

def load_workbook_info(xlsx_path):
    """
    Load workbook. 2-row header: row1=section labels, row2=column names, row3+=data.
    Returns (wb, ws, col_map, list_of_row_nums_for_index_137).
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    col_map = {}
    for cell in ws[2]:
        if cell.value is not None:
            col_map[str(cell.value).strip()] = cell.column

    idx_col = col_map.get("Index", 1)
    target_rows = []
    for r in range(3, ws.max_row + 1):
        val = ws.cell(row=r, column=idx_col).value
        try:
            if int(val) == TARGET_INDEX:
                target_rows.append(r)
        except (TypeError, ValueError):
            pass

    return wb, ws, col_map, target_rows


def read_row_values(ws, row_num, num_cols):
    """Return all cell values from a row as a list."""
    return [ws.cell(row=row_num, column=c).value for c in range(1, num_cols + 1)]


def write_doctor_rows(ws, col_map, first_row, doctor_list, practice_data):
    """
    Replace the existing row(s) for this practice with one row per doctor.
    `first_row` is the 1-based row number of the first existing row for this practice.
    `doctor_list` is a list of doctor name strings.
    `practice_data` is a dict {col_name: value} with all shared practice fields.
    """
    num_doctors = len(doctor_list)
    num_cols = ws.max_column

    # Copy formatting/styles from first_row as template
    template_values = read_row_values(ws, first_row, num_cols)
    template_styles = {}
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=first_row, column=c)
        template_styles[c] = {
            "font": cell.font.copy() if cell.font else None,
            "fill": cell.fill.copy() if cell.fill else None,
            "alignment": cell.alignment.copy() if cell.alignment else None,
            "border": cell.border.copy() if cell.border else None,
            "number_format": cell.number_format,
        }

    # If we need more rows than currently exist, insert them
    # (If fewer, extra rows will just be left as-is for now)
    if num_doctors > 1:
        ws.insert_rows(first_row + 1, num_doctors - 1)

    doc_col = col_map.get("Doctor Name", 3)

    for i, doctor_name in enumerate(doctor_list):
        row_num = first_row + i
        # Write all shared practice data
        for col_name, value in practice_data.items():
            if col_name not in col_map:
                continue
            col_idx = col_map[col_name]
            cell = ws.cell(row=row_num, column=col_idx)
            cell.value = value
            style = template_styles.get(col_idx, {})
            if style.get("font"):
                cell.font = style["font"]
            if style.get("fill"):
                cell.fill = style["fill"]
            if style.get("alignment"):
                cell.alignment = style["alignment"]
            if style.get("number_format"):
                cell.number_format = style["number_format"]

        # Write this doctor's name
        ws.cell(row=row_num, column=doc_col).value = doctor_name
        log.info("  Row %d: %s", row_num, doctor_name)


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    log.info("=" * 60)
    log.info("scrape_row_137  —  Lalor Family Dental  (Index %d)", TARGET_INDEX)
    log.info("=" * 60)

    # ── Setup cache folder ────────────────────────────────────────────────────
    cache_folder = _cache_folder()
    log.info("Cache folder: %s", cache_folder)

    # ── Scrape the website ────────────────────────────────────────────────────
    log.info("Fetching practice website (%s) ...", BASE_URL)
    pages = fetch_all_pages(cache_folder)

    site_ok = len(pages) > 0
    if not site_ok:
        log.warning("Site unreachable — falling back to KNOWN_DATA for doctors/social.")

    # Build combined text corpus
    all_text_parts = []
    for url, soup, _ in pages:
        for tag in soup(["script", "style", "noscript"]):
            tag.decompose()
        all_text_parts.append(soup.get_text(separator=" ", strip=True))
    all_text = " ".join(all_text_parts)

    # ── Extract fields ────────────────────────────────────────────────────────
    doctors         = extract_doctors(pages) or KNOWN_DOCTORS
    email           = extract_email(pages)
    social          = extract_social_links(pages)
    hygienists      = extract_hygienists(all_text)
    service_counts  = count_service_mentions(all_text)
    tech_flags      = detect_tech_flags(all_text)
    associations    = extract_associations(all_text)
    specialties     = detect_specialties(all_text)
    testimonials    = count_testimonials(pages)
    num_locations   = detect_locations(pages, all_text)

    log.info("Doctors found:  %d", len(doctors))
    log.info("Email:          %s", email)
    log.info("Social FB:      %s", social.get("facebook"))
    log.info("Hygienists:     %s", hygienists)
    log.info("Tech flags:     %s", tech_flags)
    log.info("Service counts: %s", service_counts)
    log.info("Associations:   %s", associations)
    log.info("Specialties:    %s", specialties)

    # ── Google rating ─────────────────────────────────────────────────────────
    log.info("Fetching Google rating ...")
    time.sleep(DELAY_SEC)
    g_rating, g_count = get_google_rating("Lalor Family Dental", "Elmira", "NY")
    log.info("Google: %s stars / %s reviews", g_rating or "—", g_count or "—")

    # ── Facebook followers ────────────────────────────────────────────────────
    fb_url = social.get("facebook") or KNOWN_FACEBOOK
    time.sleep(DELAY_SEC)
    fb_followers = get_facebook_followers(fb_url)
    log.info("FB followers:   %s", fb_followers)

    # ── Save result.json to cache ─────────────────────────────────────────────
    _save_result(cache_folder, {
        "doctors":          doctors,
        "email":            email,
        "facebook_url":     fb_url,
        "fb_followers":     fb_followers,
        "instagram_url":    social.get("instagram"),
        "tiktok_url":       social.get("tiktok"),
        "linkedin_url":     social.get("linkedin"),
        "hygienists":       hygienists,
        "tech_flags":       list(tech_flags),
        "service_counts":   service_counts,
        "associations":     associations,
        "specialties":      specialties,
        "testimonials":     testimonials,
        "num_locations":    num_locations,
        "google_rating":    g_rating,
        "google_reviews":   g_count,
    })

    # ── Build shared practice data dict ──────────────────────────────────────
    practice_data = {
        "Index":                          TARGET_INDEX,
        "Practice Name":                  PRACTICE_NAME,
        "Address":                        "1052 County Road 64",
        "City":                           "Elmira",
        "State":                          "NY",
        "Zip":                            14903,
        "Practice Website":               BASE_URL,
        "Facebook URL":                   fb_url,
    }
    if email:
        practice_data["Practice Email"] = email
    if hygienists is not None:
        practice_data["# of Hygienists"] = hygienists
    if social.get("instagram"):
        practice_data["Instagram URL"] = social["instagram"]
    if social.get("tiktok"):
        practice_data["TikTok URL"] = social["tiktok"]
    if social.get("linkedin"):
        practice_data["LinkedIn URL"] = social["linkedin"]
    if fb_followers is not None:
        practice_data["FB Followers"] = fb_followers
    for col_name in tech_flags:
        practice_data[col_name] = "X"
    for col_name, cnt in service_counts.items():
        if cnt > 0:
            practice_data[col_name] = cnt
    if service_counts.get("Invisalign (Mentions)", 0) == 0:
        practice_data["Invisalign (Mentions)"] = 0
        practice_data["Invisalign Tier (check manually)"] = "N/A \u2013 Not Offered"
    else:
        practice_data["Invisalign Tier (check manually)"] = "Check manually"
    if associations:
        practice_data["Associations / Memberships"] = associations
    if specialties:
        practice_data["Doctor Specialty"] = specialties
    if testimonials > 0:
        practice_data["Testimonials (Number of)"] = testimonials
    if num_locations is not None:
        practice_data["# of Locations"] = num_locations
    else:
        practice_data["# of Locations"] = 1
    if g_rating:
        try:
            practice_data["Google Reviews Ranking"] = float(g_rating)
        except ValueError:
            practice_data["Google Reviews Ranking"] = g_rating
    if g_count:
        try:
            practice_data["Total # of Google Reviews"] = int(g_count)
        except ValueError:
            practice_data["Total # of Google Reviews"] = g_count

    # ── Update xlsx: one row per doctor ──────────────────────────────────────
    log.info("Loading %s ...", XLSX_FILE)
    wb, ws, col_map, existing_rows = load_workbook_info(XLSX_FILE)

    if not existing_rows:
        log.error("Index=%d not found in xlsx!", TARGET_INDEX)
        return

    first_row = existing_rows[0]
    log.info("Writing %d doctor rows starting at row %d ...", len(doctors), first_row)
    write_doctor_rows(ws, col_map, first_row, doctors, practice_data)

    wb.save(XLSX_FILE)
    log.info("Saved: %s", XLSX_FILE)

    if not site_ok:
        log.warning(
            "\nNOTE: Website was unreachable — service/tech fields are empty.\n"
            "Re-run this script once the site is back up."
        )
    else:
        log.info("All fields filled from live site.")
    log.info("Done.  %d rows written for practice #%d.", len(doctors), TARGET_INDEX)


if __name__ == "__main__":
    main()
