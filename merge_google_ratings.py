import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border
import copy

RATINGS_FILE = "batch2/google_ratings_output.xlsx"
DEDUPED_FILE = "batch2/batch2_deduped.xlsx"
OUTPUT_FILE = "batch2/batch2_deduped_updated.xlsx"

# --- Load ratings and build index -> (ranking, total) map ---
ratings = pd.read_excel(RATINGS_FILE)
ratings_map = (
    ratings.groupby("#")[["Google Reviews Ranking", "Total # of Google Reviews"]]
    .first()
)

# --- Load deduped as raw dataframe (preserve 2-row header) ---
raw = pd.read_excel(DEDUPED_FILE, header=None)

# Row 0: section headers, Row 1: column names, Row 2+: data
INDEX_COL = 0       # column A — "Index"
RANKING_COL = 41    # column AP — "Google Reviews Ranking"
TOTAL_COL = 42      # column AQ — "Total # of Google Reviews"

updated = 0
skipped = 0

for row_i in range(2, len(raw)):
    idx = raw.iat[row_i, INDEX_COL]
    if pd.isna(idx):
        skipped += 1
        continue
    try:
        idx_int = int(idx)
    except (ValueError, TypeError):
        skipped += 1
        continue

    if idx_int in ratings_map.index:
        raw.iat[row_i, RANKING_COL] = ratings_map.at[idx_int, "Google Reviews Ranking"]
        raw.iat[row_i, TOTAL_COL] = ratings_map.at[idx_int, "Total # of Google Reviews"]
        updated += 1
    else:
        # Index not in ratings file — clear stale values
        raw.iat[row_i, RANKING_COL] = None
        raw.iat[row_i, TOTAL_COL] = None
        skipped += 1

# --- Write output preserving original formatting via openpyxl copy ---
# First write the updated data to a temp xlsx, then copy cell styles from original
raw.to_excel(OUTPUT_FILE, index=False, header=False)

# Re-apply styles from original file
wb_orig = load_workbook(DEDUPED_FILE)
wb_new = load_workbook(OUTPUT_FILE)
ws_orig = wb_orig.active
ws_new = wb_new.active

for row in ws_orig.iter_rows():
    for cell in row:
        new_cell = ws_new.cell(row=cell.row, column=cell.column)
        if cell.has_style:
            new_cell.font = copy.copy(cell.font)
            new_cell.fill = copy.copy(cell.fill)
            new_cell.border = copy.copy(cell.border)
            new_cell.alignment = copy.copy(cell.alignment)
            new_cell.number_format = cell.number_format

# Copy merged cells
for merge in ws_orig.merged_cells.ranges:
    ws_new.merge_cells(str(merge))

# Copy column widths
for col_letter, col_dim in ws_orig.column_dimensions.items():
    ws_new.column_dimensions[col_letter].width = col_dim.width

wb_new.save(OUTPUT_FILE)

print(f"Done.")
print(f"  Rows updated : {updated}")
print(f"  Rows skipped : {skipped} (no matching index in ratings file)")
print(f"  Output saved : {OUTPUT_FILE}")
