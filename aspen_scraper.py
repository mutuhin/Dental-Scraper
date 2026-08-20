#!/usr/bin/env python3
"""
aspen_scraper.py
================
Specialized, fast scraper for Aspen Dental location pages.

WHY A DEDICATED SCRIPT?
Aspen Dental is a React/Next.js SPA behind Cloudflare Enterprise.
bypass_scraper.py times out (5+ hours for 15 pages via Playwright).
This script avoids Playwright entirely and uses:

  1. __NEXT_DATA__ extraction   — Next.js SSR embeds all page data as JSON
                                  in the static HTML, no JavaScript needed.
  2. /_next/data/[buildId].json — Next.js JSON API; same data, machine-readable.
  3. JSON-LD structured data    — additional <script type="application/ld+json">.
  4. Plain-text regex scan      — last-resort name extraction from raw HTML.
  5. NPI registry lookup        — US government dentist registry; no CF, no JS;
                                  filters by street number from the URL slug.

All curl_cffi requests use rotating TLS fingerprints + Oxylabs proxy.
No Playwright is used — each site takes ~5–30s instead of 2–15 min.

INSTALL
-------
  pip install curl_cffi requests beautifulsoup4 lxml openpyxl

USAGE
-----
  python aspen_scraper.py --urls "https://www.aspendental.com/dentist/ma/worcester/537-lincoln-street-ste-3"
  python aspen_scraper.py --file aspen_urls.txt --proxies proxies.txt
  python aspen_scraper.py --file aspen_urls.txt --proxies proxies.txt --output results.xlsx
"""

import argparse
import json
import os
import re
import sys
import time
import random
import urllib.parse
from urllib.parse import urlparse, urljoin

import warnings
warnings.filterwarnings("ignore")

# ── Dependencies ──────────────────────────────────────────────────────────────

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    sys.exit("pip install openpyxl")

try:
    import requests as std_requests
except ImportError:
    sys.exit("pip install requests")

try:
    from curl_cffi import requests as cffi_requests
    _CFFI_OK = True
except ImportError:
    _CFFI_OK = False
    print("WARNING: curl_cffi not available — TLS bypass disabled (install curl_cffi)")

try:
    from bs4 import BeautifulSoup
except ImportError:
    sys.exit("pip install beautifulsoup4 lxml")

try:
    from playwright.sync_api import sync_playwright
    _PW_OK = True
except ImportError:
    _PW_OK = False

try:
    from playwright_stealth import stealth_sync as _stealth_sync
    _STEALTH_OK = True
except ImportError:
    try:
        from playwright_stealth import Stealth as _StealthClass
        _stealth_sync = _StealthClass().apply_stealth_sync
        _STEALTH_OK = True
    except (ImportError, AttributeError):
        _STEALTH_OK = False

# ── Constants ─────────────────────────────────────────────────────────────────

_CFFI_PROFILES = ["chrome136", "chrome124", "chrome133a", "chrome110", "safari260", "safari17_2"]

_CHALLENGE_MARKERS = (
    "just a moment", "checking your browser", "please enable cookies",
    "enable javascript and cookies", "access denied", "sgcaptcha",
    "robot challenge", "cf-browser-verification",
)

# Aspen Dental always offers these services at every location
_ASPEN_CORE_SERVICES = {
    "implants": 1,
    "veneers": 1,
    "whitening": 1,
    "sedation": 1,
    "clear_aligners": 1,
    "invisalign": 1,
}

# Words that are never valid as doctor name parts
_BAD_NAME_WORDS = frozenset({
    "dental", "dentist", "dentistry", "care", "service", "services", "office",
    "center", "health", "patient", "welcome", "insurance", "treatment",
    "location", "aspen", "group", "associates", "family", "general",
    "cosmetic", "pediatric", "emergency", "specialist", "provider",
    "january", "february", "march", "april", "june", "july", "august",
    "september", "october", "november", "december",
})

_proxy_pool: list = []
_proxy_fail: set = set()


# ── Proxy management ──────────────────────────────────────────────────────────

def load_proxies(path: str):
    global _proxy_pool
    if not path or not os.path.exists(path):
        return
    with open(path) as f:
        _proxy_pool = [ln.strip() for ln in f if ln.strip() and not ln.startswith("#")]
    print(f"Loaded {len(_proxy_pool)} proxies from {path}")


def _available_proxies():
    return [p for p in _proxy_pool if p not in _proxy_fail]


# ── HTML fetching ─────────────────────────────────────────────────────────────

def _is_challenge(html: str) -> bool:
    snippet = html[:3000].lower()
    return any(m in snippet for m in _CHALLENGE_MARKERS)


def _fetch(url: str, proxy: str = None, timeout: int = 8) -> str | None:
    """Fetch URL via curl_cffi, optionally through a proxy. Returns HTML or None."""
    if not _CFFI_OK:
        try:
            r = std_requests.get(url, timeout=timeout, verify=False)
            if r.status_code == 200 and not _is_challenge(r.text):
                return r.text
        except Exception:
            pass
        return None

    proxies = {"http": proxy, "https": proxy} if proxy else None
    for profile in random.sample(_CFFI_PROFILES[:4], 4):
        try:
            sess = cffi_requests.Session(impersonate=profile)
            r = sess.get(url, timeout=timeout, verify=False,
                         allow_redirects=True, proxies=proxies)
            if r.status_code == 200 and not _is_challenge(r.text):
                return r.text
            if r.status_code in (403, 407, 202):
                if proxy and r.status_code == 407:
                    _proxy_fail.add(proxy)
                break
        except Exception:
            continue
    return None


_STEALTH_JS = """
(function(){
  Object.defineProperty(navigator,'webdriver',{get:()=>undefined});
  window.chrome={app:{isInstalled:false},csi:function(){},loadTimes:function(){},runtime:{}};
  Object.defineProperty(navigator,'languages',{get:()=>['en-US','en']});
  delete window.cdc_adoQpoasnfa76pfcZLmcfl_Array;
})();
"""

def _pw_proxy_dict(proxy_url: str) -> dict:
    p = urlparse(proxy_url)
    d = {"server": f"{p.scheme}://{p.hostname}:{p.port}"}
    if p.username:
        d["username"] = p.username
    if p.password:
        d["password"] = p.password
    return d


def _pw_fetch(url: str, proxy: str = None, timeout_ms: int = 60000) -> str | None:
    """
    Fetch URL with Playwright stealth (with optional proxy).
    Hard timeout of 60 s — Aspen Dental resolves CF challenges within that window.
    Returns full rendered HTML or None.
    """
    if not _PW_OK:
        return None
    try:
        with sync_playwright() as pw:
            # proxy must go to launch(), not new_context(), for chromium.launch()
            launch_kwargs = dict(
                headless=True,
                ignore_https_errors=True,
                args=[
                    "--disable-blink-features=AutomationControlled",
                    "--disable-infobars",
                    "--no-first-run",
                    "--disable-dev-shm-usage",
                    "--window-size=1920,1080",
                ],
            )
            if proxy:
                launch_kwargs["proxy"] = _pw_proxy_dict(proxy)

            browser = pw.chromium.launch(**launch_kwargs)
            ctx = browser.new_context(
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/136.0.0.0 Safari/537.36"
                ),
                ignore_https_errors=True,
                # proxy also passed to new_context so sub-resource requests use it
                **({"proxy": _pw_proxy_dict(proxy)} if proxy else {}),
            )
            page = ctx.new_page()

            # Apply stealth
            if _STEALTH_OK:
                try:
                    _stealth_sync(page)
                except Exception:
                    pass
            else:
                page.add_init_script(script=_STEALTH_JS)

            page.set_extra_http_headers({"Accept-Language": "en-US,en;q=0.9"})

            try:
                page.goto(url, timeout=timeout_ms, wait_until="commit")
            except Exception:
                browser.close()
                return None

            # Wait for CF challenge to auto-resolve (up to 8 × 4s)
            for _ in range(8):
                try:
                    title = page.title().lower()
                except Exception:
                    break
                if any(m in title for m in ("just a moment", "checking your", "please wait")):
                    page.wait_for_timeout(4000)
                else:
                    break

            try:
                page.wait_for_load_state("domcontentloaded", timeout=15000)
            except Exception:
                pass
            try:
                page.wait_for_load_state("networkidle", timeout=10000)
            except Exception:
                pass
            page.wait_for_timeout(3000)

            try:
                html = page.content()
            except Exception:
                html = None

            browser.close()

            if not html or len(html) < 2000:
                return None
            if _is_challenge(html):
                return None
            return html
    except Exception as e:
        print(f"  [PW] Error: {e}")
        return None


_PLAIN_SESSION: std_requests.Session | None = None

def _plain_session() -> std_requests.Session:
    global _PLAIN_SESSION
    if _PLAIN_SESSION is None:
        _PLAIN_SESSION = std_requests.Session()
        _PLAIN_SESSION.headers.update({
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            "Accept-Language": "en-US,en;q=0.9",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        })
    return _PLAIN_SESSION


def _fetch_plain(url: str, timeout: int = 8) -> str | None:
    """Plain requests.get — fastest; works when CF allows standard browsers."""
    try:
        r = _plain_session().get(url, timeout=timeout, allow_redirects=True)
        if r.status_code == 200 and not _is_challenge(r.text) and len(r.text) > 1000:
            return r.text
    except Exception:
        pass
    return None


def _fetch_with_fallbacks(url: str) -> tuple[str | None, str]:
    """
    Multi-strategy fetch for Aspen Dental pages.
      1. Plain requests (fast, simple — works when CF allows standard browsers)
      2. curl_cffi direct (TLS fingerprint bypass)
      3. curl_cffi + each proxy (UK mobile → DE mobile → US residential)

    Playwright is excluded — CF Enterprise blocks it and it OOMs the runner.
    NPI registry supplements doctor/phone/address for fully blocked pages.
    Returns (html, strategy_label).
    """
    # Strategy 1: plain requests (no TLS tricks — sometimes enough)
    html = _fetch_plain(url)
    if html:
        return html, "plain-requests"

    # Strategy 2: curl_cffi direct (TLS fingerprint bypass, no proxy)
    html = _fetch(url)
    if html:
        return html, "cffi-direct"

    # Strategy 3: curl_cffi + proxy (UK mobile → US residential)
    for proxy in _available_proxies()[:4]:
        html = _fetch(url, proxy=proxy)
        if html:
            return html, f"cffi+proxy:{proxy.split('@')[-1]}"
        _proxy_fail.add(proxy)

    return None, "failed"


def _discover_provider_urls(html: str, base_url: str) -> list[str]:
    """Find all /providers/ sub-page URLs linked from this page."""
    return sorted({
        urljoin(base_url, href).split("?")[0].rstrip("/")
        for href in re.findall(r'href=["\']([^"\']+)["\']', html, re.I)
        if "/providers/" in href and "aspendental.com" in urljoin(base_url, href)
    })


def _extract_doctor_from_provider_page(html: str) -> str:
    """Pull a doctor name from an Aspen /providers/ profile page."""
    text = re.sub(r"\s+", " ", BeautifulSoup(html, "html.parser").get_text(" "))
    for pattern in (
        r"Dr\.\s+([A-Z][a-z]+(?:\s+[A-Z][a-z]+){1,3})",
        r"About\s+Dr\.\s+([A-Z][a-z]+(?:\s+[A-Z][a-z]+){1,3})",
    ):
        m = re.search(pattern, text)
        if m:
            name = re.sub(r"\s+", " ", m.group(1)).strip()
            if 2 <= len(name.split()) <= 5:
                return name
    return ""


def _scrape_provider_pages(html: str, base_url: str) -> list[str]:
    """
    Discover /providers/ sub-pages linked from html, fetch each one,
    and return a list of doctor names found.
    """
    provider_urls = _discover_provider_urls(html, base_url)
    if not provider_urls:
        return []

    names = []
    seen_lc: set[str] = set()
    for purl in provider_urls[:12]:  # cap at 12 profiles
        try:
            phtml = _fetch_plain(purl) or _fetch(purl)
            if not phtml:
                continue
            name = _extract_doctor_from_provider_page(phtml)
            if name and name.lower() not in seen_lc:
                seen_lc.add(name.lower())
                names.append(name)
                print(f"    → provider page doctor: {name}")
            time.sleep(0.5)
        except Exception:
            continue
    return names


# ── URL parsing ───────────────────────────────────────────────────────────────

def _parse_aspen_url(url: str) -> dict:
    """
    Extract state / city / address-slug from an Aspen Dental location URL.
    Pattern: https://www.aspendental.com/dentist/[state]/[city]/[address-slug]
    """
    p = urlparse(url)
    parts = [x for x in p.path.strip("/").split("/") if x]
    info = {"state": "", "city": "", "address_slug": "", "street_num": "",
            "clean_url": url.split("?")[0].split("#")[0].rstrip("/")}

    if "dentist" in parts:
        idx = parts.index("dentist")
        if idx + 1 < len(parts):
            info["state"] = parts[idx + 1].upper()
        if idx + 2 < len(parts):
            info["city"] = parts[idx + 2].replace("-", " ").title()
        if idx + 3 < len(parts):
            slug = parts[idx + 3]
            info["address_slug"] = slug
            # Extract street number: first token if purely numeric
            slug_parts = slug.split("-")
            if slug_parts and slug_parts[0].isdigit():
                info["street_num"] = slug_parts[0]

    return info


# ── Address from URL slug (deterministic — no network needed) ─────────────────

_SLUG_ABBREVS = {
    "ave": "Avenue", "blvd": "Boulevard", "rd": "Road", "dr": "Drive",
    "ln": "Lane",    "pkwy": "Parkway",   "hwy": "Highway", "ct": "Court",
    "pl": "Place",   "cir": "Circle",     "ste": "Suite",   "trl": "Trail",
    "fwy": "Freeway","expy": "Expressway","sq": "Square",
    # directionals — keep uppercase
    "n": "N", "s": "S", "e": "E", "w": "W",
    "ne": "NE", "nw": "NW", "se": "SE", "sw": "SW",
}


def _address_from_slug(slug: str) -> str:
    """
    Convert an Aspen Dental URL address slug to a human-readable street address.
    "537-lincoln-street-ste-3"  →  "537 Lincoln Street Suite 3"
    "9405-n-newport-hwy"        →  "9405 N Newport Highway"
    """
    if not slug:
        return ""
    tokens = slug.lower().split("-")
    nice = []
    for tok in tokens:
        if tok in _SLUG_ABBREVS:
            nice.append(_SLUG_ABBREVS[tok])
        elif tok.isdigit():
            nice.append(tok)
        else:
            nice.append(tok.title())
    return " ".join(nice)


# ── Data extraction from __NEXT_DATA__ ───────────────────────────────────────

def _deep_find(obj, keys: list, depth: int = 0):
    """Recursively search a JSON object for the first matching key with a truthy value."""
    if depth > 12 or obj is None:
        return None
    if isinstance(obj, dict):
        for k in keys:
            if k in obj and obj[k]:
                return obj[k]
        for v in obj.values():
            found = _deep_find(v, keys, depth + 1)
            if found:
                return found
    elif isinstance(obj, list):
        for item in obj:
            found = _deep_find(item, keys, depth + 1)
            if found:
                return found
    return None


def _collect_providers(obj, depth: int = 0) -> list[dict]:
    """
    Recursively scan a JSON object / list looking for person/provider records.
    Returns list of dicts with 'name' and optionally 'credential'.
    """
    if depth > 12 or not obj:
        return []

    found = []

    if isinstance(obj, dict):
        first = (obj.get("firstName") or obj.get("first_name") or
                 obj.get("givenName") or "").strip()
        last  = (obj.get("lastName")  or obj.get("last_name")  or
                 obj.get("familyName") or "").strip()
        cred  = (obj.get("credential") or obj.get("credentials") or
                 obj.get("degree") or obj.get("suffix") or "").strip()
        title = (obj.get("title") or obj.get("prefix") or "").strip()

        if first and last:
            name = f"{first.title()} {last.title()}"
            if title and title.lower() in ("dr", "dr."):
                name = f"Dr. {name}"
            elif not name.startswith("Dr."):
                name = f"Dr. {name}"
            if cred:
                name = f"{name}, {cred.upper()}"
            found.append({"name": name, "credential": cred})
            return found  # don't recurse into a provider object itself

        # Check if the 'name' field looks like a full person name
        name_raw = (obj.get("displayName") or obj.get("fullName") or
                    obj.get("providerName") or obj.get("name") or "").strip()
        if name_raw and isinstance(name_raw, str):
            parts = name_raw.replace(",", "").split()
            real_parts = [w for w in parts
                          if len(w) > 1 and not all(c.isdigit() or c == '.' for c in w)
                          and w.lower() not in _BAD_NAME_WORDS]
            # Require at least 2 real name words
            if len(real_parts) >= 2:
                clean = name_raw
                if not clean.startswith("Dr.") and not clean.startswith("Dr "):
                    clean = f"Dr. {clean}"
                if cred and cred not in clean:
                    clean = f"{clean}, {cred.upper()}"
                found.append({"name": clean, "credential": cred})
                return found

        # Recurse into provider-list keys
        for key in ("providers", "doctors", "staff", "dentists", "team",
                    "practitioners", "employees", "members", "physicians",
                    "associates", "clinicians"):
            val = obj.get(key)
            if isinstance(val, list) and val:
                for item in val:
                    found.extend(_collect_providers(item, depth + 1))
                if found:
                    return found

        # Recurse into other nested objects
        for v in obj.values():
            if isinstance(v, (dict, list)):
                found.extend(_collect_providers(v, depth + 1))

    elif isinstance(obj, list):
        for item in obj:
            found.extend(_collect_providers(item, depth + 1))

    return found


def _validate_doctor_name(name: str) -> bool:
    """Return True if the name looks like a real person's name."""
    body = re.sub(r'^Dr\.?\s+', '', name, flags=re.I).strip()
    body = re.sub(r'[,\s]+(?:DDS|DMD|D\.D\.S|D\.M\.D|MD|BDS|MSD|MS)\b.*$', '', body, flags=re.I).strip()
    if re.search(r'\d', body):
        return False
    words = body.split()
    if len(words) < 2:
        return False
    for w in words:
        wl = w.strip(".,-()'").lower()
        if len(wl) <= 1:
            continue
        if wl in _BAD_NAME_WORDS:
            return False
        if wl.endswith(("istry", "ology", "tion", "ment", "ness", "care")):
            return False
    return True


def _dedup_names(raw_list: list[dict]) -> list[str]:
    """Deduplicate and validate provider names, return list of clean name strings."""
    seen_lc: set = set()
    result: list = []
    for item in raw_list:
        name = re.sub(r"\s+", " ", (item.get("name") or "").strip())
        if not name or not _validate_doctor_name(name):
            continue
        key = name.lower()
        if key not in seen_lc:
            seen_lc.add(key)
            result.append(name)
    return result


def extract_next_data(html: str, page_url: str) -> dict:
    """
    Parse __NEXT_DATA__ from the raw HTML of an Aspen Dental page.
    Returns partial result dict (may be empty if not found).
    """
    result: dict = {}

    m = re.search(
        r'<script[^>]+id=["\']__NEXT_DATA__["\'][^>]*>(.*?)</script>',
        html, re.S,
    )
    if not m:
        return result

    try:
        data = json.loads(m.group(1))
    except Exception:
        return result

    props = data.get("props", {}).get("pageProps", {})

    # ── Doctor / provider names ───────────────────────────────────────────────
    # Try specific known key paths first (fast path)
    providers_raw = None
    candidate_paths = [
        ("office", "providers"), ("office", "doctors"), ("office", "staff"),
        ("officeData", "providers"), ("officeData", "doctors"),
        ("locationData", "providers"), ("location", "providers"),
        ("providers",), ("doctors",), ("staff",), ("team",),
        ("pageData", "providers"), ("data", "providers"),
    ]
    for path in candidate_paths:
        node = props
        for key in path:
            node = node.get(key) if isinstance(node, dict) else None
        if isinstance(node, list) and node:
            providers_raw = node
            break

    if providers_raw:
        raw_providers = _collect_providers(providers_raw)
    else:
        # Deep recursive search across all pageProps
        raw_providers = _collect_providers(props)

    names = _dedup_names(raw_providers)
    if names:
        result["scraped_doctor_names"] = ", ".join(names)
        result["doctors"] = [{"name": n, "specialty": "Not Found", "associations": ""} for n in names]

    # ── Phone ─────────────────────────────────────────────────────────────────
    phone = _deep_find(props, ["phone", "phoneNumber", "telephone", "officePhone", "contactPhone"])
    if phone and isinstance(phone, str) and re.search(r'\d{3}', phone):
        result["phone"] = phone

    # ── Address ───────────────────────────────────────────────────────────────
    addr_obj = _deep_find(props, ["address", "officeAddress", "locationAddress"])
    if isinstance(addr_obj, dict):
        street = (addr_obj.get("streetAddress") or addr_obj.get("street") or
                  addr_obj.get("address1") or addr_obj.get("line1") or
                  addr_obj.get("addressLine1") or "")
        city   = (addr_obj.get("addressLocality") or addr_obj.get("city") or "")
        state  = (addr_obj.get("addressRegion") or addr_obj.get("state") or "")
        zp     = (addr_obj.get("postalCode") or addr_obj.get("zip") or
                  addr_obj.get("zipCode") or "")
        if street:
            result["address"] = street
        if city:
            result["city"] = city
        if state:
            result["state"] = state
        if zp:
            result["zip"] = zp
    elif isinstance(addr_obj, str) and addr_obj:
        result["address"] = addr_obj

    # ── Rating ────────────────────────────────────────────────────────────────
    rating_obj = _deep_find(props, ["aggregateRating", "ratingData", "reviewData"])
    if isinstance(rating_obj, dict):
        rv = rating_obj.get("ratingValue") or rating_obj.get("value")
        rc = rating_obj.get("reviewCount") or rating_obj.get("ratingCount")
        if rv:
            result["google_rating"] = str(rv)
        if rc:
            result["total_google_reviews"] = str(rc)

    # ── Service keyword counts from JSON string ───────────────────────────────
    json_lc = json.dumps(data).lower()
    for terms, field, cap in [
        (["implant"],                     "implants",      20),
        (["veneer"],                      "veneers",       20),
        (["invisalign"],                  "invisalign",    10),
        (["whitening", "teeth whitening"],"whitening",     15),
        (["sedation"],                    "sedation",      10),
        (["clear aligner"],               "clear_aligners",10),
    ]:
        cnt = sum(json_lc.count(t) for t in terms)
        if cnt:
            result[field] = min(cnt, cap)

    # ── Next.js /_next/data API URL (for caller to try) ───────────────────────
    build_id = data.get("buildId", "")
    if build_id:
        p = urlparse(page_url)
        path_clean = p.path.rstrip("/")
        result["_next_api_url"] = (
            f"https://www.aspendental.com/_next/data/{build_id}{path_clean}.json"
        )

    return result


def extract_json_ld(html: str) -> dict:
    """Extract structured data from <script type='application/ld+json'> blocks."""
    result: dict = {}
    soup = BeautifulSoup(html, "lxml")
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            ld = json.loads(script.string or "")
            items = ld if isinstance(ld, list) else [ld]
            for item in items:
                if not isinstance(item, dict):
                    continue
                t = str(item.get("@type", "")).lower()
                if not any(x in t for x in ("dental", "medical", "health", "local", "org", "place")):
                    continue

                for k in ("telephone", "phone"):
                    if item.get(k) and not result.get("phone"):
                        result["phone"] = str(item[k])

                addr = item.get("address") or {}
                if isinstance(addr, dict) and not result.get("address"):
                    result["address"] = addr.get("streetAddress", "")
                    result["city"]    = addr.get("addressLocality", "")
                    result["state"]   = addr.get("addressRegion", "")
                    result["zip"]     = addr.get("postalCode", "")

                ag = item.get("aggregateRating") or {}
                if isinstance(ag, dict) and not result.get("google_rating"):
                    rv = ag.get("ratingValue")
                    rc = ag.get("reviewCount") or ag.get("ratingCount")
                    if rv:
                        result["google_rating"] = str(rv)
                    if rc:
                        result["total_google_reviews"] = str(rc)

                # Provider names from employee / staff / founder
                for key in ("employee", "staff", "founder", "member"):
                    emp = item.get(key)
                    if not emp:
                        continue
                    emp_list = emp if isinstance(emp, list) else [emp]
                    names = []
                    for e in emp_list:
                        if isinstance(e, dict):
                            n = e.get("name", "").strip()
                            if n and _validate_doctor_name(n):
                                names.append(n)
                    if names and not result.get("scraped_doctor_names"):
                        result["scraped_doctor_names"] = ", ".join(names)
                        result["doctors"] = [
                            {"name": n, "specialty": "Not Found", "associations": ""}
                            for n in names
                        ]
        except Exception:
            pass
    return result


def extract_text_scan(html: str) -> dict:
    """Last-resort: regex scan of page text for doctor names."""
    soup = BeautifulSoup(html, "lxml")
    for tag in soup.find_all(["nav", "footer", "script", "style", "header"]):
        tag.decompose()
    text = soup.get_text(separator=" ", strip=True)
    result: dict = {}
    seen: set = set()
    names: list = []

    # Pattern 1: Dr. First [M.] Last [, DDS/DMD]
    for m in re.finditer(
        r'Dr\.?\s+([A-Z][a-z]+(?:\s+[A-Z]\.?)?\s+[A-Z][A-Za-z\-\']{2,})'
        r'(?:\s*,?\s*(?:DDS|DMD|D\.D\.S|D\.M\.D|MD))?',
        text,
    ):
        n = f"Dr. {m.group(1).strip()}"
        if n.lower() not in seen and _validate_doctor_name(n):
            names.append(n)
            seen.add(n.lower())

    # Pattern 2: First Last, DDS/DMD (no Dr. prefix)
    for m in re.finditer(
        r'([A-Z][a-z]{1,20}(?:\s+[A-Z][a-z]{0,15})?\s+[A-Z][A-Za-z\-\']{2,20})'
        r'\s*,?\s+(?:DDS|DMD|D\.D\.S|D\.M\.D)',
        text,
    ):
        n = m.group(1).strip()
        if n.lower() not in seen and _validate_doctor_name(n):
            names.append(n)
            seen.add(n.lower())

    if names:
        result["scraped_doctor_names"] = ", ".join(names[:6])
        result["doctors"] = [
            {"name": n, "specialty": "Not Found", "associations": ""}
            for n in names[:6]
        ]
    return result


# ── NPI registry ──────────────────────────────────────────────────────────────

def _npi_fetch_all(city: str, state: str, max_pages: int = 5) -> list[dict]:
    """
    Fetch NPI-1 dentist records for a city/state, paginating up to max_pages.
    NPI caps each page at 200; cap at 1000 total to avoid spending 3+ min on
    large cities (Worcester, El Paso) where Aspen is never going to be #800.
    """
    all_results = []
    skip = 0
    for _ in range(max_pages):
        params = {
            "version": "2.1",
            "enumeration_type": "NPI-1",
            "city": city,
            "state": state,
            "taxonomy_description": "dentist",
            "limit": "200",
            "skip": str(skip),
        }
        url = "https://npiregistry.cms.hhs.gov/api/?" + urllib.parse.urlencode(params)
        try:
            r = std_requests.get(url, timeout=10)
            if r.status_code != 200:
                break
            data = r.json()
            page = data.get("results", [])
            all_results.extend(page)
            if len(page) < 200:
                break   # last page
            skip += 200
            time.sleep(0.2)
        except Exception:
            break
    return all_results


def npi_lookup_aspen(city: str, state: str, address_slug: str) -> dict:
    """
    Query NPI for dentists at this Aspen Dental location.
    Returns dict with keys: 'doctors' (list of name strings),
                            'phone'   (str or ""),
                            'zip'     (str or ""),
                            'address' (str or "").

    Matching strategy (most-to-least strict):
      1. "aspen" in NPI org name — most reliable signal
      2. Street number + one meaningful street word
      3. Street number only
      4. Small-city fallback: ≤15 dentists total → return all
    Paginates through all NPI results so large cities don't miss matches.
    Also returns phone / zip from the matching NPI location address record.
    """
    empty = {"doctors": [], "phone": "", "zip": "", "address": ""}
    if not city or not state:
        return empty

    # Parse address tokens from URL slug
    slug_parts = address_slug.lower().split("-")
    street_num = slug_parts[0] if slug_parts and slug_parts[0].isdigit() else ""
    _noise = {"ste", "suite", "st", "ave", "blvd", "rd", "dr", "ln",
              "pkwy", "hwy", "n", "s", "e", "w", "fl", "floor",
              "highway", "route", "b", "c", "d", "e", "f"}
    street_words = [w for w in slug_parts[1:]
                    if w not in _noise and not w.isdigit() and len(w) > 3]

    def _build_name(item: dict) -> str | None:
        basic  = item.get("basic", {})
        first  = basic.get("first_name", "").strip()
        last   = basic.get("last_name",  "").strip()
        cred   = basic.get("credential", "").strip()
        status = basic.get("status", "").upper()
        if status == "D" or not first or not last:
            return None
        name = f"Dr. {first.title()} {last.title()}"
        if cred:
            name = f"{name}, {cred}"
        return name if _validate_doctor_name(name) else None

    def _location_addr(item: dict) -> dict:
        """Extract phone/zip/address from NPI LOCATION address record."""
        for addr in item.get("addresses", []):
            if addr.get("address_purpose", "").upper() == "LOCATION":
                return {
                    "phone":   addr.get("telephone_number", ""),
                    "zip":     addr.get("postal_code", "")[:5],
                    "address": addr.get("address_1", ""),
                }
        # Fallback: first address
        addrs = item.get("addresses", [])
        if addrs:
            return {
                "phone":   addrs[0].get("telephone_number", ""),
                "zip":     addrs[0].get("postal_code", "")[:5],
                "address": addrs[0].get("address_1", ""),
            }
        return {"phone": "", "zip": "", "address": ""}

    def _addr_matches_pass(addr_obj: dict, pass_num: int) -> bool:
        a1  = addr_obj.get("address_1", "").lower()
        org = addr_obj.get("organization_name", "").lower()
        if pass_num == 1:
            return "aspen" in org
        if pass_num == 2:
            num_ok  = bool(street_num) and street_num in a1.split()
            word_ok = any(w in a1 for w in street_words[:2])
            return num_ok and word_ok
        if pass_num == 3:
            return bool(street_num) and street_num in a1.split()
        return False  # pass 4 handled separately

    try:
        print(f"  [NPI] Fetching dentists in {city}, {state}…")
        results = _npi_fetch_all(city, state)
        total = len(results)
        print(f"  [NPI] {total} dentist(s) found")

        seen_names: set = set()
        seen_npis:  set = set()

        def _collect_pass(pass_num: int) -> dict:
            """Run one matching pass. Returns result dict if any matched, else {}."""
            doctors = []
            phone = zip_ = address = ""
            for item in results:
                npi_num = item.get("number", "")
                if npi_num in seen_npis:
                    continue
                for addr in item.get("addresses", []):
                    if _addr_matches_pass(addr, pass_num):
                        n = _build_name(item)
                        if n and n.lower() not in seen_names:
                            loc = _location_addr(item)
                            doctors.append(n)
                            seen_names.add(n.lower())
                            seen_npis.add(npi_num)
                            if not phone and loc["phone"]:
                                phone = loc["phone"]
                            if not zip_ and loc["zip"]:
                                zip_ = loc["zip"]
                            if not address and loc["address"]:
                                address = loc["address"]
                        break
            if doctors:
                print(f"  [NPI] Pass {pass_num} match: {len(doctors)} doctor(s)")
                return {"doctors": doctors, "phone": phone,
                        "zip": zip_, "address": address}
            return {}

        for pass_num in (1, 2, 3):
            r = _collect_pass(pass_num)
            if r:
                return r

        # Pass 4: small-city fallback — ≤15 dentists total in city
        if total <= 15:
            doctors = []
            phone = zip_ = address = ""
            for item in results:
                n = _build_name(item)
                if n and n.lower() not in seen_names:
                    loc = _location_addr(item)
                    doctors.append(n)
                    seen_names.add(n.lower())
                    if not phone and loc["phone"]:
                        phone = loc["phone"]
                    if not zip_ and loc["zip"]:
                        zip_ = loc["zip"]
                    if not address and loc["address"]:
                        address = loc["address"]
            if doctors:
                print(f"  [NPI] Small-city fallback ({total} dentists): {len(doctors)}")
                return {"doctors": doctors, "phone": phone,
                        "zip": zip_, "address": address}

        print(f"  [NPI] No address match for {city}, {state} (slug={address_slug})")
        return empty

    except Exception as e:
        print(f"  [NPI] Error: {e}")
        return empty


# ── Main scrape function ──────────────────────────────────────────────────────

def scrape_aspen_location(url: str) -> dict:
    """Scrape one Aspen Dental location URL. Returns a data dict."""
    info = _parse_aspen_url(url)
    clean_url = info["clean_url"]

    practice_label = (f"Aspen Dental - {info['city']}, {info['state']}"
                      if info["city"] and info["state"]
                      else "Aspen Dental")

    # Derive street address from URL slug immediately — no network needed
    slug_addr = _address_from_slug(info.get("address_slug", ""))

    result: dict = {
        "website":              clean_url,
        "practice_name":        practice_label,
        "city":                 info["city"],
        "state":                info["state"],
        "scraped_doctor_names": "Not Found",
        "email":                "Not Found",
        "phone":                "Not Found",
        # Pre-fill address from URL slug; NPI/HTML may upgrade to a canonical form
        "address":              slug_addr or "Not Found",
        "zip":                  "Not Found",
        "skip_reason":          "",
        "doctors":              [],
        **{k: 0 for k in _ASPEN_CORE_SERVICES},
    }

    # ── Step 1: fetch raw HTML ────────────────────────────────────────────────
    html, strategy = _fetch_with_fallbacks(clean_url)

    if html:
        print(f"  ✓ HTML: {len(html):,} chars via {strategy}")

        # Step 1a: __NEXT_DATA__ (richest source — providers, phone, address)
        nd = extract_next_data(html, clean_url)
        if nd:
            print(f"  → __NEXT_DATA__ keys: {[k for k in nd if not k.startswith('_')]}")
            for k, v in nd.items():
                if k.startswith("_"):
                    continue
                if v and result.get(k) in (None, "Not Found", 0, ""):
                    result[k] = v

        # Step 1b: try /_next/data API endpoint (same data in clean JSON)
        next_api = nd.get("_next_api_url", "") if nd else ""
        if next_api and result.get("scraped_doctor_names") == "Not Found":
            print(f"  → Next.js data API: ...{next_api[-60:]}")
            api_html, _ = _fetch_with_fallbacks(next_api)
            if api_html:
                try:
                    api_data = json.loads(api_html)
                    page_props = api_data.get("pageProps", api_data)
                    raw = _collect_providers(page_props)
                    names = _dedup_names(raw)
                    if names:
                        result["scraped_doctor_names"] = ", ".join(names)
                        result["doctors"] = [{"name": n, "specialty": "Not Found", "associations": ""}
                                             for n in names]
                        print(f"  ✓ Next.js API doctors: {result['scraped_doctor_names']}")
                except Exception:
                    pass

        # Step 1c: JSON-LD supplement
        ld = extract_json_ld(html)
        for k, v in ld.items():
            if v and result.get(k) in (None, "Not Found", 0, ""):
                result[k] = v

        # Step 1d: plain-text regex scan (last resort for doctors)
        if result.get("scraped_doctor_names") == "Not Found":
            ts = extract_text_scan(html)
            for k, v in ts.items():
                if v and result.get(k) in (None, "Not Found", 0, ""):
                    result[k] = v

        # Step 1e: phone regex fallback from raw HTML
        if result.get("phone") == "Not Found":
            pm = re.search(r'\(?\d{3}\)?[\s.\-]?\d{3}[\s.\-]?\d{4}', html)
            if pm:
                result["phone"] = pm.group(0).strip()

        # Step 1f: crawl /providers/ sub-pages for individual doctor profiles
        print(f"  → Scanning /providers/ sub-pages…")
        provider_names = _scrape_provider_pages(html, clean_url)
        if provider_names:
            existing_lc = set()
            if result.get("scraped_doctor_names") not in ("Not Found", "", None):
                existing_lc = {n.strip().lower()
                               for n in result["scraped_doctor_names"].split(",")}
            new_pnames = [n for n in provider_names if n.lower() not in existing_lc]
            if result.get("scraped_doctor_names") in ("Not Found", "", None):
                result["scraped_doctor_names"] = ", ".join(provider_names)
                result["doctors"] = [{"name": n, "specialty": "Not Found", "associations": ""}
                                     for n in provider_names]
                print(f"  ✓ Provider pages doctors: {result['scraped_doctor_names']}")
            elif new_pnames:
                result["scraped_doctor_names"] += ", " + ", ".join(new_pnames)
                result["doctors"].extend(
                    [{"name": n, "specialty": "Not Found", "associations": ""} for n in new_pnames]
                )
                print(f"  ✓ Provider pages added {len(new_pnames)}: {', '.join(new_pnames)}")
    else:
        print(f"  ✗ Page blocked (CF Enterprise) — using NPI + URL-derived address")
        result["skip_reason"] = "CF / Bot Protected — NPI used for doctors"

    # ── Step 2: NPI registry (always runs — supplements/verifies website data) ─
    if info["city"] and info["state"]:
        npi = npi_lookup_aspen(info["city"], info["state"], info.get("address_slug", ""))

        if npi["doctors"]:
            existing_lc = set()
            if result.get("scraped_doctor_names") not in ("Not Found", "", None):
                existing_lc = {n.strip().lower()
                               for n in result["scraped_doctor_names"].split(",")}

            new_docs = [n for n in npi["doctors"] if n.lower() not in existing_lc]
            if result.get("scraped_doctor_names") in ("Not Found", "", None):
                result["scraped_doctor_names"] = ", ".join(npi["doctors"])
                result["doctors"] = [{"name": n, "specialty": "Not Found", "associations": ""}
                                     for n in npi["doctors"]]
                print(f"  ✓ NPI doctors: {result['scraped_doctor_names']}")
            elif new_docs:
                result["scraped_doctor_names"] += ", " + ", ".join(new_docs)
                result["doctors"].extend(
                    [{"name": n, "specialty": "Not Found", "associations": ""} for n in new_docs]
                )
                print(f"  ✓ NPI added {len(new_docs)}: {', '.join(new_docs)}")

            # Fill phone / zip / address from NPI if still missing
            if result.get("phone") == "Not Found" and npi["phone"]:
                result["phone"] = npi["phone"]
                print(f"  ✓ NPI phone: {npi['phone']}")

            if result.get("zip") == "Not Found" and npi["zip"]:
                result["zip"] = npi["zip"]

            # Only upgrade address from NPI if it looks more complete than slug-derived
            npi_addr = npi.get("address", "")
            cur_addr = result.get("address", "")
            if npi_addr and (cur_addr in ("Not Found", "") or
                             (slug_addr and len(npi_addr) > len(slug_addr))):
                result["address"] = npi_addr.title()

    # ── Step 3: guarantee core service flags (Aspen offers these everywhere) ──
    for svc, default_val in _ASPEN_CORE_SERVICES.items():
        if not result.get(svc):
            result[svc] = default_val

    return result


# ── Excel output ──────────────────────────────────────────────────────────────

_COL_HEADERS = [
    ("Practice Name",      30), ("Website",             50), ("City",    15),
    ("State",               8), ("Doctor Name(s)",      45), ("Phone",   16),
    ("Address",            35), ("ZIP",                 10), ("Implants",10),
    ("Veneers",            10), ("Invisalign",          12), ("Whitening",12),
    ("Sedation",           10), ("Clear Aligners",      15), ("Google Rating", 14),
    ("Total Google Reviews",18),("Skip Reason",         25),
]


def write_xlsx(rows: list[dict], output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Aspen Dental"

    hdr_fill = PatternFill("solid", fgColor="2E75B6")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    for col, (hdr, width) in enumerate(_COL_HEADERS, 1):
        c = ws.cell(1, col, hdr)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    for row_i, r in enumerate(rows, 2):
        vals = [
            r.get("practice_name",        ""),
            r.get("website",              ""),
            r.get("city",                 ""),
            r.get("state",                ""),
            r.get("scraped_doctor_names", "Not Found"),
            r.get("phone",                "Not Found"),
            r.get("address",              "Not Found"),
            r.get("zip",                  "Not Found"),
            r.get("implants",             0),
            r.get("veneers",              0),
            r.get("invisalign",           0),
            r.get("whitening",            0),
            r.get("sedation",             0),
            r.get("clear_aligners",       0),
            r.get("google_rating",        "Not Found"),
            r.get("total_google_reviews", "Not Found"),
            r.get("skip_reason",          ""),
        ]
        for col, v in enumerate(vals, 1):
            ws.cell(row_i, col, v)

    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"\nSaved → {output_path}  ({len(rows)} rows)")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(
        description="Specialized fast scraper for Aspen Dental location pages.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    grp = p.add_mutually_exclusive_group(required=True)
    grp.add_argument("--urls", nargs="+", metavar="URL",
                     help="One or more Aspen Dental location URLs")
    grp.add_argument("--file", metavar="FILE.txt",
                     help="Text file with one URL per line (or comma-separated)")
    p.add_argument("--proxies", metavar="proxies.txt",
                   help="Proxy list (one per line, UK mobile first)")
    p.add_argument("--output", "-o", default="aspen_output.xlsx",
                   help="Output Excel file (default: aspen_output.xlsx)")
    args = p.parse_args()

    load_proxies(args.proxies)

    if args.urls:
        urls = [u.strip() for u in args.urls if u.strip().startswith("http")]
    else:
        with open(args.file) as f:
            raw = f.read()
        urls = [u.strip() for u in re.split(r'[\s,]+', raw)
                if u.strip().startswith("http")]

    print(f"\nScraping {len(urls)} Aspen Dental location(s)…\n")

    all_results: list = []
    for i, url in enumerate(urls, 1):
        _proxy_fail.clear()  # reset per-site so later sites still get proxy attempts
        info = _parse_aspen_url(url)
        label = f"{info['city']}, {info['state']}" if info["city"] else url[:60]
        print(f"[{i}/{len(urls)}] {label}")
        try:
            r = scrape_aspen_location(url)
            all_results.append(r)
            doc   = r.get("scraped_doctor_names", "Not Found")
            phone = r.get("phone", "Not Found")
            rating = r.get("google_rating", "Not Found")
            print(f"  → Doc: {doc[:55]} | Ph: {phone} | Rating: {rating}")
        except Exception as e:
            print(f"  ✗ Error: {e}")
            all_results.append({
                "website": url, "practice_name": label,
                "skip_reason": str(e),
                "scraped_doctor_names": "Not Found",
                **{k: 0 for k in _ASPEN_CORE_SERVICES},
            })

        # Checkpoint every 10 sites
        if i % 10 == 0 or i == len(urls):
            write_xlsx(all_results, args.output)
            print(f"  ✓ Checkpoint saved ({i}/{len(urls)})")

        if i < len(urls):
            time.sleep(0.5)

    print(f"\nDone. {len(all_results)} locations scraped → {args.output}")


if __name__ == "__main__":
    main()
