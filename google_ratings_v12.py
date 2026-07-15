"""
google_ratings_v12.py
─────────────────────
Reads 100data.xlsx (output of dental_scraper.py).
For each unique practice website, fetches the Google Maps star rating
and review count.

Search strategy (tried in order):
  1. Google Places API  — by website domain       (fast, free, no bot risk)
  2. Google Places API  — by practice name + city (fallback when domain search misses)
  3. Google Maps via Playwright — by domain        (browser fallback)
  4. Google Maps via Playwright — by name + city  (last resort)

HOW TO GET A FREE GOOGLE PLACES API KEY
  1. Go to https://console.cloud.google.com/
  2. Create a project  →  Enable "Places API"
  3. Credentials  →  Create API Key
  4. Paste the key into GOOGLE_PLACES_API_KEY below
  Free tier: $200/month credit  ≈  11,000 Text Search calls — 100 practices costs ~$1.70

Bot protection (for Playwright fallback):
  - Visible browser + persistent profile  (reuses cookies across runs)
  - Stealth JS patches  (navigator.webdriver hidden)
  - Random 4–9 s delay between requests
  - Auto-pause on CAPTCHA — solve in browser, press Enter to resume
"""

import os
import re
import json
import time
import random
import logging
import warnings
from urllib.parse import quote_plus, urlparse

import requests
import openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import Font

warnings.filterwarnings("ignore", message="Unverified HTTPS request")

try:
    from playwright.sync_api import sync_playwright
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PLAYWRIGHT_AVAILABLE = False

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ── Configuration ──────────────────────────────────────────────────────────────
INPUT_FILE   = "batch_01_rows1_100.xlsx"
OUTPUT_FILE  = "batch_01_rows1_100.xlsx"

DELAY_MIN    = 6.0        # minimum seconds between Google requests
DELAY_MAX    = 13.0       # maximum (random jitter)
TIMEOUT      = 15
PW_TIMEOUT   = 35_000     # ms
PW_USER_DATA = "/tmp/pw_google_profile"  # persistent session — keeps cookies across runs

# After every BATCH_SIZE practices, pause for BATCH_PAUSE seconds
# (helps avoid Google detecting a pattern of rapid automated queries)
BATCH_SIZE   = 10
BATCH_PAUSE  = 45         # seconds

# ── Google Places API key (optional — leave "" to use Playwright only) ─────────
GOOGLE_PLACES_API_KEY = ""   # paste key here if you have credits

# ── Column indices in 100data.xlsx (1-based) ────────────────────────────────
# Row 1 = title banner, Row 2 = column headers, Row 3+ = data
C_INDEX    = 1
C_PRACTICE = 2
C_DOCTOR   = 3
C_ADDRESS  = 4
C_CITY     = 5
C_STATE    = 6
C_ZIP      = 7
C_WEBSITE  = 8    # "Practice Website"
C_GOOGLE_R = 42   # "Google Reviews Ranking"
C_GOOGLE_N = 43   # "Total # of Google Reviews"
HDR_ROW    = 2    # row containing column headers
DATA_START = 3    # first data row

_USER_AGENTS = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_4_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36 Edg/122.0.0.0",
]

HEADERS = {
    "User-Agent": _USER_AGENTS[0],
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Referer": "https://www.google.com/",
}

PLACEHOLDER = {"Not Found", "See Website", "See Profile", "Blocked", "ERROR", "", "N/A"}


# ── Helpers ────────────────────────────────────────────────────────────────────

def get_domain(website: str) -> str:
    """'https://www.example-dental.com/about' → 'example-dental.com'"""
    if not website or str(website).strip() in ("", "None", "nan"):
        return ""
    if not website.startswith("http"):
        website = "https://" + website
    host = urlparse(website).netloc.lower()
    return re.sub(r"^www\d*\.", "", host)


def _human_delay(pw_page=None):
    time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))
    if pw_page:
        _jitter_mouse(pw_page)


def _is_captcha(html: str) -> bool:
    lc = html.lower()
    return (
        "recaptcha" in lc
        or "captcha" in lc
        or "before you continue" in lc
        or "i'm not a robot" in lc
        or "detected unusual traffic" in lc
        or "verify you're a human" in lc
    )


def _wait_captcha(pw_page):
    log.warning("  ⚠  CAPTCHA detected — solve it in the browser window, then press Enter here…")
    input("  [Press Enter after solving CAPTCHA] ")
    pw_page.wait_for_timeout(2500)


# ── Stealth ────────────────────────────────────────────────────────────────────

_STEALTH_JS = """
(() => {
  // Hide webdriver flag
  Object.defineProperty(navigator, 'webdriver',  {get: () => undefined});
  // Fake plugin list
  Object.defineProperty(navigator, 'plugins',    {get: () => [1,2,3,4,5]});
  // Realistic language + platform
  Object.defineProperty(navigator, 'languages',  {get: () => ['en-US', 'en']});
  Object.defineProperty(navigator, 'platform',   {get: () => 'MacIntel'});
  Object.defineProperty(navigator, 'hardwareConcurrency', {get: () => 8});
  // Chrome runtime object (missing in raw Chromium)
  window.chrome = { runtime: {}, loadTimes: function(){}, csi: function(){} };
  // Permissions API — make notifications return the real permission
  const origQuery = window.navigator.permissions.query;
  window.navigator.permissions.query = (p) =>
    p.name === 'notifications'
      ? Promise.resolve({state: Notification.permission})
      : origQuery(p);
  // Remove automation-related properties from window
  delete window.__playwright;
  delete window.__pwInitScripts;
})();
"""

def apply_stealth(ctx_or_page):
    try:
        ctx_or_page.add_init_script(_STEALTH_JS)
    except Exception:
        pass


def _jitter_mouse(pw_page):
    """Move the mouse to a random position to simulate human presence."""
    try:
        x = random.randint(200, 900)
        y = random.randint(150, 600)
        pw_page.mouse.move(x, y, steps=random.randint(5, 15))
    except Exception:
        pass


def _human_scroll(pw_page):
    """Small random scroll to simulate reading the page."""
    try:
        pw_page.evaluate(
            f"window.scrollBy(0, {random.randint(80, 300)})"
        )
        time.sleep(random.uniform(0.3, 0.8))
    except Exception:
        pass


def _prewarm_browser(pw_page):
    """
    Visit Google homepage and wait a few seconds before starting searches.
    Establishes a real-looking session with cookies before any Maps queries.
    """
    try:
        log.info("  Pre-warming browser session on Google…")
        pw_page.goto("https://www.google.com/?hl=en&gl=us", timeout=PW_TIMEOUT, wait_until="domcontentloaded")
        time.sleep(random.uniform(4, 7))
        _jitter_mouse(pw_page)
        _human_scroll(pw_page)
        time.sleep(random.uniform(2, 4))
        log.info("  Browser session ready.")
    except Exception as e:
        log.debug(f"  Pre-warm failed (non-fatal): {e}")


# ── Rating extraction ──────────────────────────────────────────────────────────

def extract_rating(html: str) -> tuple:
    """Return (rating_str, count_str) from Google search/Maps HTML. Empty strings if not found."""
    soup = BeautifulSoup(html, "lxml")

    # ── Method 1: JSON-LD aggregateRating ──────────────────────────────────────
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string or "")
            if not isinstance(data, list):
                data = [data]
            for item in data:
                ar = item.get("aggregateRating") or {}
                rv = str(ar.get("ratingValue", "")).strip()
                rc = str(ar.get("reviewCount", "") or ar.get("ratingCount", "")).strip()
                if rv and re.match(r"^[1-5](\.\d)?$", rv):
                    return rv, rc
        except Exception:
            pass

    # ── Method 2: aria-label "X stars" + separate "N reviews" aria-labels ────
    found_rating, found_count = "", ""
    for tag in soup.find_all(attrs={"aria-label": True}):
        lbl = tag.get("aria-label", "")
        # Rating label: "4.8 stars" / "Rated 4.8 out of 5"
        if re.search(r"[1-5]\.\d", lbl) and re.search(r"star|rated", lbl, re.I):
            rm = re.search(r"([1-5]\.\d)", lbl)
            if rm and not found_rating:
                found_rating = rm.group(1)
                # Count sometimes in the same label
                cm = re.search(r"([\d,]+)\s*reviews?", lbl, re.I)
                if cm:
                    found_count = cm.group(1).replace(",", "")
        # Separate count label: "439 reviews" / "439 Google reviews"
        if not found_count and re.search(r"review", lbl, re.I):
            cm = re.search(r"([\d,]+)\s*reviews?", lbl, re.I)
            if cm:
                found_count = cm.group(1).replace(",", "")
    if found_rating:
        return found_rating, found_count

    # ── Method 3: data-attrid (Google knowledge panel) ────────────────────────
    _kp_rating = ""
    _kp_count  = ""
    for tag in soup.find_all(attrs={"data-attrid": True}):
        attr = tag.get("data-attrid", "")
        txt  = tag.get_text(" ", strip=True)
        if "rating" in attr and not _kp_rating:
            rm = re.search(r"([1-5]\.\d)", txt)
            if rm:
                _kp_rating = rm.group(1)
                cm = re.search(r"\((\d[\d,]+)\)", txt)
                if not cm:
                    cm = re.search(r"([\d,]+)\s*reviews?", txt, re.I)
                if cm:
                    _kp_count = cm.group(1).replace(",", "")
        # Dedicated review-count data-attrid: "kc:/location/location:user_review_count"
        if ("review" in attr or "count" in attr) and not _kp_count:
            cm = re.search(r"([\d,]+)", txt)
            if cm:
                _kp_count = cm.group(1).replace(",", "")
    if _kp_rating:
        return _kp_rating, _kp_count

    # ── Method 4: visible text patterns ───────────────────────────────────────
    text = soup.get_text(" ", strip=True)
    # Also trigger on middot (·) — Google uses "4.8 · 342 reviews" format
    rm = re.search(r"\b([1-5]\.\d)\s*(?:out of 5|stars?|\(|·|·)", text, re.I)
    _m4_rating = ""
    if rm:
        g_rating = rm.group(1)
        for pat in (
            r"\(" + re.escape(g_rating) + r"[^\)]*\)\s*([\d,]+)\s*(?:Google reviews?|reviews?)",
            r"\((\d[\d,]+)\)\s*(?:Google reviews?|reviews?)",
            r"([\d,]+)\s*(?:Google reviews?|reviews?|Google ratings?)",
            r"·\s*([\d,]+)",       # "4.8 · 342 reviews" — Google middot format
            r"·\s*([\d,]+)",            # plain ASCII middot variant
            r"\((\d[\d,]+)\)",          # bare (439) — Maps result cards, no "reviews" text
        ):
            cm = re.search(pat, text, re.I)
            if cm:
                count = cm.group(1).replace(",", "")
                if g_rating == "1.0" and int(count or "0") <= 1:
                    break
                return g_rating, count
        # Don't return yet — fall through so methods 5 & 6 can still find count
        # in a different layout (e.g. multiline or (NNN) format)
        if g_rating != "1.0":
            _m4_rating = g_rating

    # ── Method 5: "4.8 (342)" — allow newlines between rating and count ────────
    # [^\n]{0,30} was the old pattern; [\s\S]{0,60}? handles multiline layout
    rm = re.search(r"\b([1-5]\.\d)\b[\s\S]{0,60}?\((\d[\d,]+)\)", text)
    if rm:
        return rm.group(1), rm.group(2).replace(",", "")

    # ── Method 6: any (NNN) near any rating anywhere in the text ───────────────
    for rm_r in re.finditer(r"\b([1-5]\.\d)\b", text):
        window = text[rm_r.start(): rm_r.start() + 150]
        cm = re.search(r"\((\d[\d,]+)\)", window)
        if cm:
            candidate = int(cm.group(1).replace(",", ""))
            if candidate >= 1:
                return rm_r.group(1), cm.group(1).replace(",", "")

    # ── Method 7: look for count without "()" near any rating — forward & backward
    for rm_r in re.finditer(r"\b([1-5]\.\d)\b", text):
        # forward window
        window_fwd = text[rm_r.start(): rm_r.start() + 500]
        cm = re.search(r"([\d,]+)\s*(?:Google reviews?|reviews?|ratings?)", window_fwd, re.I)
        if cm:
            return rm_r.group(1), cm.group(1).replace(",", "")
        # backward window — sometimes count appears before the star rating in DOM order
        window_bwd = text[max(0, rm_r.start() - 300): rm_r.start()]
        cm = re.search(r"([\d,]+)\s*(?:Google reviews?|reviews?|ratings?)", window_bwd, re.I)
        if cm:
            return rm_r.group(1), cm.group(1).replace(",", "")

    # ── Method 8: any "N reviews" anywhere in full text — pair with nearest rating ──
    # Handles pages where count and rating are far apart in the flattened text.
    all_counts = list(re.finditer(r"([\d,]+)\s*(?:Google reviews?|reviews?)", text, re.I))
    all_ratings = list(re.finditer(r"\b([1-5]\.\d)\b", text))
    if all_counts and all_ratings:
        # Pick the count with the most reviews (most likely to be the main listing)
        best_count = max(all_counts, key=lambda m: int(m.group(1).replace(",", "")))
        count_val  = best_count.group(1).replace(",", "")
        # Nearest rating to that count position
        best_rating = min(all_ratings, key=lambda m: abs(m.start() - best_count.start()))
        return best_rating.group(1), count_val

    # If method 4 found a rating but no count format matched, return rating alone
    if _m4_rating:
        return _m4_rating, ""

    return "", ""


def extract_rating_maps_pw(pw_page) -> tuple:
    """Extract rating + count from an already-loaded Google Maps business panel."""
    try:
        pw_page.wait_for_selector('[role="main"]', timeout=8000)
        pw_page.wait_for_timeout(2000)
    except Exception:
        pass

    rating, count = "", ""

    # ── Step 1: get rating from aria-label; also check if count is in the same label ─
    for sel in ('[aria-label*="stars"]', '[aria-label*="Star"]', 'span.ceNzKf', 'span.MW4etd'):
        try:
            for el in pw_page.query_selector_all(sel):
                lbl = el.get_attribute("aria-label") or ""
                rm  = re.search(r"([1-5]\.\d)", lbl)
                if rm:
                    rating = rm.group(1)
                    # Count sometimes lives in the same aria-label
                    cm = re.search(r"([\d,]+)\s*reviews?", lbl, re.I)
                    if cm:
                        count = cm.group(1).replace(",", "")
                    break
            if rating:
                break
        except Exception:
            pass

    # ── Step 2: if count still missing, try dedicated review-count elements ──────
    if rating and not count:

        # 2a: element with aria-label mentioning "review"
        for sel in ('[aria-label*="review"]', '[aria-label*="Review"]'):
            try:
                for el in pw_page.query_selector_all(sel):
                    lbl = el.get_attribute("aria-label") or ""
                    cm  = re.search(r"([\d,]+)\s*reviews?", lbl, re.I)
                    if cm:
                        count = cm.group(1).replace(",", "")
                        break
                if count:
                    break
            except Exception:
                pass

        # 2b: rating/review button (jsaction="pane.rating…")
        if not count:
            for btn_sel in (
                'button[jsaction*="pane.rating"]',
                'button[jsaction*="review"]',
                'button[data-value*="review"]',
            ):
                try:
                    btn = pw_page.query_selector(btn_sel)
                    if btn:
                        btn_text = btn.inner_text() or ""
                        cm = re.search(r"([\d,]+)", btn_text)
                        if cm:
                            count = cm.group(1).replace(",", "")
                            break
                except Exception:
                    pass

        # 2c: text_content() of main panel — includes CSS-hidden nodes inner_text() misses
        if not count:
            try:
                panel = pw_page.query_selector('[role="main"]')
                if panel:
                    pt = panel.text_content() or ""
                    cm = re.search(r"\((\d[\d,]*)\)", pt)
                    if not cm:
                        cm = re.search(r"([\d,]+)\s*reviews?", pt, re.I)
                    if cm:
                        count = cm.group(1).replace(",", "")
            except Exception:
                pass

        # 2d: JavaScript extraction — scans every span/button for "(NNN)" or "NNN reviews"
        if not count:
            try:
                count = pw_page.evaluate("""
                    () => {
                        // Pass 1: element whose full text is exactly "(NNN)" or "NNN reviews"
                        for (const el of document.querySelectorAll('span, button, a')) {
                            const t = (el.textContent || '').trim();
                            let m = t.match(/^\\((\\d[\\d,]*)\\)$/);
                            if (m) return m[1].replace(/,/g,'');
                            m = t.match(/^(\\d[\\d,]*)\\s+(?:Google\\s+)?reviews?$/i);
                            if (m) return m[1].replace(/,/g,'');
                        }
                        // Pass 2: any element containing "NNN reviews" or "NNN Google reviews"
                        for (const el of document.querySelectorAll('span, button, a, div')) {
                            const t = (el.textContent || '').trim();
                            const m = t.match(/(\\d[\\d,]+)\\s+(?:Google\\s+)?reviews?/i);
                            if (m && t.length < 100) return m[1].replace(/,/g,'');
                        }
                        // Pass 3: aria-label on any element mentioning reviews
                        for (const el of document.querySelectorAll('[aria-label]')) {
                            const lbl = el.getAttribute('aria-label') || '';
                            const m = lbl.match(/(\\d[\\d,]+)\\s+(?:Google\\s+)?reviews?/i);
                            if (m) return m[1].replace(/,/g,'');
                        }
                        // Pass 4: full page body text
                        const body = document.body.innerText || '';
                        const m = body.match(/(\\d[\\d,]+)\\s+(?:Google\\s+)?reviews?/i);
                        return m ? m[1].replace(/,/g,'') : '';
                    }
                """) or ""
            except Exception:
                pass

    # ── Step 3: fall back to full-page HTML parse ────────────────────────────────
    # If no rating at all, let extract_rating try every method on raw HTML.
    # If we have a rating but lost the count, use extract_rating's count only.
    if not rating or not count:
        try:
            html_rating, html_count = extract_rating(pw_page.content())
            if not rating and html_rating:
                rating = html_rating
                count  = html_count
            elif rating and not count and html_count:
                count = html_count
        except Exception:
            pass

    return rating, count


# ── Search functions ───────────────────────────────────────────────────────────

def _name_score(candidate: str, target: str) -> int:
    """Rough word-overlap score between two practice names (case-insensitive)."""
    _stop = {"dentist", "dental", "family", "the", "and", "of", "dr", "dds", "dmd", "inc", "llc", "pa"}
    t_words = {w for w in re.split(r'\W+', target.lower()) if w and w not in _stop}
    c_words = {w for w in re.split(r'\W+', candidate.lower()) if w and w not in _stop}
    return len(t_words & c_words)


def _maps_navigate_and_extract(pw_page, query: str, domain: str = "",
                                practice_name: str = "") -> tuple:
    """
    Navigate Google Maps with query, pick the best-matching listing, return (rating, count).
    Prefers name-matched result over first result to avoid returning the wrong practice.
    """
    url = f"https://www.google.com/maps/search/{quote_plus(query)}?hl=en&gl=us"
    _human_delay(pw_page)
    pw_page.goto(url, timeout=PW_TIMEOUT, wait_until="domcontentloaded")
    pw_page.wait_for_timeout(random.randint(2500, 4500))
    _human_scroll(pw_page)
    html = pw_page.content()

    if _is_captcha(html):
        _wait_captcha(pw_page)
        html = pw_page.content()

    # Single-result direct panel
    if "/maps/place/" in pw_page.url:
        return extract_rating_maps_pw(pw_page)

    # List of results — pick best match by name > domain > first
    try:
        links = pw_page.query_selector_all('a[href*="/maps/place/"]')
        best_link   = None
        best_score  = -1

        for link in links[:8]:
            try:
                card_text = link.evaluate(
                    "el => el.closest('[role=\"listitem\"]')?.innerText || ''"
                ) or ""
                score = 0
                if practice_name:
                    score = _name_score(card_text, practice_name)
                elif domain and domain in card_text.lower():
                    score = 1
                # Always track first link as minimum fallback (score=-1 means not scored)
                if best_link is None:
                    best_link  = link
                    best_score = score
                elif score > best_score:
                    best_link  = link
                    best_score = score
            except Exception:
                continue

        if best_link:
            best_link.click()
            pw_page.wait_for_timeout(3000)
            return extract_rating_maps_pw(pw_page)
    except Exception as e:
        log.debug(f"  Maps list parsing error: {e}")

    return extract_rating(html)


def search_maps_by_name(practice_name: str, doctor_name: str,
                        address: str, city: str, state: str,
                        zip_code: str, pw_page) -> tuple:
    """Google Maps search using practice name + full address for precise matching."""
    clean = re.sub(r'\b(LLC|PA|PLLC|DDS|DMD|Inc\.?|Corp\.?)\b', '', practice_name, flags=re.I).strip(', ')

    # Query 1: name + street address (most precise)
    if clean and address and city:
        q1 = f"{clean} {address} {city} {state}"
        log.info(f"  [Maps/name+addr] {q1}")
        try:
            r, c = _maps_navigate_and_extract(pw_page, q1, practice_name=clean)
            if r:
                return r, c
        except Exception as e:
            log.debug(f"  Maps/name+addr error: {e}")

    # Query 2: name + city + zip
    q2 = f"{clean} dentist {city} {state} {zip_code}".strip()
    log.info(f"  [Maps/name] {q2}")
    try:
        return _maps_navigate_and_extract(pw_page, q2, practice_name=clean)
    except Exception as e:
        log.debug(f"  Maps/name error: {e}")
    return "", ""


def search_maps_by_website(domain: str, practice_name: str,
                            city: str, state: str, pw_page) -> tuple:
    """Google Maps search using practice name + city (domain not useful as Maps query)."""
    clean = re.sub(r'\b(LLC|PA|PLLC|DDS|DMD|Inc\.?|Corp\.?)\b', '', practice_name, flags=re.I).strip(', ')
    query = f"{clean} dentist {city} {state}" if clean else f"dentist {city} {state}"
    log.info(f"  [Maps/domain] {query}")
    try:
        return _maps_navigate_and_extract(pw_page, query, domain=domain, practice_name=clean)
    except Exception as e:
        log.debug(f"  Maps/domain error: {e}")
        return "", ""


def search_google_by_name(practice_name: str, city: str, state: str, pw_page) -> tuple:
    """Google Web search → knowledge panel rating (fast, no Maps needed)."""
    clean = re.sub(r'\b(LLC|PA|PLLC|DDS|DMD|Inc\.?|Corp\.?)\b', '', practice_name, flags=re.I).strip(', ')
    query = f"{clean} dentist {city} {state}"
    url   = f"https://www.google.com/search?q={quote_plus(query)}&hl=en&gl=us"
    log.info(f"  [Google/search] {query}")
    try:
        _human_delay(pw_page)
        pw_page.goto(url, timeout=PW_TIMEOUT, wait_until="domcontentloaded")
        pw_page.wait_for_timeout(random.randint(2000, 3500))
        _human_scroll(pw_page)
        html = pw_page.content()
        if _is_captcha(html):
            _wait_captcha(pw_page)
            html = pw_page.content()
        return extract_rating(html)
    except Exception as e:
        log.debug(f"  Google/search error: {e}")
        return "", ""


def _places_details(place_id: str) -> tuple:
    """Fetch accurate rating + review_count from Places Details API using a place_id."""
    if not GOOGLE_PLACES_API_KEY or not place_id:
        return "", ""
    try:
        r = requests.get(
            "https://maps.googleapis.com/maps/api/place/details/json",
            params={"place_id": place_id, "fields": "rating,user_ratings_total", "key": GOOGLE_PLACES_API_KEY},
            timeout=15,
        )
        if r.status_code == 200:
            result = r.json().get("result", {})
            rating = str(result.get("rating", "")).strip()
            raw    = result.get("user_ratings_total")
            count  = str(raw) if raw is not None else ""
            if rating:
                return rating, count
    except Exception as e:
        log.debug(f"  Places Details error: {e}")
    return "", ""


def _places_find_place(query: str, city: str, state: str) -> tuple:
    """
    findplacefromtext API — most accurate for a specific business name + location.
    Returns (rating, review_count) or ("", "").
    """
    if not GOOGLE_PLACES_API_KEY:
        return "", ""
    try:
        r = requests.get(
            "https://maps.googleapis.com/maps/api/place/findplacefromtext/json",
            params={
                "input":     query,
                "inputtype": "textquery",
                "fields":    "place_id,name,rating,user_ratings_total,formatted_address",
                "key":       GOOGLE_PLACES_API_KEY,
            },
            timeout=15,
        )
        if r.status_code == 200:
            data = r.json()
            candidates = data.get("candidates", [])
            for c in candidates:
                addr = (c.get("formatted_address") or "").lower()
                # Validate the result is in the right city/state
                if city.lower() in addr or state.lower() in addr:
                    rating = str(c.get("rating", "")).strip()
                    raw    = c.get("user_ratings_total")
                    count  = str(raw) if raw is not None else ""
                    if rating:
                        # user_ratings_total sometimes absent — get it from Details
                        if not count and c.get("place_id"):
                            _, dc = _places_details(c["place_id"])
                            if dc:
                                count = dc
                        return rating, count
                    # Have place_id but no rating inline — call Details
                    if c.get("place_id"):
                        return _places_details(c["place_id"])
            # If no city/state match, still try first candidate
            if candidates:
                c = candidates[0]
                rating = str(c.get("rating", "")).strip()
                raw    = c.get("user_ratings_total")
                count  = str(raw) if raw is not None else ""
                if rating:
                    if not count and c.get("place_id"):
                        _, dc = _places_details(c["place_id"])
                        if dc:
                            count = dc
                    return rating, count
                if c.get("place_id"):
                    return _places_details(c["place_id"])
    except Exception as e:
        log.debug(f"  findplacefromtext error: {e}")
    return "", ""


def _places_text_search(query: str, city: str = "", state: str = "") -> tuple:
    """
    Google Places Text Search API fallback.
    Validates returned address contains city or state before accepting result.
    """
    if not GOOGLE_PLACES_API_KEY:
        return "", ""
    try:
        r = requests.get(
            "https://maps.googleapis.com/maps/api/place/textsearch/json",
            params={"query": query, "key": GOOGLE_PLACES_API_KEY, "region": "us", "type": "dentist"},
            timeout=15,
        )
        if r.status_code == 200:
            data    = r.json()
            status  = data.get("status", "")
            if status not in ("OK", "ZERO_RESULTS"):
                log.warning(f"  Places API status: {status}")
            results = data.get("results", [])
            for res in results[:3]:
                addr = (res.get("formatted_address") or "").lower()
                # Accept if no city/state filter, or address matches
                if not city or city.lower() in addr or (state and state.lower() in addr):
                    rating    = str(res.get("rating", "")).strip()
                    raw_count = res.get("user_ratings_total")
                    count     = str(raw_count) if raw_count is not None else ""
                    if rating:
                        # user_ratings_total sometimes absent — get it from Details
                        if not count and res.get("place_id"):
                            _, dc = _places_details(res["place_id"])
                            if dc:
                                count = dc
                        return rating, count
    except Exception as e:
        log.debug(f"  Places textsearch error: {e}")
    return "", ""


def places_by_name(practice_name: str, city: str, state: str, zip_code: str,
                   address: str = "") -> tuple:
    """
    Places API lookup by practice name + location.
    Tries findplacefromtext (precise) first, falls back to textsearch.
    """
    if not GOOGLE_PLACES_API_KEY:
        return "", ""
    clean = re.sub(r'\b(LLC|PA|PLLC|DDS|DMD|Inc\.?|Corp\.?)\b', '', practice_name, flags=re.I).strip(', ')
    # Build query with as much location info as possible
    loc = " ".join(filter(None, [address, city, state, zip_code]))
    query = f"{clean} dentist {loc}".strip()
    log.info(f"  [Places/name] {query}")
    rating, count = _places_find_place(query, city, state)
    if not rating:
        rating, count = _places_text_search(query, city, state)
    return rating, count


def places_by_website(domain: str, city: str, state: str,
                      practice_name: str = "") -> tuple:
    """
    Places API lookup by practice name + city (domain alone is not a valid Places query).
    Falls back to textsearch with domain as a hint only if name is missing.
    """
    if not GOOGLE_PLACES_API_KEY:
        return "", ""
    if practice_name:
        clean = re.sub(r'\b(LLC|PA|PLLC|DDS|DMD|Inc\.?|Corp\.?)\b', '', practice_name, flags=re.I).strip(', ')
        query = f"{clean} dentist {city} {state}".strip()
    else:
        query = f"dentist {city} {state}".strip()
    log.info(f"  [Places/domain] {query}")
    rating, count = _places_find_place(query, city, state)
    if not rating:
        rating, count = _places_text_search(query, city, state)
    return rating, count


# ── Write helper ───────────────────────────────────────────────────────────────

def _write_results(ws, practices: dict, results: dict):
    """Write rating + count back into every row for each practice (multi-doctor rows)."""
    for col, hdr in ((C_GOOGLE_R, "Google Reviews Ranking"),
                     (C_GOOGLE_N, "Total # of Google Reviews")):
        if not ws.cell(HDR_ROW, col).value:
            ws.cell(HDR_ROW, col).value = hdr
            ws.cell(HDR_ROW, col).font  = Font(bold=True)

    for key, p in practices.items():
        if key not in results:
            continue
        rating, count = results[key]
        for r in p["rows"]:
            ws.cell(r, C_GOOGLE_R).value = rating
            ws.cell(r, C_GOOGLE_N).value = count


# ── API-only mode (no Playwright) ─────────────────────────────────────────────

def _run_api_only(practices: dict, wb, ws):
    """Use Google Places API only — no browser required."""
    results = {}
    total   = len(practices)
    for i, (key, p) in enumerate(practices.items(), 1):
        log.info(f"[{i}/{total}]  {p['practice']}  ({p['domain']})")
        if p["google_r_cur"] not in PLACEHOLDER:
            log.info(f"  Already done: {p['google_r_cur']} ★")
            results[key] = (p["google_r_cur"], p["google_n_cur"])
            continue
        rating, count = places_by_website(p["domain"], p["city"], p["state"])
        if not rating:
            rating, count = places_by_name(p["practice"], p["city"], p["state"], p["zip"])
        results[key] = (rating or "Not Found", count or "")
        log.info(f"  → {rating or 'Not Found'} ★  ({count})")
        time.sleep(0.2)   # Places API rate-limit buffer
        if i % 10 == 0:
            _write_results(ws, practices, results)
            wb.save(OUTPUT_FILE)
            log.info(f"  💾 Progress saved after {i} practices")
    _write_results(ws, practices, results)
    wb.save(OUTPUT_FILE)
    found = sum(1 for r, c in results.values() if r not in ("Not Found", ""))
    log.info(f"\n✅ Done — {found}/{len(results)} ratings found  →  {OUTPUT_FILE}")


# ── Main ──────────────────────────────────────────────────────────────────────

def run():
    if not PLAYWRIGHT_AVAILABLE and not GOOGLE_PLACES_API_KEY:
        log.error("Playwright not installed. Run:  pip install playwright && playwright install chromium")
        log.error("Or set GOOGLE_PLACES_API_KEY for API-based lookup (no browser needed).")
        return

    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active

    # Build unique-website dict  →  key = domain (or raw website if domain extraction fails)
    practices = {}
    for r in range(DATA_START, ws.max_row + 1):
        if ws.cell(r, C_INDEX).value is None:
            continue
        website = str(ws.cell(r, C_WEBSITE).value or "").strip()
        if not website or website in ("None", "nan", ""):
            continue
        domain = get_domain(website)
        key    = domain or website

        if key not in practices:
            practices[key] = {
                "practice":     str(ws.cell(r, C_PRACTICE).value or "").strip(),
                "doctor":       str(ws.cell(r, C_DOCTOR).value   or "").strip(),
                "address":      str(ws.cell(r, C_ADDRESS).value  or "").strip(),
                "city":         str(ws.cell(r, C_CITY).value     or "").strip(),
                "state":        str(ws.cell(r, C_STATE).value    or "").strip(),
                "zip":          str(ws.cell(r, C_ZIP).value      or "").strip(),
                "website":      website,
                "domain":       domain,
                "google_r_cur": str(ws.cell(r, C_GOOGLE_R).value or "").strip(),
                "google_n_cur": str(ws.cell(r, C_GOOGLE_N).value or "").strip(),
                "rows":         [r],
            }
        else:
            practices[key]["rows"].append(r)

    log.info(f"Loaded {ws.max_row - DATA_START + 1} rows | {len(practices)} unique websites")

    # ── Launch Playwright (visible — required for manual CAPTCHA solving) ──────
    # Skip if Places API key is set AND Playwright is unavailable
    if not PLAYWRIGHT_AVAILABLE and GOOGLE_PLACES_API_KEY:
        log.info("Running in Places API-only mode (no browser).")
        _run_api_only(practices, wb, ws)
        return

    log.info("Launching Playwright (visible browser)…")
    # Clear old profile if it might have cached non-English Google preferences
    _prefs_file = os.path.join(PW_USER_DATA, "Default", "Preferences")
    if os.path.exists(_prefs_file):
        try:
            import json as _json
            with open(_prefs_file, encoding="utf-8") as _pf:
                _prefs = _json.load(_pf)
            _lang = (_prefs.get("intl", {}) or {}).get("accept_languages", "en-US")
            if "en" not in _lang.lower():
                import shutil
                shutil.rmtree(PW_USER_DATA, ignore_errors=True)
                log.info("  Cleared cached non-English browser profile.")
        except Exception:
            pass
    os.makedirs(PW_USER_DATA, exist_ok=True)
    _pw = sync_playwright().__enter__()
    pw_ctx = _pw.chromium.launch_persistent_context(
        user_data_dir=PW_USER_DATA,
        headless=False,
        slow_mo=150,
        ignore_https_errors=True,
        locale="en-US",                        # force English UI
        timezone_id="America/New_York",        # appear as US East Coast
        geolocation={"latitude": 40.7128, "longitude": -74.0060},  # New York City
        permissions=["geolocation"],
        args=[
            "--disable-blink-features=AutomationControlled",
            "--no-first-run",
            "--no-default-browser-check",
            "--disable-infobars",
            "--lang=en-US",                    # Chrome language flag
        ],
        user_agent=HEADERS["User-Agent"],
    )
    # Apply stealth to every page opened from this context
    apply_stealth(pw_ctx)
    pw_page = pw_ctx.new_page()
    pw_page.set_extra_http_headers({
        "Accept-Language": "en-US,en;q=0.9",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    })

    # Pre-warm: visit Google.com first to build a real cookie session
    _prewarm_browser(pw_page)

    results = {}

    try:
        total = len(practices)
        for i, (key, p) in enumerate(practices.items(), 1):
            log.info(f"\n[{i}/{total}]  {p['practice']}  ({p['domain']})")

            # Skip if already filled
            if p["google_r_cur"] not in PLACEHOLDER:
                log.info(f"  Already done: {p['google_r_cur']} ★ ({p['google_n_cur']})")
                results[key] = (p["google_r_cur"], p["google_n_cur"])
                continue

            rating, count = "", ""

            # ── Method 1: Places API by name + address (most precise) ───────────
            if GOOGLE_PLACES_API_KEY:
                rating, count = places_by_name(
                    p["practice"], p["city"], p["state"], p["zip"], p.get("address", "")
                )
                if rating:
                    log.info(f"  ✓ Places/name: {rating} ★  ({count} reviews)")

            # ── Method 2: Places API by domain / name fallback ────────────────
            if not rating and GOOGLE_PLACES_API_KEY and p["domain"]:
                rating, count = places_by_website(
                    p["domain"], p["city"], p["state"], p["practice"]
                )
                if rating:
                    log.info(f"  ✓ Places/domain: {rating} ★  ({count} reviews)")

            # ── Method 3: Google Web Search → knowledge panel (fast) ──────────
            if not rating:
                rating, count = search_google_by_name(
                    p["practice"], p["city"], p["state"], pw_page
                )
                if rating:
                    log.info(f"  ✓ Google/search: {rating} ★  ({count} reviews)")

            # ── Method 4: Google Maps by name + address ────────────────────────
            if not rating:
                rating, count = search_maps_by_name(
                    p["practice"], p["doctor"],
                    p["address"], p["city"], p["state"], p["zip"],
                    pw_page,
                )
                if rating:
                    log.info(f"  ✓ Maps/name: {rating} ★  ({count} reviews)")

            if not rating:
                log.info("  ✗ Not found after all methods")

            results[key] = (rating or "Not Found", count or "")

            # Periodic save + batch pause every BATCH_SIZE practices
            if i % BATCH_SIZE == 0:
                _write_results(ws, practices, results)
                wb.save(OUTPUT_FILE)
                log.info(f"  💾 Progress saved after {i} practices")
                if i < total:
                    log.info(f"  ⏸  Batch pause {BATCH_PAUSE}s to avoid rate-limiting…")
                    time.sleep(BATCH_PAUSE)

    finally:
        try:
            pw_ctx.close()
            _pw.__exit__(None, None, None)
        except Exception:
            pass

    _write_results(ws, practices, results)
    wb.save(OUTPUT_FILE)

    found = sum(1 for r, c in results.values() if r not in ("Not Found", ""))
    log.info(f"\n✅ Done — {found}/{len(results)} ratings found")
    log.info(f"   Output: {OUTPUT_FILE}")


if __name__ == "__main__":
    run()
