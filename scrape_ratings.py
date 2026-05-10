"""
scrape_ratings.py
─────────────────
Reads Dental_Scrape_Output_v4.xlsx, scrapes Google rating + review count only
(Yelp and testimonials are skipped), and writes two output files:

  1. Dental_Scrape_Output_v5.xlsx  — full v4 file with Google columns updated
  2. google_ratings_output.xlsx    — compact standalone with just Google results

Scrapes PER UNIQUE PRACTICE (not per doctor row) to avoid duplicate requests.
Results are applied to every doctor row belonging to the same practice.
"""

import re
import json
import time
import logging
import warnings
from urllib.parse import quote_plus

import requests
import openpyxl
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", message="Unverified HTTPS request")

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PLAYWRIGHT_AVAILABLE = False

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

INPUT_FILE      = "/Users/mujahidulhaqtuhin/Downloads/dental/Dental_Scrape_Output_v6.xlsx"
OUTPUT_FILE     = "/Users/mujahidulhaqtuhin/Downloads/dental/Dental_Scrape_Output_v6.xlsx"
GOOGLE_OUT_FILE = "/Users/mujahidulhaqtuhin/Downloads/dental/google_ratings_output_v6.xlsx"

DELAY_SEC  = 4.0
TIMEOUT    = 15
PW_TIMEOUT = 30000

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# ── Column indices in v4 file (1-based) ──────────────────────────────────────
C_INDEX    = 1
C_PRACTICE = 2
C_DOCTOR   = 3
C_CITY     = 5
C_STATE    = 6
C_WEBSITE  = 8
C_GOOGLE_R = 42   # Google Reviews Ranking
C_GOOGLE_N = 43   # Total # of Google Reviews
TOTAL_COLS = 44   # v4 has 44 columns total


# ── HTTP helper ───────────────────────────────────────────────────────────────

def safe_get(url, retries=2):
    if not url or str(url).strip() in ("", "N/A", "None", "nan"):
        return None
    if not url.startswith("http"):
        url = "https://" + url.lstrip("/")
    from requests.exceptions import SSLError as RequestsSSLError
    for _ in range(retries):
        for verify in (True, False):
            try:
                r = requests.get(url, headers=HEADERS, timeout=TIMEOUT, verify=verify)
                if r.status_code == 200:
                    return r
                break
            except RequestsSSLError:
                if verify:
                    continue
                return None
            except Exception:
                time.sleep(1)
                break
    return None


def extract_text(html):
    soup = BeautifulSoup(html, "lxml")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    return soup.get_text(separator=" ", strip=True).lower()


# ── Business name helpers ─────────────────────────────────────────────────────

def _derive_name_from_domain(website):
    if not website:
        return None
    domain = re.sub(r'^https?://', '', website)
    domain = re.sub(r'^www{1,4}\.', '', domain)
    domain = domain.split('/')[0].split('?')[0]
    domain = domain.rsplit('.', 1)[0]
    name = domain.replace('-', ' ').replace('_', ' ')
    for kw in ['dental', 'dentist', 'dds', 'dmd', 'smile', 'family',
               'cosmetic', 'health', 'care', 'center', 'clinic', 'group',
               'studio', 'spa', 'premier', 'elite', 'bright', 'white']:
        name = re.sub(rf'(?<=[a-z])({kw})', r' \1', name, flags=re.IGNORECASE)
    return name.strip()


def _get_business_name_from_site(website, pw_page):
    if not website or not pw_page:
        return None
    url = website if website.startswith("http") else "https://" + website
    try:
        pw_page.goto(url, timeout=PW_TIMEOUT, wait_until="domcontentloaded")
        pw_page.wait_for_timeout(1500)
        soup = BeautifulSoup(pw_page.content(), "lxml")
        og = soup.find("meta", property="og:site_name")
        if og and og.get("content", "").strip():
            return og["content"].strip()
        og = soup.find("meta", property="og:title")
        if og and og.get("content", "").strip():
            t = og["content"].strip()
            return re.split(r'\s*[\|–\-]\s*', t)[0].strip()
        title = soup.find("title")
        if title:
            t = title.get_text(strip=True)
            return re.split(r'\s*[\|–\-]\s*', t)[0].strip()
    except Exception:
        pass
    return None


# ── Google search scraper ─────────────────────────────────────────────────────

_BAD_NAMES = re.compile(
    r'^(my\s+wordpress|wordpress|home|welcome|dental|dentist|dental\s+office|'
    r'dental\s+clinic|dental\s+practice|the\s+dental|dental\s+care|coming\s+soon)$',
    re.I,
)

def _is_bad_biz_name(name):
    return not name or len(name) < 4 or bool(_BAD_NAMES.match(name.strip()))


def _extract_from_html(html):
    """
    Multi-method extraction of (rating, count) from a Google search result page.
    Returns (rating_str, count_str) — either or both may be empty.
    """
    soup = BeautifulSoup(html, "lxml")
    g_rating, g_count = "", ""

    # Method 1: JSON-LD aggregateRating
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string or "")
            if not isinstance(data, list):
                data = [data]
            for item in data:
                ar = item.get("aggregateRating") or {}
                rv = str(ar.get("ratingValue", "")).strip()
                rc = str(ar.get("reviewCount", "") or ar.get("ratingCount", "")).strip()
                if rv and re.match(r'^[1-5]\.\d$', rv):
                    g_rating = rv
                    g_count  = rc
                    return g_rating, g_count
        except Exception:
            pass

    # Method 2: aria-label "Rated X out of 5"
    for tag in soup.find_all(attrs={"aria-label": True}):
        lbl = tag.get("aria-label", "")
        if "rated" in lbl.lower() and "out of 5" in lbl.lower():
            rm = re.search(r"([1-5]\.\d)", lbl)
            cm = re.search(r"\((\d[\d,]*)\)|(\d[\d,]*)\s*(?:user\s*)?reviews?", lbl, re.I)
            if rm:
                g_rating = rm.group(1)
                g_count  = (cm.group(1) or cm.group(2) or "").replace(",", "") if cm else ""
                return g_rating, g_count

    # Method 3: data-attrid="kc:/local:lu_factoid_overall_review_rating"
    for tag in soup.find_all(attrs={"data-attrid": True}):
        attrid = tag.get("data-attrid", "")
        if "review_rating" in attrid or "rating" in attrid:
            txt = tag.get_text(" ", strip=True)
            rm = re.search(r"([1-5]\.\d)", txt)
            if rm:
                g_rating = rm.group(1)
                cm = re.search(r"\((\d[\d,]+)\)", txt)
                if cm:
                    g_count = cm.group(1).replace(",", "")
                return g_rating, g_count

    # Method 4: visible text patterns
    text = soup.get_text(" ", strip=True).lower()

    # "4.8 (239)" or "4.8 · 239 reviews"
    patterns = [
        r'([1-5]\.\d)\s*\((\d[\d,]+)\)\s*(?:google\s+)?reviews?',
        r'([1-5]\.\d)\s*·\s*(\d[\d,]+)\s*(?:google\s+)?reviews?',
        r'([1-5]\.\d)\s+(\d[\d,]+)\s+(?:google\s+)?reviews?',
        r'([1-5]\.\d)\s*/?\s*5\s+\(?(\d[\d,]+)\)?\s*(?:google\s+)?reviews?',
        r'rated\s+([1-5]\.\d)\s+out\s+of\s+5.*?(\d[\d,]+)\s*(?:google\s+)?reviews?',
        r'([1-5]\.\d)\s+stars?\s+(\d[\d,]+)\s+(?:google\s+)?reviews?',
    ]
    for pat in patterns:
        m = re.search(pat, text, re.I)
        if m:
            g_rating = m.group(1)
            g_count  = m.group(2).replace(",", "")
            return g_rating, g_count

    # Rating only (no count found)
    m = re.search(r'([1-5]\.\d)\s*/?\s*5\b', text)
    if m:
        g_rating = m.group(1)

    return g_rating, g_count


def _extract_maps_rating(html):
    """Extract rating + count from a Google Maps search results page."""
    soup = BeautifulSoup(html, "lxml")

    # Method 1: JSON-LD aggregateRating
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string or "")
            if not isinstance(data, list):
                data = [data]
            for item in data:
                ar = item.get("aggregateRating") or {}
                rv = str(ar.get("ratingValue", "")).strip()
                rc = str(ar.get("reviewCount", "") or ar.get("ratingCount", "")).strip()
                if rv and re.match(r'^[1-5]\.\d$', rv):
                    return rv, rc
        except Exception:
            pass

    text = soup.get_text(" ", strip=True)

    # Method 2: aria-label with "stars" (Maps uses "4.8 stars 239 reviews")
    for tag in soup.find_all(attrs={"aria-label": True}):
        lbl = tag.get("aria-label", "")
        rm = re.search(r'([1-5]\.\d)\s+stars?', lbl, re.I)
        if rm:
            cm = re.search(r'(\d[\d,]+)\s+reviews?', lbl, re.I)
            return rm.group(1), cm.group(1).replace(",", "") if cm else ""

    # Method 3: text patterns
    patterns = [
        r'([1-5]\.\d)\s*\((\d[\d,]+)\)',
        r'([1-5]\.\d)\s+(\d[\d,]+)\s+reviews?',
        r'([1-5]\.\d)\s+stars?\s+(\d[\d,]+)',
    ]
    for pat in patterns:
        m = re.search(pat, text, re.I)
        if m:
            return m.group(1), m.group(2).replace(",", "")

    return "", ""


def _try_maps_query(query, pw_page):
    """Navigate to Google Maps search and return (rating, count, captcha_detected)."""
    url = f"https://www.google.com/maps/search/{quote_plus(query)}"
    pw_page.goto(url, timeout=PW_TIMEOUT, wait_until="domcontentloaded")
    pw_page.wait_for_timeout(4000)

    html = pw_page.content()

    # CAPTCHA detection
    if any(x in html.lower() for x in ["g-recaptcha", "recaptcha", "unusual traffic", "captcha"]):
        log.warning("  Google Maps: CAPTCHA detected")
        return "", "", True

    # On Maps search results, click the first result if it looks relevant
    try:
        # Try clicking first listing result
        first = pw_page.locator('a[href*="/maps/place/"]').first
        if first.count():
            first.click(timeout=5000)
            pw_page.wait_for_timeout(3000)
            html = pw_page.content()
    except Exception:
        pass

    rating, count = _extract_maps_rating(html)

    # Try Playwright aria-label extraction for rating+count
    if not rating or not count:
        try:
            # Rating element: "4.9 stars"
            for el in pw_page.locator("[aria-label]").all():
                try:
                    lbl = el.get_attribute("aria-label") or ""
                    rm = re.search(r'([1-5]\.\d)\s+stars?', lbl, re.I)
                    if rm and not rating:
                        rating = rm.group(1)
                    # Count element: "318 reviews" or "(318)"
                    cm = re.search(r'(\d[\d,]+)\s+reviews?', lbl, re.I)
                    if cm and not count:
                        count = cm.group(1).replace(",", "")
                    if rating and count:
                        break
                except Exception:
                    continue
        except Exception:
            pass

    # If still no count, scan all text for "N reviews" near the rating
    if rating and not count:
        try:
            full_text = pw_page.inner_text("body")
            cm = re.search(
                rf'{re.escape(rating)}\s+stars?\s+(\d[\d,]+)\s+reviews?'
                r'|(\d[\d,]+)\s+reviews?',
                full_text, re.I
            )
            if cm:
                count = (cm.group(1) or cm.group(2) or "").replace(",", "")
        except Exception:
            pass

    return rating, count, False


def _try_search_query(query, pw_page):
    """Navigate to Google Search and return (rating, count, captcha_detected)."""
    url = f"https://www.google.com/search?q={quote_plus(query)}&hl=en&gl=us"
    pw_page.goto(url, timeout=PW_TIMEOUT, wait_until="domcontentloaded")
    pw_page.wait_for_timeout(4500)

    html = pw_page.content()

    # CAPTCHA detection
    if any(x in html.lower() for x in ["g-recaptcha", "recaptcha", "unusual traffic",
                                         "captcha", "sorry, we couldn't"]):
        log.warning("  Google Search: CAPTCHA detected — waiting 12 s")
        pw_page.wait_for_timeout(12000)
        html = pw_page.content()
        if any(x in html.lower() for x in ["g-recaptcha", "recaptcha", "unusual traffic"]):
            return "", "", True

    rating, count = _extract_from_html(html)
    return rating, count, False


def scrape_google_rating(biz_name, city, state, pw_page, practice_name=None):
    """
    Try Google Maps first, then Google Search as fallback.
    Returns (g_rating, g_count).
    """
    if not pw_page:
        return "", ""

    name = biz_name if not _is_bad_biz_name(biz_name) else (practice_name or biz_name)
    if not name:
        return "", ""

    # ── Try Google Maps first ────────────────────────────────────────────────
    maps_queries = [
        f"{name} {city} {state} dentist",
        f"{name} dentist {city} {state}",
    ]
    if practice_name and practice_name != name and not _is_bad_biz_name(practice_name):
        maps_queries.append(f"{practice_name} {city} {state} dentist")

    for q in maps_queries:
        try:
            rating, count, captcha = _try_maps_query(q, pw_page)
            if captcha:
                log.warning("  Maps CAPTCHA — falling back to Search")
                break
            if rating:
                log.info(f"   → Maps hit: {q!r}")
                return rating, count
        except PlaywrightTimeout:
            log.warning(f"  Maps: timeout for {q!r}")
        except Exception as e:
            log.debug(f"  Maps error: {e}")

    # ── Fallback: Google Search ───────────────────────────────────────────────
    search_queries = [
        f"{name} {city} {state} dentist",
        f'"{name}" dentist {city} {state}',
    ]
    if practice_name and practice_name != name:
        search_queries.append(f"{practice_name} {city} {state} dentist")

    for q in search_queries:
        try:
            rating, count, captcha = _try_search_query(q, pw_page)
            if captcha:
                log.warning(f"  Search CAPTCHA — skipping {name}")
                return "", ""
            if rating:
                log.info(f"   → Search hit: {q!r}")
                return rating, count
        except PlaywrightTimeout:
            log.warning(f"  Search: timeout for {q!r}")
        except Exception as e:
            log.debug(f"  Search error: {e}")

    return "", ""


# ── Read / Write helpers ──────────────────────────────────────────────────────

def read_all_rows(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    rows = []
    for r in range(3, ws.max_row + 1):
        if ws.cell(r, C_PRACTICE).value is None:
            continue
        row = {}
        for c in range(1, TOTAL_COLS + 1):
            v = ws.cell(r, c).value
            row[c] = str(v).strip() if v is not None else ""
        rows.append(row)
    log.info(f"Loaded {len(rows)} rows from {filepath}")
    return rows


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    rows = read_all_rows(INPUT_FILE)

    pw_ctx  = None
    pw_page = None
    if PLAYWRIGHT_AVAILABLE:
        log.info("Launching Playwright with real Chrome…")
        _pw = sync_playwright().__enter__()
        try:
            browser = _pw.chromium.launch(
                headless=False,
                channel="chrome",
                args=["--disable-blink-features=AutomationControlled",
                      "--start-maximized"],
            )
        except Exception:
            # Fallback to bundled Chromium if Chrome not installed
            log.warning("Chrome not found, falling back to bundled Chromium (may hit CAPTCHA)")
            browser = _pw.chromium.launch(
                headless=False,
                args=["--disable-blink-features=AutomationControlled"],
            )
        pw_ctx = browser.new_context(
            user_agent=HEADERS["User-Agent"],
            locale="en-US",
            viewport={"width": 1280, "height": 800},
            ignore_https_errors=True,
        )
        pw_page = pw_ctx.new_page()
        pw_page.set_extra_http_headers({"Accept-Language": "en-US,en;q=0.9"})
    else:
        log.warning("Playwright not available — Google scraping will be skipped")

    try:
        cache = {}  # (practice, city, state) → {google_r, google_n}

        for row in rows:
            practice = row[C_PRACTICE]
            city     = row[C_CITY]
            state    = row[C_STATE]
            website  = row[C_WEBSITE]
            key      = (practice, city, state)

            if key in cache:
                continue

            log.info(f"▶ {practice} ({city}, {state})")

            biz_name = None
            if website:
                biz_name = _get_business_name_from_site(website, pw_page)
            if not biz_name:
                biz_name = _derive_name_from_domain(website) or practice
            log.info(f"   Biz name: {biz_name}")

            log.info("   Google search…")
            g_rating, g_count = scrape_google_rating(biz_name, city, state, pw_page, practice_name=practice)
            log.info(f"   → Google: {g_rating} ({g_count} reviews)")

            time.sleep(DELAY_SEC)
            cache[key] = {"google_r": g_rating, "google_n": g_count}

        # Apply to all rows — only overwrite if we got a non-empty result
        for row in rows:
            key = (row[C_PRACTICE], row[C_CITY], row[C_STATE])
            result = cache.get(key, {})
            # Always overwrite with freshly scraped values (even if empty — clears stale data)
            row[C_GOOGLE_R] = result.get("google_r", "")
            row[C_GOOGLE_N] = result.get("google_n", "")

    finally:
        try:
            if pw_ctx:
                pw_ctx.close()
            if 'browser' in dir() and browser:
                browser.close()
            _pw.__exit__(None, None, None)
        except Exception:
            pass

    # Write full v5 output (all 44 columns, Google updated)
    _write_v5(rows, OUTPUT_FILE)
    log.info(f"\n✅  Saved → {OUTPUT_FILE}")

    # Write compact google_ratings_output.xlsx
    _write_google_only(rows, cache, GOOGLE_OUT_FILE)
    log.info(f"✅  Saved → {GOOGLE_OUT_FILE}")
    log.info(f"    Practices scraped: {len(cache)}")


# ── Excel writers ─────────────────────────────────────────────────────────────

def _write_v5(rows, path):
    """Update v5 file in-place — only overwrite Google columns where we have new data."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    row_num = 3
    for row in rows:
        # Always write fresh scraped values (None clears the cell if scrape got nothing)
        g_r = row.get(C_GOOGLE_R, "")
        g_n = row.get(C_GOOGLE_N, "")
        ws.cell(row_num, C_GOOGLE_R).value = g_r if g_r not in ("", "None", "nan") else None
        ws.cell(row_num, C_GOOGLE_N).value = g_n if g_n not in ("", "None", "nan") else None
        row_num += 1

    wb.save(path)


def _write_google_only(rows, cache, path):
    """Write compact standalone Google ratings file."""
    wb  = Workbook()
    ws  = wb.active
    ws.title = "Google Ratings"

    thin    = Side(style="thin", color="CCCCCC")
    bdr     = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fnt = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    dat_fnt = Font(name="Arial", size=9)
    ctr     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    alt_fill = PatternFill("solid", fgColor="EBF3FB")
    wht_fill = PatternFill("solid", fgColor="FFFFFF")

    headers = ["Index", "Practice Name", "Doctor Name", "City", "State",
               "Practice Website", "Google Rating", "Google # Reviews"]

    for col_i, hdr in enumerate(headers, 1):
        c = ws.cell(1, col_i)
        c.value = hdr; c.font = hdr_fnt; c.fill = hdr_fill
        c.alignment = ctr; c.border = bdr

    seen = set()
    r_i = 2
    for row in rows:
        key = (row[C_PRACTICE], row[C_CITY], row[C_STATE])
        if key in seen:
            continue
        seen.add(key)
        rf = alt_fill if r_i % 2 == 0 else wht_fill
        vals = [
            row.get(C_INDEX, ""),
            row.get(C_PRACTICE, ""),
            row.get(C_DOCTOR, ""),
            row.get(C_CITY, ""),
            row.get(C_STATE, ""),
            row.get(C_WEBSITE, ""),
            row.get(C_GOOGLE_R, ""),
            row.get(C_GOOGLE_N, ""),
        ]
        left_cols = {2, 3, 6}
        for col_i, val in enumerate(vals, 1):
            c = ws.cell(r_i, col_i)
            c.value = val if val not in ("None", "nan") else ""
            c.font = dat_fnt; c.fill = rf; c.border = bdr
            c.alignment = lft if col_i in left_cols else ctr
        r_i += 1

    for col_i, w in enumerate([6, 28, 28, 14, 8, 32, 14, 16], 1):
        ws.column_dimensions[get_column_letter(col_i)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "B2"

    wb.save(path)


if __name__ == "__main__":
    main()
