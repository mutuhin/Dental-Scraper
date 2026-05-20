#!/usr/bin/env python3
"""
fix_missing.py
==============
Fill in missing Specialty and Memberships columns in a batch Excel output.

MODES
-----
  --show    INPUT.xlsx
        List every doctor row that is missing Specialty or Memberships,
        with the practice website URL.

  --urls    INPUT.xlsx  [--output urls.csv]
        For each practice with missing data, fetch its team/about page
        and find individual doctor bio URLs.  Outputs a CSV you can
        paste into ChatGPT or a browser.

  --patch   INPUT.xlsx  PATCH.csv  [--output OUTPUT.xlsx]
        Apply manual data from PATCH.csv and write a new Excel file.

PATCH CSV FORMAT
----------------
  Index,Doctor Name,Specialty,Memberships
  101,Dr. John Smith,"Implant Dentistry, Clear Aligners","ADA, AGD"
  102,*,"Cosmetic Dentistry",""

  • Doctor Name = "*"  →  patch ALL doctors for that Index
  • Leave Specialty or Memberships blank to keep existing value
  • Wrap values containing commas in double quotes

EXAMPLES
--------
  python fix_missing.py --show batch_05.xlsx
  python fix_missing.py --urls  batch_05.xlsx --output missing_urls.csv
  python fix_missing.py --patch batch_05.xlsx patches.csv --output batch_05_fixed.xlsx
"""

import argparse
import csv
import io
import re
import sys
import time
import os
from urllib.parse import urljoin, urlparse

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    sys.exit("pip install openpyxl")

try:
    import requests
    from bs4 import BeautifulSoup
    _REQUESTS_OK = True
except ImportError:
    _REQUESTS_OK = False

# ── Column numbers (1-based, row 2 = headers) ────────────────────────────────
COL_INDEX    = 1
COL_PRACTICE = 2
COL_DOCTOR   = 3
COL_WEBSITE  = 8
COL_ASSOC    = 40   # Associations / Memberships
COL_SPEC     = 41   # Doctor Specialty

HEADERS_ROW  = 2    # row 2 holds column labels
DATA_START   = 3    # data starts at row 3

EMPTY_VALUES = {"", "not found", "n/a", "none", "error"}

HEADERS_REQ = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def _is_empty(val):
    return str(val or "").strip().lower() in EMPTY_VALUES


def load_workbook(path):
    return openpyxl.load_workbook(path)


def _cell_val(ws, row, col):
    return ws.cell(row, col).value


def iter_data_rows(ws):
    """Yield (row_idx, index, doctor, website, assoc, specialty) for each data row."""
    for r in range(DATA_START, ws.max_row + 1):
        idx   = _cell_val(ws, r, COL_INDEX)
        doc   = _cell_val(ws, r, COL_DOCTOR)
        web   = _cell_val(ws, r, COL_WEBSITE)
        assoc = _cell_val(ws, r, COL_ASSOC)
        spec  = _cell_val(ws, r, COL_SPEC)
        if idx is None and doc is None:
            continue
        yield r, idx, doc, web, assoc, spec


def safe_get(url, timeout=10):
    if not _REQUESTS_OK:
        return None
    try:
        r = requests.get(url, headers=HEADERS_REQ, timeout=timeout, verify=False)
        return r if r.status_code == 200 else None
    except Exception:
        return None


# ── Mode: --show ──────────────────────────────────────────────────────────────

def cmd_show(path):
    wb = load_workbook(path)
    ws = wb.active
    missing = []
    for r, idx, doc, web, assoc, spec in iter_data_rows(ws):
        if _is_empty(assoc) or _is_empty(spec):
            missing.append({
                "row": r, "index": idx, "doctor": doc,
                "website": web or "",
                "assoc": assoc or "", "spec": spec or "",
            })

    if not missing:
        print("No missing Specialty/Memberships rows found.")
        return

    print(f"{'Row':<5} {'Index':<6} {'Doctor Name':<35} {'Miss':<12} {'Website'}")
    print("-" * 110)
    for m in missing:
        miss_cols = []
        if _is_empty(m["assoc"]): miss_cols.append("Memberships")
        if _is_empty(m["spec"]):  miss_cols.append("Specialty")
        miss = "+".join(miss_cols)
        print(f"{m['row']:<5} {str(m['index']):<6} {str(m['doctor'] or ''):<35} {miss:<12} {m['website']}")

    print(f"\nTotal missing: {len(missing)} rows")
    print("\nTip: run  --urls  to find individual doctor bio page links.")


# ── Mode: --urls ──────────────────────────────────────────────────────────────

_TEAM_KW = ("team", "doctor", "provider", "staff", "dentist", "meet", "our-doctor",
             "about", "physician", "specialist")
_BIO_KW  = ("bio", "profile", "doctor", "team", "provider", "dentist", "meet", "about",
             "dr-", "/dr/")
_SKIP_EXT = ('.pdf', '.jpg', '.jpeg', '.png', '.gif', '.svg', '.zip')


def _find_team_url(website):
    r = safe_get(website)
    if not r:
        return None
    soup = BeautifulSoup(r.text, "lxml")
    for a in soup.find_all("a", href=True):
        href = a["href"].lower()
        full = urljoin(website, a["href"])
        if urlparse(full).netloc != urlparse(website).netloc:
            continue
        if any(k in href for k in _TEAM_KW):
            return full
    return None


def _find_doctor_bio_urls(team_url, doctor_names):
    """Fetch team_url and return {doctor_name: bio_url} for each doctor."""
    r = safe_get(team_url)
    if not r:
        return {}
    soup = BeautifulSoup(r.text, "lxml")
    result = {}
    _DR_RE = re.compile(r'\b(Dr\.?|DDS|DMD)\b', re.I)

    for a in soup.find_all("a", href=True):
        href = a["href"]
        href_l = href.lower()
        if any(ext in href_l for ext in _SKIP_EXT):
            continue
        full = urljoin(team_url, href)
        if urlparse(full).netloc != urlparse(team_url).netloc:
            continue
        if not any(k in href_l for k in _BIO_KW):
            continue
        link_text = a.get_text(strip=True)
        if not _DR_RE.search(link_text) and not _DR_RE.search(href_l):
            continue
        # Try to match to a doctor name
        for name in doctor_names:
            if name in result:
                continue
            name_parts = re.sub(r'^Dr\.?\s+', '', name, flags=re.I).lower().split()
            last = name_parts[-1] if name_parts else ""
            if last and (last in full.lower() or last in link_text.lower()):
                result[name] = full
    return result


def cmd_urls(path, output_path):
    wb = load_workbook(path)
    ws = wb.active

    # Group by (index, website) → list of doctors missing data
    groups = {}
    for r, idx, doc, web, assoc, spec in iter_data_rows(ws):
        if not (_is_empty(assoc) or _is_empty(spec)):
            continue
        if not web:
            continue
        key = (idx, str(web).strip())
        groups.setdefault(key, []).append(str(doc or ""))

    if not groups:
        print("No missing rows with a website URL found.")
        return

    rows = []
    total = len(groups)
    for i, ((idx, website), doctors) in enumerate(groups.items(), 1):
        print(f"[{i}/{total}] Index {idx} — {website}")
        team_url = _find_team_url(website)
        if not team_url:
            team_url = website
        bio_urls = _find_doctor_bio_urls(team_url, doctors)
        for doc in doctors:
            bio = bio_urls.get(doc, team_url)
            rows.append({"Index": idx, "Doctor Name": doc,
                         "Practice Website": website, "Doctor Bio URL": bio})
        time.sleep(0.5)

    # Write output CSV
    fieldnames = ["Index", "Doctor Name", "Practice Website", "Doctor Bio URL"]
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)

    print(f"\nSaved {len(rows)} rows → {output_path}")
    print("Paste the 'Doctor Bio URL' links into ChatGPT to get Specialty & Memberships.")
    print("Then create patches.csv and run:  --patch INPUT.xlsx patches.csv")


# ── Mode: --patch ─────────────────────────────────────────────────────────────

def load_patches(patch_path):
    """
    Load patch CSV.  Returns list of dicts with keys:
      index, doctor, specialty, memberships
    Doctor "*" means apply to all doctors for that index.
    """
    patches = []
    with open(patch_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        # Normalise header names
        for row in reader:
            norm = {k.strip().lower(): (v or "").strip() for k, v in row.items()}
            patches.append({
                "index":       str(norm.get("index", "")).strip(),
                "doctor":      norm.get("doctor name", norm.get("doctor", "*")).strip(),
                "specialty":   norm.get("specialty", norm.get("doctor specialty", "")).strip(),
                "memberships": norm.get("memberships",
                               norm.get("associations / memberships",
                               norm.get("associations", ""))).strip(),
            })
    return patches


def _normalize(s):
    return re.sub(r'\s+', ' ', str(s or "").strip().lower())


def cmd_patch(input_path, patch_path, output_path):
    patches = load_patches(patch_path)
    print(f"Loaded {len(patches)} patch rows from {patch_path}")

    wb = load_workbook(input_path)
    ws = wb.active

    applied = 0
    skipped = 0

    for r, idx, doc, web, assoc, spec in iter_data_rows(ws):
        idx_str = str(idx or "").strip()
        doc_str = _normalize(doc)

        for p in patches:
            # Match on index
            if p["index"] != idx_str:
                continue
            # Match on doctor name ("*" = all doctors for this index)
            if p["doctor"] != "*" and _normalize(p["doctor"]) != doc_str:
                continue

            changed = False
            # Only fill if currently empty/not-found (don't overwrite good data)
            if p["specialty"] and _is_empty(ws.cell(r, COL_SPEC).value):
                ws.cell(r, COL_SPEC).value = p["specialty"]
                changed = True
            if p["memberships"] and _is_empty(ws.cell(r, COL_ASSOC).value):
                ws.cell(r, COL_ASSOC).value = p["memberships"]
                changed = True

            if changed:
                applied += 1
                print(f"  Row {r} | Index {idx_str} | {doc} → "
                      f"Spec={p['specialty']!r}  Assoc={p['memberships']!r}")
            else:
                skipped += 1
            break   # stop checking patches once one matched

    wb.save(output_path)
    print(f"\nApplied {applied} patches, skipped {skipped} (already had data).")
    print(f"Saved → {output_path}")


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    import warnings
    warnings.filterwarnings("ignore")

    p = argparse.ArgumentParser(
        description="Fill missing Specialty / Memberships in batch Excel output.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    p.add_argument("--show",  metavar="INPUT.xlsx",
                   help="List all rows missing Specialty or Memberships")
    p.add_argument("--urls",  metavar="INPUT.xlsx",
                   help="Find doctor bio page URLs for missing rows")
    p.add_argument("--patch", metavar="INPUT.xlsx",
                   help="Apply patches from PATCH_CSV to INPUT.xlsx")
    p.add_argument("patch_csv", nargs="?", metavar="PATCH.csv",
                   help="Patch CSV file (required with --patch)")
    p.add_argument("--output", "-o", metavar="OUTPUT.xlsx",
                   help="Output file path (default: INPUT_fixed.xlsx or missing_urls.csv)")

    args = p.parse_args()

    if args.show:
        cmd_show(args.show)

    elif args.urls:
        if not _REQUESTS_OK:
            sys.exit("pip install requests beautifulsoup4 lxml")
        out = args.output or os.path.splitext(args.urls)[0] + "_urls.csv"
        cmd_urls(args.urls, out)

    elif args.patch:
        if not args.patch_csv:
            p.error("--patch requires a PATCH.csv argument")
        out = args.output or os.path.splitext(args.patch)[0] + "_fixed.xlsx"
        cmd_patch(args.patch, args.patch_csv, out)

    else:
        p.print_help()


if __name__ == "__main__":
    main()
