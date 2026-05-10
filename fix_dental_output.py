"""
fix_dental_output.py
────────────────────
Reads Dental_Scrape_Output.xlsx and produces Dental_Scrape_Output_Fixed.xlsx

Fixes applied:
  1. Doctor row deduplication per practice
     (e.g. "Bijal Doshi DMD" + "Dr. Bijal Doshi" → one row with best name)
  2. Re-scrape # of Hygienists for rows still showing "See Website"
  3. Re-scrape Service (# of Mentions) + Technology fields for practices
     where all values are 0 / empty
  4. Blank out "Not Found" / "See Website" / "See Profile" / "Blocked" in
     non-critical display columns (TT stats, LI stats, email, Google rating)
"""

import re
import time
import logging
import warnings
from urllib.parse import urljoin

import requests
import openpyxl
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", message="Unverified HTTPS request")

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PLAYWRIGHT_AVAILABLE = False

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

INPUT_FILE  = "Dental_Scrape_Output.xlsx"
OUTPUT_FILE = "Dental_Scrape_Output_Fixed.xlsx"

DELAY_SEC  = 2.0
TIMEOUT    = 15
PW_TIMEOUT = 25000

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# ── Column index map (1-based) ────────────────────────────────────────────────
COL = {
    "index": 1, "practice_name": 2, "doctor_name": 3,
    "address": 4, "city": 5, "state": 6, "zip": 7,
    "website": 8, "email": 9, "hygienists": 10,
    "fb_url": 11, "fb_posts": 12, "fb_followers": 13,
    "ig_url": 14, "ig_posts": 15, "ig_followers": 16,
    "tt_url": 17, "tt_posts": 18, "tt_followers": 19,
    "li_url": 20, "li_posts": 21, "li_followers": 22,
    "cerec": 23, "cbct": 24, "lasers": 25, "ai": 26, "intraoral": 27,
    "invisalign": 28, "inv_tier": 29,
    "clear_aligners": 30, "veneers": 31, "implants": 32,
    "smile_make": 33, "whitening": 34, "sedation": 35,
    "holistic": 36, "cancer": 37, "locations": 38,
    "associations": 39, "specialty": 40,
    "google_rating": 41, "google_reviews": 42,
}
TOTAL_COLS = 42

SERVICE_KEYWORDS = {
    "invisalign": "Invisalign", "clear aligner": "Clear Aligners",
    "suresmile": "Clear Aligners", "clearcorrect": "Clear Aligners",
    "veneers": "Veneers", "implant": "Implants",
    "smile makeover": "Smile Makeovers", "whitening": "Teeth Whitening",
    "sedation": "Sedation Dentistry", "holistic": "Holistic Dentistry",
    "biological": "Holistic Dentistry", "cancer screening": "Cancer Screening",
}

TECH_KEYWORDS = {
    "cerec": "CEREC", "same day crown": "CEREC", "same-day crown": "CEREC",
    "cbct": "CBCT", "cone beam": "CBCT", "3d imaging": "CBCT",
    "3d x-ray": "CBCT", "i-cat": "CBCT", "dental ct": "CBCT",
    "laser": "Lasers", " ai ": "AI", "artificial intelligence": "AI",
    "intraoral scanner": "Intraoral Scanners", "itero": "Intraoral Scanners",
    "3shape": "Intraoral Scanners", "medit": "Intraoral Scanners",
}

# Columns whose "Not Found" / placeholder values should be blanked
PLACEHOLDER_BLANK = {
    "fb_posts", "fb_followers", "ig_posts", "ig_followers",
    "tt_url", "tt_posts", "tt_followers",
    "li_posts", "li_followers",
    "email", "google_rating", "google_reviews",
    "associations",   # blank "Not Found" only (keep real values)
}
PLACEHOLDER_VALUES = {
    "Not Found", "See Website", "See Profile", "Blocked",
    "N/A – Not Offered", "Not Listed", "ERROR",
}


# ── HTTP helpers ──────────────────────────────────────────────────────────────

def safe_get(url, retries=2):
    if not url or str(url).strip() in ("", "N/A", "None", "nan"):
        return None
    if not url.startswith("http"):
        url = "https://" + url.lstrip("/")
    from requests.exceptions import SSLError as RequestsSSLError
    for _ in range(retries):
        for verify in (True, False):
            try:
                r = requests.get(url, headers=HEADERS, timeout=TIMEOUT, verify=verify)
                if r.status_code == 200:
                    return r
                break
            except RequestsSSLError:
                if verify:
                    continue
                return None
            except Exception:
                time.sleep(1)
                break
    return None


def extract_text(html):
    soup = BeautifulSoup(html, "lxml")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    return soup.get_text(separator=" ", strip=True).lower()


def count_keyword(text, keyword):
    return len(re.findall(re.escape(keyword.lower()), text))


# ── Hygienist detection ───────────────────────────────────────────────────────

def find_hygienists_enhanced(text, soup):
    """
    Enhanced hygienist detection:
      1. Explicit 'X hygienists' in text
      2. Count distinct team members with RDH credential
      3. Count distinct 'Dental Hygienist' role mentions in short strings
    """
    # 1) Explicit number
    m = re.search(r"\b(\d+)\s+(?:registered\s+)?hygienists?", text, re.IGNORECASE)
    if m:
        return m.group(1)

    # 2) Count distinct RDH credential entries
    rdh_entries = set()
    for tag in soup.find_all(["h2", "h3", "h4", "h5", "p", "span", "div", "li"]):
        t = tag.get_text(strip=True)
        if len(t) > 120:
            continue
        if re.search(r"\bRDH\b|registered dental hygienist", t, re.IGNORECASE):
            key = re.sub(r"\s+", " ", t.strip().lower())[:60]
            rdh_entries.add(key)
    if rdh_entries:
        return str(len(rdh_entries))

    # 3) Count distinct 'dental hygienist' role entries in short strings
    hyg_entries = set()
    for tag in soup.find_all(["p", "span", "div", "li", "h3", "h4", "h5"]):
        t = tag.get_text(strip=True)
        if len(t) > 80:
            continue
        if re.search(r"dental\s+hygienist", t, re.IGNORECASE):
            key = re.sub(r"\s+", " ", t.strip().lower())[:60]
            hyg_entries.add(key)
    if hyg_entries:
        return str(len(hyg_entries))

    # 4) Just mentioned — leave as is (don't overwrite "See Website" with "Mentioned")
    return None


# ── Doctor name normalisation & deduplication ─────────────────────────────────

def _normalize_name(name):
    """Strip Dr./titles/credentials for comparison."""
    if not name or name in ("Not Found", "ERROR", ""):
        return ""
    n = re.sub(r"^Dr\.?\s+", "", str(name).strip(), flags=re.IGNORECASE)
    n = re.sub(
        r",?\s*(DDS|DMD|MD|M\.D\.|D\.D\.S\.|D\.M\.D\.)\s*.*$", "",
        n, flags=re.IGNORECASE,
    )
    n = re.sub(r"\s+", " ", n).strip().lower()
    return n


def _names_are_same(a, b):
    """True if two name strings likely refer to the same person."""
    na = _normalize_name(a)
    nb = _normalize_name(b)
    if not na or not nb:
        return False
    if na == nb:
        return True
    # Substring check
    if na in nb or nb in na:
        return True
    # Word-subset check: all words in shorter name appear in longer name
    wa = set(na.split())
    wb = set(nb.split())
    shorter, longer = (wa, wb) if len(wa) <= len(wb) else (wb, wa)
    if shorter and shorter.issubset(longer):
        return True
    return False


def deduplicate_doctors(group):
    """
    Remove duplicate doctor rows within a practice group.
    When duplicates are found, keep the row whose name has the most words
    (most complete name). Merges richer specialty/associations into keeper.
    """
    if len(group) <= 1:
        return group

    n = len(group)
    to_remove = set()

    for i in range(n):
        if i in to_remove:
            continue
        for j in range(n):
            if i == j or j in to_remove:
                continue
            if _names_are_same(group[i]["doctor_name"], group[j]["doctor_name"]):
                wi = len(_normalize_name(group[i]["doctor_name"]).split())
                wj = len(_normalize_name(group[j]["doctor_name"]).split())
                if wi >= wj:
                    # i is more/equally complete → remove j, merge data into i
                    to_remove.add(j)
                    _merge_richer(group[i], group[j])
                # if j is more complete, i will be removed when j is the outer loop

    kept = [row for idx, row in enumerate(group) if idx not in to_remove]
    if len(kept) < len(group):
        removed = len(group) - len(kept)
        log.info(
            f"  Deduped '{group[0]['practice_name']}': "
            f"{len(group)} → {len(kept)} rows ({removed} duplicates removed)"
        )
    return kept


def _merge_richer(target, source):
    """Copy non-empty / richer fields from source into target."""
    for field in ("specialty", "associations"):
        tv = (target.get(field) or "").strip()
        sv = (source.get(field) or "").strip()
        is_missing_t = tv in ("", "Not Found", "General Dentistry", "N/A", "ERROR")
        is_missing_s = sv in ("", "Not Found", "General Dentistry", "N/A", "ERROR")
        if is_missing_t and not is_missing_s:
            target[field] = sv
        elif not is_missing_t and not is_missing_s and len(sv) > len(tv):
            target[field] = sv  # prefer more detailed


# ── Read Excel ────────────────────────────────────────────────────────────────

def read_rows(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    rows = []
    for r in range(3, ws.max_row + 1):
        if ws.cell(r, COL["practice_name"]).value is None:
            continue
        row = {}
        for key, col in COL.items():
            v = ws.cell(r, col).value
            row[key] = str(v).strip() if v is not None else ""
        rows.append(row)
    log.info(f"Loaded {len(rows)} data rows from {filepath}")
    return rows


# ── Re-scrape helpers ─────────────────────────────────────────────────────────

def _fetch_site(website, pw_page=None):
    """
    Fetch homepage + top sub-pages for a website.
    Returns (all_text, soup) or ("", None) on failure.
    Falls back to Playwright when requests fails or gives thin content.
    """
    url = website if website.startswith("http") else "https://" + website
    all_text = ""
    soup = None

    # ── requests pass ─────────────────────────────────────────────────────────
    time.sleep(DELAY_SEC)
    r = safe_get(url)
    if r:
        soup = BeautifulSoup(r.text, "lxml")
        all_text = extract_text(r.text)

        # Fetch up to 8 relevant sub-pages — prioritise team/staff pages first
        seen_subs = set()
        all_links = soup.find_all("a", href=True)
        # Sort: team/staff/doctor links first (highest priority for hygienist count)
        def _link_priority(a):
            h = a["href"].lower()
            if any(kw in h for kw in ["team", "staff", "doctor", "provider", "meet"]):
                return 0
            if any(kw in h for kw in ["service", "technology", "treatment"]):
                return 1
            return 2
        all_links_sorted = sorted(all_links, key=_link_priority)

        for a in all_links_sorted:
            if len(seen_subs) >= 8:
                break
            href = a["href"].lower()
            if any(
                kw in href
                for kw in ["service", "about", "team", "technology", "treatment",
                            "doctor", "staff", "provider", "office", "meet"]
            ):
                sub_url = urljoin(url, a["href"])
                if sub_url.startswith("http") and sub_url not in seen_subs:
                    seen_subs.add(sub_url)
                    time.sleep(1)
                    sr = safe_get(sub_url)
                    if sr:
                        sub_soup = BeautifulSoup(sr.text, "lxml")
                        all_text += " " + extract_text(sr.text)
                        # Add sub-page soup elements for hygienist detection
                        for el in sub_soup.find_all(["div", "section", "article", "li"], limit=80):
                            soup.append(el)

    # ── Playwright enrichment: homepage + team page ───────────────────────────
    if pw_page:
        pages_to_render = [url]
        # Also try to find a team/staff page to get hygienist titles
        if soup:
            for a in soup.find_all("a", href=True):
                href = a["href"].lower()
                if any(kw in href for kw in ["team", "staff", "doctor", "provider", "meet"]):
                    team_url = urljoin(url, a["href"])
                    if team_url.startswith("http") and team_url not in pages_to_render:
                        pages_to_render.append(team_url)
                        break

        for render_url in pages_to_render[:2]:
            try:
                pw_page.goto(render_url, timeout=PW_TIMEOUT, wait_until="domcontentloaded")
                pw_page.wait_for_timeout(2500)
                pw_html = pw_page.content()
                pw_soup = BeautifulSoup(pw_html, "lxml")
                all_text += " " + extract_text(pw_html)
                if soup is None:
                    soup = pw_soup
                else:
                    for el in pw_soup.find_all(["div", "section", "article", "li"], limit=80):
                        soup.append(el)
            except Exception as e:
                log.debug(f"  Playwright fetch failed for {render_url}: {e}")

    return all_text, soup


def _rescrape_services_tech(all_text):
    """Return dict of service counts and tech flags from text."""
    svc = {
        "Invisalign": 0, "Clear Aligners": 0, "Veneers": 0,
        "Implants": 0, "Smile Makeovers": 0, "Teeth Whitening": 0,
        "Sedation Dentistry": 0, "Holistic Dentistry": 0, "Cancer Screening": 0,
    }
    for kw, cat in SERVICE_KEYWORDS.items():
        svc[cat] += count_keyword(all_text, kw)

    tech = set()
    for kw, name in TECH_KEYWORDS.items():
        if kw in all_text:
            tech.add(name)

    return {
        "invisalign":     str(svc["Invisalign"]),
        "clear_aligners": str(svc["Clear Aligners"]),
        "veneers":        str(svc["Veneers"]),
        "implants":       str(svc["Implants"]),
        "smile_make":     str(svc["Smile Makeovers"]),
        "whitening":      str(svc["Teeth Whitening"]),
        "sedation":       str(svc["Sedation Dentistry"]),
        "holistic":       str(svc["Holistic Dentistry"]),
        "cancer":         str(svc["Cancer Screening"]),
        "cerec":     "X" if "CEREC"              in tech else "",
        "cbct":      "X" if "CBCT"               in tech else "",
        "lasers":    "X" if "Lasers"             in tech else "",
        "ai":        "X" if "AI"                 in tech else "",
        "intraoral": "X" if "Intraoral Scanners" in tech else "",
    }


# ── Main fix pipeline ─────────────────────────────────────────────────────────

def fix_output():
    rows = read_rows(INPUT_FILE)

    # ── Step 1: Doctor deduplication per practice ─────────────────────────────
    deduped = []
    i = 0
    while i < len(rows):
        practice = rows[i]["practice_name"]
        group = []
        while i < len(rows) and rows[i]["practice_name"] == practice:
            group.append(rows[i])
            i += 1
        deduped.extend(deduplicate_doctors(group))

    log.info(f"After dedup: {len(deduped)} rows (was {len(rows)})")

    # ── Step 2: Blank placeholder values in display-only columns ─────────────
    for row in deduped:
        for col_name in PLACEHOLDER_BLANK:
            if row.get(col_name, "") in PLACEHOLDER_VALUES:
                row[col_name] = ""
        # Blank "See Website" for hygienists — will be re-scraped below.
        # Also treat "N/A" as unknown when there's a website (will try to verify).
        if row.get("hygienists") in ("See Website", "N/A"):
            row["hygienists"] = ""  # will be filled or left blank by step 3

    # ── Step 3: Re-scrape hygienists + missing service/tech data ──────────────
    # Launch Playwright once for all re-scraping
    pw_ctx  = None
    pw_page = None
    if PLAYWRIGHT_AVAILABLE:
        log.info("  Launching Playwright for re-scraping…")
        _pw = sync_playwright().__enter__()
        pw_ctx  = _pw.chromium.launch_persistent_context(
            user_data_dir="", headless=True,
            ignore_https_errors=True,
            args=["--disable-blink-features=AutomationControlled"],
            user_agent=HEADERS["User-Agent"],
        )
        pw_page = pw_ctx.new_page()

    try:
        website_cache = {}   # website → (all_text, soup)

        for row in deduped:
            website = row.get("website", "")
            if not website or website in ("", "None", "nan"):
                continue

            # Re-scrape hygienists for every practice that has a website
            # (blank = unknown, was "See Website" or "N/A")
            needs_hyg = row.get("hygienists", "") in ("", "Mentioned")

            # Re-scrape lasers/CBCT independently of service data —
            # a practice can have Invisalign data but still be missing tech fields
            needs_tech = (
                row.get("lasers", "") in ("", "None")
                or row.get("cbct", "") in ("", "None")
            )

            if not needs_hyg and not needs_tech:
                continue

            if website not in website_cache:
                log.info(f"  Re-scraping: {row['practice_name']} ({website})")
                all_text, soup = _fetch_site(website, pw_page)
                website_cache[website] = (all_text, soup)

            all_text, soup = website_cache[website]
            if not all_text:
                continue

            # Hygienist re-scrape
            if needs_hyg and soup:
                hyg = find_hygienists_enhanced(all_text, soup)
                if hyg:
                    log.info(f"    → Hygienists: {hyg}")
                    # Update ALL rows for this practice (same site = same count)
                    for r2 in deduped:
                        if (r2["practice_name"] == row["practice_name"]
                                and r2.get("hygienists", "") in ("", "N/A", "Mentioned")):
                            r2["hygienists"] = hyg

            # Tech re-scrape (lasers / CBCT / CEREC / AI / Intraoral)
            if needs_tech:
                updated = _rescrape_services_tech(all_text)
                log.info(
                    f"    → Tech: lasers={updated['lasers']} "
                    f"cbct={updated['cbct']} cerec={updated['cerec']}"
                )
                # Apply tech fields to ALL rows for this practice
                tech_keys = ["cerec", "cbct", "lasers", "ai", "intraoral"]
                for r2 in deduped:
                    if r2["practice_name"] == row["practice_name"]:
                        for k in tech_keys:
                            # Only fill if currently empty
                            if r2.get(k, "") in ("", "None", None):
                                r2[k] = updated[k]

    finally:
        if pw_ctx:
            try:
                pw_ctx.close()
                _pw.__exit__(None, None, None)
            except Exception:
                pass

    # ── Step 4: Write fixed output ────────────────────────────────────────────
    _write_output(deduped, OUTPUT_FILE)
    log.info(f"\n✅  Fixed output → {OUTPUT_FILE}")
    log.info(f"    Total rows: {len(deduped)}")


# ── Excel writer (same styling as original) ───────────────────────────────────

def _write_output(rows, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Scraped Data"

    hdr_font  = Font(name="Arial", bold=True, size=9)
    grp_font  = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    data_font = Font(name="Arial", size=9)
    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft  = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    fills = {
        "blue":      PatternFill("solid", fgColor="1F4E79"),
        "lt_blue":   PatternFill("solid", fgColor="BDD7EE"),
        "green":     PatternFill("solid", fgColor="375623"),
        "lt_green":  PatternFill("solid", fgColor="C6EFCE"),
        "purple":    PatternFill("solid", fgColor="7030A0"),
        "lt_purple": PatternFill("solid", fgColor="E2CFEC"),
        "orange":    PatternFill("solid", fgColor="833C00"),
        "lt_orange": PatternFill("solid", fgColor="FCE4D6"),
        "grey":      PatternFill("solid", fgColor="595959"),
        "lt_grey":   PatternFill("solid", fgColor="EDEDED"),
        "white":     PatternFill("solid", fgColor="FFFFFF"),
        "row_alt":   PatternFill("solid", fgColor="EBF3FB"),
    }

    def sc(cell, val, font=data_font, fill=fills["white"], align=ctr):
        cell.value     = val
        cell.font      = font
        cell.fill      = fill
        cell.alignment = align
        cell.border    = bdr

    # ── Group header row ──────────────────────────────────────────────────────
    groups = [
        (1,  10, "Practice Information",     "blue"),
        (11, 22, "Social Media",             "green"),
        (23, 27, "Technology in Practice",   "purple"),
        (28, 38, "Services (# of Mentions)", "orange"),
        (39, 42, "Doctor Data & Reviews",    "grey"),
    ]
    for start, end, label, color in groups:
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        sc(ws.cell(1, start), label, font=grp_font, fill=fills[color], align=ctr)

    # ── Column headers ────────────────────────────────────────────────────────
    col_headers = [
        "Index", "Practice Name", "Doctor Name", "Address", "City",
        "State", "Zip", "Practice Website", "Practice Email", "# of Hygienists",
        "Facebook URL", "FB # Posts", "FB Followers",
        "Instagram URL", "IG # Posts", "IG Followers",
        "TikTok URL", "TT # Posts", "TT Followers",
        "LinkedIn URL", "LI # Posts", "LI Followers",
        "CEREC (Same Day Crowns)", "CBCT (3D Imaging)", "Lasers", "AI",
        "Intraoral Scanners",
        "Invisalign (Mentions)", "Invisalign Tier (check manually)",
        "Clear Aligners", "Veneers", "Implants",
        "Smile Makeovers", "Teeth Whitening", "Sedation Dentistry",
        "Holistic Dentistry", "Cancer Screening", "# of Locations",
        "Associations / Memberships", "Doctor Specialty",
        "Google Rating", "Google # Reviews",
    ]

    fill_col = {}
    for c in range(1,  11): fill_col[c] = "lt_blue"
    for c in range(11, 23): fill_col[c] = "lt_green"
    for c in range(23, 28): fill_col[c] = "lt_purple"
    for c in range(28, 39): fill_col[c] = "lt_orange"
    for c in range(39, 43): fill_col[c] = "lt_grey"

    for col, hdr in enumerate(col_headers, 1):
        cell = ws.cell(2, col)
        sc(cell, hdr, font=hdr_font,
           fill=fills[fill_col.get(col, "white")], align=ctr)

    # ── Data rows ─────────────────────────────────────────────────────────────
    left_cols = {2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 14, 17, 20, 40, 41}

    def _v(row, key):
        v = row.get(key, "")
        return "" if v in ("None", "nan") else v

    for r_idx, row in enumerate(rows, start=3):
        rf = fills["row_alt"] if r_idx % 2 == 0 else fills["white"]

        row_vals = [
            _v(row, "index"), _v(row, "practice_name"), _v(row, "doctor_name"),
            _v(row, "address"), _v(row, "city"), _v(row, "state"),
            _v(row, "zip"), _v(row, "website"), _v(row, "email"),
            _v(row, "hygienists"),
            # Social
            _v(row, "fb_url"),   _v(row, "fb_posts"),   _v(row, "fb_followers"),
            _v(row, "ig_url"),   _v(row, "ig_posts"),   _v(row, "ig_followers"),
            _v(row, "tt_url"),   _v(row, "tt_posts"),   _v(row, "tt_followers"),
            _v(row, "li_url"),   _v(row, "li_posts"),   _v(row, "li_followers"),
            # Technology
            _v(row, "cerec"), _v(row, "cbct"), _v(row, "lasers"),
            _v(row, "ai"), _v(row, "intraoral"),
            # Services
            _v(row, "invisalign"), _v(row, "inv_tier"),
            _v(row, "clear_aligners"), _v(row, "veneers"), _v(row, "implants"),
            _v(row, "smile_make"), _v(row, "whitening"), _v(row, "sedation"),
            _v(row, "holistic"), _v(row, "cancer"), _v(row, "locations"),
            # Doctor data
            _v(row, "associations"), _v(row, "specialty"),
            _v(row, "google_rating"), _v(row, "google_reviews"),
        ]

        for c_idx, val in enumerate(row_vals, 1):
            cell  = ws.cell(r_idx, c_idx)
            align = lft if c_idx in left_cols else ctr
            sc(cell, val, font=data_font, fill=rf, align=align)

    # ── Column widths ─────────────────────────────────────────────────────────
    widths = {
        1: 6, 2: 28, 3: 30, 4: 30, 5: 14, 6: 7, 7: 8, 8: 32, 9: 28, 10: 12,
        11: 30, 12: 10, 13: 12, 14: 30, 15: 10, 16: 12,
        17: 18, 18: 10, 19: 12, 20: 30, 21: 10, 22: 12,
        23: 18, 24: 16, 25: 10, 26: 8, 27: 16,
        28: 14, 29: 26, 30: 14, 31: 10, 32: 10,
        33: 14, 34: 14, 35: 16, 36: 14, 37: 14, 38: 10,
        39: 38, 40: 30, 41: 14, 42: 14,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 40
    ws.freeze_panes = "B3"

    wb.save(path)


if __name__ == "__main__":
    fix_output()
