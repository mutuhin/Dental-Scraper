"""
patch_services.py  v3
──────────────────────
Re-scrapes 10 problem indices from  batch 1/100.xlsx.
Fixes:
  • Playwright-rendered crawl for ALL sites (not just blocked ones)
  • Collects EVERY nav-menu link, then all keyword-matched links
  • Probes known tech URL patterns (/technology, /cerec, /cbct, etc.)
  • Inserts extra doctor rows when a practice has more doctors than xlsx rows
  • Outputs to  batch 1/batch1_patched.xlsx

Does NOT touch: social media (11-22), Google ratings (42-43), Yelp (44-45)
"""

import os, sys, re, time, json, shutil, logging, warnings
warnings.filterwarnings("ignore", message="Unverified HTTPS request")

from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Alignment
from urllib.parse import urljoin, urlparse

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import dental_scraper as ds

ds.DELAY_SEC  = 1.5
ds.TIMEOUT    = 15
ds.PW_TIMEOUT = 25000
ds.CACHE_DIR  = "page_cache"

logging.basicConfig(level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger(__name__)

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE  = os.path.join(BASE_DIR, "batch 1", "100.xlsx")
OUTPUT_DIR  = os.path.join(BASE_DIR, "batch 1")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "batch1_patched.xlsx")

TARGETS = {
    20:  {"extra_urls": []},
    30:  {"extra_urls": []},
    35:  {"extra_urls": []},
    40:  {"base_url_override": "https://nofeardentist.com", "extra_urls": []},
    62:  {"extra_urls": []},
    79:  {"extra_urls": ["https://lapetitedent.com/meet-the-all-star-team-lpd/"]},
    84:  {"extra_urls": []},
    90:  {"extra_urls": ["https://snitzerdental.com/services",
                         "https://snitzerdental.com/general-dentistry",
                         "https://snitzerdental.com/cosmetic-dentistry"]},
    98:  {"extra_urls": []},
    100: {"extra_urls": ["https://alicedmd.com/services/"]},
}

# Column map (1-based) — matches 100.xlsx layout
C_IDX=1; C_NAME=2; C_DOC=3; C_ADDR=4; C_CITY=5; C_STATE=6; C_ZIP=7
C_SITE=8; C_EMAIL=9; C_HYG=10
C_CEREC=23; C_CBCT=24; C_LASER=25; C_AI=26; C_INTRA=27
C_INV=28; # 29=InvTier skipped
C_CLEAR=30; C_VENEER=31; C_IMPL=32; C_SMILE=33; C_WHITE=34
C_SEDAT=35; C_HOLIST=36; C_DPLAN=37; C_CANCER=38
C_LOC=39; C_ASSOC=40; C_SPEC=41
# 42=GoogleRating 43=GoogleReviews  ← DO NOT TOUCH
# 44=YelpRating   45=YelpReviews    ← DO NOT TOUCH
C_TESTI=46
DATA_START=3

_BLANK = frozenset({None,"","None","Not Found","0",0,"N/A","Not Offered",
                    "N/A - Not Offered","N/A – Not Offered","General Dentistry"})

_SKIP_EXTS = (".pdf",".jpg",".jpeg",".png",".gif",".svg",
              ".zip",".doc",".docx",".mp4",".mp3",".webp",".ico")
_SUB_KW = [
    "service","about","team","technology","treatment","doctor","provider",
    "procedure","cosmetic","implant","laser","invisalign","crown","cerec",
    "cbct","sedation","whitening","veneer","aligner","holistic","biological",
    "membership","plan","digital","3d","xray","x-ray","imaging","scan",
    "specialist","staff","meet","smile","restoration","restorative","general",
    "care","cancer","oral","teeth","bleach","advanced","equipment","innovation",
    "patient","contact",
]
# Tech-specific URL slugs to try even if not in nav
_TECH_PATHS = [
    "/technology","/our-technology","/dental-technology","/advanced-technology",
    "/digital-dentistry","/technology-dentistry","/state-of-the-art",
    "/cerec","/same-day-crowns","/same-day-crown",
    "/cbct","/3d-imaging","/cone-beam","/3d-xray",
    "/laser","/laser-dentistry","/laser-treatments",
    "/intraoral-scanner","/digital-impressions","/itero",
    "/office-tour","/our-office","/equipment",
    "/services","/treatments",
]
_TEST_CLASS_RE = re.compile(
    r"(testimonial|review|quote|patient.story|patient.review|"
    r"feedback|client.say|what.people|slider|swiper|slick|"
    r"rating.block|star.review)", re.I)
_TEST_ATTR_RE = re.compile(r"(testimonial|review|quote|patient)", re.I)
_STEALTH = """
(() => {
  Object.defineProperty(navigator,'webdriver',{get:()=>undefined});
  Object.defineProperty(navigator,'plugins',  {get:()=>[1,2,3,4,5]});
  Object.defineProperty(navigator,'languages',{get:()=>['en-US','en']});
  window.chrome={runtime:{},loadTimes:function(){},csi:function(){}};
  delete window.__playwright;
})();
"""


# ── Playwright helpers ──────────────────────────────────────────────────────

def _launch_pw():
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        return None, None, None
    _pw = sync_playwright().__enter__()
    ctx = _pw.chromium.launch_persistent_context(
        user_data_dir="/tmp/pw_patch_v3",
        headless=True, slow_mo=60, locale="en-US",
        args=["--disable-blink-features=AutomationControlled","--lang=en-US"],
        user_agent=(
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
    )
    try: ctx.add_init_script(_STEALTH)
    except Exception: pass
    page = ctx.new_page()
    page.set_extra_http_headers({"Accept-Language": "en-US,en;q=0.9"})
    return _pw, ctx, page


def _pw_get(page, url, wait_ms=2500):
    """Render url via Playwright. Returns (html, final_url) or (None, None)."""
    try:
        page.goto(url, timeout=28000, wait_until="domcontentloaded")
        page.wait_for_timeout(wait_ms)
        html = page.content()
        return (html, page.url) if len(html) > 500 else (None, None)
    except Exception as e:
        log.debug("PW %s: %s", url, e)
        return None, None


def _url_variants(url):
    parsed = urlparse(url if url.startswith("http") else "https://"+url)
    h, p = parsed.netloc, parsed.path or "/"
    if h.startswith("www."):
        b = h[4:]
        return [f"https://{b}{p}", f"http://{b}{p}", f"http://{h}{p}"]
    return [f"https://www.{h}{p}", f"http://{h}{p}", f"http://www.{h}{p}"]


def _same_domain(url, base):
    try:
        return urlparse(url).netloc == urlparse(base).netloc
    except Exception:
        return False


def _collect_all_links(soup, base_url, seen):
    """Collect ALL same-domain links from the rendered page."""
    urls = []
    for a in soup.find_all("a", href=True):
        full = urljoin(base_url, a["href"])
        if not full.startswith("http"):
            continue
        if not _same_domain(full, base_url):
            continue
        full = full.split("#")[0].rstrip("/") or full
        if full in seen:
            continue
        if any(full.lower().endswith(ext) for ext in _SKIP_EXTS):
            continue
        seen.add(full)
        urls.append(full)
    return urls


def _collect_kw_links(soup, base_url, seen):
    """Collect keyword-matched same-domain links."""
    urls = []
    for a in soup.find_all("a", href=True):
        full = urljoin(base_url, a["href"])
        if not full.startswith("http") or not _same_domain(full, base_url):
            continue
        full = full.split("#")[0].rstrip("/") or full
        if full in seen or any(full.lower().endswith(e) for e in _SKIP_EXTS):
            continue
        href_l = a["href"].lower()
        if any(kw in href_l for kw in _SUB_KW):
            seen.add(full)
            urls.append(full)
    return urls


# ── Main scraper ────────────────────────────────────────────────────────────

def _scrape_site(idx, name, website, pw_page):
    overrides = TARGETS.get(idx, {})
    if overrides.get("base_url_override"):
        base_url = overrides["base_url_override"]
    else:
        base_url = website.split("?")[0].rstrip("/")
        if not base_url.startswith("http"):
            base_url = "https://" + base_url

    extra_urls = overrides.get("extra_urls", [])
    log.info("\n[%d] %s  =>  %s", idx, name, base_url)

    cache_folder = ds._cache_dir(idx, name)
    counter = [0]
    all_text, all_soup, all_soups = "", None, []

    def _save(ptype, url, html):
        ds._cache_html(cache_folder, ptype, url, html)

    # ── 1. Homepage ──────────────────────────────────────────────────────
    # Try requests first; if blocked, use Playwright
    r = ds.safe_get(base_url)
    if not r:
        for alt in _url_variants(base_url):
            r = ds.safe_get(alt)
            if r: base_url = alt; break

    if r and len(r.text) > 800:
        all_soup = BeautifulSoup(r.text, "lxml")
        all_text = ds.extract_text(r.text)
        _save("homepage", r.url, r.text)
        log.info("  Homepage (req): %d chars", len(r.text))
    elif pw_page:
        log.info("  Requests blocked — using Playwright for homepage")
        for try_url in [base_url] + _url_variants(base_url):
            html, final = _pw_get(pw_page, try_url)
            if html:
                all_soup = BeautifulSoup(html, "lxml")
                all_text = ds.extract_text(html)
                base_url = try_url
                _save("homepage", final or try_url, html)
                log.info("  Homepage (PW): %d chars", len(html))
                break

    if not all_soup:
        log.warning("  Could not load homepage for idx %d", idx)
        return {}

    seen = {base_url, base_url.rstrip("/")+"/"}

    # ── 2. Playwright render of homepage for JS-rendered nav links ───────
    pw_nav_soup = None
    if pw_page:
        html_pw, _ = _pw_get(pw_page, base_url)
        if html_pw and len(html_pw) > 800:
            pw_nav_soup = BeautifulSoup(html_pw, "lxml")
            all_text += " " + ds.extract_text(html_pw)
            all_soups.append(("pw_home", pw_nav_soup))
            _save("pw_homepage", base_url, html_pw)

    use_soup = pw_nav_soup or all_soup

    # ── 3. Collect sub-pages (nav + keyword-matched from rendered page) ──
    nav_links = _collect_all_links(use_soup, base_url, seen)
    kw_links  = _collect_kw_links(all_soup, base_url, seen)
    all_links = nav_links + kw_links
    log.info("  Found %d nav links, %d kw links to crawl", len(nav_links), len(kw_links))

    for sub_url in all_links[:80]:
        time.sleep(ds.DELAY_SEC)
        sr = ds.safe_get(sub_url)
        if sr and len(sr.text) > 300:
            all_text += " " + ds.extract_text(sr.text)
            ssoup = BeautifulSoup(sr.text, "lxml")
            all_soups.append(("sub", ssoup))
            counter[0] += 1
            _save(f"sub_{counter[0]:02d}", sub_url, sr.text)

    # ── 4. Probe tech URL patterns not yet visited ────────────────────────
    tech_hits = 0
    for path in _TECH_PATHS:
        probe = base_url.rstrip("/") + path
        if probe in seen:
            continue
        seen.add(probe)
        time.sleep(0.8)
        pr = ds.safe_get(probe)
        if pr and len(pr.text) > 800:
            all_text += " " + ds.extract_text(pr.text)
            all_soups.append(("tech_probe", BeautifulSoup(pr.text, "lxml")))
            counter[0] += 1
            _save(f"sub_{counter[0]:02d}", probe, pr.text)
            tech_hits += 1
        elif pw_page:
            html_t, ft = _pw_get(pw_page, probe, wait_ms=1500)
            if html_t and len(html_t) > 800 and ft and ft.rstrip("/") not in seen:
                all_text += " " + ds.extract_text(html_t)
                all_soups.append(("tech_probe_pw", BeautifulSoup(html_t, "lxml")))
                counter[0] += 1
                _save(f"sub_{counter[0]:02d}", probe, html_t)
                seen.add(ft.rstrip("/"))
                tech_hits += 1
    if tech_hits:
        log.info("  Tech probe: %d new pages", tech_hits)

    # ── 5. Extra known URLs ───────────────────────────────────────────────
    for eurl in extra_urls:
        if eurl in seen: continue
        seen.add(eurl)
        log.info("  Extra: %s", eurl)
        time.sleep(ds.DELAY_SEC)
        er = ds.safe_get(eurl)
        if er and len(er.text) > 200:
            all_text += " " + ds.extract_text(er.text)
            all_soups.append(("extra", BeautifulSoup(er.text, "lxml")))
            counter[0] += 1
            _save(f"sub_{counter[0]:02d}", eurl, er.text)
        elif pw_page:
            html_e, _ = _pw_get(pw_page, eurl)
            if html_e:
                all_text += " " + ds.extract_text(html_e)
                all_soups.append(("extra_pw", BeautifulSoup(html_e, "lxml")))
                counter[0] += 1
                _save(f"sub_{counter[0]:02d}", eurl, html_e)

    if not all_text.strip():
        log.warning("  No content for idx %d", idx)
        return {}

    log.info("  Total: %d chars from %d pages", len(all_text), counter[0]+1)
    result = {}

    # ── Tech detection ────────────────────────────────────────────────────
    _tt = [all_text]
    mpath = os.path.join(cache_folder, "manifest.json")
    if os.path.exists(mpath):
        try: m = json.load(open(mpath))
        except: m = {}
        for _, pi in m.get("pages",{}).items():
            fp = os.path.join(cache_folder, pi.get("file",""))
            if os.path.exists(fp):
                _tt.append(ds.extract_text(open(fp, encoding="utf-8",
                                                 errors="replace").read()))
    ct = " ".join(_tt)
    tf = set()
    for kw, tn in ds.TECH_KEYWORDS.items():
        if kw in ct: tf.add(tn)
    if "AI" not in tf and re.search(r"\bai\b", ct): tf.add("AI")

    result["cerec"]    = "X" if "CEREC"              in tf else ""
    result["cbct"]     = "X" if "CBCT"               in tf else ""
    result["lasers"]   = "X" if "Lasers"             in tf else ""
    result["ai"]       = "X" if "AI"                 in tf else ""
    result["intraoral"]= "X" if "Intraoral Scanners" in tf else ""
    log.info("  Tech: CEREC=%s CBCT=%s Laser=%s AI=%s Intra=%s",
             result["cerec"] or "-", result["cbct"] or "-",
             result["lasers"] or "-", result["ai"] or "-",
             result["intraoral"] or "-")

    # ── Service detection ─────────────────────────────────────────────────
    svc = {k: 0 for k in [
        "Invisalign","Clear Aligners","Veneers","Implants","Smile Makeovers",
        "Teeth Whitening","Sedation Dentistry","Holistic Dentistry",
        "Dental Plan","Cancer Screening",
    ]}
    _sp = []
    if os.path.exists(mpath):
        try: m = json.load(open(mpath))
        except: m = {}
        _seen_u: set = set()
        for _, pi in m.get("pages",{}).items():
            pu = pi.get("url","").split("#")[0]
            if pu in _seen_u: continue
            _seen_u.add(pu)
            fp = os.path.join(cache_folder, pi.get("file",""))
            if os.path.exists(fp):
                raw = open(fp, encoding="utf-8", errors="replace").read()
                _sp.append((ds.extract_body_text(raw), ds.extract_text(raw)))
    if not _sp:
        _sp = [(all_text, all_text)]
    else:
        _sp.append((all_text, all_text))

    _b = {k: 0 for k in svc}; _f = {k: 0 for k in svc}
    for bt, ft in _sp:
        for kw, cat in ds.SERVICE_KEYWORDS.items():
            _b[cat] += ds.count_keyword_capped(bt, kw, cap=5)
            _f[cat] += ds.count_keyword_capped(ft, kw, cap=3)
    for cat in svc:
        svc[cat] = _b[cat] if _b[cat] > 0 else _f[cat]

    result["invisalign"]      = svc["Invisalign"]
    result["clear_aligners"]  = svc["Clear Aligners"]
    result["veneers"]         = svc["Veneers"]
    result["implants"]        = svc["Implants"]
    result["smile_makeovers"] = svc["Smile Makeovers"]
    result["whitening"]       = svc["Teeth Whitening"]
    result["sedation"]        = svc["Sedation Dentistry"]
    result["holistic"]        = svc["Holistic Dentistry"]
    result["dental_plan"]     = "Mentioned" if svc["Dental Plan"] > 0 else ""
    result["cancer_screening"]= svc["Cancer Screening"]
    log.info("  Svc: Inv=%s Cl=%s Ven=%s Impl=%s White=%s Sed=%s Hol=%s Plan=%s Canc=%s",
             result["invisalign"], result["clear_aligners"], result["veneers"],
             result["implants"], result["whitening"], result["sedation"],
             result["holistic"], result["dental_plan"] or "-", result["cancer_screening"])

    result["associations"]    = ds.find_associations(all_text)
    result["specialty"]       = ds.find_specialty(all_text)
    result["locations_count"] = ds.find_locations_count(
        all_text, all_soup or BeautifulSoup("","lxml"))
    log.info("  Assoc=%s  Spec=%s  Locs=%s",
             result["associations"], result["specialty"], result["locations_count"])

    # ── Testimonials ──────────────────────────────────────────────────────
    ts, tt = set(), 0
    for sp in ([all_soup] if all_soup else []) + [s for _,s in all_soups]:
        for blk in sp.find_all(["div","section","article","blockquote","li"],
                               class_=_TEST_CLASS_RE):
            k = blk.get_text(separator=" ",strip=True)[:80]
            if k and k not in ts: ts.add(k); tt += 1
        for blk in sp.find_all(lambda t: any(
            _TEST_ATTR_RE.search(str(v))
            for k,v in t.attrs.items()
            if k.startswith("data-") and isinstance(v,str)
        )):
            k = blk.get_text(separator=" ",strip=True)[:80]
            if k and k not in ts: ts.add(k); tt += 1
    if tt == 0:
        for sp in ([all_soup] if all_soup else []) + [s for _,s in all_soups]:
            for bq in sp.find_all("blockquote"):
                k = bq.get_text(separator=" ",strip=True)[:80]
                if k and k not in ts: ts.add(k); tt += 1
    result["testimonials"] = str(tt) if tt > 0 else "0"
    log.info("  Testimonials: %s", result["testimonials"])

    # ── Doctors ───────────────────────────────────────────────────────────
    if all_soup:
        doctors, _ = ds.scrape_doctors_full(
            all_soup, base_url, all_text, None,
            all_soups_for_team=all_soups)
        result["doctors"] = doctors
        log.info("  Doctors: %s", [d["name"] for d in doctors[:6]])
    else:
        result["doctors"] = []

    return result


# ── Read xlsx ───────────────────────────────────────────────────────────────

def _read_xlsx(filepath):
    """Returns (idx_to_rows, idx_info, all_rows_data)."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    idx_to_rows = {}   # idx -> [row_number, ...]
    idx_info    = {}   # idx -> {name, website}
    all_rows    = {}   # row_number -> dict of col->value (for later copy)
    rn = DATA_START
    for row_cells in ws.iter_rows(min_row=DATA_START, values_only=False):
        raw = row_cells[C_IDX-1].value
        if raw is None:
            rn += 1; continue
        try: idx = int(raw)
        except: rn += 1; continue
        if idx in TARGETS:
            idx_to_rows.setdefault(idx, []).append(rn)
            if idx not in idx_info:
                idx_info[idx] = {
                    "name":    str(row_cells[C_NAME-1].value or ""),
                    "website": str(row_cells[C_SITE-1].value or ""),
                }
        # Store ALL column values for this row (for copying to new doctor rows)
        all_rows[rn] = {c+1: row_cells[c].value for c in range(len(row_cells))}
        rn += 1
    wb.close()
    return idx_to_rows, idx_info, all_rows


# ── Write patches ───────────────────────────────────────────────────────────

def _write_patches(filepath, idx_to_rows, idx_to_result, all_rows_data):
    """
    Patch existing rows AND insert extra doctor rows where needed.
    Social (11-22), Google (42-43), Yelp (44-45) are never touched.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # Process in REVERSE order of row number so inserts don't shift earlier rows
    sorted_indices = sorted(idx_to_result.keys(),
                            key=lambda i: idx_to_rows[i][0], reverse=True)

    for idx in sorted_indices:
        result   = idx_to_result[idx]
        row_nums = idx_to_rows.get(idx, [])
        if not row_nums: continue

        doctors   = result.get("doctors", [])
        n_existing = len(row_nums)
        n_doctors  = len(doctors)
        updates   = 0

        def _set(rn, col, new_val):
            nonlocal updates
            # Never touch social (11-22) or ratings (42-45)
            if 11 <= col <= 22 or 42 <= col <= 45:
                return
            cur = ws.cell(rn, col).value
            if cur in _BLANK or str(cur).strip() in _BLANK:
                if new_val not in (None, "", 0):
                    ws.cell(rn, col).value = new_val
                    updates += 1

        # Write practice-level fields to ALL existing rows
        for rn in row_nums:
            _set(rn, C_CEREC,  result.get("cerec",    ""))
            _set(rn, C_CBCT,   result.get("cbct",     ""))
            _set(rn, C_LASER,  result.get("lasers",   ""))
            _set(rn, C_AI,     result.get("ai",       ""))
            _set(rn, C_INTRA,  result.get("intraoral",""))
            _set(rn, C_INV,    result.get("invisalign",      0))
            _set(rn, C_CLEAR,  result.get("clear_aligners",  0))
            _set(rn, C_VENEER, result.get("veneers",         0))
            _set(rn, C_IMPL,   result.get("implants",        0))
            _set(rn, C_SMILE,  result.get("smile_makeovers", 0))
            _set(rn, C_WHITE,  result.get("whitening",       0))
            _set(rn, C_SEDAT,  result.get("sedation",        0))
            _set(rn, C_HOLIST, result.get("holistic",        0))
            _set(rn, C_DPLAN,  result.get("dental_plan",     ""))
            _set(rn, C_CANCER, result.get("cancer_screening", 0))
            _set(rn, C_LOC,    result.get("locations_count", ""))
            _set(rn, C_ASSOC,  result.get("associations",    ""))
            _set(rn, C_SPEC,   result.get("specialty",       ""))
            _set(rn, C_TESTI,  result.get("testimonials",    ""))

        # Write per-doctor name/assoc/spec to each existing row (in order)
        for i, rn in enumerate(row_nums):
            if i < n_doctors:
                doc = doctors[i]
                _set(rn, C_DOC,   doc["name"])
                _set(rn, C_ASSOC, doc.get("associations", ""))
                _set(rn, C_SPEC,  doc.get("specialty",    ""))

        # ── Insert extra rows for additional doctors ──────────────────────
        if n_doctors > n_existing:
            extra_count = n_doctors - n_existing
            last_rn = row_nums[-1]
            log.info("  [%d] Inserting %d extra doctor row(s) after row %d",
                     idx, extra_count, last_rn)

            # Insert rows immediately after the last existing row
            ws.insert_rows(last_rn + 1, extra_count)

            # Copy practice-level data from the first existing row to new rows
            src_data = all_rows_data.get(row_nums[0], {})
            for offset, doc_idx in enumerate(range(n_existing, n_doctors)):
                new_rn = last_rn + 1 + offset
                doc = doctors[doc_idx]
                # Copy all columns from template row
                for col, val in src_data.items():
                    ws.cell(new_rn, col).value = val
                # Override doctor-specific fields
                ws.cell(new_rn, C_DOC).value   = doc["name"]
                ws.cell(new_rn, C_ASSOC).value = doc.get("associations", "")
                ws.cell(new_rn, C_SPEC).value  = doc.get("specialty",    "")
                # Also write patched practice fields to new rows
                for col, key in [
                    (C_CEREC,"cerec"),(C_CBCT,"cbct"),(C_LASER,"lasers"),
                    (C_AI,"ai"),(C_INTRA,"intraoral"),
                    (C_INV,"invisalign"),(C_CLEAR,"clear_aligners"),
                    (C_VENEER,"veneers"),(C_IMPL,"implants"),
                    (C_SMILE,"smile_makeovers"),(C_WHITE,"whitening"),
                    (C_SEDAT,"sedation"),(C_HOLIST,"holistic"),
                    (C_DPLAN,"dental_plan"),(C_CANCER,"cancer_screening"),
                    (C_LOC,"locations_count"),(C_TESTI,"testimonials"),
                ]:
                    v = result.get(key, "")
                    if v not in (None, "", 0):
                        ws.cell(new_rn, col).value = v
                updates += 1

        log.info("  [%d] %d cell(s) updated / added across %d row(s)",
                 idx, updates, len(row_nums))

    wb.save(filepath)
    log.info("\nSaved => %s", filepath)


# ── Entry point ─────────────────────────────────────────────────────────────

def main():
    log.info("="*60)
    log.info("  patch_services.py v3")
    log.info("  Input : %s", INPUT_FILE)
    log.info("  Output: %s", OUTPUT_FILE)
    log.info("="*60)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    shutil.copy2(INPUT_FILE, OUTPUT_FILE)
    log.info("Copied input to output.")

    idx_to_rows, idx_info, all_rows_data = _read_xlsx(OUTPUT_FILE)
    log.info("Found indices: %s", sorted(idx_to_rows.keys()))

    _pw, ctx, pw_page = _launch_pw()
    log.info("Playwright: %s", "ready" if pw_page else "unavailable")

    idx_to_result = {}
    try:
        for idx in sorted(TARGETS.keys()):
            if idx not in idx_to_rows:
                log.warning("Index %d not in xlsx -- skipping", idx)
                continue
            info   = idx_info[idx]
            result = _scrape_site(idx, info["name"], info["website"], pw_page)
            if result:
                idx_to_result[idx] = result
            time.sleep(1.5)
    finally:
        if ctx:
            try: ctx.close(); _pw.__exit__(None,None,None)
            except: pass

    if not idx_to_result:
        log.error("Nothing scraped.")
        return

    log.info("\n--- Writing patches ---")
    _write_patches(OUTPUT_FILE, idx_to_rows, idx_to_result, all_rows_data)
    log.info("Done.")


if __name__ == "__main__":
    main()
