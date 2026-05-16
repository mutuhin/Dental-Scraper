"""
Dental Practice Web Scraper v2
================================
Reads practices from Sample_List_for_Web_Scraping.xlsx
Uses requests + BeautifulSoup for static content.
Uses Playwright (browser) for Facebook, Instagram, Invisalign Tier, and
JS-rendered emails.
Outputs filled data to Dental_Scrape_Output.xlsx

HOW TO RUN:
    pip install requests beautifulsoup4 openpyxl lxml playwright
    playwright install chromium
    python dental_scraper.py

IMPORTANT:
    - Run on YOUR computer (not a restricted server)
    - Adds delays between requests to be polite
    - Some sites block bots — those cells will show "Blocked"
    - Facebook/Instagram may block or require login for full stats
    - Invisalign Tier is scraped from the official JS-rendered locator
"""

import time
import re
import sys
import os
import json
import random
import logging
import warnings
import unicodedata
from urllib.parse import urljoin, urlparse, quote_plus, unquote
from requests.exceptions import SSLError as RequestsSSLError

# Suppress urllib3 InsecureRequestWarning globally (we log our own warning)
warnings.filterwarnings("ignore", message="Unverified HTTPS request")

import requests
from bs4 import BeautifulSoup

# curl_cffi — bypasses Cloudflare TLS fingerprint checks (optional, graceful fallback)
try:
    from curl_cffi import requests as cffi_requests
    _CFFI_AVAILABLE = True
except ImportError:
    _CFFI_AVAILABLE = False
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PLAYWRIGHT_AVAILABLE = False
    print(
        "WARNING: playwright not installed.\n"
        "  Run: pip install playwright && playwright install chromium\n"
        "  Falling back to requests-only mode (limited FB/IG/Invisalign data).\n"
    )

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION  — edit these values to control behaviour
# ─────────────────────────────────────────────────────────────────────────────
INPUT_FILE   = "/Users/mujahidulhaqtuhin/Downloads/dental/py files/6000 Data COMPLETE.xlsx"
OUTPUT_FILE  = "/Users/mujahidulhaqtuhin/Downloads/dental/py files/100data.xlsx"
SKIPPED_DIR  = "skipped"   # folder + file for bot-blocked / unreachable sites
# In GitHub Actions (CI=true) use tighter limits so 100 practices finish in ~2.5h
IS_CI        = os.environ.get("CI", "").lower() in ("true", "1")
DELAY_SEC    = 1.5  if IS_CI else 2.5
TIMEOUT      = 10   if IS_CI else 15
PW_TIMEOUT   = 20000 if IS_CI else 25000
# Sub-page crawl limits.  Nav links are ALWAYS fetched in full (no cap).
# These limits apply only to keyword-matched and remaining pages beyond nav.
L1_LIMIT     = 40   if IS_CI else 60   # keyword links beyond nav (per practice)
L2_LIMIT     = 30   if IS_CI else 40   # sub-pages of L1 pages
L3_LIMIT     = 20   if IS_CI else 30   # any remaining same-domain links


# ── Row-range control (0-based index into the practices list) ──
# Examples:
#   Rows  1-10 : START_IDX=0,  END_IDX=10
#   Rows 11-20 : START_IDX=10, END_IDX=20
#   All rows   : START_IDX=0,  END_IDX=None
START_IDX = 0    # ← change this
END_IDX   = 100  # ← change this (None = no limit)

# Set False to skip all browser-based scraping (faster but less data)
USE_PLAYWRIGHT = True

# ── Page cache directory ──────────────────────────────────────────────────────
# Every scraped HTML page is saved here so you can reprocess without
# hitting the web again.  Set to "" to disable caching.
CACHE_DIR = "page_cache"

# ─────────────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

_SOCIAL_UA_LIST = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_4_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
]
_CFFI_PROFILES  = ["chrome124", "chrome136", "chrome133a", "safari260"]
_CFFI_IG_PROF   = ["chrome124", "chrome136", "safari260"]

SERVICE_KEYWORDS = {
    # ── Invisalign ────────────────────────────────────────────────────────────
    "invisalign":            "Invisalign",

    # ── Clear Aligners ────────────────────────────────────────────────────────
    "clear aligner":         "Clear Aligners",
    "suresmile":             "Clear Aligners",
    "clearcorrect":          "Clear Aligners",
    "spark aligner":         "Clear Aligners",
    "byte aligner":          "Clear Aligners",
    "candidpro":             "Clear Aligners",
    "candid pro":            "Clear Aligners",
    "invisible brace":       "Clear Aligners",
    "clear brace":           "Clear Aligners",
    "smiledirectclub":       "Clear Aligners",
    "smile direct club":     "Clear Aligners",
    "reveal aligner":        "Clear Aligners",
    "evo aligner":           "Clear Aligners",
    "ulab aligner":          "Clear Aligners",
    "angelalign":            "Clear Aligners",
    "3m clarity":            "Clear Aligners",
    "ormco aligner":         "Clear Aligners",

    # ── Veneers ───────────────────────────────────────────────────────────────
    "veneers":               "Veneers",
    "veneer":                "Veneers",
    "porcelain veneer":      "Veneers",
    "composite veneer":      "Veneers",
    "lumineers":             "Veneers",
    "dental laminate":       "Veneers",
    "dental laminates":      "Veneers",
    "tooth laminates":       "Veneers",

    # ── Implants ──────────────────────────────────────────────────────────────
    "implant":               "Implants",

    # ── Smile Makeovers ───────────────────────────────────────────────────────
    "smile makeover":        "Smile Makeovers",
    "smile transformation":  "Smile Makeovers",
    "smile design":          "Smile Makeovers",
    "smile redesign":        "Smile Makeovers",
    "complete smile":        "Smile Makeovers",
    "full smile":            "Smile Makeovers",
    "total smile":           "Smile Makeovers",
    "smile restoration":     "Smile Makeovers",
    "smile enhancement":     "Smile Makeovers",
    "cosmetic makeover":     "Smile Makeovers",
    "aesthetic dentistry":   "Smile Makeovers",

    # ── Teeth Whitening ───────────────────────────────────────────────────────
    "whitening":             "Teeth Whitening",
    "bleach":                "Teeth Whitening",
    "zoom whitening":        "Teeth Whitening",
    "zoom! whitening":       "Teeth Whitening",
    "philips zoom":          "Teeth Whitening",
    "opalescence":           "Teeth Whitening",
    "bright smile":          "Teeth Whitening",
    "kor whitening":         "Teeth Whitening",
    "glo whitening":         "Teeth Whitening",
    "enlighten whitening":   "Teeth Whitening",
    "pola whitening":        "Teeth Whitening",
    "nusmile":               "Teeth Whitening",
    "teeth brightening":     "Teeth Whitening",

    # ── Sedation Dentistry ────────────────────────────────────────────────────
    "sedation":              "Sedation Dentistry",
    "sleep dentistry":       "Sedation Dentistry",
    "nitrous oxide":         "Sedation Dentistry",
    "laughing gas":          "Sedation Dentistry",
    "iv sedation":           "Sedation Dentistry",
    "oral conscious":        "Sedation Dentistry",
    "conscious sedation":    "Sedation Dentistry",
    "general anesthesia":    "Sedation Dentistry",
    "twilight sedation":     "Sedation Dentistry",
    "minimal sedation":      "Sedation Dentistry",
    "moderate sedation":     "Sedation Dentistry",
    "dental anxiety":        "Sedation Dentistry",
    "relaxation dentistry":  "Sedation Dentistry",
    "comfort dentistry":     "Sedation Dentistry",

    # ── Holistic Dentistry ────────────────────────────────────────────────────
    "holistic":              "Holistic Dentistry",
    "biological dentist":    "Holistic Dentistry",
    "mercury-free":          "Holistic Dentistry",
    "mercury free":          "Holistic Dentistry",
    "mercury safe":          "Holistic Dentistry",
    "fluoride-free":         "Holistic Dentistry",
    "biocompatible":         "Holistic Dentistry",
    "ozone therapy":         "Holistic Dentistry",
    "zirconia implant":      "Holistic Dentistry",
    "metal-free":            "Holistic Dentistry",
    "smart protocol":        "Holistic Dentistry",
    "naturopathic dent":     "Holistic Dentistry",
    "non-toxic dentist":     "Holistic Dentistry",
    "bioregulatory":         "Holistic Dentistry",
    "ceramic implant":       "Holistic Dentistry",
    "zurich protocol":       "Holistic Dentistry",
    "natural dentistry":     "Holistic Dentistry",

    # ── Cancer Screening ──────────────────────────────────────────────────────
    "cancer screening":      "Cancer Screening",
    "oral cancer":           "Cancer Screening",
    "velscope":              "Cancer Screening",
    "identafi":              "Cancer Screening",
    "vizilite":              "Cancer Screening",
    "oral id":               "Cancer Screening",
    "oralid":                "Cancer Screening",
    "fluorescence screening":"Cancer Screening",
    "early detection oral":  "Cancer Screening",
    "tissue staining":       "Cancer Screening",

    # ── Dental Plan (Membership) — only specific plan phrases, no generic terms
    "membership plan":       "Dental Plan",
    "dental membership":     "Dental Plan",
    "annual membership":     "Dental Plan",
    "in-house plan":         "Dental Plan",
    "in house plan":         "Dental Plan",
    "in-house membership":   "Dental Plan",
    "in house membership":   "Dental Plan",
    "dental savings":        "Dental Plan",
    "dental subscription":   "Dental Plan",
    "uninsured patients":    "Dental Plan",
    "no dental insurance":   "Dental Plan",
    "in-office plan":        "Dental Plan",
    "office membership":     "Dental Plan",
    "patient membership":    "Dental Plan",
    "dental club":           "Dental Plan",
    "monthly dental plan":   "Dental Plan",
    "annual dental plan":    "Dental Plan",
    "wellness membership":   "Dental Plan",
    # removed: "savings plan", "discount plan", "wellness plan", "care plan",
    # "no insurance", "without insurance", "preventive plan" — too generic
}

TECH_KEYWORDS = {
    # ── CEREC / Same-day crowns ───────────────────────────────────────────────
    "cerec":                     "CEREC",
    "same day crown":            "CEREC",
    "same-day crown":            "CEREC",
    "same day restoration":      "CEREC",
    "same-day restoration":      "CEREC",
    "single-visit crown":        "CEREC",
    "single visit crown":        "CEREC",
    "one visit crown":           "CEREC",
    "one-visit crown":           "CEREC",
    "in-office crown":           "CEREC",
    "in office crown":           "CEREC",
    "chairside crown":           "CEREC",
    "chairside milling":         "CEREC",
    "in-office milling":         "CEREC",
    "milled crown":              "CEREC",
    "cad/cam crown":             "CEREC",
    "cadcam crown":              "CEREC",
    "cad cam crown":             "CEREC",
    "cad/cam":                   "CEREC",
    "cad cam":                   "CEREC",
    "chairside restoration":     "CEREC",
    "one appointment crown":     "CEREC",
    "one-appointment crown":     "CEREC",
    "crown in one":              "CEREC",
    "crown in a day":            "CEREC",
    "crown same day":            "CEREC",
    "digital crown":             "CEREC",
    "same-day dentistry":        "CEREC",
    "same day dentistry":        "CEREC",
    "e4d":                       "CEREC",
    "omnicam":                   "CEREC",
    # primescan is the CEREC scanner — listed under Intraoral Scanners too
    # but milling/same-day crown context is the distinguisher; keep here
    "cerec primescan":           "CEREC",
    "cerec omnicam":             "CEREC",
    "cerec mc":                  "CEREC",
    "cerec ac":                  "CEREC",
    # General crown service phrases — websites list "Dental Crowns" as a service
    # without always naming the brand; map to CEREC since column = Same Day Crowns
    "crowns":                    "CEREC",
    "dental crowns":             "CEREC",
    "dental crown":              "CEREC",
    "tooth crowns":              "CEREC",
    "tooth crown":               "CEREC",
    "porcelain crown":           "CEREC",
    "porcelain crowns":          "CEREC",
    "zirconia crown":            "CEREC",
    "zirconia crowns":           "CEREC",
    "ceramic crown":             "CEREC",
    "ceramic crowns":            "CEREC",
    "crown restoration":         "CEREC",
    "crown restorations":        "CEREC",
    "crown and bridge":          "CEREC",
    "crowns and bridges":        "CEREC",
    "full crown":                "CEREC",
    "crown placement":           "CEREC",
    "crown procedure":           "CEREC",
    "dental caps":               "CEREC",

    # ── CBCT / 3D Imaging ─────────────────────────────────────────────────────
    "cbct":                      "CBCT",
    "cone beam":                 "CBCT",
    "cone-beam":                 "CBCT",
    "3d scanning":               "CBCT",
    "3-d imaging":               "CBCT",
    "3-d scan":                  "CBCT",
    "dental 3d":                 "CBCT",
    "3d imaging":                "CBCT",
    "3d x-ray":                  "CBCT",
    "3d xray":                   "CBCT",
    "3d x ray":                  "CBCT",
    "3d scan":                   "CBCT",
    "3d radiograph":             "CBCT",
    "3d cone beam":              "CBCT",
    "3d dental imaging":         "CBCT",
    "3d dental scan":            "CBCT",
    "volumetric tomography":     "CBCT",
    "digital tomography":        "CBCT",
    "i-cat":                     "CBCT",
    "dental ct":                 "CBCT",
    # planmeca removed (too generic — makes 2D X-rays and intraoral scanners too)
    # use only CBCT-specific Planmeca model names:
    "planmeca promax 3d":        "CBCT",
    "planmeca cbct":             "CBCT",
    "vatech cbct":               "CBCT",
    "vatech green":              "CBCT",
    "kavo cbct":                 "CBCT",
    "acteon cbct":               "CBCT",
    "galileos":                  "CBCT",
    "prexion":                   "CBCT",
    "cs 9600":                   "CBCT",
    "cs9600":                    "CBCT",
    "cs 9300":                   "CBCT",
    "cs9300":                    "CBCT",
    "orthophos":                 "CBCT",
    "accuitomo":                 "CBCT",
    "newtom":                    "CBCT",
    "j. morita":                 "CBCT",
    "3d tomograph":              "CBCT",

    # ── Lasers ────────────────────────────────────────────────────────────────
    "laser":                     "Lasers",
    "waterlase":                 "Lasers",
    "biolase":                   "Lasers",
    "solea laser":               "Lasers",
    "diode laser":               "Lasers",
    "erbium laser":              "Lasers",
    "fotona":                    "Lasers",
    "lightwalker":               "Lasers",
    "dental laser":              "Lasers",
    "laser dentistry":           "Lasers",
    "laser therapy":             "Lasers",
    "laser treatment":           "Lasers",
    "nv microlaser":             "Lasers",
    "epic x laser":              "Lasers",
    "soft tissue laser":         "Lasers",
    "hard tissue laser":         "Lasers",
    "lanap":                     "Lasers",
    "lightscalpel":              "Lasers",
    "periodontal laser":         "Lasers",
    "gum laser":                 "Lasers",
    "laser gum":                 "Lasers",
    "laser periodon":            "Lasers",
    "laser whitening":           "Lasers",
    "laser surgery":             "Lasers",

    # ── AI ────────────────────────────────────────────────────────────────────
    # " ai " (space-bounded) kept but supplemented with regex check below
    "artificial intelligence":   "AI",
    "overjet":                   "AI",
    "pearl ai":                  "AI",
    "diagnocat":                 "AI",
    "dental ai":                 "AI",
    "ai-powered":                "AI",
    "ai powered":                "AI",
    "ai-based":                  "AI",
    "ai-assisted":               "AI",
    "ai assisted":               "AI",
    "ai detection":              "AI",
    "ai diagnostic":             "AI",
    "ai technology":             "AI",
    "ai analysis":               "AI",
    "videa":                     "AI",
    "dexis ai":                  "AI",
    "dentsply ai":               "AI",
    "medtrics":                  "AI",
    "dental intel":              "AI",
    "denti.ai":                  "AI",
    "second.opinion":            "AI",
    "machine learning":          "AI",
    "dexis clarity":             "AI",
    "apteryx ai":                "AI",

    # ── Intraoral Scanners ────────────────────────────────────────────────────
    "intraoral scanner":         "Intraoral Scanners",
    "intra-oral scanner":        "Intraoral Scanners",
    "digital impression":        "Intraoral Scanners",
    "digital impressions":       "Intraoral Scanners",
    "optical impression":        "Intraoral Scanners",
    "no messy impressions":      "Intraoral Scanners",
    "no more impressions":       "Intraoral Scanners",
    "no impressions":            "Intraoral Scanners",
    "eliminate impressions":     "Intraoral Scanners",
    "skip the impressions":      "Intraoral Scanners",
    "without impressions":       "Intraoral Scanners",
    "wireless impression":       "Intraoral Scanners",
    "digital intraoral":         "Intraoral Scanners",
    "3d intraoral":              "Intraoral Scanners",
    "digital dental scan":       "Intraoral Scanners",
    "trios":                     "Intraoral Scanners",
    "medit":                     "Intraoral Scanners",
    "itero":                     "Intraoral Scanners",
    "itero element":             "Intraoral Scanners",
    "3shape":                    "Intraoral Scanners",
    "trios scanner":             "Intraoral Scanners",
    "medit scanner":             "Intraoral Scanners",
    "medit i-500":               "Intraoral Scanners",
    "medit i500":                "Intraoral Scanners",
    "primescan":                 "Intraoral Scanners",
    "true definition scanner":   "Intraoral Scanners",
    "3m true definition":        "Intraoral Scanners",
    "planmeca emerald":          "Intraoral Scanners",
    "carestream cs 3600":        "Intraoral Scanners",
    "carestream cs3600":         "Intraoral Scanners",
    # removed: "digital scan" (too broad), "planmeca scanner" (now specific models),
    # "dental scanner" (too vague), "carestream dental" (too broad)
}

SOCIAL_PLATFORMS = ["facebook", "instagram", "tiktok", "linkedin"]

# ─────────────────────────────────────────────────────────────────────────────
# DOCTOR NAME NORMALISATION  (for deduplication)
# ─────────────────────────────────────────────────────────────────────────────

_CRED_RE = re.compile(
    r'\b(Dr\.?|D\.?D\.?S\.?|D\.?M\.?D\.?|M\.?D\.?|RDH|PhD|Ph\.D\.?|'
    r'DPM|DO|NMD|FAGD|MAGD|FICOI|FACD|FICD|AACD|ABGD|Fellow)\b',
    re.IGNORECASE,
)

def _normalize_name_for_dedup(name: str) -> str:
    """Strip titles/credentials; return lowercase word-sorted string for comparison."""
    n = _CRED_RE.sub('', name)
    n = re.sub(r'[,.\-]', ' ', n)
    n = re.sub(r'\s+', ' ', n).strip().lower()
    # Normalize doubled trailing characters in each word (e.g., "Robertss" → "Roberts")
    # This catches names garbled by CSS truncation or HTML artifacts on web pages.
    words = n.split()
    fixed = []
    for w in words:
        if len(w) > 3 and w[-1] == w[-2]:
            w = w[:-1]
        fixed.append(w)
    n = " ".join(fixed)
    return n

def _is_duplicate_doctor(name: str, kept_norms: list) -> bool:
    """
    Return True if 'name' is a variant already represented in kept_norms.
    Uses subset-word matching so:
      "Dr. Stacy Wince"  ≡  "Stacy Wince, DDS"  ≡  "Dr. Stacy L. Wince"
    The longer (more complete) name wins.
    """
    norm = _normalize_name_for_dedup(name)
    words = set(norm.split())
    if not words:
        return False
    for kept in kept_norms:
        kw = set(kept.split())
        if not kw:
            continue
        # One is a word-subset of the other → same person
        if words <= kw or kw <= words:
            return True
    return False

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CACHE  — save every scraped HTML so reprocess.py can re-extract later
# ─────────────────────────────────────────────────────────────────────────────

def _cache_dir(practice_idx, practice_name: str) -> str:
    """Return (and create) the cache directory for this practice."""
    if not CACHE_DIR:
        return ""
    slug = re.sub(r'[^\w]', '_', str(practice_name))[:30].strip('_')
    folder = os.path.join(CACHE_DIR, f"{int(practice_idx):03d}_{slug}")
    os.makedirs(folder, exist_ok=True)
    return folder

def _cache_html(folder: str, page_type: str, url: str, html: str):
    """Save one HTML page to the cache folder and update manifest.json."""
    if not folder:
        return
    fname = f"{page_type}.html"
    fpath = os.path.join(folder, fname)
    with open(fpath, "w", encoding="utf-8", errors="replace") as f:
        f.write(html)
    manifest_path = os.path.join(folder, "manifest.json")
    manifest = {}
    if os.path.exists(manifest_path):
        with open(manifest_path, encoding="utf-8") as f:
            try:
                manifest = json.load(f)
            except Exception:
                manifest = {}
    manifest.setdefault("pages", {})[page_type] = {
        "url": url,
        "file": fname,
        "saved_at": time.strftime("%Y-%m-%d %H:%M:%S"),
    }
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2, ensure_ascii=False)

def _cache_result(folder: str, practice_info: dict, result: dict):
    """Save the scraped non-HTML result dict to result.json inside the cache folder."""
    if not folder:
        return
    payload = {
        "practice": practice_info,
        "scraped_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "result": {k: v for k, v in result.items() if k != "doctors"},
        "doctors": result.get("doctors", []),
    }
    with open(os.path.join(folder, "result.json"), "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)


def _load_cache_result(folder: str) -> dict | None:
    """Load a previously saved result.json. Returns scraped dict or None if not found."""
    if not folder:
        return None
    path = os.path.join(folder, "result.json")
    if not os.path.exists(path):
        return None
    try:
        with open(path, encoding="utf-8") as f:
            payload = json.load(f)
        result = dict(payload.get("result", {}))
        result["doctors"] = payload.get("doctors", [])
        return result
    except Exception:
        return None

INVISALIGN_TIERS = [
    "Diamond Plus", "Diamond",
    "Platinum Plus", "Platinum",
    "Gold Plus", "Gold",
    "Silver Plus", "Silver",
    "Bronze Plus", "Bronze",
]


# ─────────────────────────────────────────────────────────────────────────────
# HTTP UTILITIES
# ─────────────────────────────────────────────────────────────────────────────

def _is_cloudflare_block(r) -> bool:
    """Return True if the response looks like a Cloudflare challenge/block page."""
    if r.status_code in (403, 429, 503):
        return True
    # Cloudflare error pages (520, 521, 522, 524 etc.) or JS challenge
    if r.status_code >= 500:
        ct = r.headers.get("Content-Type", "")
        if "text/html" in ct:
            snip = r.text[:2000].lower()
            if any(kw in snip for kw in (
                "cloudflare", "cf-ray", "just a moment", "checking your browser",
                "enable javascript", "web server is returning an unknown error",
            )):
                return True
    return False


def _cffi_get(url) -> "requests.Response | None":
    """Try fetching url via curl_cffi (Chrome TLS impersonation) to bypass Cloudflare."""
    if not _CFFI_AVAILABLE:
        return None
    try:
        r = cffi_requests.get(url, impersonate="chrome120", timeout=TIMEOUT)
        if r.status_code == 200:
            log.info(f"   curl_cffi bypass succeeded: {url}")
            return r
        log.warning(f"   curl_cffi got {r.status_code}: {url}")
    except Exception as e:
        log.warning(f"   curl_cffi failed for {url}: {e}")
    return None


def safe_get(url, retries=2):
    """GET a URL with retries. Falls back to curl_cffi on Cloudflare blocks. Returns Response or None."""
    if not url or str(url).strip() in ("", "N/A", "None"):
        return None
    if not url.startswith("http"):
        url = "https://" + url.lstrip("/")
    for attempt in range(retries):
        for verify in (True, False):   # second pass bypasses SSL cert check
            try:
                r = requests.get(url, headers=HEADERS, timeout=TIMEOUT, verify=verify)
                if r.status_code == 200:
                    if _is_cloudflare_block(r):
                        log.warning(f"  Cloudflare JS challenge on 200 response: {url}")
                        cf_r = _cffi_get(url)
                        return cf_r if cf_r else None
                    if not verify:
                        log.warning(f"  SSL bypassed (verify=False): {url}")
                    return r
                if _is_cloudflare_block(r):
                    log.warning(f"  Cloudflare block ({r.status_code}): {url} — trying curl_cffi…")
                    cf_r = _cffi_get(url)
                    if cf_r:
                        return cf_r
                    return None
                if r.status_code == 403:
                    log.warning(f"  403 Forbidden: {url}")
                    return None
                break  # non-200/403: don't retry verify=False for same status
            except RequestsSSLError:
                if verify:
                    continue   # try again without verification
                log.warning(f"  SSL error even with verify=False: {url}")
                return None
            except Exception as e:
                log.warning(f"  Attempt {attempt+1} failed for {url}: {e}")
                time.sleep(1)
                break
    return None


def extract_text(html):
    """Return clean lowercase plain text from HTML."""
    soup = BeautifulSoup(html, "lxml")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    return soup.get_text(separator=" ", strip=True).lower()


def extract_body_text(html):
    """Return clean lowercase plain text, stripping nav/header/footer to avoid
    repeated keyword inflation from navigation menus and footers."""
    soup = BeautifulSoup(html, "lxml")
    for tag in soup(["script", "style", "noscript", "nav", "header", "footer"]):
        tag.decompose()
    # Also strip common nav/menu class containers
    for tag in soup.find_all(class_=re.compile(
        r'\b(menu|navigation|navbar|nav-bar|topbar|top-bar|site-header|footer|breadcrumb)\b', re.I
    )):
        tag.decompose()
    return soup.get_text(separator=" ", strip=True).lower()


def extract_augmented_text(html: str, page_url: str = "") -> str:
    """
    Extract keyword-bearing text that extract_text() misses:
      - <meta name="description"> / <meta name="keywords"> content
      - JSON-LD structured data strings (schema.org availableService, etc.)
      - URL path tokens from the page URL and all internal link hrefs
      - title="" and aria-label="" attribute values

    Returns lowercase string; safe to concatenate with extract_text() output.
    """
    import json as _json

    soup = BeautifulSoup(html, "lxml")
    parts: list = []

    # Meta description / keywords
    for m in soup.find_all("meta"):
        name = (m.get("name") or m.get("property") or "").lower()
        if any(x in name for x in ("description", "keyword", "og:description")):
            content = m.get("content", "")
            if content:
                parts.append(content)

    # JSON-LD — service/tech names often listed here
    for script in soup.find_all("script", type=re.compile(r"application/ld\+json", re.I)):
        try:
            ld = _json.loads(script.string or "")
            def _walk(obj):
                if isinstance(obj, str):
                    parts.append(obj)
                elif isinstance(obj, list):
                    for item in obj:
                        _walk(item)
                elif isinstance(obj, dict):
                    for v in obj.values():
                        _walk(v)
            _walk(ld)
        except Exception:
            pass

    # Page URL path (e.g. /services/laser-dentistry → "laser dentistry")
    if page_url:
        path = unquote(urlparse(page_url).path)
        parts.append(path.replace("-", " ").replace("/", " ").replace("_", " "))

    # All internal link href paths
    for a in soup.find_all("a", href=True):
        href = str(a["href"])
        if href.startswith("/") or "://" in href:
            try:
                path = unquote(urlparse(href).path)
                parts.append(path.replace("-", " ").replace("/", " ").replace("_", " "))
            except Exception:
                pass

    # title, aria-label, alt, placeholder attributes — all carry keyword-rich text
    for tag in soup.find_all(True):
        for attr in ("title", "aria-label", "alt", "placeholder", "data-title", "data-label"):
            val = tag.get(attr, "")
            if val and isinstance(val, str) and len(val) > 2:
                parts.append(val)

    # og:title / og:keywords meta (description already covered above)
    for m in soup.find_all("meta"):
        prop = (m.get("property") or "").lower()
        if prop in ("og:title", "og:keywords", "twitter:title", "twitter:description"):
            content = m.get("content", "")
            if content:
                parts.append(content)

    return " ".join(parts).lower()


def count_keyword(text, keyword):
    """Count occurrences of keyword in lowercase text."""
    return len(re.findall(re.escape(keyword.lower()), text))


def count_keyword_capped(text, keyword, cap=3):
    """Count keyword occurrences capped at `cap` per call.
    Used for per-page service counting: avoids nav/footer inflation while
    still capturing real body-text mentions (cap=3 means 'mentioned on this page')."""
    return min(len(re.findall(re.escape(keyword.lower()), text)), cap)


# ─────────────────────────────────────────────────────────────────────────────
# FIELD EXTRACTORS — requests-based
# ─────────────────────────────────────────────────────────────────────────────

def _decode_cloudflare_email(encoded):
    """
    Decode a Cloudflare-obfuscated email from a cdn-cgi/l/email-protection URL fragment.
    e.g. /cdn-cgi/l/email-protection#bcd5d2dad3fcd0d9cfded9... → info@example.com
    """
    try:
        hex_str = encoded.lstrip("#")
        if len(hex_str) < 4:
            return None
        key = int(hex_str[:2], 16)
        result = ""
        for i in range(2, len(hex_str) - 1, 2):
            result += chr(int(hex_str[i:i+2], 16) ^ key)
        if "@" in result and "." in result.split("@")[-1]:
            return result
    except Exception:
        pass
    return None


def find_email(text, soup):
    """Find email from mailto links, Cloudflare protection links, then plain text."""
    # Cloudflare email protection links
    for a in soup.find_all("a", href=re.compile(r"cdn-cgi/l/email-protection", re.I)):
        href = a.get("href", "")
        fragment = href.split("#")[-1] if "#" in href else ""
        decoded = _decode_cloudflare_email(fragment)
        if decoded:
            return decoded

    for a in soup.find_all("a", href=re.compile(r"mailto:", re.I)):
        addr = a["href"].replace("mailto:", "").split("?")[0].strip()
        if "@" in addr:
            return addr

    match = re.search(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", text)
    if match:
        return match.group(0)

    for section in soup.find_all(
        ["div", "section", "footer"],
        class_=re.compile(r"(contact|footer|info)", re.I),
    ):
        m = re.search(
            r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}",
            section.get_text(),
        )
        if m:
            return m.group(0)

    return "Not Found"


def find_email_pw(website, page):
    """
    Navigate to the contact/about page with Playwright and look for email.
    Used as fallback when requests-based extraction fails.
    """
    if not page or not website:
        return "Not Found"
    for path in ["/contact", "/contact-us", "/contact-us/", "/about",
                 "/about-us", "/reach-us", "/get-in-touch", "/"]:
        try:
            url = website.rstrip("/") + path
            page.goto(url, timeout=PW_TIMEOUT, wait_until="domcontentloaded")
            page.wait_for_timeout(3000)
            content = page.content()
            # Search in raw HTML first
            m = re.search(
                r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", content
            )
            if m:
                addr = m.group(0)
                # Skip common false positives
                if not any(skip in addr.lower() for skip in
                           ["example.", "yourdomain.", "email@", "name@", "user@"]):
                    return addr
            # Cloudflare email protection decode from Playwright HTML
            soup_pw = BeautifulSoup(content, "lxml")
            for a in soup_pw.find_all("a", href=re.compile(r"cdn-cgi/l/email-protection", re.I)):
                href = a.get("href", "")
                fragment = href.split("#")[-1] if "#" in href else ""
                decoded = _decode_cloudflare_email(fragment)
                if decoded:
                    return decoded
            # Also check innerText (catches obfuscated mailto links)
            try:
                inner = page.evaluate("() => document.body.innerText")
                m2 = re.search(
                    r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", inner
                )
                if m2:
                    addr2 = m2.group(0)
                    if not any(skip in addr2.lower() for skip in
                               ["example.", "yourdomain.", "email@", "name@", "user@"]):
                        return addr2
            except Exception:
                pass
        except Exception:
            pass
    return "Not Found"


_SKIP_WORDS = (
    "patient", "office", "staff", "appointment", "service",
    "insurance", "dental", "care", "treatment", "contact",
    "welcome", "schedule", "hour", "location",
    "difference", "experience", "results", "overview",
    "directions", "meet", "learn", "blog", "menu", "home",
    "today", "read", "more", "view", "skip", "once",
    "providing", "providing", "hills", "beverly",
)

# Words that are never valid as a name component (last name / first name)
_INVALID_NAME_WORDS = frozenset({
    # months
    "january", "february", "march", "april", "may", "june",
    "july", "august", "september", "october", "november", "december",
    # days
    "monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday",
    # pronouns — never valid name components
    "his", "her", "him", "its", "our", "ours", "your", "yours",
    "their", "theirs", "we", "us", "my", "me", "he", "she", "it", "they",
    "who", "whom", "which",
    # question words — appear in headings like "Dr. Smith Why..." or "Dr. Jones How..."
    "how", "why",
    # common non-name words that follow "Dr. LastName"
    "with", "for", "and", "the", "new", "all", "has",
    "will", "can", "are", "was", "were", "been", "have",
    "this", "that", "from", "about", "also", "only",
    "after", "before", "during", "when", "what", "where",
    "open", "closed", "hours", "call", "visit", "book", "now",
})

def _is_valid_doctor_name(name: str) -> bool:
    """Return False if any word in the name is clearly not a real name component."""
    # Strip leading 'Dr.' / 'Dr' prefix for word analysis
    stripped = re.sub(r'^Dr\.?\s+', '', name, flags=re.IGNORECASE).strip()
    # Strip trailing credentials
    stripped = re.sub(
        r'[,\s]+(?:DDS|DMD|MD|MS|FAGD|MAGD|FICOI|FACD|FICD|AACD|ABGD|ABPD|ABOD|ABCD|Ph\.?D\.?).*$',
        '', stripped, flags=re.IGNORECASE
    ).strip()
    words = stripped.split()
    for w in words:
        # Skip single-letter initials like "M."
        clean_w = w.strip('.,')
        if len(clean_w) <= 1:
            continue
        if clean_w.lower() in _INVALID_NAME_WORDS:
            return False
    return True

_NAME_PART      = r'[A-Z][A-Za-z\u00C0-\u024F]+(?:-[A-Z][A-Za-z\u00C0-\u024F]+)?'  # handles unicode + hyphens, NO digits
_NAME_PART_CI   = r'[A-Za-z][A-Za-z\u00C0-\u024F]+(?:-[A-Za-z][A-Za-z\u00C0-\u024F]+)?'  # case-insensitive variant, NO digits
# Credential block: DDS, DMD, MS, FAGD, MAGD, FICOI, FACD, FICD, AACD, ABGD, ABPD, etc.
_CRED_SINGLE    = r'(?:DDS|DMD|MD|M\.D\.|MS|FAGD|MAGD|FICOI|FACD|FICD|AACD|ABGD|ABPD|ABOD|ABCD|Ph\.?D\.?)'
_CRED_SUFFIX    = rf'(?:,?\s*{_CRED_SINGLE}(?:,?\s*{_CRED_SINGLE})*)?'
_DOCTOR_PATTERNS = [
    # "Dr. First [M.] Last, DDS, MS" — comma before credentials allowed
    rf'Dr\.?\s+{_NAME_PART}(?:\s+[A-Z]\.?)?\s+{_NAME_PART}{_CRED_SUFFIX}',
    # "First [M.] Last, DDS" without Dr. prefix
    rf'{_NAME_PART}(?:\s+[A-Z]\.?)?\s+{_NAME_PART},?\s+{_CRED_SINGLE}(?:,?\s*{_CRED_SINGLE})*',
    # "Dr. First Last" short form — still capture credentials if present
    rf'Dr\.?\s+{_NAME_PART}(?:\s+{_NAME_PART})?{_CRED_SUFFIX}',
    # Case-insensitive variant — catches "dani keepes, dds" style (mixed-case team pages)
    rf'{_NAME_PART_CI}(?:\s+[A-Za-z]\.?)?\s+{_NAME_PART_CI},?\s+(?:[Dd][Dd][Ss]|[Dd][Mm][Dd])',
]
# Used to extend a matched name with trailing credentials (e.g. ", MS" after ", DDS")
_CRED_TAIL_RE   = re.compile(
    rf'^(?:,?\s*{_CRED_SINGLE})+',
    re.IGNORECASE,
)

_TEAM_LINK_TEXT = (
    "our dentist", "our doctor", "meet the team", "meet our team",
    "meet the doctor", "our team", "our provider", "our staff",
    "the team", "about the doctor", "our doctors",
)
_TEAM_HREF_KW = ("doctor", "dentist", "team", "provider", "staff", "meet-the",
                 "about-us", "about_us", "who-we-are", "our-doctor", "our-dentist")


def _ascii_normalize(text: str) -> str:
    """NFKD-normalize unicode → ASCII (ē→e, ñ→n, etc.) for regex matching."""
    return unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")


def _extract_names_from_soup(soup):
    """Return a set of doctor name strings from a BeautifulSoup object."""
    names = set()
    raw = " ".join(soup.stripped_strings)
    raw_ascii = _ascii_normalize(raw)
    for pattern in _DOCTOR_PATTERNS:
        for m in re.finditer(pattern, raw_ascii):
            clean = re.sub(r"\s+", " ", m.group(0).strip())
            # Extend with trailing credentials after the match
            _after = raw_ascii[m.end():]
            _ext = _CRED_TAIL_RE.match(_after)
            if _ext:
                clean = re.sub(r"\s+", " ", (clean + _ext.group(0)).strip())
            if (len(clean.split()) >= 2
                    and not re.search(r'\d', clean)
                    and not any(w in clean.lower() for w in _SKIP_WORDS)
                    and _is_valid_doctor_name(clean)):
                names.add(clean)
    return names


def _extract_names_from_soup_strict(soup):
    """
    Stricter doctor-name extraction used as a last resort when heading-based
    parsing found zero sections.  Scans ONLY heading and strong/label tags
    (not full page text) and requires at least First + Last name words after
    the 'Dr.' prefix so testimonial snippets ('Thank you Dr. Smith!') and
    shortened running-text mentions are excluded.
    """
    _DR_PREFIX = re.compile(r'^Dr\.?\s+', re.I)
    names = set()
    # Only scan tags that are likely to contain a doctor's full name
    for tag in soup.find_all(["h1", "h2", "h3", "h4", "h5", "strong", "b", "dt", "th"]):
        text = tag.get_text(separator=" ", strip=True)
        if len(text) > 120:      # skip long blocks; headings are short
            continue
        text_ascii = _ascii_normalize(text)
        for pattern in _DOCTOR_PATTERNS:
            m = re.search(pattern, text_ascii)
            if not m:
                continue
            clean = re.sub(r"\s+", " ", m.group(0).strip())
            _after = text_ascii[m.end():]
            _ext = _CRED_TAIL_RE.match(_after)
            if _ext:
                clean = re.sub(r"\s+", " ", (clean + _ext.group(0)).strip())
            if re.search(r'\d', clean):
                continue
            if any(w in clean.lower() for w in _SKIP_WORDS):
                continue
            if not _is_valid_doctor_name(clean):
                continue
            # Require at least 2 real name words after stripping 'Dr.' prefix
            _name_body = _DR_PREFIX.sub("", clean).strip()
            _name_body = re.sub(
                r'[,\s]+(?:DDS|DMD|MD|MS|FAGD|MAGD|FICOI|FACD|FICD|AACD|ABGD|ABPD|ABOD|ABCD|Ph\.?D\.?).*$',
                '', _name_body, flags=re.I
            ).strip()
            _real_words = [w for w in _name_body.split() if len(w.strip('.,')) > 1]
            if len(_real_words) < 2:
                continue   # "Dr. Smith" alone is too ambiguous — skip
            names.add(clean)
            break  # one match per tag
    return names


def _extract_doctor_scoped_text(soup, name_core):
    """
    Find the heading on `soup` whose text contains `name_core`, then return
    the narrowest ancestor container that belongs only to this doctor (does not
    contain other doctor headings).  Falls back to sibling traversal.
    Returns lowercased scoped text, or "" if nothing found.

    Used instead of full-page text so that site-wide keywords ("orthodontic" in
    navigation / banners) don't contaminate per-doctor specialty extraction.
    """
    _DR_RE2 = re.compile(r'\b(Dr\.?|DDS|DMD|Doctor)\b', re.I)
    name_words = name_core.split()
    _last = name_words[-1] if name_words else ""
    _first = name_words[0] if name_words else ""

    def _matches(h_lower):
        return (
            name_core in h_lower
            or (len(_last) >= 5 and _last in h_lower)
            or (len(_first) >= 4 and len(_last) >= 5
                and _first in h_lower and _last in h_lower)
        )

    for _h in soup.find_all(["h1", "h2", "h3", "h4", "h5"]):
        if not _matches(_h.get_text().lower()):
            continue
        # Walk ancestors for narrowest non-shared container
        for anc in _h.parents:
            if anc.name not in ["div", "section", "article", "li", "td", "tr"]:
                continue
            _others = [h for h in anc.find_all(["h2", "h3", "h4", "h5"])
                       if h is not _h and _DR_RE2.search(h.get_text())]
            if _others:
                break  # too broad — fall through to sibling traversal
            anc_text = anc.get_text(separator=" ", strip=True)
            if len(anc_text) > 30:
                return anc_text.lower()
        # Sibling traversal fallback
        parts = [_h.get_text(separator=" ", strip=True)]
        for sib in _h.next_siblings:
            if not hasattr(sib, "name"):
                continue
            if sib.name in ["h2", "h3", "h4", "h5"] and _DR_RE2.search(sib.get_text()):
                break
            parts.append(sib.get_text(separator=" ", strip=True))
        result = " ".join(parts).lower()
        if len(result) > 10:
            return result
    return ""


def _parse_team_page_for_doctors(soup):
    """
    Parse a team/doctor page to find individual doctor sections.
    For each doctor heading found, collects the surrounding container text as bio
    and the bio_url (link to that doctor's detail page, if any).
    Returns list of {"name": str, "text": str, "bio_url": str}.
    """
    doctors = []
    seen_names = set()

    for heading in soup.find_all(["h1", "h2", "h3", "h4", "h5"]):
        heading_text = heading.get_text(separator=" ", strip=True)
        heading_ascii = _ascii_normalize(heading_text)
        for pattern in _DOCTOR_PATTERNS:
            m = re.search(pattern, heading_ascii)
            if not m:
                continue
            name = re.sub(r"\s+", " ", m.group(0).strip())
            # Extend with any additional credentials that follow the match (e.g. ", MS")
            _after = heading_ascii[m.end():]
            _cred_ext = _CRED_TAIL_RE.match(_after)
            if _cred_ext:
                name = name + _cred_ext.group(0).rstrip()
                name = re.sub(r"\s+", " ", name.strip())
            if re.search(r'\d', name):
                continue
            if not _is_valid_doctor_name(name):
                continue
            if any(w.lower() in name.lower() for w in _SKIP_WORDS):
                continue
            if name.lower() in seen_names:
                continue
            seen_names.add(name.lower())

            # Prefer a semantically labelled parent container
            parent = heading.find_parent(
                ["div", "section", "article", "li"],
                class_=re.compile(
                    r"(team|doctor|provider|dentist|member|staff|bio|card|person|profile)",
                    re.I,
                ),
            )
            if not parent:
                parent = heading.find_parent(["div", "article", "li"])

            # Extract bio link from the container (doctor detail page).
            # Prefer links whose href/text suggests a bio/profile page;
            # fall back to the first valid same-domain link.
            _BIO_HREF_RE = re.compile(
                r'team|doctor|provider|staff|bio|profile|meet|about|physician',
                re.I,
            )
            _BIO_TEXT_RE = re.compile(
                r'bio|learn\s+more|meet|read\s+more|profile|about\s+dr|about\s+the',
                re.I,
            )
            bio_url = ""
            _bio_fallback = ""
            search_node = parent or heading
            # Also search heading's immediate parent siblings (split-column layouts)
            _bio_nodes = [search_node]
            if heading.parent and heading.parent is not search_node:
                _bio_nodes.append(heading.parent)
            for _bnode in _bio_nodes:
                for a in _bnode.find_all("a", href=True):
                    href = a.get("href", "")
                    if not href or href.startswith(("#", "mailto:", "tel:", "javascript:")):
                        continue
                    link_text = a.get_text(strip=True)
                    if _BIO_HREF_RE.search(href) or _BIO_TEXT_RE.search(link_text):
                        bio_url = href
                        break
                    if not _bio_fallback:
                        _bio_fallback = href
                if bio_url:
                    break
            if not bio_url:
                bio_url = _bio_fallback

            # ── Bio text: walk ancestors to find the narrowest container that
            # (a) contains only THIS doctor's heading (not other doctors), AND
            # (b) has meaningful content beyond just the heading text.
            # This handles card layouts, flat layouts, and split-column layouts.
            _DR_RE = re.compile(r'\b(Dr\.?|DDS|DMD|Doctor)\b', re.I)
            _MIN_EXTRA = 40   # chars beyond the heading to count as "has bio"

            def _sibling_traversal(pivot, stop_re):
                """Text of siblings after pivot, stopping at another doctor heading."""
                parts = [pivot.get_text(separator=" ", strip=True)]
                for sib in pivot.next_siblings:
                    if not hasattr(sib, "name"):
                        continue
                    if sib.name in ["h2", "h3", "h4", "h5"] and stop_re.search(sib.get_text()):
                        break
                    parts.append(sib.get_text(separator=" ", strip=True))
                return " ".join(parts)

            bio_text = ""
            _broad_anc = None   # first ancestor that contains other doctors

            for anc in heading.parents:
                if anc.name not in ["div", "section", "article", "li", "td", "tr"]:
                    continue
                _others = [h for h in anc.find_all(["h2", "h3", "h4", "h5"])
                           if h is not heading and _DR_RE.search(h.get_text())]
                if _others:
                    _broad_anc = anc
                    break  # Too broad — we'll do sibling traversal
                # Not too broad. Does it have real bio content?
                anc_text = anc.get_text(separator=" ", strip=True)
                extra = len(anc_text) - len(heading_text)
                if extra >= _MIN_EXTRA:
                    bio_text = anc_text.lower()
                    break

            if not bio_text:
                # Either no container found or every container was too broad.
                # Strategy: sibling traversal from the heading (catches flat layouts),
                # PLUS sibling traversal from the heading's PARENT element (catches
                # split-column layouts where heading and bio are in adjacent sibling divs).
                candidate = _sibling_traversal(heading, _DR_RE)

                if _broad_anc is not None:
                    # Also look at heading's immediate parent's siblings within
                    # the broad ancestor — catches bio text in adjacent columns.
                    h_parent = heading.parent
                    if h_parent and h_parent is not _broad_anc:
                        extra_parts = [candidate]
                        for psib in h_parent.next_siblings:
                            if not hasattr(psib, "find_all"):
                                continue
                            # Stop if this sibling contains another doctor heading
                            if any(_DR_RE.search(h.get_text())
                                   for h in psib.find_all(["h2","h3","h4","h5"])):
                                break
                            extra_parts.append(psib.get_text(separator=" ", strip=True))
                        candidate = " ".join(extra_parts)

                bio_text = candidate.lower()

            doctors.append({"name": name, "text": bio_text, "bio_url": bio_url})
            break

    # Deduplicate using normalised name matching
    # Sort by word count descending so the most complete name is kept first
    doctors.sort(key=lambda d: len(d["name"].split()), reverse=True)
    kept_norms: list = []
    result = []
    for d in doctors:
        if not _is_duplicate_doctor(d["name"], kept_norms):
            kept_norms.append(_normalize_name_for_dedup(d["name"]))
            result.append(d)
    return result[:30]


def _count_hygienists_from_team(soup):
    """Count distinct team members with hygienist titles on a team page."""
    _HYG_TITLE_RE = re.compile(
        r'\bR\.?D\.?H\.?\b|RDHAP|BSDH|'
        r'registered\s+dental\s+hygienist|'
        r'licensed\s+dental\s+hygienist|'
        r'dental\s+hygienist',
        re.IGNORECASE,
    )
    _NAME_RE = re.compile(r'([A-Z][a-z]+(?:\s+[A-Z]\.?)?\s+[A-Z][a-z]+)')
    seen_names: set = set()
    seen_keys: set = set()
    count = 0
    for tag in soup.find_all(["h1", "h2", "h3", "h4", "h5", "p", "span", "div", "li", "a", "strong", "b"]):
        text = tag.get_text(strip=True)
        if len(text) > 200:
            continue
        if not _HYG_TITLE_RE.search(text):
            continue
        # Prefer name-based dedup so the same person in two tags counts once
        nm = _NAME_RE.search(text)
        if nm:
            key = nm.group(1).strip().lower()
            if key not in seen_names:
                seen_names.add(key)
                count += 1
        else:
            # No extractable name — use a short text window as key
            key = re.sub(r"\s+", " ", text.strip().lower())[:60]
            if key not in seen_keys:
                seen_keys.add(key)
                count += 1
    return count


def _find_team_urls(homepage_soup, base_url):
    """Scan homepage links for pages that likely list doctors / team members."""
    team_urls = []
    seen = set()
    for a in homepage_soup.find_all("a", href=True):
        href = a["href"]
        text = a.get_text(strip=True).lower()
        href_l = href.lower()

        is_team = (
            any(kw in text for kw in _TEAM_LINK_TEXT)
            or any(kw in href_l for kw in _TEAM_HREF_KW)
        )
        if not is_team:
            continue

        full = urljoin(
            base_url if base_url.startswith("http") else "https://" + base_url,
            href,
        )
        if full.startswith("http") and full not in seen:
            seen.add(full)
            team_urls.append(full)

    return team_urls[:6]


def scrape_doctors_full(homepage_soup, base_url, all_text, pw_page=None,
                        all_soups_for_team=None):
    """
    Find the practice's team/doctor page and return one dict per doctor.
    Each dict: {"name": str, "specialty": str, "associations": str}
    Also returns hygienist_count (int or None).

    all_soups_for_team: list of (page_type, soup) from previously scraped sub-pages,
    used to pick the best team soup without re-fetching.

    Strategy:
      1. From all_soups_for_team, pick the page with the most doctor sections.
      2. Find team-page links; fetch any not already scraped.
      3. Parse individual doctor sections; follow each doctor's bio link.
      4. Fallback to names-only extraction if no sections found.
    """
    # Build domain-derived skip words: words from the netloc that are not dental/generic.
    # This prevents "Dr. Woo Hartsdale" when the domain is hartsdaledentalcares.com.
    _DOMAIN_GENERIC = {
        "www", "com", "net", "org", "io", "dental", "dentistry", "dentist", "smile",
        "smiles", "care", "cares", "health", "wellness", "family", "general",
        "associates", "group", "center", "centre", "office", "clinic", "studio",
        "practice", "the", "and", "of",
    }
    _netloc_lower = urlparse(base_url).netloc.lower() if base_url else ""
    _domain_parts = {
        w for w in re.split(r'[\W_]+', _netloc_lower) if len(w) >= 5
    } - _DOMAIN_GENERIC

    def _is_location_false_name(name: str) -> bool:
        """Return True if the doctor name's last word looks like a location from the domain.
        Uses substring check so 'Hartsdale' matches within 'hartsdaledentalcares.com'."""
        words = name.split()
        if not words:
            return False
        last = words[-1].lower()
        # Must be a meaningful length and appear as a substring in the netloc
        if len(last) < 5 or len(words) > 3:
            return False
        # Check substring presence in the domain netloc
        if last in _netloc_lower:
            return True
        # Also check against split domain parts (handles multi-word domains)
        return last in _domain_parts

    # ── Pick best team soup from already-scraped pages ────────────────────────
    team_soup = None
    best_section_count = 0
    all_soups_for_team = all_soups_for_team or []
    for _, sp in all_soups_for_team:
        secs = _parse_team_page_for_doctors(sp)
        if len(secs) > best_section_count:
            best_section_count = len(secs)
            team_soup = sp

    # ── Also check team-specific URLs not yet in our soup list ────────────────
    # Skip network fetching when base_url is empty (offline/reprocess mode)
    team_urls = _find_team_urls(homepage_soup, base_url) if base_url else []
    already_fetched_urls = set()

    # requests pass for any team URL not already covered
    for url in team_urls:
        if url in already_fetched_urls:
            continue
        log.info(f"   Doctor page: {url}")
        time.sleep(DELAY_SEC)
        r = safe_get(url)
        if r:
            ts = BeautifulSoup(r.text, "lxml")
            secs = _parse_team_page_for_doctors(ts)
            if len(secs) > best_section_count:
                best_section_count = len(secs)
                team_soup = ts
            already_fetched_urls.add(url)
            if best_section_count >= 2:
                break

    # ── Playwright pass (JS-rendered team pages) ──────────────────────────────
    if pw_page and team_urls and best_section_count < 2:
        for url in team_urls[:2]:
            try:
                log.info(f"   Doctor page (Playwright): {url}")
                pw_page.goto(url, timeout=PW_TIMEOUT, wait_until="domcontentloaded")
                pw_page.wait_for_timeout(2500)
                pw_html = pw_page.content()
                ts = BeautifulSoup(pw_html, "lxml")
                secs = _parse_team_page_for_doctors(ts)
                if len(secs) > best_section_count:
                    best_section_count = len(secs)
                    team_soup = ts
                if best_section_count >= 2:
                    break
            except Exception as e:
                log.debug(f"   Playwright team-page error: {e}")

    # ── Playwright fallback — re-render base_url when still few doctors ───────
    if pw_page and best_section_count < 2:
        urls_to_pw = ([base_url] if base_url.startswith("http") else
                      [f"https://{base_url}"]) + team_urls
        for pw_url in urls_to_pw[:2]:
            try:
                log.info(f"   Doctor page (Playwright base): {pw_url}")
                pw_page.goto(pw_url, timeout=PW_TIMEOUT, wait_until="domcontentloaded")
                pw_page.wait_for_timeout(3000)
                pw_html = pw_page.content()
                pw_soup = BeautifulSoup(pw_html, "lxml")
                secs = _parse_team_page_for_doctors(pw_soup)
                if len(secs) > best_section_count:
                    best_section_count = len(secs)
                    team_soup = pw_soup
                    log.info(f"   Playwright found {best_section_count} doctors on {pw_url}")
                if best_section_count >= 2:
                    break
            except Exception as e:
                log.debug(f"   Playwright base-url error: {e}")

    # ── Hygienist count — search team page + ALL scraped soups ───────────────
    hygienist_count = None
    _soups_to_check = []
    if team_soup:
        _soups_to_check.append(team_soup)
    for _, _sp in (all_soups_for_team or []):
        if _sp is not team_soup:
            _soups_to_check.append(_sp)
    if homepage_soup and homepage_soup not in _soups_to_check:
        _soups_to_check.append(homepage_soup)
    _HYG_TITLE_RE2 = re.compile(
        r'\bR\.?D\.?H\.?\b|RDHAP|BSDH|'
        r'registered\s+dental\s+hygienist|'
        r'licensed\s+dental\s+hygienist|'
        r'dental\s+hygienist',
        re.IGNORECASE,
    )
    _NAME_RE2 = re.compile(r'([A-Z][a-z]+(?:\s+[A-Z]\.?)?\s+[A-Z][a-z]+)')
    _hyg_names: set = set()
    _hyg_keys: set = set()
    _hyg_total = 0
    # Also match single first-name staff like "Dawn, RDH"
    _SINGLE_NAME_RE2 = re.compile(
        r'\b([A-Z][a-z]{2,})\s*,?\s+'
        r'(?:R\.?D\.?H\.?|BSDH|RDHAP|'
        r'registered\s+dental\s+hygienist|'
        r'licensed\s+dental\s+hygienist|'
        r'dental\s+hygienist)',
        re.IGNORECASE,
    )
    for _sp in _soups_to_check:
        for _tag in _sp.find_all(["h1", "h2", "h3", "h4", "h5", "p", "span", "div", "li", "a", "strong", "b"]):
            _t = _tag.get_text(strip=True)
            if not _HYG_TITLE_RE2.search(_t):
                continue
            if len(_t) > 200:
                # Tag too long to use as a unit — regex-scan all credential
                # snippets within it to avoid missing hygienists in big blocks
                for _cm in _HYG_TITLE_RE2.finditer(_t):
                    _snippet = _t[max(0, _cm.start()-80): _cm.start()+80]
                    _nm2 = _NAME_RE2.search(_snippet)
                    if _nm2:
                        _nk2 = _nm2.group(1).strip().lower()
                        if _nk2 not in _hyg_names:
                            _hyg_names.add(_nk2)
                            _hyg_total += 1
                    else:
                        _sn2 = _SINGLE_NAME_RE2.search(_snippet)
                        if _sn2:
                            _nk2 = _sn2.group(1).strip().lower()
                            if _nk2 not in _hyg_names:
                                _hyg_names.add(_nk2)
                                _hyg_total += 1
                        else:
                            _key2 = re.sub(r"\s+", " ", _snippet.strip().lower())[:80]
                            if _key2 not in _hyg_keys:
                                _hyg_keys.add(_key2)
                                _hyg_total += 1
                continue
            _nm = _NAME_RE2.search(_t)
            if _nm:
                _nk = _nm.group(1).strip().lower()
                if _nk not in _hyg_names:
                    _hyg_names.add(_nk)
                    _hyg_total += 1
            else:
                # Try single-first-name match before falling back to key dedup
                _sn = _SINGLE_NAME_RE2.search(_t)
                if _sn:
                    _nk = _sn.group(1).strip().lower()
                    if _nk not in _hyg_names:
                        _hyg_names.add(_nk)
                        _hyg_total += 1
                else:
                    _key = re.sub(r"\s+", " ", _t.strip().lower())[:60]
                    if _key not in _hyg_keys:
                        _hyg_keys.add(_key)
                        _hyg_total += 1
    if _hyg_total > 0:
        hygienist_count = _hyg_total

    # ── Try per-doctor section parsing first ──────────────────────────────────
    use_soup = team_soup or homepage_soup
    sections = [
        s for s in _parse_team_page_for_doctors(use_soup)
        if not _is_location_false_name(s["name"])
    ]

    # Also collect sections from ALL scraped soups and merge (catches doctors
    # that appear on service/about pages but not on the main team page)
    all_sections_combined = list(sections)
    combined_norms: list = [_normalize_name_for_dedup(s["name"]) for s in sections]
    for _, sp in all_soups_for_team:
        if sp is use_soup:
            continue
        for sec in _parse_team_page_for_doctors(sp):
            if _is_location_false_name(sec["name"]):
                continue
            if not _is_duplicate_doctor(sec["name"], combined_norms):
                combined_norms.append(_normalize_name_for_dedup(sec["name"]))
                all_sections_combined.append(sec)

    # ── Supplementary names: text-scan ONLY the best team page, and ONLY when
    # heading-based parsing found zero sections.  Scanning full page text on
    # service/blog/testimonial pages (the old "second chance" / "last resort"
    # loops) was the primary source of false-positive doctor names.
    if not all_sections_combined and use_soup:
        all_names = _extract_names_from_soup_strict(use_soup)
        for n in sorted(all_names, key=lambda x: len(x.split()), reverse=True):
            if _is_location_false_name(n):
                continue
            if not _is_duplicate_doctor(n, combined_norms):
                combined_norms.append(_normalize_name_for_dedup(n))
                all_sections_combined.append({"name": n, "text": "", "bio_url": ""})

    if all_sections_combined:
        # ── Follow each doctor's bio link for richer data ─────────────────────
        # Strategy (in priority order):
        #   1. bio_text from _parse_team_page_for_doctors (properly scoped)
        #   2. Doctor's individual bio URL page (safe_get, no pw_page needed)
        #   3. Search already-scraped soups for a page whose h1/h2 names this doctor
        #   4. If multi-doctor and still no data: output "" → "Not Found"
        #      (never copy shared context; that causes all doctors to duplicate)
        #   5. If single doctor: fall back to all_text (whole site = 1 doctor)
        _multi_doctor = len(all_sections_combined) > 1
        doctors = []
        for sec in all_sections_combined:
            bio_text = sec["text"]
            bio_url  = sec.get("bio_url", "")

            # Precompute name_core once (shared by Steps 2 & 3)
            _name_core = re.sub(r'^Dr\.?\s+', '', sec["name"], flags=re.I).strip()
            _name_core = re.sub(
                r'[,\s]+(?:DDS|DMD|MD|MS|FAGD|MAGD|FICOI|FACD|FICD|AACD|Ph\.?D\.?)\b.*$',
                '', _name_core, flags=re.I,
            ).strip().lower()

            # Step 1b — attempt specialty extraction from card text immediately.
            # Card titles like "Orthodontist" or "Prosthodontist" are a reliable
            # signal even when the card text is only 30-50 chars.  If found here,
            # we skip the noisier full-page augmentation steps below.
            _specialty = find_specialty(bio_text) if bio_text else ""
            _assoc     = find_associations(bio_text) if bio_text else ""

            # Step 2 — follow bio URL for richer individual-page data.
            # Only fetch when card didn't already yield a specialty OR bio is short.
            # Parse the fetched page and scope to this doctor's section to avoid
            # contamination from site-wide navigation/banner keywords.
            if bio_url and base_url and (not _specialty or len(bio_text) < 300):
                full_bio = urljoin(base_url, bio_url)
                if urlparse(full_bio).netloc == urlparse(base_url).netloc:
                    try:
                        log.info(f"   Doctor bio: {full_bio}")
                        time.sleep(DELAY_SEC)
                        bio_r = safe_get(full_bio)
                        if bio_r and len(bio_r.text) > 500:
                            _bio_soup = BeautifulSoup(bio_r.text, "lxml")
                            # Try scoped extraction first; fall back to main content
                            _scoped = _extract_doctor_scoped_text(_bio_soup, _name_core)
                            if len(_scoped) > 80:
                                bio_text = (bio_text + " " + _scoped).strip().lower()
                            else:
                                # Remove nav/header/footer to avoid site-wide noise
                                for _noise in _bio_soup.find_all(["nav", "header", "footer"]):
                                    _noise.decompose()
                                _main = (_bio_soup.find(["main", "article"])
                                         or _bio_soup.find("body") or _bio_soup)
                                bio_text = (bio_text + " " + _main.get_text(
                                    separator=" ", strip=True
                                )[:4000]).strip().lower()
                            _specialty = find_specialty(bio_text)
                            _assoc     = find_associations(bio_text)
                    except Exception:
                        pass

            # Step 3 — search already-scraped soups for this doctor's bio page.
            # Use SCOPED text (heading container only) instead of full page text
            # so that practice-wide keywords in navigation don't bleed across doctors.
            if _multi_doctor and not _specialty and len(bio_text) < 100 and all_soups_for_team:
                for _, _sp in all_soups_for_team:
                    _scoped = _extract_doctor_scoped_text(_sp, _name_core)
                    if len(_scoped) > 30:
                        bio_text = (bio_text + " " + _scoped).strip().lower()
                        _specialty = find_specialty(bio_text)
                        _assoc     = find_associations(bio_text)
                        break

            # Step 4 — output per-doctor result.
            # Blank if multi-doctor site and we found nothing specific to this doctor.
            if _multi_doctor and not _specialty and not _assoc and len(bio_text) < 50:
                doctors.append({
                    "name":         sec["name"],
                    "specialty":    "",
                    "associations": "",
                })
            else:
                doctors.append({
                    "name":         sec["name"],
                    "specialty":    _specialty,
                    "associations": _assoc,
                })
        return doctors, hygienist_count

    # ── Fallback: names only (should rarely reach here) ────────────────────────
    deduped_names: list = []
    deduped_norms: list = []
    for n in sorted(all_names, key=lambda x: len(x.split()), reverse=True):
        if not _is_duplicate_doctor(n, deduped_norms):
            deduped_norms.append(_normalize_name_for_dedup(n))
            deduped_names.append(n)

    if deduped_names:
        return [
            {"name": n, "specialty": "", "associations": ""}
            for n in deduped_names[:12]
        ], hygienist_count

    return [], hygienist_count


def find_locations_count(text, soup):
    """
    Detect number of practice locations.
    Looks for explicit 'X locations' or 'X offices' language.
    Does NOT count raw address patterns (avoids false positives).
    """
    explicit = re.search(
        r"\b(\d+)\s+(?:convenient\s+)?(?:locations?|offices?|clinics?)\b",
        text,
        re.IGNORECASE,
    )
    if explicit:
        n = int(explicit.group(1))
        return str(n) if n > 1 else "1"

    multi_phrases = [
        "multiple locations", "multiple offices", "two locations",
        "three locations", "all locations", "all offices", "our locations",
        "find a location", "find our offices",
    ]
    for phrase in multi_phrases:
        if phrase in text.lower():
            return "Multiple"

    # Check for a nav item labelled "Locations"
    if soup.find("a", string=re.compile(r"^locations?$", re.I)):
        return "Multiple"

    return "1"


def find_associations(text):
    """Find dental associations / memberships in page text."""
    assoc_map = {
        "AACD":  "American Academy of Cosmetic Dentistry",
        "AACA":  "American Academy of Clear Aligners",
        "AGD":   "Academy of General Dentistry",
        "ADA":   "American Dental Association",
        "AAID":  "American Academy of Implant Dentistry",
        "ABGD":  "American Board of General Dentistry",
        "AAPD":  "American Academy of Pediatric Dentistry",
        "AAED":  "American Academy of Esthetic Dentistry",
        "AAOMS": "American Association of Oral and Maxillofacial Surgeons",
        "ABOMS": "American Board of Oral and Maxillofacial Surgery",
        "ICOI":  "International Congress of Oral Implantologists",
        "MICOI": "International Congress of Oral Implantologists",
        "FICOI": "Fellow of the International Congress of Oral Implantologists",
        "FAGD":  "Fellow of the Academy of General Dentistry",
        "MAGD":  "Master of the Academy of General Dentistry",
        "FACD":  "Fellow of the American College of Dentists",
        "FICD":  "Fellow of the International College of Dentists",
        "ABCD":  "American Board of Cosmetic Dentistry",
        "ABOD":  "American Board of Oral and Maxillofacial Surgery",
        "ABPD":  "American Board of Pediatric Dentistry",
        "AO":    "Academy of Osseointegration",
        "ITI":   "International Team for Implantology",
    }
    text_upper = text.upper()
    found = []
    for abbr, full in assoc_map.items():
        # Match abbreviation as a whole word (avoid ADA matching "AAID", etc.)
        if re.search(rf'\b{re.escape(abbr)}\b', text_upper) or full.upper() in text_upper:
            found.append(abbr)
    return ", ".join(found) if found else ""


def find_specialty(text):
    """
    Detect all specialties mentioned in page text and return them joined with ' / '.
    Uses short label names to produce output like:
      "Cosmetic / Restorative / Laser"
      "Cosmetic / Family (Invisalign Specialist)"
      "Pediatric / General"
    """
    # Ordered from most specific to most general so the list reads naturally
    specialty_map = [
        ("Cosmetic",        ["cosmetic dent", "esthetic dent", "smile makeover", "cosmetic smile"]),
        ("Restorative",     ["restorative dent", "dental restoration", "full mouth restoration",
                             "full mouth reconstruction", "dental rebuild"]),
        ("Implants",        ["dental implant", "implant specialist", "implant dentist",
                             "tooth implant", "all-on-4", "all on 4", "all-on-x"]),
        ("Orthodontics",    ["orthodontist", "orthodontic"]),
        ("Invisalign Specialist", ["invisalign specialist", "invisalign provider",
                                   "invisalign diamond", "invisalign platinum", "invisalign gold"]),
        ("Pediatric",       ["pediatric dent", "children's dent", "kids dent", "child dent"]),
        ("Periodontics",    ["periodontist", "periodontal", "gum disease specialist"]),
        ("Endodontics",     ["endodontist", "root canal specialist"]),
        ("Oral Surgery",    ["oral surgeon", "oral surgery", "wisdom teeth removal",
                             "jaw surgery", "maxillofacial"]),
        ("Prosthodontics",  ["prosthodontist", "prosthodontic"]),
        ("TMJ / Sleep",     ["tmj", "sleep apnea", "sleep dentistry", "snoring treatment"]),
        ("Laser",           ["laser dent", "laser treatment", "laser therapy",
                             "soft tissue laser", "laser whitening", "laser technology",
                             "laser procedure", "diode laser", "erbium laser",
                             "biolase", "waterlase", "dental laser"]),
        ("Sedation",        ["sedation dent", "sedation specialist", "sleep dent",
                             "iv sedation", "nitrous oxide"]),
        ("Sports Dentistry",["sports dent", "athletic mouthguard", "sports mouthguard"]),
        ("Holistic / Biological", ["holistic dent", "biological dent", "mercury-free",
                                   "mercury free", "biocompatible"]),
        ("Women's Health",  ["women's health", "women's dental", "prenatal dental",
                             "pregnancy dental"]),
        ("Community Health",["community health center", "federally qualified",
                             "fqhc", "community clinic"]),
        ("Family",          ["family dent", "family practice", "comprehensive dental",
                             "general and family"]),
        ("General",         ["general dent", "general dentist"]),
    ]

    text_lower = text.lower()
    found = []
    seen_labels = set()

    for label, keywords in specialty_map:
        if label in seen_labels:
            continue
        if any(kw in text_lower for kw in keywords):
            found.append(label)
            seen_labels.add(label)

    if not found:
        return ""

    return " / ".join(found)


def find_social_links(soup):
    """Extract social media URLs from anchor tags."""
    found = {p: "" for p in SOCIAL_PLATFORMS}
    for a in soup.find_all("a", href=True):
        href = a["href"].lower()
        for platform in SOCIAL_PLATFORMS:
            if platform + ".com" in href and not found[platform]:
                found[platform] = a["href"]
    return found


_SOCIAL_URL_SKIP = (
    "sharer", "/share", "intent/tweet", "api.", "/policy",
    "/help", "/login", "/signup", "/apps", "oauth",
)


def find_social_links_regex(html):
    """
    Scan raw HTML source for social media URLs using regex.
    Catches social links inside JS strings, data-attributes, and
    non-<a> elements that BeautifulSoup misses.
    """
    found = {p: "" for p in SOCIAL_PLATFORMS}
    for platform in SOCIAL_PLATFORMS:
        pattern = (
            rf'https?://(?:www\.)?{re.escape(platform)}\.com'
            rf'/[A-Za-z0-9._\-/@%]+'
        )
        for m in re.findall(pattern, html, re.IGNORECASE):
            clean = m.split("?")[0].rstrip("/")
            if not any(s in clean.lower() for s in _SOCIAL_URL_SKIP):
                found[platform] = clean
                break   # first non-share URL wins
    return found


def find_email_from_fb_about(fb_url, page):
    """
    Try to find an email address in the Facebook Page's About section
    using Playwright.
    """
    if not fb_url or not page:
        return None
    about_url = fb_url.rstrip("/") + "/about"
    try:
        page.goto(about_url, timeout=25000, wait_until="domcontentloaded")
        page.wait_for_timeout(3000)
        content = page.content()
        m = re.search(
            r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", content
        )
        if m:
            return m.group(0)
    except Exception:
        pass
    return None


def find_facebook_url_via_search(practice_name, city, state, pw_page):
    """
    Use Playwright to Google-search for the practice's Facebook page.
    Called as a last resort when the website doesn't link to FB directly.
    Returns FB URL string or "".
    """
    if not pw_page:
        return ""
    query = f'"{practice_name}" {city} {state} site:facebook.com'
    url = f"https://www.google.com/search?q={quote_plus(query)}&hl=en"
    try:
        pw_page.goto(url, timeout=PW_TIMEOUT, wait_until="domcontentloaded")
        pw_page.wait_for_timeout(2000)
        content = pw_page.content()
        _FB_SKIP = ("share", "sharer", "login", "signup", "help", "policy",
                    "dialog", "tr.facebook", "l.facebook")
        for m in re.finditer(
            r'https?://(?:www\.)?facebook\.com/[A-Za-z0-9._\-/@%]+', content
        ):
            fb = m.group(0).split('"')[0].split("'")[0].rstrip("/")
            if not any(s in fb.lower() for s in _FB_SKIP):
                log.info(f"   Facebook found via search: {fb}")
                return fb
    except Exception:
        pass
    return ""


def count_testimonials(soup):
    """Count testimonial / review blocks on the website."""
    blocks = soup.find_all(
        ["div", "section", "article", "blockquote"],
        class_=re.compile(r"(testimonial|review|quote|patient.story|patient.review)", re.I),
    )
    count = len(blocks)
    if count == 0:
        count = len(soup.find_all("blockquote"))
    # Also check star-rating widgets as proxy for reviews
    if count == 0:
        stars = soup.find_all(class_=re.compile(r"(star|rating|review)", re.I))
        count = min(len(stars), 50)  # cap at 50 to avoid noise
    return str(count) if count > 0 else "0"


def find_hygienists(text):
    """
    Try to detect hygienist count from website text.
    Returns a number string if found, or "" if not determinable.

    Three-stage approach:
      1. Explicit numeric statement ("3 dental hygienists", "2 RDH on staff")
      2. Count distinct named credential holders ("Jane Smith, RDH" x N)
      3. Count raw credential mentions as a last resort (deduplicated by proximity)
    """
    # Stage 1 — explicit count statements
    _HYG_PATTERNS = [
        r"\b(\d+)\s+(?:registered\s+)?(?:dental\s+)?hygienists?",
        r"\b(\d+)\s+r\.?d\.?h\.?s?\b",
        r"hygienists?\s*(?:on\s+(?:staff|our\s+team|the\s+team))?[:\s]+(\d+)",
        r"\b(\d+)\s+(?:licensed\s+)?dental\s+hygiene\s+(?:specialists?|therapists?|professionals?)",
    ]
    for pat in _HYG_PATTERNS:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            for g in m.groups():
                if g and g.isdigit():
                    return g

    # Stage 2 — count distinct named RDH / hygienist credential holders
    # Matches "First [M.] Last, RDH" or "First Last RDH" with common credential variants
    _CRED_RE = re.compile(
        r'([A-Z][a-z]+(?:\s+[A-Z]\.?)?\s+[A-Z][a-z]+)'   # First [M.] Last
        r'\s*,?\s*'
        r'(?:R\.?D\.?H\.?|BSDH|RDHAP|'
        r'Registered\s+Dental\s+Hygienist|'
        r'Licensed\s+Dental\s+Hygienist|'
        r'Dental\s+Hygienist)',
        re.IGNORECASE,
    )
    named = {m.group(1).strip().lower() for m in _CRED_RE.finditer(text)}
    if named:
        return str(len(named))

    # Stage 2b — single-first-name staff (e.g. "Dawn, RDH" or "Dawn RDH")
    # Only used when Stage 2 found nothing; avoids over-counting body text.
    _SINGLE_RE = re.compile(
        r'\b([A-Z][a-z]{2,})\s*,?\s+'
        r'(?:R\.?D\.?H\.?|BSDH|RDHAP|'
        r'Registered\s+Dental\s+Hygienist|'
        r'Licensed\s+Dental\s+Hygienist|'
        r'Dental\s+Hygienist)',
        re.IGNORECASE,
    )
    single_named = {m.group(1).strip().lower() for m in _SINGLE_RE.finditer(text)}
    if single_named:
        return str(len(single_named))

    # Stage 3 — raw credential count: use a 60-char window (wider than before)
    # around each RDH / "dental hygienist" mention as a dedup key so the same
    # person listed on nav + body doesn't inflate the count.
    windows: set = set()
    for m in re.finditer(
        r'\bR\.?D\.?H\.?\b|(?:registered|licensed)\s+dental\s+hygienist',
        text, re.IGNORECASE,
    ):
        start = max(0, m.start() - 60)
        key = text[start: m.start() + 60].strip().lower()
        windows.add(key)
    if windows:
        return str(len(windows))

    return ""


# ─────────────────────────────────────────────────────────────────────────────
# GOOGLE & YELP RATINGS — requests-based
# ─────────────────────────────────────────────────────────────────────────────

def google_search_rating(practice_name, city, state):
    """
    Search Google for the practice and extract star rating + review count
    from search snippet text.
    Returns (rating_str, count_str).
    """
    query = f"{practice_name} {city} {state} dentist reviews"
    url   = f"https://www.google.com/search?q={quote_plus(query)}&hl=en"
    time.sleep(DELAY_SEC)
    r = safe_get(url)
    if not r:
        return "Not Found", "Not Found"
    text = extract_text(r.text)

    rating_match = re.search(r"\b([45]\.\d)\b", text)
    count_match  = re.search(r"(\d{1,5})\s*(?:google\s*)?reviews?", text)

    rating = rating_match.group(1) if rating_match else "Not Found"
    count  = count_match.group(1)  if count_match  else "Not Found"
    return rating, count


def yelp_search_rating(practice_name, city, state):
    """
    Search Yelp and extract star rating + review count.
    Returns (rating_str, count_str).
    """
    query = f"{practice_name} {city} {state}"
    url = (
        f"https://www.yelp.com/search"
        f"?find_desc={quote_plus(query)}"
        f"&find_loc={quote_plus(city + ', ' + state)}"
    )
    time.sleep(DELAY_SEC)
    r = safe_get(url)
    if not r:
        return "Not Found", "Not Found"
    text = extract_text(r.text)

    rating_match = re.search(r"\b([1-5]\.\d)\b", text)
    count_match  = re.search(r"(\d{1,5})\s*reviews?", text)

    rating = rating_match.group(1) if rating_match else "Not Found"
    count  = count_match.group(1)  if count_match  else "Not Found"
    return rating, count


# ─────────────────────────────────────────────────────────────────────────────
# PLAYWRIGHT — Facebook
# ─────────────────────────────────────────────────────────────────────────────

def get_facebook_stats_pw(url, page):
    """
    Navigate to a Facebook Page and extract follower count + visible post count.
    Returns (posts_str, followers_str).
    """
    if not url or not page:
        return "Not Found", "Not Found"
    try:
        page.goto(url, timeout=PW_TIMEOUT, wait_until="domcontentloaded")
        page.wait_for_timeout(4000)

        content = page.content()
        text = extract_text(content)

        # Follower / like count — "X followers" or "X people follow this"
        follower_match = re.search(
            r"([\d,]+(?:\.\d+)?[KMB]?)\s*(?:people\s+)?follow(?:ers?)?",
            text, re.IGNORECASE
        )
        like_match = re.search(
            r"([\d,]+(?:\.\d+)?[KMB]?)\s*(?:people\s+)?like",
            text, re.IGNORECASE
        )
        followers = "Not Found"
        if follower_match:
            followers = follower_match.group(1).replace(",", "")
        elif like_match:
            followers = like_match.group(1).replace(",", "")

        # Post count — count visible post/article elements
        try:
            post_count = page.locator('[role="article"]').count()
            posts = str(post_count) if post_count > 0 else "See Page"
        except Exception:
            posts = "See Page"

        return posts, followers

    except PlaywrightTimeout:
        log.warning(f"  Facebook timeout: {url}")
        return "Blocked", "Blocked"
    except Exception as e:
        log.warning(f"  Facebook Playwright error: {e}")
        return "Blocked", "Blocked"


def get_facebook_stats_requests(url):
    """Fallback requests-based Facebook scrape."""
    if not url:
        return "Not Found", "Not Found"
    time.sleep(DELAY_SEC)
    r = safe_get(url)
    if not r:
        return "Blocked", "Blocked"
    text = extract_text(r.text)
    m = re.search(r"([\d,]+)\s*(?:people\s*)?(?:like|follow)", text)
    followers = m.group(1).replace(",", "") if m else "Blocked"
    return "Login Required", followers


# ─────────────────────────────────────────────────────────────────────────────
# PLAYWRIGHT — Instagram
# ─────────────────────────────────────────────────────────────────────────────

def get_instagram_stats_pw(url, page):
    """
    Navigate to an Instagram profile page and extract post + follower counts.
    Returns (posts_str, followers_str).
    """
    if not url or not page:
        return "Not Found", "Not Found"
    try:
        page.goto(url, timeout=PW_TIMEOUT, wait_until="domcontentloaded")
        page.wait_for_timeout(4000)

        content = page.content()

        # 1) Try embedded JSON data (Instagram embeds stats in JS)
        posts_match     = re.search(r'"edge_owner_to_timeline_media".*?"count":(\d+)', content)
        followers_match = re.search(r'"edge_followed_by".*?"count":(\d+)', content)
        if posts_match and followers_match:
            return posts_match.group(1), followers_match.group(1)

        # 2) Try meta description  (e.g. "1,234 Followers, 567 Following, 89 Posts")
        soup = BeautifulSoup(content, "lxml")
        meta_desc = soup.find("meta", attrs={"name": "description"})
        if meta_desc:
            desc = meta_desc.get("content", "")
            pm = re.search(r"([\d,]+)\s*Posts?", desc, re.IGNORECASE)
            fm = re.search(r"([\d,]+)\s*Followers?", desc, re.IGNORECASE)
            if pm and fm:
                return pm.group(1).replace(",", ""), fm.group(1).replace(",", "")

        # 3) Plain text search
        text = extract_text(content)
        pm2 = re.search(r"([\d,]+)\s+posts?", text, re.IGNORECASE)
        fm2 = re.search(r"([\d,]+)\s+followers?", text, re.IGNORECASE)
        posts     = pm2.group(1).replace(",", "") if pm2 else "Blocked"
        followers = fm2.group(1).replace(",", "") if fm2 else "Blocked"
        return posts, followers

    except PlaywrightTimeout:
        log.warning(f"  Instagram timeout: {url}")
        return "Blocked", "Blocked"
    except Exception as e:
        log.warning(f"  Instagram Playwright error: {e}")
        return "Blocked", "Blocked"


def get_instagram_stats_requests(url):
    """Fallback requests-based Instagram scrape."""
    if not url:
        return "Not Found", "Not Found"
    time.sleep(DELAY_SEC)
    r = safe_get(url)
    if not r:
        return "Blocked", "Blocked"
    text = r.text
    pm = re.search(r'"edge_owner_to_timeline_media".*?"count":(\d+)', text)
    fm = re.search(r'"edge_followed_by".*?"count":(\d+)', text)
    posts     = pm.group(1) if pm else "Blocked"
    followers = fm.group(1) if fm else "Blocked"
    return posts, followers


def get_instagram_stats_api(ig_url: str) -> tuple[str, str]:
    """
    Fetch Instagram posts + followers via the internal web API (no login needed
    for public profiles).  Uses curl_cffi TLS impersonation.
    Returns (posts, followers) or ("", "") on failure.
    """
    if not ig_url or not _CFFI_AVAILABLE:
        return "", ""
    try:
        path = urlparse(ig_url).path.strip("/")
        username = path.split("/")[0] if path else ""
        if not username:
            return "", ""
    except Exception:
        return "", ""

    api_url = f"https://www.instagram.com/api/v1/users/web_profile_info/?username={username}"
    headers = {
        "x-ig-app-id":      "936619743392459",
        "User-Agent":       random.choice(_SOCIAL_UA_LIST),
        "Accept":           "*/*",
        "Accept-Language":  "en-US,en;q=0.9",
        "Referer":          f"https://www.instagram.com/{username}/",
        "X-Requested-With": "XMLHttpRequest",
    }
    try:
        sess = cffi_requests.Session(impersonate=random.choice(_CFFI_IG_PROF))
        r = sess.get(api_url, headers=headers, timeout=15)
        if r.status_code == 200:
            data = r.json()
            user = data.get("data", {}).get("user", {})
            if user:
                posts_raw     = user.get("edge_owner_to_timeline_media", {}).get("count", "")
                followers_raw = user.get("edge_followed_by", {}).get("count", "")
                posts_str     = str(int(posts_raw))     if posts_raw     != "" else ""
                followers_str = str(int(followers_raw)) if followers_raw != "" else ""
                return posts_str, followers_str
        log.debug(f"  IG API status {r.status_code} for @{username}")
    except Exception as e:
        log.debug(f"  IG API error for @{username}: {e}")
    return "", ""


def get_tiktok_stats(tt_url: str) -> tuple[str, str]:
    """
    Fetch TikTok video count + follower count via curl_cffi.
    Parses the __UNIVERSAL_DATA_FOR_REHYDRATION__ JSON embedded in the page.
    Returns (videos, followers) or ("", "") on failure.
    """
    if not tt_url or not _CFFI_AVAILABLE:
        return "", ""
    try:
        path = urlparse(tt_url).path.strip("/")
        username = path.split("/")[0].lstrip("@")
        if not username:
            return "", ""
    except Exception:
        return "", ""

    url = f"https://www.tiktok.com/@{username}"
    headers = {
        "User-Agent":      random.choice(_SOCIAL_UA_LIST),
        "Accept-Language": "en-US,en;q=0.9",
        "Accept":          "text/html,application/xhtml+xml,*/*;q=0.8",
        "Referer":         "https://www.tiktok.com/",
    }
    for profile in random.sample(_CFFI_PROFILES, len(_CFFI_PROFILES)):
        try:
            sess = cffi_requests.Session(impersonate=profile)
            r = sess.get(url, headers=headers, timeout=20, allow_redirects=True)
            if r.status_code != 200:
                continue
            html = r.text
            # Method 1: embedded JSON
            m = re.search(
                r'id=["\']__UNIVERSAL_DATA_FOR_REHYDRATION__["\'][^>]*>(.*?)</script>',
                html, re.S
            )
            if m:
                try:
                    data  = json.loads(m.group(1))
                    stats = data["__DEFAULT_SCOPE__"]["webapp.user-detail"]["userInfo"].get("stats", {})
                    followers = str(int(stats["followerCount"])) if "followerCount" in stats else ""
                    videos    = str(int(stats["videoCount"]))    if "videoCount"    in stats else ""
                    if followers or videos:
                        return videos, followers
                except Exception:
                    pass
            # Method 2: raw regex fallback
            mf = re.search(r'"followerCount"\s*:\s*(\d+)', html)
            mv = re.search(r'"videoCount"\s*:\s*(\d+)', html)
            followers = mf.group(1) if mf else ""
            videos    = mv.group(1) if mv else ""
            if followers or videos:
                return videos, followers
        except Exception as e:
            log.debug(f"  TikTok error ({profile}): {e}")
        time.sleep(0.5)
    return "", ""


# ─────────────────────────────────────────────────────────────────────────────
# PLAYWRIGHT — Invisalign Tier
# ─────────────────────────────────────────────────────────────────────────────

def get_invisalign_tier_pw(practice_name, city, state, zip_code, page):
    """
    Search the Invisalign Find-a-Doctor locator (JS-rendered) with Playwright.
    Returns tier string like "Gold", "Platinum Plus", etc. or "Not Listed".
    """
    if not page:
        return _invisalign_api_fallback(practice_name, zip_code)

    search_term = zip_code if zip_code and zip_code not in ("", "None") else f"{city} {state}"
    url = f"https://www.invisalign.com/find-a-doctor#q={quote_plus(search_term)}"

    try:
        page.goto(url, timeout=PW_TIMEOUT * 2, wait_until="networkidle")
        page.wait_for_timeout(6000)

        # Try typing into the search box if it exists
        for selector in [
            'input[placeholder*="zip" i]',
            'input[placeholder*="location" i]',
            'input[placeholder*="search" i]',
            'input[type="search"]',
            'input[type="text"]',
        ]:
            try:
                el = page.locator(selector).first
                if el.is_visible(timeout=2000):
                    el.fill(search_term)
                    el.press("Enter")
                    page.wait_for_timeout(4000)
                    break
            except Exception:
                pass

        content = page.content()
        name_words = [w for w in practice_name.lower().split() if len(w) > 3]

        # Search tier labels near the practice name
        for tier in INVISALIGN_TIERS:
            indices = [
                m.start()
                for m in re.finditer(re.escape(tier.lower()), content.lower())
            ]
            for idx in indices:
                context = content[max(0, idx - 600) : idx + 600].lower()
                if any(w in context for w in name_words):
                    return tier

        # Fallback: return the first tier found on the page if any
        for tier in INVISALIGN_TIERS:
            if tier.lower() in content.lower():
                return f"{tier} (verify name match)"

        return "Not Listed"

    except PlaywrightTimeout:
        log.warning("  Invisalign locator timed out.")
        return _invisalign_api_fallback(practice_name, zip_code)
    except Exception as e:
        log.warning(f"  Invisalign Playwright error: {e}")
        return _invisalign_api_fallback(practice_name, zip_code)


def _invisalign_api_fallback(practice_name, zip_code):
    """Try the Invisalign REST API as a fallback."""
    api_url = "https://www.invisalign.com/api/locator/search"
    params  = {
        "query":   practice_name,
        "country": "US",
        "zip":     str(zip_code) if zip_code else "",
        "lang":    "en",
    }
    try:
        r = requests.get(api_url, params=params, headers=HEADERS, timeout=TIMEOUT)
        if r.status_code == 200:
            data      = r.json()
            providers = data.get("providers", data.get("results", []))
            for p in providers:
                name = (p.get("name", "") + " " + p.get("practiceName", "")).lower()
                if any(w in name for w in practice_name.lower().split()[:2]):
                    tier = p.get("tier", p.get("providerLevel", ""))
                    return tier if tier else "Not Listed"
    except Exception:
        pass
    return f"Check: invisalign.com/find-a-doctor (zip:{zip_code})"


# ─────────────────────────────────────────────────────────────────────────────
# MAIN PER-PRACTICE SCRAPER
# ─────────────────────────────────────────────────────────────────────────────

def scrape_practice(row, pw_page=None):
    """
    Scrape all data for one practice.
    pw_page — a live Playwright Page object, or None for requests-only mode.
    Returns a flat dict of all scraped fields.
    """
    name    = str(row.get("Practice Name", "")).strip()
    website = str(row.get("Website",       "")).strip()
    city    = str(row.get("City",          "")).strip()
    state   = str(row.get("State",         "")).strip()
    zip_c   = str(row.get("Zip",           "")).strip()
    log.info(f"▶ Scraping: {name}")

    result = {
        # Practice info
        "skip_reason":           "",   # set when site is bot-blocked or unreachable
        "email":                 "Not Found",
        "scraped_doctor_names":  "Not Found",
        "hygienists":            "N/A",
        "locations_count":       "1",
        # Social URLs
        "facebook_url":          "",
        "facebook_posts":        "Not Found",
        "facebook_followers":    "Not Found",
        "instagram_url":         "",
        "instagram_posts":       "Not Found",
        "instagram_followers":   "Not Found",
        "tiktok_url":            "",
        "tiktok_posts":          "Not Found",
        "tiktok_followers":      "Not Found",
        "linkedin_url":          "",
        "linkedin_posts":        "Not Found",
        "linkedin_followers":    "Not Found",
        # Technology
        "cerec":                 "",
        "cbct":                  "",
        "lasers":                "",
        "ai":                    "",
        "intraoral":             "",
        # Services
        "invisalign":            0,
        "invisalign_tier":       "N/A",   # disabled — check manually
        "clear_aligners":        0,
        "veneers":               0,
        "implants":              0,
        "smile_makeovers":       0,
        "whitening":             0,
        "sedation":              0,
        "holistic":              0,
        "dental_plan":           "",
        "cancer_screening":      0,
        # Doctor data — "doctors" holds per-doctor list for multi-row output
        "doctors":               [],
        "associations":          "Not Found",
        "specialty":             "Not Found",
        # Patient experience
        "google_rating":         "Not Found",
        "google_reviews":        "Not Found",
        "yelp_rating":           "Not Found",
        "yelp_reviews":          "Not Found",
        "testimonials":          "0",
    }

    # ── helpers ───────────────────────────────────────────────────────────────
    practice_idx  = row.get("Index", 0)
    practice_name = row.get("Practice Name", "unknown")
    _cache_folder = _cache_dir(practice_idx, practice_name)
    _sub_counter  = [0]   # mutable int for sub-page numbering inside closures

    def _cache(page_type: str, url: str, html: str):
        _cache_html(_cache_folder, page_type, url, html)

    base_url = website if website.startswith("http") else "https://" + website

    def _url_variations(url):
        """Return alternative URL forms to try when original fails (DNS/SSL)."""
        parsed = urlparse(url if url.startswith("http") else "https://" + url)
        host = parsed.netloc
        path = parsed.path or "/"
        alts = []
        if host.startswith("www."):
            bare = host[4:]
            alts += [f"https://{bare}{path}", f"http://{bare}{path}",
                     f"http://{host}{path}"]
        else:
            alts += [f"https://www.{host}{path}", f"http://{host}{path}",
                     f"http://www.{host}{path}"]
        # If domain ends with "dent" (not "dental"), also try with "al" appended
        # e.g. hallfamilydent.com → hallfamilydental.com
        bare_check = host[4:] if host.startswith("www.") else host
        if re.search(r'dent\.', bare_check) and not re.search(r'dental\.', bare_check):
            fixed = re.sub(r'dent\.', 'dental.', bare_check, count=1)
            alts += [f"https://{fixed}{path}", f"https://www.{fixed}{path}"]
        return alts

    def _merge_socials(html_or_soup, raw_html=None):
        """Extract social links from soup AND raw HTML; merge into result."""
        # DOM-based (catches explicit <a> tags)
        if html_or_soup:
            for platform, url_val in find_social_links(html_or_soup).items():
                if url_val and not result[f"{platform}_url"]:
                    result[f"{platform}_url"] = url_val
        # Regex-based (catches JS strings, data-attrs, non-<a> elements)
        if raw_html:
            for platform, url_val in find_social_links_regex(raw_html).items():
                if url_val and not result[f"{platform}_url"]:
                    result[f"{platform}_url"] = url_val

    def _check_mailto(soup_obj):
        """Return first mailto or Cloudflare-protected email from a soup, or None."""
        # Cloudflare email obfuscation
        for a in soup_obj.find_all("a", href=re.compile(r"cdn-cgi/l/email-protection", re.I)):
            href = a.get("href", "")
            fragment = href.split("#")[-1] if "#" in href else ""
            decoded = _decode_cloudflare_email(fragment)
            if decoded:
                return decoded
        # Standard mailto
        for a in soup_obj.find_all("a", href=re.compile(r"mailto:", re.I)):
            addr = a["href"].replace("mailto:", "").split("?")[0].strip()
            if "@" in addr:
                return addr
        return None

    # ── 1. Scrape practice website ────────────────────────────────────────────
    if website and website not in ("", "None", "nan"):
        log.info(f"   Fetching website: {website}")
        time.sleep(DELAY_SEC)

        all_text = ""
        all_soup = None
        pw_loaded_homepage = False

        # ── a) Try requests (with SSL bypass built-in) ────────────────────────
        r = safe_get(website)
        # If original URL fails, try www/non-www and http/https variants
        if not r:
            for alt in _url_variations(base_url):
                log.info(f"   Retrying with URL variant: {alt}")
                r = safe_get(alt)
                if r:
                    base_url = alt   # use the working URL for sub-page joins
                    break
        if r:
            all_soup = BeautifulSoup(r.text, "lxml")
            all_text = extract_text(r.text)
            _merge_socials(all_soup, r.text)
            _cache("homepage", r.url, r.text)

        # ── b) Playwright fallback when requests completely fails ─────────────
        if not all_soup and pw_page:
            log.info("   Requests failed — using Playwright for homepage…")
            urls_to_try = [base_url] + _url_variations(base_url)
            for try_url in urls_to_try:
                try:
                    pw_page.goto(try_url, timeout=PW_TIMEOUT, wait_until="domcontentloaded")
                    pw_page.wait_for_timeout(2500)
                    pw_html = pw_page.content()
                    # Only accept if page loaded real content (not a DNS error page)
                    if len(pw_html) > 2000:
                        base_url = try_url
                        all_soup = BeautifulSoup(pw_html, "lxml")
                        all_text = extract_text(pw_html)
                        pw_loaded_homepage = True
                        _merge_socials(all_soup, pw_html)
                        _cache("homepage", try_url, pw_html)
                        log.info(f"   Playwright loaded: {try_url}")
                        break
                except Exception as e:
                    log.debug(f"   Playwright attempt failed for {try_url}: {e}")
            if not all_soup:
                log.warning(f"   All URL variants failed for: {name}")
                # Determine WHY: quick HEAD to classify block vs. unreachable
                try:
                    _probe = requests.head(
                        base_url, headers=HEADERS, timeout=8, verify=False,
                        allow_redirects=True,
                    )
                    _sc = _probe.status_code
                    if _sc in (403, 429, 503) or (_sc >= 520 and _sc < 530):
                        result["skip_reason"] = f"Bot Protection (HTTP {_sc})"
                    elif _sc == 404:
                        result["skip_reason"] = "Domain Not Found (404)"
                    else:
                        result["skip_reason"] = f"Access Blocked / JS Challenge (HTTP {_sc})"
                except requests.exceptions.ConnectionError:
                    result["skip_reason"] = "Connection Failed / Domain Unreachable"
                except requests.exceptions.Timeout:
                    result["skip_reason"] = "Connection Timeout"
                except Exception:
                    result["skip_reason"] = "Inaccessible (unknown)"
                log.warning(f"   Skip reason: {result['skip_reason']}")

        # ── c) Scrape sub-pages ───────────────────────────────────────────────
        # Sub-page keywords — every link whose href contains any of these is visited
        _SUB_KW = [
            "service", "about", "about-us", "about_us", "our-story",
            "team", "technology", "treatment", "contact", "contact-us",
            "doctor", "provider", "location", "office", "staff",
            "cosmetic", "procedure", "care", "a-i-", "ai-assisted",
            "implant", "laser", "invisalign", "crown",
            # additional keywords for broader coverage
            "sedation", "whitening", "veneer", "aligner", "cbct",
            "3d-imaging", "3d-xray", "3d-scan", "3d", "x-ray", "xray",
            "cerec", "digital", "hygiene", "patient", "faq",
            "smile", "restoration", "restorative", "general",
            # priority field keywords — technology / services / doctor
            "holistic", "biological", "biomimetic", "membership", "plan",
            "advanced", "innovation", "equipment", "state-of-the-art",
            "specialist", "specialty", "association", "membership",
            "perio", "ortho", "endo", "oral-surgery", "oral-health",
            "scan", "scanner", "imaging", "cbct", "cone-beam",
            "itero", "3shape", "medit", "planmeca",
            "overjet", "pearl", "diagnocat", "videa",
            "waterlase", "biolase", "solea", "fotona",
            "membership-plan", "dental-plan", "savings-plan", "wellness-plan",
            "cancer", "oral-cancer", "velscope",
        ]

        _SKIP_EXTS = ('.pdf', '.jpg', '.jpeg', '.png', '.gif', '.svg',
                      '.zip', '.doc', '.docx', '.mp4', '.mp3')

        def _collect_nav_links(soup_obj, already_seen):
            """Collect ALL same-domain links from nav/header menu containers.
            This ensures every navigation item is visited regardless of keyword."""
            links = []
            _nav_containers = (
                soup_obj.find_all(["nav", "header"]) +
                soup_obj.find_all(class_=re.compile(
                    r'\b(main.?menu|nav.?menu|site.?nav|primary.?nav|top.?nav|'
                    r'navigation|navbar|megamenu|header.?menu)\b', re.I
                ))
            )
            for container in _nav_containers:
                for a in container.find_all("a", href=True):
                    href = a["href"]
                    if not href or href.startswith(("#", "mailto:", "tel:", "javascript:")):
                        continue
                    full = urljoin(base_url, href)
                    if not full.startswith("http"):
                        continue
                    if urlparse(full).netloc != urlparse(base_url).netloc:
                        continue
                    if any(full.lower().endswith(ext) for ext in _SKIP_EXTS):
                        continue
                    if full in already_seen:
                        continue
                    already_seen.add(full)
                    links.append(full)
            return links

        def _collect_subpage_links(soup_obj, already_seen):
            """Return new sub-page URLs matching _SUB_KW not already in already_seen."""
            links = []
            for a in soup_obj.find_all("a", href=True):
                href_l = a["href"].lower()
                full = urljoin(base_url, a["href"])
                if not full.startswith("http"):
                    continue
                # Only same-domain links
                if urlparse(full).netloc != urlparse(base_url).netloc:
                    continue
                if full in already_seen:
                    continue
                if any(ext in full.lower() for ext in _SKIP_EXTS):
                    continue
                if any(kw in href_l for kw in _SUB_KW):
                    links.append(full)
                    already_seen.add(full)
            return links

        # Track all scraped soups for doctor extraction
        all_scraped_soups: list = []   # list of (page_type, soup)

        if all_soup:
            sub_pages_found = set([base_url])
            # ── Nav links — always fetch ALL menu items, no cap ───────────────
            # These are the most important pages (services, team, technology,
            # about) and must never be skipped due to a page count limit.
            nav_urls = _collect_nav_links(all_soup, sub_pages_found)
            # ── Keyword-matched links not already in nav ──────────────────────
            kw_urls  = _collect_subpage_links(all_soup, sub_pages_found)

            lvl2_candidates = []

            def _fetch_subpage(sub_url, page_label="sub"):
                """Fetch one sub-page, merge text/soups/socials, cache it."""
                nonlocal all_text
                log.info(f"   Sub-page ({page_label}): {sub_url}")
                time.sleep(DELAY_SEC)
                sub_r = safe_get(sub_url)
                if sub_r:
                    sub_html = sub_r.text
                    all_text += " " + extract_text(sub_html)
                    sub_soup = BeautifulSoup(sub_html, "lxml")
                    all_scraped_soups.append((page_label, sub_soup))
                    _merge_socials(sub_soup, sub_html)
                    if result["email"] == "Not Found":
                        found_mail = _check_mailto(sub_soup)
                        if found_mail:
                            result["email"] = found_mail
                    _sub_counter[0] += 1
                    _cache(f"sub_{_sub_counter[0]:02d}", sub_url, sub_html)
                    return sub_soup
                return None

            # ── Step 1: ALL nav/menu links (uncapped) ─────────────────────────
            for sub_url in nav_urls:
                sub_soup = _fetch_subpage(sub_url, "nav")
                if sub_soup:
                    lvl2_candidates += _collect_subpage_links(sub_soup, sub_pages_found)

            # ── Step 2: Keyword-matched links beyond nav (L1_LIMIT cap) ───────
            for sub_url in kw_urls[:L1_LIMIT]:
                sub_soup = _fetch_subpage(sub_url, "kw")
                if sub_soup:
                    lvl2_candidates += _collect_subpage_links(sub_soup, sub_pages_found)

            # ── Level-2 sub-pages (up to L2_LIMIT, new unique pages only) ───────
            for sub_url in lvl2_candidates[:L2_LIMIT]:
                _fetch_subpage(sub_url, "sub_l2")

            # ── Level-3 priority pass — crawl every remaining same-domain link
            # not yet visited in L1/L2. Capped at L3_LIMIT to stay reasonable.
            _l3_seen = set(sub_pages_found)
            _l3_urls = []
            for _, _sp in [(None, all_soup)] + all_scraped_soups:
                for _a in _sp.find_all("a", href=True):
                    _full = urljoin(base_url, _a["href"])
                    if (
                        _full.startswith("http")
                        and urlparse(_full).netloc == urlparse(base_url).netloc
                        and _full not in _l3_seen
                        and not any(ext in _full.lower() for ext in _SKIP_EXTS)
                    ):
                        _l3_urls.append(_full)
                        _l3_seen.add(_full)
            if _l3_urls:
                log.info(f"   Priority L3 pass: {len(_l3_urls)} uncrawled links — fetching up to {L3_LIMIT}")
            for _l3_url in _l3_urls[:L3_LIMIT]:
                _fetch_subpage(_l3_url, "sub_l3")

        # ── d) Extract fields that need soup (testimonials, email) ─────────────
        if all_soup:
            # Email: homepage → all_text regex → Playwright contact pages
            if result["email"] == "Not Found":
                result["email"] = find_email(all_text, all_soup)
            if result["email"] == "Not Found" and pw_page:
                log.info("   Email not found via requests — trying Playwright…")
                result["email"] = find_email_pw(website, pw_page)

            result["locations_count"] = find_locations_count(all_text, all_soup)
            # Testimonials: count across homepage + all sub-pages, deduplicate by text
            _test_seen = set()
            _test_total = 0
            _TEST_CLASS_RE = re.compile(
                r"(testimonial|review|quote|patient.story|patient.review|"
                r"feedback|client.say|what.people|slider.item|carousel.item|"
                r"swiper.slide|slick.slide|rating.block|star.review)", re.I
            )
            _TEST_ATTR_RE  = re.compile(
                r"(testimonial|review|quote|patient)", re.I
            )
            for _tsp in [all_soup] + [sp for _, sp in all_scraped_soups]:
                for _blk in _tsp.find_all(
                    ["div", "section", "article", "blockquote", "li"],
                    class_=_TEST_CLASS_RE,
                ):
                    _key = _blk.get_text(separator=" ", strip=True)[:80]
                    if _key and _key not in _test_seen:
                        _test_seen.add(_key)
                        _test_total += 1
                # Also check data-* attributes (e.g. data-testimonial, data-type="review")
                for _blk in _tsp.find_all(
                    lambda tag: any(
                        _TEST_ATTR_RE.search(str(v))
                        for k, v in tag.attrs.items()
                        if k.startswith("data-") and isinstance(v, str)
                    )
                ):
                    _key = _blk.get_text(separator=" ", strip=True)[:80]
                    if _key and _key not in _test_seen:
                        _test_seen.add(_key)
                        _test_total += 1
            if _test_total == 0:
                for _tsp in [all_soup] + [sp for _, sp in all_scraped_soups]:
                    for _bq in _tsp.find_all("blockquote"):
                        _key = _bq.get_text(separator=" ", strip=True)[:80]
                        if _key and _key not in _test_seen:
                            _test_seen.add(_key)
                            _test_total += 1
            result["testimonials"] = str(_test_total) if _test_total > 0 else "0"

        # ── e) Playwright pass — always render homepage to capture JS content ─
        # This enriches all_text with dynamically loaded text (social links,
        # Invisalign mentions, associations, specialty keywords that live in JS).
        if pw_page and not pw_loaded_homepage:
            log.info(f"   Rendering homepage with Playwright for JS content…")
            try:
                pw_page.goto(base_url, timeout=PW_TIMEOUT, wait_until="domcontentloaded")
                pw_page.wait_for_timeout(2500)
                pw_html = pw_page.content()
                pw_soup = BeautifulSoup(pw_html, "lxml")
                # Merge JS-rendered text into all_text
                all_text += " " + extract_text(pw_html)
                # Merge social links found in JS-rendered page
                _merge_socials(pw_soup, pw_html)
                _cache("pw_homepage", base_url, pw_html)
                # Email fallback from rendered page
                if result["email"] == "Not Found":
                    m = re.search(
                        r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", pw_html
                    )
                    if m:
                        result["email"] = m.group(0)
            except Exception as e:
                log.debug(f"   Playwright homepage pass failed: {e}")

        # ── e2) Playwright render of service/treatment sub-pages (JS-rendered sites) ─
        # Some sites load service lists dynamically; render up to 3 pages to enrich
        # all_text so service keyword counts are accurate.
        if pw_page and all_soup and base_url:
            _svc_pw_urls = []
            _seen_svc = set(sub_pages_found) if 'sub_pages_found' in dir() else set()
            for _a in all_soup.find_all("a", href=True):
                _href_l = _a["href"].lower()
                if any(kw in _href_l for kw in ("service", "treatment", "procedure", "cosmetic")):
                    _full = urljoin(base_url, _a["href"])
                    if (urlparse(_full).netloc == urlparse(base_url).netloc
                            and _full not in _seen_svc):
                        _svc_pw_urls.append(_full)
                        _seen_svc.add(_full)
            for _svc_url in _svc_pw_urls[:10]:
                try:
                    log.info(f"   Rendering service page (PW): {_svc_url}")
                    pw_page.goto(_svc_url, timeout=PW_TIMEOUT, wait_until="domcontentloaded")
                    pw_page.wait_for_timeout(2000)
                    _pw_svc_html = pw_page.content()
                    all_text += " " + extract_text(_pw_svc_html)
                    _sub_counter[0] += 1
                    _cache(f"pw_svc_{_sub_counter[0]:02d}", _svc_url, _pw_svc_html)
                except Exception:
                    pass

        # ── f) Text-based field extraction — runs AFTER all_text is enriched ─
        if all_text:
            result["associations"] = find_associations(all_text)
            result["specialty"]    = find_specialty(all_text)
            # Hygienist: search all_text + each cached page individually for number patterns
            hyg_text = find_hygienists(all_text)
            if not hyg_text and _cache_folder and os.path.exists(
                os.path.join(_cache_folder, "manifest.json")
            ):
                with open(os.path.join(_cache_folder, "manifest.json"), encoding="utf-8") as _hf:
                    try:
                        _hm = json.load(_hf)
                    except Exception:
                        _hm = {}
                for _ptype, _pinfo in _hm.get("pages", {}).items():
                    _fpath = os.path.join(_cache_folder, _pinfo.get("file", ""))
                    if os.path.exists(_fpath):
                        with open(_fpath, encoding="utf-8", errors="replace") as _hf2:
                            _page_text = extract_text(_hf2.read())
                        hyg_text = find_hygienists(_page_text)
                        if hyg_text:
                            break
            result["hygienists"] = hyg_text

        # ── g) Per-doctor scraping (team page) ────────────────────────────────
        if all_soup:
            doctors, hyg_count = scrape_doctors_full(
                all_soup, base_url, all_text, pw_page,
                all_soups_for_team=all_scraped_soups
            )
            result["doctors"] = doctors
            # Doctor names fallback for display (comma-joined)
            if doctors:
                result["scraped_doctor_names"] = ", ".join(d["name"] for d in doctors)
            # Override hygienist count if team page gave a concrete number
            if hyg_count is not None and hyg_count > 0:
                result["hygienists"] = str(hyg_count)

        # ── Count service keywords (per-page, body-text only to avoid nav inflation) ──
        svc_counts = {
            "Invisalign": 0, "Clear Aligners": 0, "Veneers": 0,
            "Implants": 0, "Smile Makeovers": 0, "Teeth Whitening": 0,
            "Sedation Dentistry": 0, "Holistic Dentistry": 0, "Dental Plan": 0, "Cancer Screening": 0,
        }
        # Build body_text from cached HTML pages (nav/header/footer stripped)
        # Build two text versions per deduplicated page:
        #   body_text = nav/header/footer stripped (accurate, but misses services listed only in nav)
        #   full_text = only script/style stripped (catches nav-listed services as fallback)
        _svc_pages: list = []  # list of (body_text, full_text)
        _cache_manifest = os.path.join(_cache_folder, "manifest.json") if _cache_folder else ""
        if _cache_folder and os.path.exists(_cache_manifest):
            with open(_cache_manifest, encoding="utf-8") as _f:
                _manifest = json.load(_f)
            _seen_page_urls: set = set()
            for _ptype, _pinfo in _manifest.get("pages", {}).items():
                # Deduplicate by URL (strip fragment) — avoids counting anchor-variant pages twice
                _page_url = _pinfo.get("url", "")
                _page_url_base = _page_url.split("#")[0] if _page_url else _ptype
                if _page_url_base in _seen_page_urls:
                    continue
                _seen_page_urls.add(_page_url_base)
                _fpath = os.path.join(_cache_folder, _pinfo.get("file", ""))
                if os.path.exists(_fpath):
                    with open(_fpath, encoding="utf-8", errors="replace") as _f:
                        _raw = _f.read()
                    # Full text augmented with meta/JSON-LD/URL signals
                    _aug = extract_augmented_text(_raw, _page_url)
                    _svc_pages.append((extract_body_text(_raw), extract_text(_raw) + " " + _aug))
        else:
            _svc_pages = [(all_text, all_text)]

        # Always include all_text as supplementary source: captures Playwright-rendered
        # JS content (homepage + any PW-rendered sub-pages) not in static cache files
        if all_text:
            _svc_pages.append((all_text, all_text))

        # Primary count: body text only (nav stripped), capped at 5 per page
        # Full text includes meta/JSON-LD/URL path signals, capped at 3 per page
        _svc_body = {k: 0 for k in svc_counts}
        _svc_full = {k: 0 for k in svc_counts}
        for _bt, _ft in _svc_pages:
            for keyword, category in SERVICE_KEYWORDS.items():
                _svc_body[category] += count_keyword_capped(_bt, keyword, cap=5)
                _svc_full[category] += count_keyword_capped(_ft, keyword, cap=3)
        # Take max of body and full text — never leave a category blank if
        # the keyword appears anywhere (meta, URL, JSON-LD, or body text)
        for _cat in svc_counts:
            svc_counts[_cat] = max(_svc_body[_cat], _svc_full[_cat])

        result["invisalign"]      = svc_counts["Invisalign"]
        result["clear_aligners"]  = svc_counts["Clear Aligners"]
        result["veneers"]         = svc_counts["Veneers"]
        result["implants"]        = svc_counts["Implants"]
        result["smile_makeovers"] = svc_counts["Smile Makeovers"]
        result["whitening"]       = svc_counts["Teeth Whitening"]
        result["sedation"]        = svc_counts["Sedation Dentistry"]
        result["holistic"]        = svc_counts["Holistic Dentistry"]
        result["dental_plan"]     = "Mentioned" if svc_counts["Dental Plan"] > 0 else ""
        result["cancer_screening"]= svc_counts["Cancer Screening"]

        # ── Technology detection — search every cached page + all_text ─────
        tech_found = set()
        # Build combined text: all_text already has all pages; also check each
        # cached page's raw text to catch keywords that may have been stripped
        _tech_texts = [all_text]
        if _cache_folder and os.path.exists(os.path.join(_cache_folder, "manifest.json")):
            with open(os.path.join(_cache_folder, "manifest.json"), encoding="utf-8") as _tf:
                try:
                    _tm = json.load(_tf)
                except Exception:
                    _tm = {}
            for _ptype, _pinfo in _tm.get("pages", {}).items():
                _fpath = os.path.join(_cache_folder, _pinfo.get("file", ""))
                if os.path.exists(_fpath):
                    with open(_fpath, encoding="utf-8", errors="replace") as _tf2:
                        _raw2 = _tf2.read()
                    # Include meta/JSON-LD/URL signals so tech in structured data isn't missed
                    _aug2 = extract_augmented_text(_raw2, _pinfo.get("url", ""))
                    _tech_texts.append(extract_text(_raw2) + " " + _aug2)
        _combined_tech = " ".join(_tech_texts)
        # Normalize hyphens → spaces so "cone-beam" matches "cone beam" keyword etc.
        _combined_tech_n = _combined_tech.replace("-", " ")
        for keyword, tech_name in TECH_KEYWORDS.items():
            if keyword in _combined_tech_n:
                tech_found.add(tech_name)
        # AI: also check for standalone word "ai" with word boundaries
        if "AI" not in tech_found and re.search(r'\bai\b', _combined_tech):
            tech_found.add("AI")

        result["cerec"]     = "X" if "CEREC"              in tech_found else ""
        result["cbct"]      = "X" if "CBCT"               in tech_found else ""
        result["lasers"]    = "X" if "Lasers"             in tech_found else ""
        result["ai"]        = "X" if "AI"                 in tech_found else ""
        result["intraoral"] = "X" if "Intraoral Scanners" in tech_found else ""

    else:
        log.warning(f"   No website for: {name}")

    # ── 2. Social media stats ─────────────────────────────────────────────────
    if result["facebook_url"]:
        log.info("   Fetching Facebook stats…")
        if pw_page and USE_PLAYWRIGHT:
            _p, f = get_facebook_stats_pw(result["facebook_url"], pw_page)
        else:
            _p, f = get_facebook_stats_requests(result["facebook_url"])
        # facebook_posts intentionally not captured (per user request)
        result["facebook_followers"] = f

    if result["instagram_url"]:
        log.info("   Fetching Instagram stats…")
        ig_posts, ig_foll = get_instagram_stats_api(result["instagram_url"])
        if not ig_posts and not ig_foll:
            if pw_page and USE_PLAYWRIGHT:
                ig_posts, ig_foll = get_instagram_stats_pw(result["instagram_url"], pw_page)
            else:
                ig_posts, ig_foll = get_instagram_stats_requests(result["instagram_url"])
        result["instagram_posts"]     = ig_posts or "Not Found"
        result["instagram_followers"] = ig_foll  or "Not Found"

    if result["tiktok_url"]:
        log.info("   Fetching TikTok stats…")
        tt_videos, tt_foll = get_tiktok_stats(result["tiktok_url"])
        result["tiktok_posts"]     = tt_videos or "Not Found"
        result["tiktok_followers"] = tt_foll   or "Not Found"

    if result["linkedin_url"]:
        log.info("   LinkedIn found — manual verification needed")
        # linkedin_posts intentionally not captured
        result["linkedin_followers"] = "See Profile"

    # ── 2b. Facebook URL fallback — search Google if not found on website ────────
    if not result["facebook_url"] and pw_page and USE_PLAYWRIGHT:
        log.info("   Facebook not found on site — searching Google…")
        fb_found = find_facebook_url_via_search(name, city, state, pw_page)
        if fb_found:
            result["facebook_url"] = fb_found
            # Fetch followers for the newly found FB URL (posts not captured)
            _p, f = get_facebook_stats_pw(fb_found, pw_page)
            result["facebook_followers"] = f

    # ── 2c. Email fallback — Facebook About page ──────────────────────────────
    if result["email"] == "Not Found" and result["facebook_url"] and pw_page:
        log.info("   Trying Facebook About page for email…")
        found = find_email_from_fb_about(result["facebook_url"], pw_page)
        if found:
            result["email"] = found

    # ── 3. Invisalign Tier ────────────────────────────────────────────────────
    if result["invisalign"] and int(result["invisalign"]) > 0:
        log.info("   Looking up Invisalign tier…")
        # Try REST API first (fast, no browser required)
        tier = _invisalign_api_fallback(name, zip_c)
        # If API returned a "Check:" URL (not a real tier), try Playwright
        if tier.startswith("Check:") and pw_page and USE_PLAYWRIGHT:
            tier = get_invisalign_tier_pw(name, city, state, zip_c, pw_page)
        result["invisalign_tier"] = tier
    else:
        result["invisalign_tier"] = "N/A – Not Offered"

    # ── 4. Google rating ──────────────────────────────────────────────────────
    log.info("   Fetching Google rating…")
    result["google_rating"], result["google_reviews"] = google_search_rating(
        name, city, state
    )

    # ── Yelp — DISABLED (Yelp blocks automated requests) ─────────────────────
    # log.info("   Fetching Yelp rating…")
    # result["yelp_rating"], result["yelp_reviews"] = yelp_search_rating(
    #     name, city, state
    # )

    log.info(f"   ✓ Done: {name}\n")
    return result


# ─────────────────────────────────────────────────────────────────────────────
# READ INPUT EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def read_practices(filepath, start_idx=0, end_idx=None):
    """
    Read practices from Excel.
    start_idx / end_idx are 0-based indices into the full practice list.
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # Auto-detect header row: "6000 Data COMPLETE" has headers on row 1,
    # older sample files have headers on row 2.
    first_cell = str(ws.cell(1, 1).value or "").strip().lower()
    if first_cell in ("id", "index", "#"):
        hdr_row  = 1
        data_row = 2
    else:
        hdr_row  = 2
        data_row = 3

    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[hdr_row]]
    col_map = {h: i for i, h in enumerate(headers)}

    practices = []
    for row in ws.iter_rows(min_row=data_row, values_only=True):
        if row[0] is None:
            continue
        # "Office Name" in 6000-data file, "Practice Name" in older files
        name_col = col_map.get("Office Name", col_map.get("Practice Name", 2))
        practice = {
            "Index":         row[col_map.get("ID",           col_map.get("Index", 0))],
            "Practice Name": row[name_col],
            "Doctor Name":   row[col_map.get("Doctor Name",  10)],
            "Street":        row[col_map.get("Street",       col_map.get("Address", 3))],
            "City":          row[col_map.get("City",         4)],
            "State":         row[col_map.get("State",        5)],
            "Zip":           row[col_map.get("Zip",          6)],
            "Website":       row[col_map.get("Website",      col_map.get("Practice Website", 9))],
        }
        practices.append(practice)

    total = len(practices)
    sliced = practices[start_idx:end_idx]
    log.info(
        f"Total practices in file: {total} | "
        f"Scraping rows {start_idx + 1}–{(end_idx or total)}: {len(sliced)} practices"
    )
    return sliced


# ─────────────────────────────────────────────────────────────────────────────
# WRITE OUTPUT EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def write_output(practices_data, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Scraped Data"

    # ── Styles ────────────────────────────────────────────────────────────────
    hdr_font  = Font(name="Arial", bold=True, size=9)
    grp_font  = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    data_font = Font(name="Arial", size=9)
    thin      = Side(style="thin", color="CCCCCC")
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft       = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    fills = {
        "blue":      PatternFill("solid", fgColor="1F4E79"),
        "lt_blue":   PatternFill("solid", fgColor="BDD7EE"),
        "green":     PatternFill("solid", fgColor="375623"),
        "lt_green":  PatternFill("solid", fgColor="C6EFCE"),
        "purple":    PatternFill("solid", fgColor="7030A0"),
        "lt_purple": PatternFill("solid", fgColor="E2CFEC"),
        "orange":    PatternFill("solid", fgColor="833C00"),
        "lt_orange": PatternFill("solid", fgColor="FCE4D6"),
        "grey":      PatternFill("solid", fgColor="595959"),
        "lt_grey":   PatternFill("solid", fgColor="EDEDED"),
        "white":     PatternFill("solid", fgColor="FFFFFF"),
        "row_alt":   PatternFill("solid", fgColor="EBF3FB"),
    }

    def sc(cell, val, font=data_font, fill=fills["white"], align=ctr):
        cell.value      = val
        cell.font       = font
        cell.fill       = fill
        cell.alignment  = align
        cell.border     = bdr

    # ── Group header row (row 1) ───────────────────────────────────────────
    # Columns:
    #  1-10  : Practice Information
    # 11-22  : Social Media
    # 23-27  : Technology in Practice
    # 28-39  : Services (# of Mentions)  — now includes Dental Plan col 37
    # 40-44  : Doctor Data & Reviews
    groups = [
        (1,  10, "Practice Information",    "blue"),
        (11, 22, "Social Media",            "green"),
        (23, 27, "Technology in Practice",  "purple"),
        (28, 39, "Services (# of Mentions)","orange"),
        (40, 46, "Doctor Data & Reviews",   "grey"),
    ]
    for start, end, label, color in groups:
        ws.merge_cells(
            start_row=1, start_column=start,
            end_row=1,   end_column=end
        )
        sc(ws.cell(1, start), label, font=grp_font, fill=fills[color], align=ctr)

    # ── Column headers (row 2) ─────────────────────────────────────────────
    col_headers = [
        # Practice Information (1-10)
        "Index", "Practice Name", "Doctor Name", "Address", "City",
        "State", "Zip", "Practice Website", "Practice Email",
        "# of Hygienists",
        # Social Media (11-22)
        "Facebook URL",   "FB # Posts",  "FB Followers",
        "Instagram URL",  "IG # Posts",  "IG Followers",
        "TikTok URL",     "TT # Posts",  "TT Followers",
        "LinkedIn URL",   "LI # Posts",  "LI Followers",
        # Technology (23-27)
        "CEREC (Same Day Crowns)",
        "CBCT (3D Imaging)",
        "Lasers",
        "AI",
        "Intraoral Scanners",
        # Services (28-39)
        "Invisalign (Mentions)", "Invisalign Tier (check manually)",
        "Clear Aligners", "Veneers", "Implants",
        "Smile Makeovers", "Teeth Whitening", "Sedation Dentistry",
        "Holistic Dentistry", "Dental Plan (Membership Plan)", "Cancer Screening",
        "# of Locations",
        # Doctor Data & Reviews (40-44)
        "Associations / Memberships",
        "Doctor Specialty",
        "Google Reviews Ranking", "Total # of Google Reviews",
        "Yelp Rating", "Total # of Yelp Reviews",
        "Testimonials (Number of)",
    ]

    fill_col = {}
    for c in range(1,  11): fill_col[c] = "lt_blue"
    for c in range(11, 23): fill_col[c] = "lt_green"
    for c in range(23, 28): fill_col[c] = "lt_purple"
    for c in range(28, 40): fill_col[c] = "lt_orange"
    for c in range(40, 47): fill_col[c] = "lt_grey"

    for col, hdr in enumerate(col_headers, 1):
        cell = ws.cell(2, col)
        sc(cell, hdr,
           font=hdr_font,
           fill=fills[fill_col.get(col, "white")],
           align=ctr)

    # ── Data x — one row per doctor ────────────────────────────────────
    r_idx = 3
    for (inp, s) in practices_data:
        # Build doctor list: prefer per-doctor list; fall back to single row
        doctors = list(s.get("doctors") or [])
        if not doctors:
            fallback_name = (
                s.get("scraped_doctor_names")
                or inp.get("Doctor Name")
                or "Not Found"
            )
            if fallback_name in ("Not Found", "", None):
                fallback_name = inp.get("Doctor Name") or "Not Found"
            doctors = [{
                "name":         fallback_name,
                "specialty":    "",
                "associations": "",
            }]

        # Ensure the practice owner (from "Practice Name" column) appears in
        # the doctor list IF the column value looks like a personal name.
        # "Last, First" format → convert to "First Last".
        # Guard: skip company/practice names (3+ real words without credentials,
        # e.g. "Secure Dental East Peoria") — only personal names are inserted.
        raw_owner = str(inp.get("Practice Name") or "").strip()
        if raw_owner:
            owner_clean = _CRED_RE.sub("", raw_owner).strip(" ,.-")
            # "Last, First" → "First Last"
            if "," in owner_clean:
                parts = [p.strip() for p in owner_clean.split(",", 1)]
                owner_name = f"{parts[1]} {parts[0]}".strip() if parts[1] else parts[0]
            else:
                owner_name = owner_clean
            # Guard: only treat as a personal name if:
            # - original value had "Last, First" comma format, OR
            # - exactly 2 words remain and none is a dental/business term
            _BIZ_WORDS = frozenset({
                "dental", "dentistry", "dentist", "care", "group", "center",
                "centre", "clinic", "practice", "office", "associates",
                "health", "wellness", "smile", "smiles", "studio", "family",
                "general", "orthodontic", "implant", "cosmetic", "pediatric",
            })
            _own_words = [w for w in owner_name.split()
                          if len(w.strip('.,')) > 1 and not re.match(r'^[A-Z]\.$', w)]
            _has_comma_format = "," in _CRED_RE.sub("", raw_owner)
            _no_biz_words = not any(w.lower() in _BIZ_WORDS for w in _own_words)
            _is_person = (
                _is_valid_doctor_name(owner_name)
                and _no_biz_words
                and 2 <= len(_own_words) <= 3
                and (_has_comma_format or len(_own_words) == 2)
            )
            if _is_person:
                owner_key = _normalize_name_for_dedup(owner_name)
                already_present = any(
                    _normalize_name_for_dedup(d.get("name", "")) == owner_key
                    for d in doctors
                )
                if not already_present:
                    doctors.insert(0, {
                        "name":         owner_name,
                        "specialty":    "",
                        "associations": "",
                    })

        for doc in doctors:
            rf = fills["row_alt"] if r_idx % 2 == 0 else fills["white"]

            row_vals = [
                # Practice info
                inp.get("Index"),
                inp.get("Practice Name"),
                doc["name"],
                inp.get("Street", ""),
                inp.get("City", ""),
                inp.get("State", ""),
                inp.get("Zip", ""),
                inp.get("Website", ""),
                s["email"],
                s["hygienists"],
                # Social
                s["facebook_url"],   s["facebook_posts"],   s["facebook_followers"],
                s["instagram_url"],  s["instagram_posts"],  s["instagram_followers"],
                s["tiktok_url"],     s["tiktok_posts"],     s["tiktok_followers"],
                s["linkedin_url"],   s["linkedin_posts"],   s["linkedin_followers"],
                # Technology
                s["cerec"], s["cbct"], s["lasers"], s["ai"], s["intraoral"],
                # Services
                s["invisalign"],      s["invisalign_tier"],
                s["clear_aligners"],  s["veneers"],
                s["implants"],        s["smile_makeovers"],
                s["whitening"],       s["sedation"],
                s["holistic"],        s["dental_plan"],     s["cancer_screening"],
                s["locations_count"],
                # Doctor data & reviews (per-doctor specialty/associations)
                # Do NOT fall back to s["associations"]/s["specialty"] here —
                # those are practice-level values that would make every doctor
                # appear to have the first doctor's data when no per-doctor bio exists.
                doc["associations"] or "Not Found",
                doc["specialty"] or "Not Found",
                s["google_rating"],  s["google_reviews"],
                s["yelp_rating"],    s["yelp_reviews"],
                s["testimonials"],
            ]

            left_cols = {2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 14, 17, 20, 41, 42}
            for c_idx, val in enumerate(row_vals, 1):
                cell  = ws.cell(r_idx, c_idx)
                align = lft if c_idx in left_cols else ctr
                sc(cell, val, font=data_font, fill=rf, align=align)

            r_idx += 1

    # ── Column widths ──────────────────────────────────────────────────────
    widths = {
        1: 6,  2: 28, 3: 30, 4: 30, 5: 14, 6: 7,  7: 8,  8: 32, 9: 28, 10: 12,
        11: 30, 12: 10, 13: 12,
        14: 30, 15: 10, 16: 12,
        17: 18, 18: 10, 19: 12,
        20: 30, 21: 10, 22: 12,
        23: 18, 24: 16, 25: 10, 26: 8,  27: 16,
        28: 14, 29: 26, 30: 14, 31: 10, 32: 10,
        33: 14, 34: 14, 35: 16, 36: 14, 37: 24, 38: 14, 39: 10,
        40: 38, 41: 30, 42: 14, 43: 14, 44: 12, 45: 18, 46: 14,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 40
    ws.freeze_panes = "B3"

    # ── Legend sheet ───────────────────────────────────────────────────────
    ns = wb.create_sheet("Legend & Notes")
    notes = [
        ("LEGEND", ""),
        ("X in Technology columns",    "Confirmed on practice website"),
        ("0 in Services columns",      "Keyword not found (may still be offered)"),
        ("Not Found",                  "Could not be extracted automatically"),
        ("Blocked",                    "Platform blocked automated access — check manually"),
        ("See Page / See Profile",     "Platform requires manual review"),
        ("N/A",                        "Not applicable for this practice"),
        ("", ""),
        ("DATA SOURCES", ""),
        ("Practice website",           "Homepage + services/about/team/contact sub-pages"),
        ("Email",                      "mailto links → plain text → Playwright contact page"),
        ("Social URLs",                "From anchor tags on practice website"),
        ("FB/IG Stats",                "Playwright browser scrape (public pages)"),
        ("Invisalign Tier",            "Playwright scrape of invisalign.com/find-a-doctor"),
        ("Google Rating",              "Google search result snippet"),
        ("Yelp Rating",                "Yelp search result page"),
        ("Testimonials",               "Count of review/testimonial blocks on website"),
        ("", ""),
        ("ROW RANGE", ""),
        ("START_IDX / END_IDX",        "Edit these values in the script to select which rows to scrape"),
        ("", ""),
        ("IMPORTANT NOTES", ""),
        ("Accuracy",                   "Service counts = exact keyword hits. A service may be offered but not mentioned."),
        ("Google / Yelp ratings",      "Scraped from search snippets — verify on platform directly"),
        ("Facebook post count",        "Counts visible articles on public page — login may reveal more"),
        ("Invisalign Tier",            "If not found: check https://www.invisalign.com/find-a-doctor"),
    ]
    ns.column_dimensions["A"].width = 34
    ns.column_dimensions["B"].width = 72
    bold_rows = {"LEGEND", "DATA SOURCES", "ROW RANGE", "IMPORTANT NOTES"}
    for r_i, (a, b) in enumerate(notes, 1):
        ca = ns.cell(r_i, 1, a)
        cb = ns.cell(r_i, 2, b)
        is_bold = a in bold_rows
        for cell in (ca, cb):
            cell.font      = Font(name="Arial", bold=is_bold, size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if is_bold and a:
                cell.fill = PatternFill("solid", fgColor="1F4E79")
                cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        ns.row_dimensions[r_i].height = 22 if not b else 28

    wb.save(output_path)
    log.info(f"\n✅ Output saved → {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# SKIPPED SITES OUTPUT
# ─────────────────────────────────────────────────────────────────────────────

def write_skipped_sites(skipped: list, out_dir: str = SKIPPED_DIR):
    """
    Write an xlsx listing every practice whose website was blocked or unreachable.
    skipped = list of (inp_dict, scraped_dict) tuples where scraped["skip_reason"] != "".
    """
    if not skipped:
        log.info("  No skipped sites to write.")
        return

    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "skipped_sites.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Skipped Sites"

    hdr_font  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    ctr       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft       = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    data_font = Font(name="Arial", size=10)

    headers = ["#", "Practice Name", "Website", "City", "State", "Zip", "Skip Reason"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = ctr

    alt_fill = PatternFill("solid", fgColor="DCE6F1")
    for r_idx, (inp, s) in enumerate(skipped, 2):
        row_fill = alt_fill if r_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        vals = [
            inp.get("Index", ""),
            inp.get("Practice Name", ""),
            inp.get("Website", ""),
            inp.get("City", ""),
            inp.get("State", ""),
            inp.get("Zip", ""),
            s.get("skip_reason", ""),
        ]
        for c_idx, val in enumerate(vals, 1):
            cell           = ws.cell(r_idx, c_idx, val)
            cell.font      = data_font
            cell.fill      = row_fill
            cell.alignment = lft if c_idx in (2, 3, 7) else ctr

    col_widths = {1: 6, 2: 30, 3: 38, 4: 16, 5: 7, 6: 8, 7: 40}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    wb.save(out_path)
    log.info(f"  Skipped sites saved → {out_path}  ({len(skipped)} practices)")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

EMPTY_SCRAPED = {
    "skip_reason": "", "email": "ERROR", "scraped_doctor_names": "ERROR",
    "hygienists": "ERROR", "locations_count": "ERROR",
    "facebook_url": "", "facebook_posts": "ERROR", "facebook_followers": "ERROR",
    "instagram_url": "", "instagram_posts": "ERROR", "instagram_followers": "ERROR",
    "tiktok_url": "", "tiktok_posts": "ERROR", "tiktok_followers": "ERROR",
    "linkedin_url": "", "linkedin_posts": "ERROR", "linkedin_followers": "ERROR",
    "cerec": "", "cbct": "", "lasers": "", "ai": "", "intraoral": "",
    "invisalign": 0, "invisalign_tier": "N/A",
    "clear_aligners": 0, "veneers": 0, "implants": 0,
    "smile_makeovers": 0, "whitening": 0, "sedation": 0,
    "holistic": 0, "cancer_screening": 0,
    "doctors": [], "associations": "ERROR", "specialty": "ERROR",
    "google_rating": "ERROR", "google_reviews": "ERROR",
    "yelp_rating": "ERROR", "yelp_reviews": "ERROR",
    "testimonials": "ERROR",
}


def main():
    log.info("=" * 60)
    log.info("  Dental Practice Scraper v2")
    log.info(f"  Rows: {START_IDX + 1} – {END_IDX or 'end'}")
    log.info(f"  Playwright: {'ON' if USE_PLAYWRIGHT and PLAYWRIGHT_AVAILABLE else 'OFF'}")
    log.info("=" * 60)

    try:
        practices = read_practices(INPUT_FILE, start_idx=START_IDX, end_idx=END_IDX)
    except FileNotFoundError:
        log.error(f"Input file not found: {INPUT_FILE}")
        log.error("Place this script in the same folder as the Excel file and retry.")
        sys.exit(1)

    if not practices:
        log.warning("No practices found for the given row range. Check START_IDX / END_IDX.")
        sys.exit(0)

    all_results = []

    # Open one Playwright browser for all practices (reuse = faster + less detection)
    pw_context = None
    pw_page    = None

    if USE_PLAYWRIGHT and PLAYWRIGHT_AVAILABLE:
        log.info("  Launching Playwright browser (Chromium headless)…")
        _pw  = sync_playwright().__enter__()
        pw_context = _pw.chromium.launch_persistent_context(
            user_data_dir="",                # ephemeral profile
            headless=True,
            ignore_https_errors=True,        # handles SSL cert mismatches
            args=["--disable-blink-features=AutomationControlled"],
            user_agent=HEADERS["User-Agent"],
        )
        pw_page = pw_context.new_page()
        pw_page.set_extra_http_headers({"Accept-Language": "en-US,en;q=0.9"})

    try:
        for i, practice in enumerate(practices, 1):
            log.info(f"[{i}/{len(practices)}] {practice.get('Practice Name')}")

            # Resume: load from cache if already scraped in a previous run
            folder = _cache_dir(practice.get("Index", i), practice.get("Practice Name", ""))
            cached = _load_cache_result(folder)
            if cached is not None:
                log.info(f"  ✓ Loaded from cache — skipping re-scrape")
                all_results.append((practice, cached))
                if i % 10 == 0:
                    write_output(all_results, OUTPUT_FILE)
                continue

            try:
                scraped = scrape_practice(practice, pw_page=pw_page)
            except Exception as e:
                log.error(f"  ERROR: {practice.get('Practice Name')}: {e}", exc_info=True)
                scraped = dict(EMPTY_SCRAPED)
                # Recover Playwright if its page was closed/crashed by the error
                if pw_context:
                    try:
                        if pw_page is None or pw_page.is_closed():
                            log.warning("  Playwright page died — reopening…")
                            pw_page = pw_context.new_page()
                            pw_page.set_extra_http_headers({"Accept-Language": "en-US,en;q=0.9"})
                    except Exception:
                        try:
                            log.warning("  Playwright context dead — relaunching browser…")
                            pw_context.close()
                        except Exception:
                            pass
                        _pw2  = sync_playwright().__enter__()
                        pw_context = _pw2.chromium.launch_persistent_context(
                            user_data_dir="", headless=True,
                            ignore_https_errors=True,
                            args=["--disable-blink-features=AutomationControlled"],
                            user_agent=HEADERS["User-Agent"],
                        )
                        pw_page = pw_context.new_page()
                        pw_page.set_extra_http_headers({"Accept-Language": "en-US,en;q=0.9"})

            all_results.append((practice, scraped))

            # Save result.json to cache for this practice
            _cache_result(folder, practice, scraped)

            # Save progress every 10 practices
            if i % 10 == 0:
                log.info(f"  💾 Saving progress after {i} practices…")
                write_output(all_results, OUTPUT_FILE)

    finally:
        if pw_context:
            try:
                pw_context.close()
                _pw.__exit__(None, None, None)
            except Exception:
                pass

    write_output(all_results, OUTPUT_FILE)

    # Write skipped sites (bot-blocked / unreachable) to a separate file
    skipped = [(inp, s) for (inp, s) in all_results if s.get("skip_reason")]
    write_skipped_sites(skipped, SKIPPED_DIR)

    log.info(f"\n🎉 Done!  {len(all_results)} practices scraped.")
    log.info(f"   Output: {OUTPUT_FILE}")
    if skipped:
        skipped_indices = ", ".join(str(inp.get("Index", "?")) for inp, s in skipped)
        log.info(f"   Skipped: {len(skipped)} blocked/unreachable → {SKIPPED_DIR}/skipped_sites.xlsx")
        log.info(f"   Skipped indices: {skipped_indices}")


if __name__ == "__main__":
    main()
