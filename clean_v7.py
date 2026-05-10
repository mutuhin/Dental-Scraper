"""
clean_v7.py
-----------
Cleans Dental_Scrape_Output_v7.xlsx → Dental_Scrape_Output_v7_cleaned.xlsx

Fixes applied:
  1. Duplicate practice removal — Fariborz/Rodef (same practice, two input rows)
  2. Doctor deduplication within each practice
  3. Noise doctor name filtering
  4. Blank placeholder values in non-critical display columns
"""

import re
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_FILE  = "Dental_Scrape_Output_v12.xlsx"
OUTPUT_FILE = "Dental_Scrape_Output_v13_cleaned.xlsx"

# ── Column map (1-based) ──────────────────────────────────────────────────────
COL = {
    "index": 1, "practice_name": 2, "doctor_name": 3,
    "address": 4, "city": 5, "state": 6, "zip": 7,
    "website": 8, "email": 9, "hygienists": 10,
    "fb_url": 11, "fb_posts": 12, "fb_followers": 13,
    "ig_url": 14, "ig_posts": 15, "ig_followers": 16,
    "tt_url": 17, "tt_posts": 18, "tt_followers": 19,
    "li_url": 20, "li_posts": 21, "li_followers": 22,
    "cerec": 23, "cbct": 24, "lasers": 25, "ai": 26, "intraoral": 27,
    "invisalign": 28, "inv_tier": 29,
    "clear_aligners": 30, "veneers": 31, "implants": 32,
    "smile_make": 33, "whitening": 34, "sedation": 35,
    "holistic": 36, "dental_plan": 37, "cancer": 38,
    "locations": 39,
    "associations": 40, "specialty": 41,
    "google_rating": 42, "google_reviews": 43, "testimonials": 44,
}

# Columns where "Not Found" / similar placeholders should be blanked
PLACEHOLDER_BLANK_COLS = {
    "fb_posts", "fb_followers", "ig_posts", "ig_followers",
    "tt_url", "tt_posts", "tt_followers",
    "li_posts", "li_followers",
    "email", "google_rating", "google_reviews", "associations",
}
PLACEHOLDER_VALUES = {
    "Not Found", "See Website", "See Profile", "Blocked",
    "N/A – Not Offered", "Not Listed", "ERROR", "None", "nan",
}

# Known per-practice doctor aliases to drop (normalised, no Dr./credentials)
KNOWN_DROP = {
    "Moradzadeh, Alexander": {"alex moradzadeh"},
    "Khatchaturian, Victor": {"khachaturian"},           # typo variant
    "Kong, Vu":              {"uhand"},                  # typo of "uhland"
    "Fandel, Jay Edward":    {"mitchell bloom", "glenn ludwig", "azadeh akhavan", "fandel", "jay"},  # wrong site + fragments
    "Yunusov, Regina":       {"jenna jesko"},            # not on cedarparkdentalwellness.com
    "Thind, Sukhjinder":     {"received her dds"},       # sentence fragment
    "Paul Decker DDS PLLC":  {"and his dmd"},           # sentence fragment
    "Wells, Loren Brett":    {"jd", "jdwas"},            # fragment / noise
    "Troy, Deborah A":       {"debbie"},                 # nickname fragment of Deborah Troy
    "Locke, Virginia":       {"virignia locke", "virgina locke"},       # OCR/typo variants
    "Tchakarova, Ludmila":   {"ludmila tchakarovas", "tchakarovas", "tchakarova ludmila"},  # typos/reversed
    # v11 fixes
    "Premier Dental Of Oakwood":    {"photos"},          # "Photos" link scraped as doctor name
    "Premier Dental Of Johnstown":  {"photos"},          # same site, same issue
    "Wong, Lennie":                 {"earned her", "mk"},# "earned her DMD" fragment + "MK" initials
    "Speed, Joelle Taves":          {"speed doing"},     # "Dr. Speed doing..." sentence fragment
    "Dinning, Ronald":              {"the best"},        # "the best DDS" sentence fragment
}

# Noise name prefixes — names starting with these are structural noise
_NOISE_PREFIXES = {"suite", "dr suite"}


# Noise words — a doctor name containing any of these is not a real name
_NOISE_WORDS = {
    "meet", "today", "recommend", "welcome", "patient",
    "appointment", "insurance", "care", "treatment",
    "contact", "schedule", "location", "existing", "team",
    "difference", "experience", "results", "overview",
    "directions", "learn", "blog", "menu", "home",
    "read", "more", "view", "skip", "once", "providing",
    "hills", "beverly",
    "building", "if", "many", "often", "there", "last",
    "receiving", "received", "comprehensive", "medicine", "sleep", "call", "was",
    "director", "and", "clinical",
    "earned", "photos",
}


# ── Name helpers ──────────────────────────────────────────────────────────────

def _normalize_name(name):
    """Strip Dr./credentials/noise for comparison."""
    if not name or str(name).strip() in ("", "Not Found", "ERROR", "None", "nan"):
        return ""
    n = re.sub(r"^(Drs?\.?\s+)", "", str(name).strip(), flags=re.IGNORECASE)
    n = re.sub(
        r",?\s*(DDS|DMD|MD|MS|M\.D\.|D\.D\.S\.|D\.M\.D\.)\s*.*$",
        "", n, flags=re.IGNORECASE,
    )
    n = re.sub(r"\s+", " ", n).strip().lower()
    return n


def _depos(word):
    """Strip a trailing possessive 's' from a word (e.g. 'nguyens' → 'nguyen')."""
    if len(word) > 3 and word.endswith("s"):
        return word[:-1]
    return word


def _is_same_doctor(a, b):
    """True when two name strings likely refer to the same person."""
    na = _normalize_name(a)
    nb = _normalize_name(b)
    if not na or not nb:
        return False
    if na == nb:
        return True
    if na in nb or nb in na:
        return True
    wa = set(na.split())
    wb = set(nb.split())
    shorter = wa if len(wa) <= len(wb) else wb
    longer  = wb if len(wa) <= len(wb) else wa
    if bool(shorter) and shorter.issubset(longer):
        return True
    # Possessive check — strip trailing 's' from each word and retry
    wa_dep = {_depos(w) for w in wa}
    wb_dep = {_depos(w) for w in wb}
    shorter_dep = wa_dep if len(wa_dep) <= len(wb_dep) else wb_dep
    longer_dep  = wb_dep if len(wa_dep) <= len(wb_dep) else wa_dep
    if bool(shorter_dep) and shorter_dep.issubset(longer_dep):
        return True
    # Prefix/nickname check — handles "Ben" ≡ "Benjamin", strips initials & punctuation
    # Requires ≥2 significant words in the shorter name to avoid false positives
    # (e.g. "mat" from a noise fragment should not match "matthew moadel")
    wa_c = {re.sub(r'[^a-z]', '', w) for w in wa}
    wb_c = {re.sub(r'[^a-z]', '', w) for w in wb}
    wa_c = {w for w in wa_c if len(w) >= 3}   # drop single-letter initials
    wb_c = {w for w in wb_c if len(w) >= 3}
    shorter_c = wa_c if len(wa_c) <= len(wb_c) else wb_c
    longer_c  = wb_c if len(wa_c) <= len(wb_c) else wa_c
    if len(shorter_c) >= 2 and all(
        any(lw.startswith(sw) or sw.startswith(lw) for lw in longer_c)
        for sw in shorter_c
    ):
        return True
    return False


def _is_noise_name(name, city=""):
    """Return True if the name looks like a non-doctor noise string."""
    if not name or len(str(name).strip()) < 3:
        return True
    nl = str(name).lower()
    # Repeated "dr" prefix
    if re.search(r"\bdr\.?\s+dr\.?\b", nl):
        return True
    # Address/suite fragments scraped as names (e.g. "Dr Suite C Elkridge")
    norm_stripped = re.sub(r"^drs?\.?\s+", "", nl).strip()
    first_word = norm_stripped.split()[0] if norm_stripped.split() else ""
    if first_word in _NOISE_PREFIXES or first_word in {"suite", "ste", "floor", "unit", "room"}:
        return True
    # Matches the practice city (city name scraped as a doctor name)
    if city:
        city_norm = city.strip().lower()
        if city_norm and len(city_norm) > 3 and city_norm in nl:
            return True
    # Noise trigger words
    words = set(re.sub(r"[^a-z\s]", " ", nl).split())
    if words & _NOISE_WORDS:
        return True
    # Looks like a practice / business name rather than a person
    if re.search(
        r"\b(dental|dentistry|clinic|center|associates|group|office|pc|pllc|dds)\b",
        nl,
    ):
        # Allow "DDS" as part of a real name suffix (handled by normalize), but
        # a standalone dental-business word means this isn't a person.
        pure_biz = re.sub(
            r"\b(dds|dmd|dr\.?)\b", "", nl, flags=re.IGNORECASE
        ).strip()
        if re.search(
            r"\b(dental|dentistry|clinic|center|associates|group|office)\b",
            pure_biz,
        ):
            return True
    return False


def _norm_practice_key(name):
    """Canonical key that makes 'Fariborz, Rodef' == 'Rodef, Fariborz'."""
    words = re.sub(r"[^a-z\s]", "", str(name).lower()).split()
    return " ".join(sorted(words))


# ── Read ──────────────────────────────────────────────────────────────────────

def read_rows(filepath):
    wb  = openpyxl.load_workbook(filepath)
    ws  = wb.active
    rows = []
    for r in range(3, ws.max_row + 1):
        v = ws.cell(r, COL["practice_name"]).value
        if v is None:
            continue
        row = {}
        for key, col in COL.items():
            cv = ws.cell(r, col).value
            row[key] = str(cv).strip() if cv is not None else ""
        rows.append(row)
    print(f"Loaded {len(rows)} rows from {filepath}")
    return rows


# ── Clean ─────────────────────────────────────────────────────────────────────

def clean_rows(rows):
    # ── 1. Group consecutive rows by practice name ────────────────────────────
    groups: list[list[dict]] = []
    seen_prac_keys: dict[str, int] = {}   # norm_key → first group index

    i = 0
    while i < len(rows):
        prac_name = rows[i]["practice_name"]
        group: list[dict] = []
        while i < len(rows) and rows[i]["practice_name"] == prac_name:
            group.append(rows[i])
            i += 1

        key = _norm_practice_key(prac_name)
        if key in seen_prac_keys:
            first_name = groups[seen_prac_keys[key]][0]["practice_name"]
            print(f"  [DUP PRACTICE] Skipped '{prac_name}' "
                  f"(duplicate of '{first_name}')")
            continue

        seen_prac_keys[key] = len(groups)
        groups.append(group)

    # ── 2. Per-group: dedup + noise filter ────────────────────────────────────
    result: list[dict] = []
    for group in groups:
        prac_name = group[0]["practice_name"]
        drop_norms = KNOWN_DROP.get(prac_name, set())

        kept: list[dict] = []
        kept_norms: list[str] = []

        # Sort by name completeness (most words first) so full names like
        # "Dr. Molly Gunsaulis" are processed before fragments "Dr. Molly"
        group_sorted = sorted(
            group,
            key=lambda r: len(_normalize_name(r["doctor_name"]).split()),
            reverse=True,
        )

        for row in group_sorted:
            dname = row["doctor_name"]
            dnorm = _normalize_name(dname)

            if dnorm in drop_norms:
                print(f"  [KNOWN DROP] '{dname}' for {prac_name}")
                continue

            city = group[0].get("city", "")
            if _is_noise_name(dname, city=city):
                print(f"  [NOISE]      '{dname}' for {prac_name}")
                continue

            is_dup = any(_is_same_doctor(dname, kn) for kn in kept_norms)
            if is_dup:
                print(f"  [DUP DOC]    '{dname}' for {prac_name}")
                continue

            kept_norms.append(dnorm or dname.lower())
            kept.append(row)

        # ── Reconstruct full name when practice = "LastName, FirstName" and name fragments exist
        # e.g. practice "Gunsaulis, Molly" with "Dr. Molly" + "Dr. Gunsaulis" alongside "Dr. David C. Page"
        # → merge only the fragments into "Dr. Molly Gunsaulis", keep other doctors
        if re.match(r'^[A-Z][a-z]+,\s+[A-Z][a-z]+$', prac_name):
            prac_parts = re.split(r',\s*', prac_name)  # ["LastName", "FirstName"]
            if len(prac_parts) == 2:
                last_name, first_name = prac_parts[0], prac_parts[1]
                prac_words = {last_name.lower(), first_name.lower()}
                frag_rows = [
                    r for r in kept
                    if len(_normalize_name(r["doctor_name"]).split()) == 1
                    and _normalize_name(r["doctor_name"]) in prac_words
                ]
                non_frag_rows = [r for r in kept if r not in frag_rows]
                if len(frag_rows) >= 2:
                    full_name = f"Dr. {first_name} {last_name}"
                    print(f"  [RECONSTRUCT] '{prac_name}' fragments → '{full_name}'")
                    merged = dict(frag_rows[0])
                    merged["doctor_name"] = full_name
                    kept = [merged] + non_frag_rows

        # If everything was filtered, keep the first row unchanged
        if not kept:
            kept = [group[0]]

        result.extend(kept)

    # ── 3. Blank placeholder values ───────────────────────────────────────────
    for row in result:
        for col_name in PLACEHOLDER_BLANK_COLS:
            if row.get(col_name, "") in PLACEHOLDER_VALUES:
                row[col_name] = ""

    print(f"After cleaning: {len(result)} rows")
    return result


# ── Write Excel ───────────────────────────────────────────────────────────────

def write_clean(rows, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Scraped Data"

    hdr_font  = Font(name="Arial", bold=True, size=9)
    grp_font  = Font(name="Arial", bold=True, size=9, color="FFFFFF")
    data_font = Font(name="Arial", size=9)
    thin  = Side(style="thin", color="CCCCCC")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    fills = {
        "blue":      PatternFill("solid", fgColor="1F4E79"),
        "lt_blue":   PatternFill("solid", fgColor="BDD7EE"),
        "green":     PatternFill("solid", fgColor="375623"),
        "lt_green":  PatternFill("solid", fgColor="C6EFCE"),
        "purple":    PatternFill("solid", fgColor="7030A0"),
        "lt_purple": PatternFill("solid", fgColor="E2CFEC"),
        "orange":    PatternFill("solid", fgColor="833C00"),
        "lt_orange": PatternFill("solid", fgColor="FCE4D6"),
        "grey":      PatternFill("solid", fgColor="595959"),
        "lt_grey":   PatternFill("solid", fgColor="EDEDED"),
        "white":     PatternFill("solid", fgColor="FFFFFF"),
        "row_alt":   PatternFill("solid", fgColor="EBF3FB"),
    }

    def sc(cell, val, font=data_font, fill=fills["white"], align=ctr):
        cell.value     = val
        cell.font      = font
        cell.fill      = fill
        cell.alignment = align
        cell.border    = bdr

    # ── Group header row ──────────────────────────────────────────────────────
    groups_hdr = [
        (1,  10, "Practice Information",     "blue"),
        (11, 22, "Social Media",             "green"),
        (23, 27, "Technology in Practice",   "purple"),
        (28, 39, "Services (# of Mentions)", "orange"),
        (40, 44, "Doctor Data & Reviews",    "grey"),
    ]
    for start, end, label, color in groups_hdr:
        ws.merge_cells(start_row=1, start_column=start,
                       end_row=1,   end_column=end)
        sc(ws.cell(1, start), label,
           font=grp_font, fill=fills[color], align=ctr)

    # ── Column headers ────────────────────────────────────────────────────────
    col_headers = [
        # Practice Information (1-10)
        "Index", "Practice Name", "Doctor Name", "Address", "City",
        "State", "Zip", "Practice Website", "Practice Email",
        "# of Hygienists",
        # Social Media (11-22)
        "Facebook URL",   "FB # Posts",  "FB Followers",
        "Instagram URL",  "IG # Posts",  "IG Followers",
        "TikTok URL",     "TT # Posts",  "TT Followers",
        "LinkedIn URL",   "LI # Posts",  "LI Followers",
        # Technology (23-27)
        "CEREC (Same Day Crowns)", "CBCT (3D Imaging)",
        "Lasers", "AI", "Intraoral Scanners",
        # Services (28-39)
        "Invisalign (Mentions)", "Invisalign Tier (check manually)",
        "Clear Aligners", "Veneers", "Implants",
        "Smile Makeovers", "Teeth Whitening", "Sedation Dentistry",
        "Holistic Dentistry", "Dental Plan (Membership Plan)",
        "Cancer Screening", "# of Locations",
        # Doctor Data & Reviews (40-44)
        "Associations / Memberships", "Doctor Specialty",
        "Google Reviews Ranking", "Total # of Google Reviews",
        "Testimonials (Number of)",
    ]

    fill_col = {}
    for c in range(1,  11): fill_col[c] = "lt_blue"
    for c in range(11, 23): fill_col[c] = "lt_green"
    for c in range(23, 28): fill_col[c] = "lt_purple"
    for c in range(28, 40): fill_col[c] = "lt_orange"
    for c in range(40, 45): fill_col[c] = "lt_grey"

    for col, hdr in enumerate(col_headers, 1):
        sc(ws.cell(2, col), hdr,
           font=hdr_font, fill=fills[fill_col.get(col, "white")], align=ctr)

    # ── Data rows ─────────────────────────────────────────────────────────────
    left_cols = {2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 14, 17, 20, 41, 42}

    def _v(row, key):
        v = row.get(key, "")
        return "" if v in ("None", "nan") else v

    for r_idx, row in enumerate(rows, start=3):
        rf = fills["row_alt"] if r_idx % 2 == 0 else fills["white"]

        row_vals = [
            _v(row, "index"),       _v(row, "practice_name"),  _v(row, "doctor_name"),
            _v(row, "address"),     _v(row, "city"),            _v(row, "state"),
            _v(row, "zip"),         _v(row, "website"),         _v(row, "email"),
            _v(row, "hygienists"),
            # Social
            _v(row, "fb_url"),      _v(row, "fb_posts"),        _v(row, "fb_followers"),
            _v(row, "ig_url"),      _v(row, "ig_posts"),        _v(row, "ig_followers"),
            _v(row, "tt_url"),      _v(row, "tt_posts"),        _v(row, "tt_followers"),
            _v(row, "li_url"),      _v(row, "li_posts"),        _v(row, "li_followers"),
            # Technology
            _v(row, "cerec"),       _v(row, "cbct"),            _v(row, "lasers"),
            _v(row, "ai"),          _v(row, "intraoral"),
            # Services
            _v(row, "invisalign"),  _v(row, "inv_tier"),
            _v(row, "clear_aligners"), _v(row, "veneers"),      _v(row, "implants"),
            _v(row, "smile_make"),  _v(row, "whitening"),       _v(row, "sedation"),
            _v(row, "holistic"),    _v(row, "dental_plan"),     _v(row, "cancer"),
            _v(row, "locations"),
            # Doctor data
            _v(row, "associations"), _v(row, "specialty"),
            _v(row, "google_rating"), _v(row, "google_reviews"),
            _v(row, "testimonials"),
        ]

        for c_idx, val in enumerate(row_vals, 1):
            cell  = ws.cell(r_idx, c_idx)
            align = lft if c_idx in left_cols else ctr
            sc(cell, val, font=data_font, fill=rf, align=align)

    # ── Column widths ─────────────────────────────────────────────────────────
    widths = {
        1: 6,  2: 28, 3: 30, 4: 30, 5: 14, 6: 7,  7: 8,  8: 32, 9: 28, 10: 12,
        11: 30, 12: 10, 13: 12,
        14: 30, 15: 10, 16: 12,
        17: 18, 18: 10, 19: 12,
        20: 30, 21: 10, 22: 12,
        23: 18, 24: 16, 25: 10, 26: 8, 27: 16,
        28: 14, 29: 26, 30: 14, 31: 10, 32: 10,
        33: 14, 34: 14, 35: 16, 36: 14, 37: 24, 38: 14, 39: 10,
        40: 38, 41: 30, 42: 14, 43: 14, 44: 14,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 40
    ws.freeze_panes = "B3"

    wb.save(path)
    print(f"✅ Saved → {path}  ({len(rows)} rows)")


# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    rows    = read_rows(INPUT_FILE)
    cleaned = clean_rows(rows)
    write_clean(cleaned, OUTPUT_FILE)
