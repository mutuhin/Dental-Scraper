#!/usr/bin/env python3
"""
bypass_scraper.py
=================
Re-scrape dental practices that were blocked by bot protection in the main scraper.

BYPASS STRATEGIES (tried in order per request):
  1. curl_cffi — rotates through 6 Chrome/Safari TLS fingerprint profiles
  2. curl_cffi + proxy rotation — same profiles but through a proxy
  3. Playwright + stealth JS — hides automation markers, no proxy
  4. Playwright + stealth JS + proxy — last resort

USAGE
-----
  # Re-scrape specific URLs:
  python bypass_scraper.py --urls https://site1.com https://site2.com

  # Auto-detect bot-blocked rows in a batch output Excel and re-scrape:
  python bypass_scraper.py --batch batch_05_deduped.xlsx

  # With a proxy list:
  python bypass_scraper.py --batch batch_05_deduped.xlsx --proxies proxies.txt

  # Custom output file:
  python bypass_scraper.py --batch batch_05_deduped.xlsx --output retried.xlsx

PROXIES FILE (proxies.txt) — one per line:
  http://host:port
  http://user:pass@host:port
  socks5://host:port

INSTALL
-------
  pip install curl_cffi playwright openpyxl requests beautifulsoup4 lxml
  playwright install chromium
"""

import argparse
import random
import sys
import time
import re
import os
from urllib.parse import urlparse, urljoin

import warnings
warnings.filterwarnings("ignore")

# ── Dependencies ──────────────────────────────────────────────────────────────

try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl")

try:
    import requests as std_requests
except ImportError:
    sys.exit("pip install requests")

try:
    from curl_cffi import requests as cffi_requests
    _CFFI_OK = True
except ImportError:
    _CFFI_OK = False
    print("WARNING: curl_cffi not installed — proxy+TLS bypass disabled")
    print("         pip install curl_cffi")

try:
    from playwright.sync_api import sync_playwright
    _PW_OK = True
except ImportError:
    _PW_OK = False
    print("WARNING: playwright not installed — JS-site bypass disabled")
    print("         pip install playwright && playwright install chromium")

try:
    # playwright-stealth v1.x API
    from playwright_stealth import stealth_sync as _stealth_sync
    _STEALTH_LIB_OK = True
    print("playwright-stealth: v1 API (stealth_sync)")
except ImportError:
    try:
        # playwright-stealth v2.x renamed the function
        from playwright_stealth import Stealth as _StealthClass
        _stealth_sync = _StealthClass().apply_stealth_sync
        _STEALTH_LIB_OK = True
        print("playwright-stealth: v2 API (Stealth().apply_stealth_sync)")
    except (ImportError, AttributeError) as _e:
        _STEALTH_LIB_OK = False   # falls back to manual stealth JS below
        print(f"playwright-stealth not available ({_e}) — using manual stealth JS")

# ── Import dental_scraper functions ──────────────────────────────────────────

try:
    import dental_scraper as ds
except ImportError:
    sys.exit("dental_scraper.py must be in the same folder")

# Tell dental_scraper to allow Playwright on CF-gated sites — bypass scraper's
# stealth+proxy Playwright can actually solve challenges, unlike the plain scraper.
ds.BYPASS_MODE = True

# On CI, dental_scraper uses very short Playwright waits (1 s) and timeouts (15 s)
# which aren't enough for JS-heavy / CF-protected sites.  Override them here so
# the bypass scraper always uses the longer "local" timings regardless of CI env.
ds.IS_CI                   = False
ds.PW_TIMEOUT              = 30000   # 30 s per page.goto (was 15 s in CI)
ds._PROXY_CAP_PER_PRACTICE = 40      # bypass sites need more proxy calls (was 15)

# ── Stealth JS (injected into every Playwright page) ─────────────────────────

_STEALTH_JS = """
(function(){
  Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
  window.chrome = {
    app:{ isInstalled:false, getDetails:function(){}, getIsInstalled:function(){}, runningState:function(){} },
    csi:function(){}, loadTimes:function(){},
    runtime:{ OnInstalledReason:{}, PlatformArch:{}, PlatformNaclArch:{}, PlatformOs:{}, RequestUpdateCheckStatus:{} }
  };
  const fp=[
    {name:'Chrome PDF Plugin',  filename:'internal-pdf-viewer',              description:'Portable Document Format'},
    {name:'Chrome PDF Viewer',  filename:'mhjfbmdgcfjbbpaeojofohoefgiehjai', description:''},
    {name:'Native Client',      filename:'internal-nacl-plugin',              description:''},
  ];
  Object.defineProperty(navigator,'plugins',{get:()=>{
    const a=fp.map(p=>Object.assign(Object.create(Plugin.prototype),p));
    Object.setPrototypeOf(a,PluginArray.prototype); return a;
  }});
  Object.defineProperty(navigator,'languages',{get:()=>['en-US','en']});
  try{
    const oq=window.navigator.permissions.query.bind(navigator.permissions);
    window.navigator.permissions.query=(p)=>
      p.name==='notifications'?Promise.resolve({state:Notification.permission}):oq(p);
  }catch(e){}
  try{
    const gp=WebGLRenderingContext.prototype.getParameter;
    WebGLRenderingContext.prototype.getParameter=function(p){
      if(p===37445)return 'Intel Inc.';
      if(p===37446)return 'Intel Iris OpenGL Engine';
      return gp.call(this,p);
    };
  }catch(e){}
  delete window.cdc_adoQpoasnfa76pfcZLmcfl_Array;
  delete window.cdc_adoQpoasnfa76pfcZLmcfl_Promise;
  delete window.cdc_adoQpoasnfa76pfcZLmcfl_Symbol;
})();
"""

# curl_cffi browser profiles to rotate through
_CFFI_PROFILES = ["chrome136", "chrome124", "chrome133a", "chrome110", "safari260", "safari17_2"]

# Markers that indicate a bot-challenge page was returned (even at HTTP 200)
_CHALLENGE_MARKERS = (
    "sgcaptcha", "robot challenge", "just a moment", "checking your browser",
    "access denied", "please enable cookies", "enable javascript and cookies",
)

# ── Skyvern cloud-browser integration ────────────────────────────────────────
# Skyvern runs real Chrome with UK residential IPs that are not in Cloudflare's
# blocklist.  Set SKYVERN_API_KEY (GitHub Secret) to enable this fallback.
_SKYVERN_API = "https://api.skyvern.com/v1"
_SKYVERN_KEY = os.environ.get("SKYVERN_API_KEY", "")


def _skyvern_session(proxy_location="RESIDENTIAL_GB"):
    """Create a Skyvern browser session; return session_id or None."""
    if not _SKYVERN_KEY:
        return None
    try:
        r = std_requests.post(
            f"{_SKYVERN_API}/browser-sessions",
            json={"proxy_location": proxy_location, "timeout": 10},
            headers={"x-api-key": _SKYVERN_KEY},
            timeout=30,
        )
        r.raise_for_status()
        return r.json().get("session_id")
    except Exception as e:
        print(f"  [Skyvern] session create failed: {e}")
        return None


def _skyvern_close(session_id):
    """Close a Skyvern session (best-effort)."""
    if not session_id:
        return
    try:
        std_requests.delete(
            f"{_SKYVERN_API}/browser-sessions/{session_id}",
            headers={"x-api-key": _SKYVERN_KEY},
            timeout=10,
        )
    except Exception:
        pass


def _skyvern_navigate(session_id, url, timeout_ms=60000):
    """Navigate Skyvern browser to url; return True on success."""
    try:
        r = std_requests.post(
            f"{_SKYVERN_API}/browser-sessions/{session_id}/navigate",
            json={"url": url, "wait_until": "networkidle", "timeout": timeout_ms},
            headers={"x-api-key": _SKYVERN_KEY},
            timeout=timeout_ms / 1000 + 30,
        )
        return r.status_code < 400
    except Exception as e:
        print(f"  [Skyvern] navigate error: {e}")
        return False


def _skyvern_get_text(session_id):
    """Return visible body text from the current Skyvern page, or ''."""
    try:
        r = std_requests.post(
            f"{_SKYVERN_API}/browser-sessions/{session_id}/evaluate",
            json={"expression": "document.body.innerText"},
            headers={"x-api-key": _SKYVERN_KEY},
            timeout=30,
        )
        result = r.json()
        return str(result.get("result", ""))
    except Exception:
        return ""


def _skyvern_get_html(session_id):
    """Return full page HTML from the current Skyvern page, or ''."""
    try:
        r = std_requests.post(
            f"{_SKYVERN_API}/browser-sessions/{session_id}/evaluate",
            json={"expression": "document.documentElement.outerHTML"},
            headers={"x-api-key": _SKYVERN_KEY},
            timeout=30,
        )
        result = r.json()
        return str(result.get("result", ""))
    except Exception:
        return ""


def _skyvern_get_page_links(session_id, base_url):
    """Return same-domain href links found on current Skyvern page."""
    try:
        r = std_requests.post(
            f"{_SKYVERN_API}/browser-sessions/{session_id}/evaluate",
            json={"expression": (
                "JSON.stringify([...new Set("
                "  Array.from(document.querySelectorAll('a[href]'))"
                "  .map(a => a.href)"
                "  .filter(h => h.startsWith('" + base_url + "'))"
                ")])"
            )},
            headers={"x-api-key": _SKYVERN_KEY},
            timeout=20,
        )
        import json as _json
        raw = r.json().get("result", "[]")
        return _json.loads(raw) if raw else []
    except Exception:
        return []


# Sub-page keywords that likely contain doctor / team / service data
_SUB_PAGE_KW = re.compile(
    r"(doctor|team|staff|provider|dentist|about|service|treatment|"
    r"procedure|implant|cosmetic|specialist|meet|our-dentist)",
    re.I,
)


def _skyvern_full_scrape(url):
    """
    Use Skyvern cloud browser (UK residential IP) to:
      1. Load the main URL
      2. Extract ALL visible text + HTML
      3. Find same-domain sub-pages with dental/doctor keywords
      4. Load up to 5 sub-pages and collect their text too
    Returns dict: {main_html, main_text, sub_texts: [(url, text), ...]}
    or None on failure.
    """
    if not _SKYVERN_KEY:
        return None

    print(f"  [Skyvern] Starting cloud browser session (UK residential)…")
    session_id = _skyvern_session()
    if not session_id:
        return None

    try:
        # ── Load main page ────────────────────────────────────────────────────
        print(f"  [Skyvern] Navigating → {url[:80]}")
        ok = _skyvern_navigate(session_id, url, timeout_ms=60000)
        if not ok:
            print(f"  [Skyvern] Navigation failed")
            return None

        main_html = _skyvern_get_html(session_id)
        main_text = _skyvern_get_text(session_id)

        if len(main_text) < 200:
            print(f"  [Skyvern] Page appears empty ({len(main_text)} chars)")
            return None

        print(f"  [Skyvern] Main page: {len(main_text):,} chars of text")

        # ── Discover sub-pages ────────────────────────────────────────────────
        parsed = urlparse(url)
        base = f"{parsed.scheme}://{parsed.netloc}"
        all_links = _skyvern_get_page_links(session_id, base)
        sub_links = [
            lnk for lnk in all_links
            if lnk.rstrip("/") != url.rstrip("/")
            and _SUB_PAGE_KW.search(lnk)
        ][:5]  # cap at 5 sub-pages

        # ── Load sub-pages ────────────────────────────────────────────────────
        sub_texts = []
        for sub_url in sub_links:
            print(f"  [Skyvern] Sub-page → {sub_url[:80]}")
            ok2 = _skyvern_navigate(session_id, sub_url, timeout_ms=30000)
            if not ok2:
                continue
            txt = _skyvern_get_text(session_id)
            if txt and len(txt) > 100:
                sub_texts.append((sub_url, txt))
                print(f"    {len(txt):,} chars")

        return {
            "main_html": main_html,
            "main_text": main_text,
            "sub_texts": sub_texts,
        }

    finally:
        _skyvern_close(session_id)
        print(f"  [Skyvern] Session closed")


def _is_challenge_html(r):
    """Return True if the response is a bot-challenge page regardless of status code."""
    if r.status_code in (202, 403):
        return True
    url_lower = r.url.lower()
    if "sgcaptcha" in url_lower or "captcha" in url_lower or "challenge" in url_lower:
        return True
    snippet = r.text[:2000].lower()
    return any(m in snippet for m in _CHALLENGE_MARKERS)

# SPA shell patterns — React/Vue/Angular pages that render entirely in JS.
# Static requests return a near-empty HTML shell; Playwright must render them.
_SPA_ROOT_RE = re.compile(
    r'<(?:div|main|section)\s[^>]*\bid=["\'](?:root|app|__next|ng-app|react-root|'
    r'ember-application|nuxt|svelte)["\'][^>]*>\s*<\/(?:div|main|section)>',
    re.I,
)

def _is_js_shell(r) -> bool:
    """
    Return True when the HTTP response is a JavaScript SPA shell —
    a 200-OK page whose body has no meaningful text content because all
    rendering happens client-side.  Returning True causes _bypass_safe_get
    to yield None so scrape_practice's Playwright fallback renders the page.
    """
    if len(r.text) > 50_000:        # big pages have real content even in SPAs
        return False
    html = r.text
    # Empty SPA root element: <div id="root"></div> with nothing inside
    if _SPA_ROOT_RE.search(html):
        # Confirm there's genuinely no text — extract visible chars
        visible = ds.extract_text(html)
        if len(visible) < 300:
            return True
    return False

# ── Proxy pool ────────────────────────────────────────────────────────────────

_proxy_pool: list = []
_proxy_fail:  set  = set()   # proxies that failed — skip for a while


def load_proxies(path):
    global _proxy_pool
    if not path or not os.path.exists(path):
        return
    with open(path) as f:
        _proxy_pool = [ln.strip() for ln in f if ln.strip() and not ln.startswith("#")]
    print(f"Loaded {len(_proxy_pool)} proxies from {path}")


def _next_proxy():
    """Return the first working proxy in the pool (order matters — UK is first)."""
    available = [p for p in _proxy_pool if p not in _proxy_fail]
    return available[0] if available else None


# ── Enhanced safe_get (monkey-patched over ds.safe_get) ───────────────────────

_orig_safe_get = None


def _bypass_safe_get(url, retries=2):
    """
    Drop-in replacement for ds.safe_get that tries:
      1. curl_cffi with rotating browser profiles (no proxy)
      2. curl_cffi with proxy rotation
      3. Original safe_get as final fallback
    """
    if not url or str(url).strip() in ("", "N/A", "None"):
        return None
    if not url.startswith("http"):
        url = "https://" + url.lstrip("/")

    # ── Strategy 1: curl_cffi, no proxy ──────────────────────────────────────
    # Do NOT pass ds.HEADERS — curl_cffi sets headers that match the impersonated
    # browser profile; a mismatched User-Agent breaks the TLS fingerprint and gets blocked.
    if _CFFI_OK:
        profiles = random.sample(_CFFI_PROFILES, len(_CFFI_PROFILES))
        for profile in profiles:
            try:
                sess = cffi_requests.Session(impersonate=profile)
                r = sess.get(url, timeout=15, verify=False, allow_redirects=True)
                if r.status_code == 200 and not _is_challenge_html(r):
                    if _is_js_shell(r):
                        ds.log.info(f"   [bypass] curl_cffi/{profile}: JS SPA shell detected — deferring to Playwright")
                        return None
                    ds.log.info(f"   [bypass] curl_cffi/{profile} OK: {url}")
                    return r
                # Hard block (403/407/202) OR 200 with challenge body — try proxy
                if r.status_code in (403, 407, 202) or _is_challenge_html(r):
                    break
            except Exception:
                continue

    # ── Strategy 2: curl_cffi + proxy ────────────────────────────────────────
    # Proxies are ordered (UK first) — do NOT shuffle; geo-order is intentional.
    if _CFFI_OK and _proxy_pool:
        available = [p for p in _proxy_pool if p not in _proxy_fail]
        for proxy in available[:6]:
            profile = random.choice(_CFFI_PROFILES[:3])
            try:
                sess = cffi_requests.Session(impersonate=profile)
                r = sess.get(url, timeout=25, verify=False, allow_redirects=True,
                             proxies={"http": proxy, "https": proxy})
                if r.status_code == 200 and not _is_challenge_html(r):
                    if _is_js_shell(r):
                        ds.log.info(f"   [bypass] curl_cffi/{profile}+proxy: JS SPA shell — deferring to Playwright")
                        return None
                    ds.log.info(f"   [bypass] curl_cffi/{profile}+proxy OK: {url}")
                    return r
                if r.status_code in (407, 403):
                    _proxy_fail.add(proxy)
            except Exception:
                _proxy_fail.add(proxy)
                continue

    # ── Strategy 3: original safe_get ────────────────────────────────────────
    # IMPORTANT: if safe_get returns a Cloudflare challenge page (status 200
    # but challenge HTML), return None so that scrape_practice's Playwright
    # fallback is triggered.  Without this, all_soup gets set to CF HTML and
    # Playwright is never attempted.
    _fallback = _orig_safe_get(url, retries=retries)
    if _fallback and (_is_challenge_html(_fallback) or _is_js_shell(_fallback)):
        ds.log.info(f"   [bypass] safe_get returned challenge/JS-shell — deferring to Playwright")
        return None
    return _fallback


# ── Playwright with stealth + optional proxy ──────────────────────────────────

def _pw_proxy_dict(proxy_url):
    """Convert http://user:pass@host:port to Playwright proxy dict."""
    p = urlparse(proxy_url)
    d = {"server": f"{p.scheme}://{p.hostname}:{p.port}"}
    if p.username:
        d["username"] = p.username
    if p.password:
        d["password"] = p.password
    return d


def _apply_stealth(page):
    """Apply playwright-stealth if available, else fall back to manual JS patch."""
    if _STEALTH_LIB_OK:
        try:
            _stealth_sync(page)
            return
        except Exception:
            pass
    # Manual fallback — less comprehensive but doesn't require the package
    try:
        page.add_init_script(script=_STEALTH_JS)
    except Exception:
        pass


def _make_pw_context(pw, proxy_url=None):
    kwargs = dict(
        user_data_dir="",
        headless=True,
        ignore_https_errors=True,
        args=[
            "--disable-blink-features=AutomationControlled",
            "--disable-infobars",
            "--window-size=1920,1080",
            "--no-first-run",
            "--disable-dev-shm-usage",
        ],
        user_agent=ds.HEADERS["User-Agent"],
    )
    if proxy_url:
        kwargs["proxy"] = _pw_proxy_dict(proxy_url)
    ctx = pw.chromium.launch_persistent_context(**kwargs)
    # Inject stealth on context level — applies to every page opened on this context
    if not _STEALTH_LIB_OK:
        ctx.add_init_script(script=_STEALTH_JS)
    return ctx


def _pw_get_html(url, proxy_url=None, pw_context=None):
    """
    Fetch a URL using Playwright with stealth. Returns HTML string or None.
    If pw_context (an already-open browser context) is given, reuse it via a
    new tab instead of spinning up a second sync_playwright() instance —
    Playwright's sync API does not support nested instances on the same
    thread and raises if you try.
    """
    if not _PW_OK:
        return None

    def _load(page):
        if _STEALTH_LIB_OK:
            try:
                _stealth_sync(page)
            except Exception:
                pass
        page.set_extra_http_headers({"Accept-Language": "en-US,en;q=0.9"})
        # Use "commit" (not "domcontentloaded") — fires as soon as the server
        # sends the first bytes.  "domcontentloaded" never fires when CF holds
        # the connection open waiting for the JS challenge to run, causing a
        # full 30-45 s timeout even though the server IS responding.
        try:
            page.goto(url, timeout=15000, wait_until="commit")
        except Exception:
            # Even "commit" timed out — server is completely unreachable/hung
            return None
        # Wait for Cloudflare JS challenge to auto-resolve (up to 8 × 5s = 40s)
        for _ in range(8):
            try:
                title = page.title().lower()
            except Exception:
                break
            if any(m in title for m in ("just a moment", "checking your", "please wait", "one moment")):
                page.wait_for_timeout(5000)
            else:
                break
        # After CF resolves (or no CF), wait for JS to render page content.
        try:
            page.wait_for_load_state("domcontentloaded", timeout=15000)
        except Exception:
            pass
        try:
            page.wait_for_load_state("networkidle", timeout=10000)
        except Exception:
            pass   # fine if it times out
        page.wait_for_timeout(3000)   # extra grace for React/SPA hydration
        try:
            html = ds._safe_content(page, timeout_secs=20)
        except Exception:
            return None
        if len(html) < 2000:
            return None
        snippet = html[:3000].lower()
        if any(m in snippet for m in _CHALLENGE_MARKERS):
            return None
        return html

    if pw_context is not None:
        try:
            page = pw_context.new_page()
            try:
                return _load(page)
            finally:
                page.close()
        except Exception as e:
            ds.log.warning(f"   [bypass] Playwright failed for {url}: {e}")
            return None

    # Try without proxy first, then with proxy if that fails
    for use_proxy in [False, True]:
        _proxy = proxy_url if use_proxy else None
        if use_proxy and not _proxy_pool:
            break
        if use_proxy and not _proxy:
            _proxy = _next_proxy()
        try:
            with sync_playwright() as pw:
                ctx  = _make_pw_context(pw, _proxy)
                page = ctx.new_page()
                html = _load(page)
                ctx.close()
                if html:
                    return html
        except Exception as e:
            ds.log.warning(f"   [bypass] Playwright{'(+proxy)' if use_proxy else ''} failed for {url}: {e}")
    return None


# ── Detect bot-blocked rows in a batch Excel output ───────────────────────────

# Batch output column positions (1-based), matching write_output col_headers
_B_INDEX    = 1
_B_NAME     = 2
_B_DOCTOR   = 3
_B_CITY     = 5
_B_STATE    = 6
_B_ZIP      = 7
_B_WEBSITE  = 8
_B_EMAIL    = 9


def _is_blocked_row(ws, r):
    """Return True if this row likely failed due to bot protection."""
    web = str(ws.cell(r, _B_WEBSITE).value or "").strip()
    if not web or web.lower() in ("not found", "n/a", "none", ""):
        return False   # no website — not a bot-block issue
    doc   = str(ws.cell(r, _B_DOCTOR).value  or "").strip().lower()
    email = str(ws.cell(r, _B_EMAIL).value   or "").strip().lower()
    # If doctor name AND email are both missing/not-found, likely bot-blocked
    doc_missing   = doc   in ("not found", "", "none", "n/a")
    email_missing = email in ("not found", "", "none", "n/a")
    return doc_missing and email_missing


def load_blocked_from_batch(path):
    """
    Read a batch output Excel, return list of row dicts for bot-blocked practices.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for r in range(3, ws.max_row + 1):   # row 1 = group header, row 2 = col headers
        if not _is_blocked_row(ws, r):
            continue
        rows.append({
            "Index":         ws.cell(r, _B_INDEX).value,
            "Practice Name": ws.cell(r, _B_NAME).value  or "",
            "Website":       ws.cell(r, _B_WEBSITE).value or "",
            "City":          ws.cell(r, _B_CITY).value   or "",
            "State":         ws.cell(r, _B_STATE).value  or "",
            "Zip":           ws.cell(r, _B_ZIP).value    or "",
        })
    wb.close()
    return rows


# ── Supplementary doctor extraction for network / FQHC sites ─────────────────
# Some websites (e.g. clinicas.org) list all dental providers on a separate
# sub-page with doctor names in <a> tags instead of headings.  The main
# dental_scraper pipeline misses these.  After the main scrape, if no doctors
# were found, this supplementary crawl finds and fetches dental sub-pages and
# extracts names using a direct regex scan on the full page text.

def _find_dental_subpage_urls(html, base_url):
    """
    Find URLs of dental / provider sub-pages linked from the given page HTML.
    Returns up to 5 absolute URLs on the same domain.
    """
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html, "lxml")
    base_netloc = urlparse(base_url).netloc
    found = []
    seen  = set()
    _DENTAL_LINK_RE = re.compile(
        r'\b(dental|dentist|dentistry|provider|our\s+team|meet\s+the\s+team|'
        r'our\s+staff|our\s+doctors|physicians|specialists)\b', re.I,
    )
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if not href or href.startswith(("#", "mailto:", "tel:", "javascript:")):
            continue
        text = a.get_text(strip=True)
        if not (_DENTAL_LINK_RE.search(href) or _DENTAL_LINK_RE.search(text)):
            continue
        full = urljoin(base_url, href).split("?")[0].rstrip("/")
        if urlparse(full).netloc != base_netloc:
            continue
        if full in seen or full == base_url.rstrip("/"):
            continue
        seen.add(full)
        found.append(full)
        if len(found) >= 5:
            break
    return found


_ROLE_WORDS_EXT = frozenset({
    "dentist", "doctor", "hygienist", "orthodontist", "periodontist",
    "endodontist", "prosthodontist", "surgeon", "specialist", "director",
    "coordinator", "assistant", "manager", "receptionist", "founder",
    "owner", "provider", "physician", "associate", "general", "pediatric",
    "cosmetic", "family", "lead", "attending", "chief", "principal",
})


def _extract_doctors_from_html(html):
    """
    Scan page HTML for doctor names matching _DOCTOR_PATTERNS.
    Extends leftward to recover first names that precede the regex match
    (e.g. "Ayesha" before "Raza Waseer, DDS").
    Returns list of {"name", "specialty", "associations"}.
    """
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html, "lxml")
    for t in soup.find_all(["nav", "header", "footer", "script", "style"]):
        t.decompose()
    text = soup.get_text(separator=" ", strip=True)

    seen    = set()
    doctors = []
    for pat in ds._DOCTOR_PATTERNS:
        for m in re.finditer(pat, text):
            raw_name = re.sub(r"\s+", " ", m.group(0).strip())
            # Extend leftward to recover additional first-name words that were
            # not captured by the pattern (e.g. 3-word names like "Ayesha Raza Waseer")
            start = m.start()
            prefix = text[max(0, start - 45):start].strip()
            if prefix:
                extra = []
                for w in reversed(prefix.split()):
                    # Only plain capitalized words that are not role titles
                    if re.match(r'^[A-Z][a-z]{1,20}$', w) and w.lower() not in _ROLE_WORDS_EXT:
                        extra.insert(0, w)
                    else:
                        break   # stop at any non-name word
                if extra:
                    raw_name = " ".join(extra) + " " + raw_name
                    raw_name = re.sub(r"\s+", " ", raw_name.strip())
            key = raw_name.lower()
            if key in seen:
                continue
            if re.search(r"\d", raw_name) or len(raw_name) < 5 or len(raw_name) > 70:
                continue
            if any(w in raw_name.lower() for w in ds._SKIP_WORDS):
                continue
            if not ds._is_valid_doctor_name(raw_name):
                continue
            seen.add(key)
            spec  = ds.find_specialty(text[:8000])
            assoc = ds.find_associations(text[:8000])
            doctors.append({
                "name":         raw_name,
                "specialty":    spec  or "Not Found",
                "associations": assoc or "Not Found",
            })
    return doctors


def _fetch_html_bypass(url, pw_context=None):
    """Fetch a URL with bypass strategies, return HTML text or None."""
    # Strategy 1: curl_cffi without proxy (fast)
    _cffi_blocked = False
    if _CFFI_OK:
        for profile in random.sample(_CFFI_PROFILES[:4], 4):
            try:
                sess = cffi_requests.Session(impersonate=profile)
                r = sess.get(url, timeout=15, verify=False, allow_redirects=True)
                if r.status_code == 200 and not _is_challenge_html(r):
                    return r.text
                # Hard block (403/407/202) OR 200 with challenge body — try proxy
                _cffi_blocked = True
                break
            except Exception:
                # Network/timeout error — mark for proxy retry, try next profile
                _cffi_blocked = True
                continue

    # Strategy 2: curl_cffi WITH proxy (when no-proxy was blocked)
    if _CFFI_OK and _cffi_blocked and _proxy_pool:
        available = [p for p in _proxy_pool if p not in _proxy_fail]
        for proxy in available[:3]:
            profile = random.choice(_CFFI_PROFILES[:3])
            try:
                sess = cffi_requests.Session(impersonate=profile)
                r = sess.get(url, timeout=25, verify=False, allow_redirects=True,
                             proxies={"http": proxy, "https": proxy})
                if r.status_code == 200 and not _is_challenge_html(r):
                    return r.text
                if r.status_code in (407, 403):
                    _proxy_fail.add(proxy)
            except Exception:
                _proxy_fail.add(proxy)
                continue

    # Strategy 3: Playwright stealth (uses pw_context which has proxy)
    html = _pw_get_html(url, pw_context=pw_context)
    if html:
        return html
    # Skyvern cloud browser (UK residential) as final fallback
    if _SKYVERN_KEY:
        session_id = _skyvern_session()
        if session_id:
            try:
                if _skyvern_navigate(session_id, url, timeout_ms=45000):
                    sk_html = _skyvern_get_html(session_id)
                    if sk_html and len(sk_html) > 2000:
                        return sk_html
            finally:
                _skyvern_close(session_id)
    return None


def _lookup_npi_dentists(city, state, street_address=""):
    """
    Query the public NPI registry for individual dentists (NPI-1) in a city/state.
    Filters by practice address when street_address is provided.
    Returns list of dicts: {name, credentials, specialty}.

    NPI registry is a US government public API with no auth/rate limits.
    Useful as fallback when the practice website is CF-blocked.
    """
    try:
        import urllib.parse, json as _json
        params = {
            "version": "2.1",
            "enumeration_type": "NPI-1",
            "city": city,
            "state": state,
            "taxonomy_description": "dentist",
            "limit": "50",
        }
        url = "https://npiregistry.cms.hhs.gov/api/?" + urllib.parse.urlencode(params)
        resp = std_requests.get(url, timeout=15)
        if resp.status_code != 200:
            return []
        data = resp.json()
        results = data.get("results", [])
        if not results:
            return []

        doctors = []
        norm_street = street_address.lower().replace(",", "").replace(".", "").replace("  ", " ").strip()

        for r in results:
            basic = r.get("basic", {})
            first = basic.get("first_name", "").strip()
            last  = basic.get("last_name", "").strip()
            cred  = basic.get("credential", "").strip()
            if not first or not last:
                continue

            # Try to match practice address — filter dentists at this specific location
            if norm_street:
                addrs = r.get("addresses", [])
                matched = False
                for a in addrs:
                    a1 = a.get("address_1", "").lower().replace(",", "").replace(".", "").strip()
                    if a1 and any(part in a1 for part in norm_street.split() if len(part) > 3):
                        matched = True
                        break
                if not matched:
                    continue

            spec = ""
            taxs = r.get("taxonomies", [])
            if taxs:
                spec = taxs[0].get("desc", "")
                if spec.lower() == "dentist":
                    spec = ""   # too generic; skip obvious default

            name = f"Dr. {first.title()} {last.title()}"
            doctors.append({"name": name, "credentials": cred, "specialty": spec})

        return doctors

    except Exception as e:
        print(f"  → NPI lookup error: {e}")
        return []


def _supplement_doctors(result, website, pw_context=None):
    """
    Fetch dental sub-pages and extract doctor names via full-text regex,
    merging any new names with whatever the main scrape already found.
    Runs even when some doctors were already captured — catches practices
    where the main scrape found 1 doctor but 3 more are on sub-pages.
    """
    existing_doctors = result.get("doctors") or []
    existing_names = {
        d["name"] for d in existing_doctors
        if d.get("name") and d["name"] not in ("Not Found", "")
    }

    # Don't bother if we already have a large list (5+ doctors)
    if len(existing_names) >= 5:
        return

    label = "supplementary" if not existing_names else "supplementary-merge"
    print(f"  → Running {label} dental sub-page crawl (have {len(existing_names)} doctor(s))…")

    main_html = _fetch_html_bypass(website, pw_context=pw_context)
    if not main_html:
        print("  → Could not fetch main page for sub-page discovery")
        return

    sub_urls = _find_dental_subpage_urls(main_html, website)
    if not sub_urls:
        print("  → No dental sub-pages found in navigation")
        return

    print(f"  → Dental sub-pages: {sub_urls}")
    found_doctors = list(existing_doctors)   # start with already-found list
    found_names_lc = {n.lower() for n in existing_names}

    for url in sub_urls:
        print(f"     Fetching: {url}")
        html = _fetch_html_bypass(url, pw_context=pw_context)
        if not html:
            continue
        drs = _extract_doctors_from_html(html)
        new_here = []
        for d in drs:
            if d["name"].lower() not in found_names_lc:
                found_doctors.append(d)
                found_names_lc.add(d["name"].lower())
                new_here.append(d["name"])
        if new_here:
            print(f"     New doctor(s): {', '.join(new_here)}")

    new_total = [d for d in found_doctors if d["name"] not in existing_names]
    if new_total:
        result["doctors"] = found_doctors
        result["scraped_doctor_names"] = ", ".join(d["name"] for d in found_doctors)
        print(f"  → After supplementary crawl: {result['scraped_doctor_names']}")
    elif not existing_names:
        print(f"  → Supplementary crawl: still no doctors found")


def _parse_skyvern_data(scraped_data, result):
    """
    Parse the full-text dump from Skyvern into the result dict.

    Extracts from ALL collected text (main page + sub-pages):
      - Doctor names (Dr. First Last patterns)
      - Phone numbers
      - Services keywords
      - Rating / review count
    Updates `result` in-place; returns True if any useful data was found.
    """
    from bs4 import BeautifulSoup

    all_text_parts = [scraped_data.get("main_text", "")]
    for _, txt in scraped_data.get("sub_texts", []):
        all_text_parts.append(txt)
    full_text = "\n".join(all_text_parts)

    if not full_text.strip():
        return False

    improved = False

    # ── Doctor names ──────────────────────────────────────────────────────────
    doc_val = str(result.get("scraped_doctor_names", "Not Found")).strip()
    if doc_val in ("Not Found", "", "None", "N/A"):
        # Pattern: Dr. FirstName LastName (optional credentials)
        dr_pattern = re.compile(
            r"Dr\.?\s+([A-Z][a-z]+(?:\s+[A-Z]\.?)?\s+[A-Z][A-Za-z'\-]+)"
            r"(?:\s*,?\s*(DDS|DMD|DDS/DMD|DMD/DDS|MD|BDS|MSD|MS|PhD|FAGD|ABGD|FICOI|AACD|FICCDE)"
            r"(?:[,\s/]+(?:DDS|DMD|DDS/DMD|DMD/DDS|MD|BDS|MSD|MS|PhD|FAGD|ABGD|FICOI|AACD))*)?",
            re.I,
        )
        found_doctors = {}
        for m in dr_pattern.finditer(full_text):
            name = "Dr. " + m.group(1).strip()
            cred = (m.group(2) or "").strip().upper()
            if name not in found_doctors:
                found_doctors[name] = cred

        # Also look for name patterns following "Owner", "Associate", specialty titles
        role_pattern = re.compile(
            r"(Dr\.?\s+[A-Z][a-z]+\s+[A-Z][A-Za-z'\-]+)\s*\n?\s*"
            r"(Owner|Associate|Endodontist|Orthodontist|Oral Surgeon|Periodontist|"
            r"Pediatric|General Dentist|Prosthodontist|Cosmetic|Family Dentist)",
            re.I,
        )
        for m in role_pattern.finditer(full_text):
            name = m.group(1).strip()
            if not name.startswith("Dr."):
                name = "Dr. " + name.split("Dr.")[-1].strip()
            if name not in found_doctors:
                found_doctors[name] = ""

        if found_doctors:
            doctors_list = [
                {"name": n, "credentials": c, "specialty": ""}
                for n, c in found_doctors.items()
            ]
            result["doctors"] = doctors_list
            result["scraped_doctor_names"] = ", ".join(found_doctors.keys())
            print(f"  [Skyvern] Doctors found: {result['scraped_doctor_names']}")
            improved = True

    # ── Phone ─────────────────────────────────────────────────────────────────
    phone_val = str(result.get("phone", "Not Found")).strip()
    if phone_val in ("Not Found", "", "None"):
        phone_m = re.search(
            r"\(?\d{3}\)?[\s.\-]?\d{3}[\s.\-]?\d{4}", full_text
        )
        if phone_m:
            result["phone"] = phone_m.group(0).strip()
            improved = True

    # ── Rating ────────────────────────────────────────────────────────────────
    rating_val = str(result.get("google_rating", "Not Found")).strip()
    if rating_val in ("Not Found", "", "None"):
        # Look for JSON-LD in main HTML first
        main_html = scraped_data.get("main_html", "")
        if main_html:
            import json as _json
            for script_match in re.finditer(
                r'<script[^>]+type=["\']application/ld\+json["\'][^>]*>(.*?)</script>',
                main_html, re.S | re.I
            ):
                try:
                    ld = _json.loads(script_match.group(1))
                    ag = ld.get("aggregateRating") or {}
                    rv = ag.get("ratingValue") or ld.get("ratingValue")
                    rc = ag.get("reviewCount") or ag.get("ratingCount")
                    if rv:
                        result["google_rating"] = str(rv)
                        improved = True
                    if rc:
                        result["total_google_reviews"] = str(rc)
                        improved = True
                    if rv:
                        break
                except Exception:
                    pass

        # Plain-text fallback: "4.4 stars" or "4.4 out of 5"
        if result.get("google_rating", "Not Found") in ("Not Found", "", "None"):
            rat_m = re.search(r"(\d\.\d)\s*(?:stars?|out of 5|★)", full_text, re.I)
            if rat_m:
                result["google_rating"] = rat_m.group(1)
                improved = True

    # ── Services ─────────────────────────────────────────────────────────────
    svc_val = str(result.get("services", "Not Found")).strip()
    if svc_val in ("Not Found", "", "None"):
        svc_keywords = [
            "Dentures", "Dental implants", "Dental cleaning", "Root canal",
            "Teeth whitening", "Invisalign", "Clear aligners", "Crowns",
            "Bridges", "Veneers", "Extractions", "Emergency dental",
            "Sedation", "Pediatric", "Orthodontics", "Oral surgery",
            "Wisdom tooth", "Cosmetic dentistry", "Fillings", "Bonding",
            "Night guards", "Sleep apnea", "Gum disease", "Periodontal",
            "Implant dentures", "Full arch", "All-on-4",
        ]
        found_svcs = [s for s in svc_keywords if s.lower() in full_text.lower()]
        if found_svcs:
            result["services"] = ", ".join(found_svcs)
            improved = True

    # ── Address ───────────────────────────────────────────────────────────────
    addr_val = str(result.get("address", "Not Found")).strip()
    if addr_val in ("Not Found", "", "None"):
        # Simple US street address pattern
        addr_m = re.search(
            r"\d+\s+[A-Z][a-zA-Z0-9\s\.\,]+(?:Blvd|Ave|St|Rd|Dr|Way|Ln|Pkwy|Suite|Ste)"
            r"[^\n]{0,60}(?:[A-Z]{2}\s+\d{5})?",
            full_text,
        )
        if addr_m:
            result["address"] = addr_m.group(0).strip()
            improved = True

    return improved


# ── Scrape one practice with all bypass strategies ────────────────────────────

_EMPTY_VALS = {"not found", "", "none", "n/a", "0"}


def _is_empty_result(result):
    """Return True when scrape_practice captured nothing useful.

    Triggers ALL supplementary passes whenever doctor + email + key
    services are all missing — regardless of the skip_reason.  The old
    check (require a CF-block note in skip_reason) missed sites that
    were reachable but had silent extraction failures in CI.
    """
    if not isinstance(result, dict):
        return True
    doc   = str(result.get("scraped_doctor_names", "Not Found")).strip().lower()
    email = str(result.get("email",                "Not Found")).strip().lower()
    no_doc   = doc   in _EMPTY_VALS
    no_email = email in _EMPTY_VALS
    # Check key service columns — if all zero the page content wasn't captured
    svc_total = (
        int(result.get("implants",      0) or 0) +
        int(result.get("veneers",       0) or 0) +
        int(result.get("invisalign",    0) or 0) +
        int(result.get("whitening",     0) or 0) +
        int(result.get("sedation",      0) or 0)
    )
    return no_doc and no_email and svc_total == 0


# Email addresses to ignore during extraction
_EMAIL_SKIP = frozenset(["noreply", "example", "sentry", "wordpress", "gravatar",
                          "schema.org", "w3.org", "png", "jpg", "gif", "svg"])

# Social handles to ignore
_FB_SKIP_HANDLES  = frozenset(["sharer", "share", "login", "dialog", "pages",
                                 "groups", "events", "marketplace", "watch"])
_IG_SKIP_HANDLES  = frozenset(["p", "reel", "explore", "stories", "accounts",
                                 "tv", "direct", "tags"])

# Service keyword → result field → count cap
_DIRECT_SVC = [
    (["implant", "dental implant"],    "implants",       99),
    (["veneer"],                        "veneers",        99),
    (["invisalign"],                    "invisalign",     20),
    (["whitening", "teeth whitening"],  "whitening",      30),
    (["sedation"],                      "sedation",       30),
    (["clear aligner"],                 "clear_aligners", 20),
    (["smile makeover"],                "smile_makeovers",20),
    (["holistic"],                      "holistic",       10),
    (["cancer screening"],              "cancer_screening",10),
    (["cerec"],                         "cerec",           5),
    (["laser"],                         "lasers",          5),
    (["3d imaging", "cbct", "cone beam"],"cbct",           5),
    (["intraoral scanner"],             "intraoral_scanners",5),
]


def _fill_missing_data(result, html, label="direct"):
    """
    Parse raw HTML and fill any fields in `result` that are still
    'Not Found' / empty / zero.  Runs AFTER all other passes as a
    safety net so we capture at minimum email, social links, and
    service counts from the homepage HTML.
    """
    if not html or len(html) < 500:
        return

    from bs4 import BeautifulSoup as _BS4
    import json as _json

    soup = _BS4(html, "lxml")
    # Plain text for regex + keyword counting
    body_text = soup.get_text(separator=" ", strip=True)

    filled = []

    # ── Email ──────────────────────────────────────────────────────────────────
    if str(result.get("email", "")).strip().lower() in _EMPTY_VALS:
        for em in re.findall(
            r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}', html
        ):
            el = em.lower()
            if not any(s in el for s in _EMAIL_SKIP):
                result["email"] = em
                filled.append(f"email={em}")
                break

    # ── Facebook URL ───────────────────────────────────────────────────────────
    if not str(result.get("facebook_url", "")).strip():
        fb_matches = re.findall(
            r'https?://(?:www\.)?facebook\.com/([A-Za-z0-9_.%\-]{3,80})', html
        )
        for h in fb_matches:
            if h.lower().rstrip("/") not in _FB_SKIP_HANDLES:
                result["facebook_url"] = f"https://www.facebook.com/{h}"
                filled.append(f"fb={h}")
                break

    # ── Instagram URL ──────────────────────────────────────────────────────────
    if not str(result.get("instagram_url", "")).strip():
        ig_matches = re.findall(
            r'https?://(?:www\.)?instagram\.com/([A-Za-z0-9_.%\-]{2,60})(?:[/?]|$|["\'])', html
        )
        for h in ig_matches:
            if h.lower().rstrip("/") not in _IG_SKIP_HANDLES:
                result["instagram_url"] = f"https://www.instagram.com/{h}"
                filled.append(f"ig={h}")
                break

    # ── Service keyword counts ──────────────────────────────────────────────────
    text_lc = body_text.lower()
    for terms, field, cap in _DIRECT_SVC:
        cur = int(result.get(field, 0) or 0)
        if cur == 0:
            cnt = sum(text_lc.count(t) for t in terms)
            if cnt > 0:
                result[field] = min(cnt, cap)
                filled.append(f"{field}={result[field]}")

    # ── Doctor names ────────────────────────────────────────────────────────────
    # Always run: merge any newly found names with existing ones.
    existing_doc_val = str(result.get("scraped_doctor_names", "Not Found")).strip()
    existing_names_list = (
        [n.strip() for n in existing_doc_val.split(",") if n.strip().lower() not in _EMPTY_VALS]
        if existing_doc_val.lower() not in _EMPTY_VALS else []
    )
    found_names = list(existing_names_list)  # start with already-found names
    found_names_lc = {n.lower() for n in found_names}

    def _add_doctor(name):
        """Validate and add a doctor name if not already present."""
        name = re.sub(r"\s+", " ", name.strip())
        if not name or re.search(r'\d', name):
            return
        if len(name.split()) < 2:
            return  # require at least first + last name
        if any(w in name.lower() for w in ds._SKIP_WORDS):
            return
        if not ds._is_valid_doctor_name(name):
            return
        if name.lower() not in found_names_lc:
            found_names.append(name)
            found_names_lc.add(name.lower())

    # Pattern 1: "Dr. FirstName LastName" (with optional middle initial)
    for m in re.finditer(
        r'Dr\.?\s+([A-Z][a-z]+(?:\s+[A-Z]\.?)?\s+[A-Z][A-Za-z\-\']{2,})',
        body_text,
    ):
        _add_doctor(f"Dr. {m.group(1).strip()}")

    # Pattern 2: "FirstName LastName, DDS/DMD" (no Dr. prefix)
    for m in re.finditer(
        r'([A-Z][a-z]{1,20}(?:\s+[A-Z][a-z]{0,15})?\s+[A-Z][A-Za-z\-\']{2,20})'
        r'\s*,?\s+(?:DDS|DMD|D\.D\.S|D\.M\.D)',
        body_text,
    ):
        _add_doctor(m.group(1).strip())

    # Only update result when we found NEW names
    if len(found_names) > len(existing_names_list):
        result["scraped_doctor_names"] = ", ".join(found_names[:8])
        result["doctors"] = [
            {"name": n, "specialty": "Not Found", "associations": "Not Found"}
            for n in found_names[:8]
        ]
        new_names = found_names[len(existing_names_list):]
        filled.append(f"doctors+={', '.join(new_names)}")

    # ── Google rating from JSON-LD ──────────────────────────────────────────────
    rating_val = str(result.get("google_rating", "Not Found")).strip()
    if rating_val.lower() in _EMPTY_VALS:
        for script in soup.find_all("script", type="application/ld+json"):
            try:
                ld = _json.loads(script.string or "")
                ag = ld.get("aggregateRating") or {}
                rv = str(ag.get("ratingValue") or ld.get("ratingValue") or "").strip()
                rc = str(ag.get("reviewCount") or ag.get("ratingCount") or "").strip()
                if rv:
                    result["google_rating"] = rv
                    if rc:
                        result["total_google_reviews"] = rc
                    filled.append(f"rating={rv}({rc})")
                    break
            except Exception:
                pass

    if filled:
        print(f"  → [{label}] Filled: {', '.join(filled)}")


def scrape_with_bypass(row, pw_page=None, pw_page_noproxy=None):
    """
    Monkey-patch ds.safe_get, run the full scraper pipeline, restore original.

    Strategy:
      1. Run scrape_practice with the proxy Playwright page.
      2. If the site was CF-blocked and nothing was captured, retry the
         Playwright section with a no-proxy stealth page (pw_page_noproxy).
         Some CF Enterprise configs block known proxy ranges but pass
         stealth-only headless browsers through a Turnstile challenge.
      3. Supplementary sub-page crawl when doctors still not found.
    """
    global _orig_safe_get
    _orig_safe_get = ds.safe_get
    ds.safe_get = _bypass_safe_get
    try:
        result = ds.scrape_practice(row, pw_page=pw_page)
    finally:
        ds.safe_get = _orig_safe_get

    # ── Retry with no-proxy stealth Playwright when proxy was blocked ─────────
    if _is_empty_result(result) and pw_page_noproxy is not None:
        print(f"  → Proxy blocked — retrying with no-proxy stealth Playwright…")
        _orig_safe_get = ds.safe_get
        ds.safe_get = _bypass_safe_get
        try:
            result2 = ds.scrape_practice(row, pw_page=pw_page_noproxy)
        finally:
            ds.safe_get = _orig_safe_get
        # Use the no-proxy result only if it captured more
        if not _is_empty_result(result2):
            print(f"  → No-proxy stealth succeeded!")
            result = result2
        else:
            print(f"  → No-proxy stealth also blocked — site requires real browser")

    # ── Skyvern cloud browser fallback (UK residential IP, bypasses CF Enterprise) ──
    # Triggered when both proxy and no-proxy Playwright fail.
    # Skyvern uses real Chrome with residential IPs not in CF's blocklist.
    # Requires SKYVERN_API_KEY environment variable / GitHub Secret.
    if _is_empty_result(result) and _SKYVERN_KEY:
        website = row.get("Website", "")
        if website:
            print(f"  → All local strategies blocked — trying Skyvern cloud browser…")
            skyvern_data = _skyvern_full_scrape(website)
            if skyvern_data and isinstance(result, dict):
                improved = _parse_skyvern_data(skyvern_data, result)
                if improved:
                    print(f"  → Skyvern: data captured successfully")
                    # Feed the full HTML into the regular scraper pipeline too,
                    # so doctor/service extractors in dental_scraper.py can run
                    # over properly rendered content.
                    if skyvern_data.get("main_html") and len(skyvern_data["main_html"]) > 2000:
                        try:
                            from bs4 import BeautifulSoup
                            sk_soup = BeautifulSoup(skyvern_data["main_html"], "lxml")
                            sk_text = ds.extract_text(skyvern_data["main_html"])
                            # Re-run doctor extraction on the rendered HTML
                            sk_doctors, _ = ds.scrape_doctors_full(
                                sk_soup, website, sk_text, pw_page=None
                            )
                            if sk_doctors:
                                result["doctors"] = sk_doctors
                                result["scraped_doctor_names"] = ", ".join(
                                    d["name"] for d in sk_doctors
                                )
                                print(f"  → Skyvern+pipeline doctors: "
                                      f"{result['scraped_doctor_names']}")
                        except Exception as e:
                            print(f"  → Skyvern pipeline re-parse skipped: {e}")
                else:
                    print(f"  → Skyvern reached page but extracted no usable data")
            elif not _SKYVERN_KEY:
                pass  # key not set — skip silently
            else:
                print(f"  → Skyvern: could not load page")

    # ── Sub-page hint: if the original input URL was a sub-page (team, about,
    # services, etc.), fetch it now so its content supplements what the homepage
    # scrape found.  This captures doctors listed only on the team page, or
    # service keywords that only appear on a dedicated services page.
    _extra_sub = row.get("_extra_sub_url", "")
    if _extra_sub and isinstance(result, dict):
        try:
            from bs4 import BeautifulSoup as _BS4
            _sub_html = ""
            # Try Playwright first (handles JS-rendered pages)
            if pw_page is not None:
                try:
                    pw_page.goto(_extra_sub, timeout=30000, wait_until="domcontentloaded")
                    pw_page.wait_for_timeout(2000)
                    _sub_html = ds._safe_content(pw_page)
                except Exception:
                    pass
            # Fallback: bypass safe_get (proxy requests)
            if not _sub_html:
                _sr = _bypass_safe_get(_extra_sub)
                if _sr:
                    _sub_html = _sr.text
            if _sub_html and len(_sub_html) > 500:
                _sub_text = ds.extract_text(_sub_html)
                _sub_soup = _BS4(_sub_html, "lxml")
                print(f"  → Sub-page hint scraped: {_extra_sub} ({len(_sub_text)} chars)")
                # Supplement doctors if still missing
                _cur_doc = str(result.get("scraped_doctor_names", "Not Found")).strip()
                if _cur_doc in ("Not Found", "", "None", "N/A"):
                    _sub_docs, _ = ds.scrape_doctors_full(
                        _sub_soup, row.get("Website", ""), _sub_text, pw_page=None
                    )
                    if _sub_docs:
                        result["doctors"] = _sub_docs
                        result["scraped_doctor_names"] = ", ".join(
                            d["name"] for d in _sub_docs
                        )
                        print(f"  → Sub-page doctors: {result['scraped_doctor_names']}")
                # Supplement services from sub-page text
                from dental_scraper import SERVICE_KEYWORDS, count_keyword_capped
                _svc_keys = [
                    "invisalign", "clear_aligners", "veneers", "implants",
                    "smile_makeovers", "whitening", "sedation", "holistic",
                    "cancer_screening",
                ]
                _cat_map = {
                    "Invisalign": "invisalign", "Clear Aligners": "clear_aligners",
                    "Veneers": "veneers", "Implants": "implants",
                    "Smile Makeovers": "smile_makeovers", "Teeth Whitening": "whitening",
                    "Sedation Dentistry": "sedation", "Holistic Dentistry": "holistic",
                    "Cancer Screening": "cancer_screening",
                }
                for kw, cat in SERVICE_KEYWORDS.items():
                    if cat in _cat_map:
                        _fk = _cat_map[cat]
                        _cur = result.get(_fk, 0)
                        _new = count_keyword_capped(_sub_text, kw, cap=5)
                        if _new > (_cur or 0):
                            result[_fk] = _new
        except Exception as _sub_e:
            print(f"  → Sub-page hint error: {_sub_e}")

    # Supplementary pass: for network/FQHC sites where doctors are on a
    # separate dental-services page (not the main location/homepage)
    website = row.get("Website", "")
    if website and isinstance(result, dict):
        pw_context = pw_page.context if pw_page is not None else None
        _supplement_doctors(result, website, pw_context=pw_context)

    # ── NPI registry fallback when website is completely inaccessible ─────────
    # The NPI registry is a US government public API (no bot protection, no auth).
    # When all website-based methods fail we query it for individual dentists
    # (NPI-1) at this practice address to recover doctor names and credentials.
    if isinstance(result, dict):
        doc_val = str(result.get("scraped_doctor_names", "Not Found")).strip()
        still_no_doctors = doc_val in ("Not Found", "", "None", "N/A")
        if still_no_doctors:
            city    = row.get("City",  "")
            state   = row.get("State", "")
            address = row.get("Address", row.get("Street Address", ""))
            if city and state:
                print(f"  → Trying NPI registry fallback for {city}, {state}…")
                npi_docs = _lookup_npi_dentists(city, state, street_address=address)
                if npi_docs:
                    result["doctors"] = npi_docs
                    result["scraped_doctor_names"] = ", ".join(d["name"] for d in npi_docs)
                    print(f"  → NPI registry found {len(npi_docs)} doctor(s): "
                          f"{result['scraped_doctor_names']}")
                    # Mark source so downstream knows this came from NPI, not the website
                    result["npi_doctors"] = True
                else:
                    print(f"  → NPI registry: no matching dentists at this address")

    # ── Final safety net: direct HTML extraction ──────────────────────────────
    # When ALL previous passes left the result empty (no email, no doctors,
    # no services), fetch the homepage one more time and extract what we can
    # directly from the raw HTML — bypassing the full scrape_practice() pipeline.
    # This catches sites where scrape_practice() silently failed in CI but the
    # page is actually reachable.
    if _is_empty_result(result) and isinstance(result, dict):
        website = row.get("Website", "")
        if website:
            print(f"  → Result still empty — running direct HTML extraction…")
            pw_ctx_direct = pw_page.context if pw_page is not None else None
            _direct_html = _fetch_html_bypass(website, pw_context=pw_ctx_direct)
            if _direct_html:
                _fill_missing_data(result, _direct_html, label="homepage")
            # Always scan known team/about sub-pages and MERGE doctors across all of them
            _parsed = urlparse(website)
            _base = f"{_parsed.scheme}://{_parsed.netloc}"
            for _slug in ["/about/", "/our-team/", "/meet-the-doctor/",
                           "/team/", "/doctors/", "/staff/", "/about-us/",
                           "/our-doctors/", "/meet-our-team/", "/providers/",
                           "/meet-the-team/", "/our-dentists/"]:
                _sub_url = _base + _slug
                try:
                    _sub_html = _fetch_html_bypass(
                        _sub_url, pw_context=pw_ctx_direct
                    )
                    if _sub_html and len(_sub_html) > 500:
                        _fill_missing_data(result, _sub_html, label=_slug)
                except Exception:
                    pass

    # ── Even when the result is NOT empty, run _fill_missing_data on team sub-pages
    # to capture additional doctors missed by the main scrape pipeline.
    elif isinstance(result, dict):
        _doc_val = str(result.get("scraped_doctor_names", "Not Found")).strip()
        _want_more_doctors = _doc_val.lower() in _EMPTY_VALS or len(_doc_val.split(",")) < 3
        if _want_more_doctors:
            website = row.get("Website", "")
            if website:
                pw_ctx_direct = pw_page.context if pw_page is not None else None
                _parsed = urlparse(website)
                _base = f"{_parsed.scheme}://{_parsed.netloc}"
                for _slug in ["/our-team/", "/team/", "/doctors/",
                               "/meet-the-doctor/", "/meet-our-team/",
                               "/our-doctors/", "/providers/", "/staff/"]:
                    _sub_url = _base + _slug
                    try:
                        _sub_html = _fetch_html_bypass(
                            _sub_url, pw_context=pw_ctx_direct
                        )
                        if _sub_html and len(_sub_html) > 500:
                            _fill_missing_data(result, _sub_html, label=_slug)
                    except Exception:
                        pass

    return result


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    p = argparse.ArgumentParser(
        description="Re-scrape bot-blocked dental practice websites.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    grp = p.add_mutually_exclusive_group(required=True)
    grp.add_argument("--urls",  nargs="+", metavar="URL",
                     help="One or more website URLs to scrape")
    grp.add_argument("--batch", metavar="BATCH.xlsx",
                     help="Batch output Excel — auto-detects bot-blocked rows")
    p.add_argument("--proxies", metavar="proxies.txt",
                   help="Proxy list file (one proxy per line)")
    p.add_argument("--output", "-o", default="bypass_output.xlsx",
                   help="Output Excel file (default: bypass_output.xlsx)")
    args = p.parse_args()

    load_proxies(args.proxies)

    # ── Build practice rows ───────────────────────────────────────────────────
    if args.urls:
        rows = []
        for i, url in enumerate(args.urls, 1):
            if not url.startswith("http"):
                url = "https://" + url
            # Strip path/query/fragment so scrape_practice() always starts from the
            # homepage — sub-page or UTM-param URLs only contain partial site content
            # and miss services, tech, and social links that live on the homepage.
            from urllib.parse import urlparse as _urlparse
            _p = _urlparse(url)
            homepage = f"{_p.scheme}://{_p.netloc}/"
            practice_label = _p.netloc.replace("www.", "")  # human-readable name fallback
            row_entry = {
                "Index": i, "Practice Name": practice_label,
                "Website": homepage, "City": "", "State": "", "Zip": "",
            }
            # If the original URL is a sub-page (has a meaningful path), keep it
            # as a hint so the bypass scraper also fetches it for doctor/team data.
            if _p.path and _p.path not in ("/", ""):
                row_entry["_extra_sub_url"] = url
            rows.append(row_entry)
    else:
        rows = load_blocked_from_batch(args.batch)
        if not rows:
            print("No bot-blocked rows found in the batch file.")
            print("Criteria: website present but Doctor Name + Email both missing.")
            sys.exit(0)
        print(f"Found {len(rows)} likely bot-blocked rows in {args.batch}")

    print(f"\nScraping {len(rows)} practice(s) with bypass strategies…")
    if _proxy_pool:
        print(f"Proxy pool: {len(_proxy_pool)} proxies")
    print()

    # ── Set up Playwright (one browser for all practices) ─────────────────────
    pw_ctx        = None   # proxy context (UK/mobile)
    pw_ctx_plain  = None   # no-proxy stealth context (fallback)
    pw_page       = None
    pw_page_plain = None
    _pw_mgr       = None

    if _PW_OK:
        proxy_url = _next_proxy()
        _pw_mgr  = sync_playwright().__enter__()

        # Primary: proxy context (UK mobile/residential — bypasses geo-blocks)
        pw_ctx   = _make_pw_context(_pw_mgr, proxy_url)
        pw_page  = pw_ctx.new_page()
        _apply_stealth(pw_page)
        pw_page.set_extra_http_headers({"Accept-Language": "en-US,en;q=0.9"})

        # Fallback: no-proxy stealth context (for sites where proxy itself is blocked
        # but the CF Turnstile challenge can be solved by the stealth browser alone)
        pw_ctx_plain   = _make_pw_context(_pw_mgr, None)
        pw_page_plain  = pw_ctx_plain.new_page()
        _apply_stealth(pw_page_plain)
        pw_page_plain.set_extra_http_headers({"Accept-Language": "en-US,en;q=0.9"})

        stealth_label = "(playwright-stealth)" if _STEALTH_LIB_OK else "(manual stealth)"
        print(f"Playwright ready — proxy: {proxy_url.split('@')[-1] if proxy_url else 'none'} "
              f"{stealth_label}\n")

    all_results = []
    try:
        for i, row in enumerate(rows, 1):
            name = row.get("Practice Name") or row.get("Website")
            print(f"[{i}/{len(rows)}] {name}")
            result = scrape_with_bypass(row, pw_page=pw_page,
                                        pw_page_noproxy=pw_page_plain)
            all_results.append((row, result))
            # Save checkpoint every 5 practices
            if i % 5 == 0 or i == len(rows):
                ds.write_output(all_results, args.output)
                print(f"  ✓ Checkpoint saved → {args.output}")
            time.sleep(1)
    finally:
        if pw_ctx:       pw_ctx.close()
        if pw_ctx_plain: pw_ctx_plain.close()
        if _pw_mgr:
            try: _pw_mgr.__exit__(None, None, None)
            except Exception: pass

    ds.write_output(all_results, args.output)
    print(f"\nDone. {len(all_results)} practices saved → {args.output}")


if __name__ == "__main__":
    main()
