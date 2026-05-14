"""
refresh_tech_services.py
========================
Re-extracts Technology in Practice, Services (# of Mentions), Testimonials,
and # of Hygienists from the page_cache/ AND via live requests crawl as a
fallback when cache is missing or fields are still blank.

Rules when writing back (NEVER REDUCE):
  • Tech fields  (CEREC/CBCT/Laser/AI/Intraoral) : keep "X" if EITHER old or new is "X"
  • Service counts (numeric)                      : keep MAX(old, new)
  • Dental Plan                                   : keep "Mentioned" if EITHER has it
  • Testimonials / Hygienists                     : keep MAX(old, new)

Usage:
    python3 refresh_tech_services.py <batch_deduped.xlsx>
    python3 refresh_tech_services.py <batch_deduped.xlsx> --cache-dir /path/to/page_cache

Outputs:
    <input>_refreshed.xlsx   — patched xlsx
    <input>_comparison.xlsx  — before/after report (changed cells highlighted)
"""

import os, sys, re, shutil, glob, time, random, logging, warnings, threading
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse

warnings.filterwarnings("ignore", message="Unverified HTTPS request")

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import dental_scraper as ds

logging.basicConfig(level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger(__name__)

# ── Column positions (1-based), matching dental_scraper.py ───────────────────
C_IDX     = 1;  C_NAME    = 2;  C_WEBSITE = 8
C_HYG     = 10
C_CEREC   = 23; C_CBCT    = 24; C_LASER   = 25; C_AI      = 26; C_INTRA   = 27
C_INV     = 28  # col 29 = InvTier (skip)
C_CLEAR   = 30; C_VEN     = 31; C_IMPL    = 32; C_SMILE   = 33; C_WHITE   = 34
C_SED     = 35; C_HOL     = 36; C_PLAN    = 37; C_CANC    = 38
C_TESTI   = 46

TECH_COLS = {C_CEREC, C_CBCT, C_LASER, C_AI, C_INTRA}

# (xlsx_col, result_key, display_label)
FIELD_MAP = [
    (C_HYG,   "hygienists",      "# Hygienists"),
    (C_CEREC, "cerec",           "CEREC"),
    (C_CBCT,  "cbct",            "CBCT"),
    (C_LASER, "lasers",          "Lasers"),
    (C_AI,    "ai",              "AI"),
    (C_INTRA, "intraoral",       "Intraoral"),
    (C_INV,   "invisalign",      "Invisalign"),
    (C_CLEAR, "clear_aligners",  "Clear Aligners"),
    (C_VEN,   "veneers",         "Veneers"),
    (C_IMPL,  "implants",        "Implants"),
    (C_SMILE, "smile_makeovers", "Smile Makeovers"),
    (C_WHITE, "whitening",       "Whitening"),
    (C_SED,   "sedation",        "Sedation"),
    (C_HOL,   "holistic",        "Holistic"),
    (C_PLAN,  "dental_plan",     "Dental Plan"),
    (C_CANC,  "cancer_screening","Cancer Screening"),
    (C_TESTI, "testimonials",    "Testimonials"),
]

DATA_START = 3

# Testimonial block patterns — class, id, or data-* attributes
_TEST_RE = re.compile(
    r"(testimonial|review|quote|patient.story|patient.review|"
    r"feedback|client.say|what.people|slider.item|carousel.item|"
    r"swiper.slide|slick.slide|rating.block|star.review|"
    r"review.card|review.item|review.block|review.section|"
    r"patient.comment|patient.feedback|google.review|"
    r"review.widget|rating.widget|review.container)", re.I
)

# Paths that suggest service / technology / team pages worth crawling
_SVC_PATH_RE  = re.compile(
    r'/(service|treatment|technology|tech|procedure|offer|speciali|invisalign|'
    r'implant|veneer|whitening|cerec|laser|sedation|cbct|intraoral|aligner)',
    re.I,
)
_TEAM_PATH_RE = re.compile(
    r'/(team|staff|about|hygien|doctor|provider|meet)',
    re.I,
)
_TEST_PATH_RE = re.compile(
    r'/(testimonial|review|patient|feedback)',
    re.I,
)

# Live-crawl settings
_LIVE_MAX_PAGES    = 10
_LIVE_DELAY_MIN    = 1.2
_LIVE_DELAY_MAX    = 3.0
_LIVE_PW_TIMEOUT   = 90   # seconds — hard wall-clock limit per practice live crawl

# URLs whose content is always loaded via JS APIs that Playwright cannot
# reach (HMO portals, Kyruus SPAs, large aggregator sites).  Skip live crawl.
_SKIP_LIVE_RE = re.compile(
    r'(kaiser|healthgrades|zocdoc|vitals\.com|yelp\.com|google\.com|'
    r'doctors-locations|find-a-physician|find-a-doctor|kyruus|'
    r'adventhealth\.com/find|mdvip|webmd)',
    re.I,
)
_LIVE_HEADERS      = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept":          "text/html,application/xhtml+xml,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Connection":      "keep-alive",
}


# ── Cache loader ──────────────────────────────────────────────────────────────

def _find_cache_folder(idx: int, cache_dir: str):
    prefix = f"{idx:03d}_"
    try:
        for name in sorted(os.listdir(cache_dir)):
            if name.startswith(prefix) and os.path.isdir(os.path.join(cache_dir, name)):
                return os.path.join(cache_dir, name)
    except FileNotFoundError:
        pass
    return None


def _load_pages(folder: str) -> list:
    """
    Load all cached HTML pages for a practice.
    Reads every .html file in the folder (not just manifest entries).
    Returns list of (page_type, url, html_text).
    """
    import json

    url_map = {}
    manifest_path = os.path.join(folder, "manifest.json")
    if os.path.exists(manifest_path):
        try:
            m = json.load(open(manifest_path, encoding="utf-8"))
            for ptype, info in m.get("pages", {}).items():
                url_map[info.get("file", "")] = (ptype, info.get("url", ""))
        except Exception:
            pass

    pages = []
    seen_files = set()
    for fpath in sorted(glob.glob(os.path.join(folder, "*.html"))):
        fname = os.path.basename(fpath)
        if fname in seen_files:
            continue
        seen_files.add(fname)
        try:
            html = open(fpath, encoding="utf-8", errors="replace").read()
        except Exception:
            continue
        if len(html) < 200:
            continue
        ptype, url = url_map.get(fname, (fname.replace(".html", ""), ""))
        pages.append((ptype, url, html))

    return pages


# ── Live crawl ────────────────────────────────────────────────────────────────

def _priority(path: str) -> int:
    if _SVC_PATH_RE.search(path):
        return 1
    if _TEAM_PATH_RE.search(path):
        return 2
    if _TEST_PATH_RE.search(path):
        return 3
    return 9


def _collect_links(html: str, base_url: str, base_netloc: str,
                   visited: set, queue: list):
    """Parse all useful internal links from an HTML page into the priority queue."""
    from heapq import heappush
    root = base_netloc.lower().removeprefix("www.")
    soup = BeautifulSoup(html, "lxml")
    for a in soup.find_all("a", href=True):
        href = str(a["href"]).split("#")[0].rstrip("/")
        if not href or href.startswith("mailto") or href.startswith("tel"):
            continue
        try:
            full = urljoin(base_url, href)
            parsed = urlparse(full)
            if parsed.netloc.lower().removeprefix("www.") != root:
                continue
            p = _priority(parsed.path)
            if p < 9 and full not in visited:
                heappush(queue, (p, full))
        except Exception:
            pass


def _live_crawl(website_url: str, practice_name: str = "") -> list:
    """
    Crawl the practice website to extract keyword-bearing HTML.
    Tries Playwright first (renders JS — catches React/Vue/Wix/Squarespace sites).
    Falls back to requests for static sites if Playwright is unavailable.
    Returns list of ("live", url, html) triples.
    """
    if not website_url or website_url in ("None", "nan", "N/A", ""):
        return []
    if not website_url.startswith("http"):
        website_url = "https://" + website_url

    # Skip HMO portals / Kyruus SPAs — content is loaded via JS APIs
    if _SKIP_LIVE_RE.search(website_url):
        log.info("  skipping live crawl (known SPA/portal): %s", website_url)
        return []

    try:
        base_netloc = urlparse(website_url).netloc
    except Exception:
        return []

    # Run Playwright with a hard wall-clock timeout so a hanging page can't
    # freeze the entire workflow.
    pages = _live_crawl_playwright_timed(website_url, base_netloc)
    if not pages:
        pages = _live_crawl_requests(website_url, base_netloc)
    return pages


def _live_crawl_playwright_timed(website_url: str, base_netloc: str) -> list:
    """Thread-wrapper that kills _live_crawl_playwright after _LIVE_PW_TIMEOUT seconds."""
    result: list = []

    def _worker():
        try:
            result.extend(_live_crawl_playwright(website_url, base_netloc))
        except Exception:
            pass

    t = threading.Thread(target=_worker, daemon=True)
    t.start()
    t.join(timeout=_LIVE_PW_TIMEOUT)
    if t.is_alive():
        log.warning("  live crawl timed out after %ds — skipping %s",
                    _LIVE_PW_TIMEOUT, base_netloc)
    return result


def _live_crawl_playwright(website_url: str, base_netloc: str) -> list:
    """
    Playwright-based crawl — fully renders JS so React/Vue/Wix/Squarespace
    sites show their actual content (services lists, technology pages, etc.).
    Returns [] if Playwright is not installed.
    """
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        return []

    from heapq import heappush, heappop

    pages   = []
    visited = set()
    queue   = []
    heappush(queue, (0, website_url))

    try:
        with sync_playwright() as pw:
            browser = pw.chromium.launch(
                headless=True,
                args=[
                    "--no-sandbox", "--disable-setuid-sandbox",
                    "--disable-dev-shm-usage", "--disable-gpu",
                    "--disable-blink-features=AutomationControlled",
                ],
            )
            ctx = browser.new_context(
                locale="en-US",
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/120.0.0.0 Safari/537.36"
                ),
            )
            page = ctx.new_page()
            page.set_extra_http_headers({"Accept-Language": "en-US,en;q=0.9"})

            while queue and len(pages) < _LIVE_MAX_PAGES:
                pri, url = heappop(queue)
                url_norm = url.split("#")[0].rstrip("/") or url
                if url_norm in visited:
                    continue
                visited.add(url_norm)

                try:
                    time.sleep(random.uniform(_LIVE_DELAY_MIN, _LIVE_DELAY_MAX))
                    page.goto(url, timeout=20000,
                              wait_until="domcontentloaded")
                    # Extra wait for JS-rendered content to appear
                    try:
                        page.wait_for_load_state("networkidle", timeout=5000)
                    except Exception:
                        pass
                    # Scroll in 300 px steps so IntersectionObserver fires for
                    # every Wix/Squarespace/React lazy-loaded section
                    try:
                        page.evaluate("""
                            (async () => {
                                const delay = ms => new Promise(r => setTimeout(r, ms));
                                const step = 300;
                                let pos = 0;
                                let maxSteps = 30;
                                while (pos < document.body.scrollHeight + step && maxSteps-- > 0) {
                                    window.scrollTo(0, pos);
                                    await delay(180);
                                    pos += step;
                                }
                                await delay(1200);
                            })()
                        """)
                        time.sleep(2.5)
                    except Exception:
                        pass
                    html = page.content()
                    if len(html) < 300:
                        continue
                except Exception as exc:
                    log.debug("  PW fetch error %s: %s", url, exc)
                    continue

                pages.append(("live", url, html))
                log.debug("  PW page %d: %s", len(pages), url)
                _collect_links(html, url, base_netloc, visited, queue)

            ctx.close()
            browser.close()
    except Exception as exc:
        log.debug("  Playwright crawl error: %s", exc)

    log.info("  live crawl (Playwright): %d pages from %s", len(pages), base_netloc)
    return pages


def _live_crawl_requests(website_url: str, base_netloc: str) -> list:
    """
    Requests-based crawl — fallback when Playwright is unavailable.
    Works well for WordPress / static sites; misses JS-rendered content.
    """
    try:
        import requests as _req
    except ImportError:
        log.warning("requests not installed — skipping live crawl")
        return []

    from heapq import heappush, heappop

    pages   = []
    visited = set()
    queue   = []
    heappush(queue, (0, website_url))

    while queue and len(pages) < _LIVE_MAX_PAGES:
        pri, url = heappop(queue)
        url_norm = url.split("#")[0].rstrip("/") or url
        if url_norm in visited:
            continue
        visited.add(url_norm)

        try:
            time.sleep(random.uniform(_LIVE_DELAY_MIN, _LIVE_DELAY_MAX))
            r = _req.get(url, headers=_LIVE_HEADERS, timeout=18,
                         verify=False, allow_redirects=True)
            ct = r.headers.get("content-type", "")
            if r.status_code != 200 or "text/html" not in ct:
                continue
            html = r.text
            if len(html) < 300:
                continue
        except Exception as exc:
            log.debug("  requests fetch error %s: %s", url, exc)
            continue

        pages.append(("live", url, html))
        log.debug("  requests page %d: %s", len(pages), url)
        _collect_links(html, url, base_netloc, visited, queue)

    log.info("  live crawl (requests): %d pages from %s", len(pages), base_netloc)
    return pages


# ── Extraction ────────────────────────────────────────────────────────────────

def _extract(pages: list) -> dict:
    """
    Extract all target fields from a list of (ptype, url, html) pages.
    Works on both cache pages and live-crawled pages.
    """
    all_text = ""
    per_page = []    # (body_text, full_text_augmented)
    all_soups = []

    for _ptype, url, html in pages:
        ft  = ds.extract_text(html)
        bt  = ds.extract_body_text(html)
        aug = ds.extract_augmented_text(html, url)
        ft_aug = ft + " " + aug
        all_text += " " + ft_aug
        per_page.append((bt, ft_aug))
        all_soups.append(BeautifulSoup(html, "lxml"))

    if not all_text.strip():
        return {}

    # ── Technology ────────────────────────────────────────────────────────
    # Normalize hyphens → spaces so "cone-beam" matches keyword "cone beam" etc.
    all_text_n = all_text.replace("-", " ")
    tf = set()
    for kw, tn in ds.TECH_KEYWORDS.items():
        if kw in all_text_n:
            tf.add(tn)
    if "AI" not in tf and re.search(r"\bai\b", all_text, re.I):
        tf.add("AI")

    # ── Services (per-page, body primary / full+augmented fallback) ───────
    svc_b = dict.fromkeys(set(ds.SERVICE_KEYWORDS.values()), 0)
    svc_f = dict.fromkeys(set(ds.SERVICE_KEYWORDS.values()), 0)
    seen_urls: set = set()
    for i, (bt, ft_aug) in enumerate(per_page):
        url_key = pages[i][1] or str(i)
        if url_key in seen_urls:
            continue
        seen_urls.add(url_key)
        for kw, cat in ds.SERVICE_KEYWORDS.items():
            svc_b[cat] += ds.count_keyword_capped(bt, kw, cap=5)
            svc_f[cat] += ds.count_keyword_capped(ft_aug, kw, cap=3)
    svc = {cat: max(svc_b[cat], svc_f[cat]) for cat in svc_b}

    # ── Testimonials ──────────────────────────────────────────────────────
    seen_t, tt = set(), 0
    for soup in all_soups:
        # By class / id attribute
        for blk in soup.find_all(
            ["div", "section", "article", "blockquote", "li", "p"],
            class_=_TEST_RE,
        ):
            k = blk.get_text(separator=" ", strip=True)[:100]
            if k and k not in seen_t:
                seen_t.add(k); tt += 1
        # By data-* attributes
        for blk in soup.find_all(lambda t: any(
            _TEST_RE.search(str(v))
            for attr, v in t.attrs.items()
            if attr.startswith("data-") and isinstance(v, str)
        )):
            k = blk.get_text(separator=" ", strip=True)[:100]
            if k and k not in seen_t:
                seen_t.add(k); tt += 1
        # By id attribute
        for blk in soup.find_all(True, id=_TEST_RE):
            k = blk.get_text(separator=" ", strip=True)[:100]
            if k and k not in seen_t:
                seen_t.add(k); tt += 1
        # schema.org Review / itemprop
        for blk in soup.find_all(True, attrs={"itemprop": re.compile(r"review", re.I)}):
            k = blk.get_text(separator=" ", strip=True)[:100]
            if k and k not in seen_t:
                seen_t.add(k); tt += 1
    # Fallback: plain <blockquote> tags
    if tt == 0:
        for soup in all_soups:
            for bq in soup.find_all("blockquote"):
                k = bq.get_text(separator=" ", strip=True)[:100]
                if k and k not in seen_t:
                    seen_t.add(k); tt += 1

    # ── Hygienists ────────────────────────────────────────────────────────
    hyg = ""
    # 1. Pattern + name-counting search across all combined text
    h = ds.find_hygienists(all_text)
    if h:
        hyg = h
    # 2. Structured tag scan across ALL pages (no class-name gating)
    if not hyg or hyg == "0":
        _HYG_RE = re.compile(
            r'\bR\.?D\.?H\.?\b|RDHAP|BSDH|'
            r'registered\s+dental\s+hygienist|'
            r'licensed\s+dental\s+hygienist|'
            r'dental\s+hygienist',
            re.I,
        )
        _NM_RE = re.compile(r'([A-Z][a-z]+(?:\s+[A-Z]\.?)?\s+[A-Z][a-z]+)')
        rdh_names: set = set()
        rdh_keys: set = set()
        for soup in all_soups:
            for tag in soup.find_all(
                ["h1", "h2", "h3", "h4", "h5", "p", "span", "div", "li", "a", "strong", "b"]
            ):
                text = tag.get_text(separator=" ", strip=True)
                if len(text) > 200:
                    continue
                if not _HYG_RE.search(text):
                    continue
                nm = _NM_RE.search(text)
                if nm:
                    rdh_names.add(nm.group(1).strip().lower())
                else:
                    rdh_keys.add(re.sub(r"\s+", " ", text.strip().lower())[:60])
        total = len(rdh_names) + len(rdh_keys)
        if total:
            hyg = str(total)

    return {
        C_HYG:   hyg,
        C_CEREC: "X" if "CEREC"              in tf else "",
        C_CBCT:  "X" if "CBCT"               in tf else "",
        C_LASER: "X" if "Lasers"             in tf else "",
        C_AI:    "X" if "AI"                 in tf else "",
        C_INTRA: "X" if "Intraoral Scanners" in tf else "",
        C_INV:   svc.get("Invisalign",        0),
        C_CLEAR: svc.get("Clear Aligners",    0),
        C_VEN:   svc.get("Veneers",           0),
        C_IMPL:  svc.get("Implants",          0),
        C_SMILE: svc.get("Smile Makeovers",   0),
        C_WHITE: svc.get("Teeth Whitening",   0),
        C_SED:   svc.get("Sedation Dentistry",0),
        C_HOL:   svc.get("Holistic Dentistry",0),
        C_PLAN:  "Mentioned" if svc.get("Dental Plan", 0) > 0 else "",
        C_CANC:  svc.get("Cancer Screening",  0),
        C_TESTI: str(tt),
    }


# ── Merge logic — never reduce ────────────────────────────────────────────────

def _merge(col: int, old, new):
    """
    Combine two values. Never reduces — always keeps the better/larger value.
    """
    if col in TECH_COLS:
        return "X" if (str(old or "").strip() == "X" or str(new or "").strip() == "X") else ""

    if col == C_PLAN:
        return "Mentioned" if (
            str(old or "").strip() == "Mentioned" or
            str(new or "").strip() == "Mentioned"
        ) else ""

    if col == C_HYG:
        def _n(v):
            try:
                n = int(str(v or "").strip())
                return n if n >= 0 else -1
            except Exception:
                return -1
        on, nn = _n(old), _n(new)
        if on >= 0 and nn >= 0:
            return str(max(on, nn))
        if on >= 0:
            return str(on)
        if nn >= 0:
            return str(nn)
        # Neither is a number — keep any non-blank non-N/A value
        os_ = str(old or "").strip()
        ns_ = str(new or "").strip()
        for v in (os_, ns_):
            if v and v not in ("N/A", "None", "ERROR", "Not Found"):
                return v
        return os_ or ns_

    # Numeric fields (service counts, testimonials) — take max
    try:
        old_n = int(str(old or 0).replace(",", ""))
    except (ValueError, TypeError):
        old_n = 0
    try:
        new_n = int(str(new or 0).replace(",", ""))
    except (ValueError, TypeError):
        new_n = 0
    return max(old_n, new_n)


def _needs_live_crawl(extracted: dict) -> bool:
    """
    Return True if any important field is still blank/zero after cache extraction.
    We always supplement with live crawl when ANY tech field or Invisalign is missing
    — not just when everything is blank.
    """
    any_tech_blank = any(extracted.get(c, "") == "" for c in TECH_COLS)
    inv_zero       = extracted.get(C_INV, 0) in (0, "", "0")
    hyg_blank      = not extracted.get(C_HYG, "")
    testi_zero     = extracted.get(C_TESTI, "0") in ("0", "", 0)
    return any_tech_blank or inv_zero or hyg_blank or testi_zero


# ── Helpers ───────────────────────────────────────────────────────────────────

def _str(v) -> str:
    if v is None or str(v).strip() in ("", "None", "Not Found", "ERROR"):
        return ""
    return str(v).strip()


# ── Main ──────────────────────────────────────────────────────────────────────

def refresh(input_xlsx: str, cache_dir: str = "page_cache"):
    base, ext = os.path.splitext(input_xlsx)
    out_path  = base + "_refreshed" + ext
    comp_path = base + "_comparison.xlsx"

    shutil.copy2(input_xlsx, out_path)
    log.info("Input  : %s", input_xlsx)
    log.info("Output : %s", out_path)
    log.info("Cache  : %s", cache_dir)

    # ── Read xlsx ──────────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(out_path, data_only=True)
    ws = wb.active

    idx_rows    = {}   # idx -> [row_num, ...]
    idx_before  = {}   # idx -> {col: old_value}
    idx_name    = {}   # idx -> practice name
    idx_website = {}   # idx -> website url

    for row in ws.iter_rows(min_row=DATA_START, values_only=False):
        raw = row[C_IDX - 1].value
        if raw is None:
            continue
        try:
            idx = int(raw)
        except (TypeError, ValueError):
            continue
        rn = row[0].row
        idx_rows.setdefault(idx, []).append(rn)
        if idx not in idx_before:
            idx_before[idx]  = {col: ws.cell(rn, col).value for col, _, _ in FIELD_MAP}
            idx_name[idx]    = str(row[C_NAME    - 1].value or "")
            website_raw      = str(row[C_WEBSITE - 1].value or "").strip()
            idx_website[idx] = website_raw if website_raw not in ("None", "nan", "") else ""

    wb.close()
    log.info("Practices in xlsx: %d", len(idx_rows))

    # ── Extract per practice ───────────────────────────────────────────────
    merged_vals = {}
    n_cache = n_live = n_missing = 0

    for idx in sorted(idx_rows):
        name = idx_name.get(idx, "")
        url  = idx_website.get(idx, "")

        # ── 1. Cache ──────────────────────────────────────────────────────
        cache_extracted = {}
        folder = _find_cache_folder(idx, cache_dir)
        if folder:
            pages = _load_pages(folder)
            if pages:
                cache_extracted = _extract(pages)
                n_cache += 1

        # ── 2. Live crawl fallback ────────────────────────────────────────
        # Trigger when: no cache at all, OR cache gave all-blank results,
        # OR hygienist is still missing and a URL is available.
        live_extracted = {}
        need_live = url and (
            not cache_extracted or
            _needs_live_crawl(cache_extracted)
        )
        if need_live:
            log.info("  [%03d] %-28s  → live crawl (%s)",
                     idx, name[:28], "no cache" if not cache_extracted else "fields blank")
            live_pages = _live_crawl(url, name)
            if live_pages:
                live_extracted = _extract(live_pages)
                n_live += 1

        # ── 3. Merge cache + live ─────────────────────────────────────────
        if not cache_extracted and not live_extracted:
            n_missing += 1
            continue

        best: dict = {}
        for col, _, _ in FIELD_MAP:
            best[col] = _merge(col,
                               cache_extracted.get(col),
                               live_extracted.get(col))

        # ── 4. Merge best with existing xlsx values ───────────────────────
        old = idx_before[idx]
        merged: dict = {}
        for col, _, _ in FIELD_MAP:
            merged[col] = _merge(col, old.get(col), best.get(col))
        merged_vals[idx] = merged

        log.info("  [%03d] %-28s  CEREC=%-2s  Inv=%-3s  Impl=%-3s  Hyg=%-3s  Testi=%s",
                 idx, name[:28],
                 merged[C_CEREC] or "-",
                 merged[C_INV],
                 merged[C_IMPL],
                 merged.get(C_HYG, "-") or "-",
                 merged[C_TESTI])

    log.info("Cache: %d  Live: %d  No data: %d  /  %d total",
             n_cache, n_live, n_missing, len(idx_rows))

    if not merged_vals:
        log.warning("No data extracted — nothing to patch.")
        return out_path, None

    # ── Write merged values ────────────────────────────────────────────────
    wb2 = openpyxl.load_workbook(out_path, data_only=True)
    ws2 = wb2.active
    updates = 0

    for idx, vals in merged_vals.items():
        for rn in idx_rows.get(idx, []):
            for col, v in vals.items():
                ws2.cell(rn, col).value = v
                updates += 1

    wb2.save(out_path)
    log.info("Wrote %d cell updates → %s", updates, out_path)

    # ── Comparison report ──────────────────────────────────────────────────
    _write_comparison(comp_path, idx_rows, idx_name, idx_before, merged_vals)
    return out_path, comp_path


# ── Comparison report ─────────────────────────────────────────────────────────

def _write_comparison(path, idx_rows, idx_name, before, after):
    from openpyxl import Workbook

    wb  = Workbook()
    ws  = wb.active
    ws.title = "Tech & Services Comparison"

    YELLOW = PatternFill("solid", fgColor="FFFACD")
    GREEN  = PatternFill("solid", fgColor="C6EFCE")
    GREY   = PatternFill("solid", fgColor="E8E8E8")
    BLUE   = PatternFill("solid", fgColor="DDEEFF")
    HDR    = Font(bold=True, size=9)
    DATA   = Font(size=9)
    ctr    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft    = Alignment(horizontal="left",   vertical="center")

    labels = [lbl for _, _, lbl in FIELD_MAP]
    n      = len(FIELD_MAP)

    # Row 1: group headers
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    ws.cell(1, 1, "Practice").font = HDR; ws.cell(1, 1).alignment = ctr

    ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=3 + n)
    c = ws.cell(1, 4, "BEFORE"); c.font = HDR; c.fill = GREY; c.alignment = ctr

    ws.merge_cells(start_row=1, start_column=4 + n, end_row=1, end_column=3 + 2 * n)
    c = ws.cell(1, 4 + n, "AFTER (refreshed)"); c.font = HDR; c.fill = BLUE; c.alignment = ctr

    # Row 2: column headers
    for c_idx, h in enumerate(["Index", "Practice Name", "# Improved"] + labels + labels, 1):
        cell = ws.cell(2, c_idx, h)
        cell.font = HDR; cell.alignment = ctr
        if 4 <= c_idx <= 3 + n:
            cell.fill = GREY
        elif c_idx >= 4 + n:
            cell.fill = BLUE

    r = 3
    n_improved_total = 0

    for idx in sorted(idx_rows):
        if idx not in after:
            continue

        b = before.get(idx, {})
        a = after[idx]

        bv = [_str(b.get(col, "")) for col, _, _ in FIELD_MAP]
        av = [_str(a.get(col, "")) for col, _, _ in FIELD_MAP]
        improved = [av[i] != bv[i] and av[i] not in ("", "0") for i in range(n)]
        n_imp = sum(improved)

        if n_imp > 0:
            n_improved_total += 1

        row_data = [idx, idx_name.get(idx, ""), n_imp] + bv + av
        for c_idx, v in enumerate(row_data, 1):
            cell = ws.cell(r, c_idx, v)
            cell.font = DATA
            cell.alignment = lft if c_idx == 2 else ctr
            fi = c_idx - 4
            if 4 <= c_idx <= 3 + n and improved[fi]:
                cell.fill = YELLOW
            elif c_idx >= 4 + n and improved[c_idx - 4 - n]:
                cell.fill = GREEN

        r += 1

    # Column widths
    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 10
    for ci in range(4, 4 + 2 * n):
        ws.column_dimensions[get_column_letter(ci)].width = 11

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 32
    ws.freeze_panes = "D3"

    log.info("Comparison: %d practices, %d with improvements → %s",
             r - 3, n_improved_total, path)
    wb.save(path)


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    args      = sys.argv[1:]
    cache_dir = "page_cache"
    xlsx_file = None
    i = 0
    while i < len(args):
        if args[i] == "--cache-dir" and i + 1 < len(args):
            cache_dir = args[i + 1]; i += 2
        elif xlsx_file is None:
            xlsx_file = args[i]; i += 1
        else:
            i += 1

    if not xlsx_file:
        print("Usage: python3 refresh_tech_services.py <batch.xlsx> [--cache-dir dir]")
        sys.exit(1)
    if not os.path.exists(xlsx_file):
        log.error("File not found: %s", xlsx_file)
        sys.exit(1)

    refresh(xlsx_file, cache_dir)


if __name__ == "__main__":
    main()
